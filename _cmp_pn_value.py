# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from pathlib import Path

out = []
files = [
    ("good", Path(r"d:\CodeUp-工资哥All项目\转换\账单\now\Taiwan\PN_CoralSea_N-C 20260311.xlsx")),
    ("tpl", Path(r"d:\CodeUp-工资哥All项目\转换\templates\taiwan\template.xlsx")),
    ("conv", sorted(Path(r"d:\CodeUp-工资哥All项目\转换\输出").glob("PN_CoralSea_N-C_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)[0]),
]
for label, p in files:
    if not p.exists():
        out.append(f"missing {label} {p}")
        continue
    out.append(f"\n==== {label} {p.name} ====")
    wb = load_workbook(p, data_only=False)
    ws = wb["PN"]
    for row in range(14, 30):
        for col in range(1, 7):
            cell = ws.cell(row, col)
            if cell.value is not None:
                out.append(f"  {cell.coordinate}: {repr(cell.value)[:200]}")
    # TW key cells
    if "TW" in wb.sheetnames:
        tw = wb["TW"]
        out.append("  -- TW --")
        for coord in ("B2", "C6", "B9", "B10", "B11", "J9", "J10", "J11", "Z9", "Z10", "Z11", "AB9", "AB10", "AB11", "G6"):
            out.append(f"  TW!{coord}: {repr(tw[coord].value)[:120]}")
    wb.close()

    wb2 = load_workbook(p, data_only=True)
    if "PN" in wb2.sheetnames:
        ws = wb2["PN"]
        out.append("  -- PN data_only --")
        for row in range(14, 28):
            a, e, f = ws.cell(row, 1).value, ws.cell(row, 5).value, ws.cell(row, 6).value
            if a is not None or e is not None or f is not None:
                out.append(f"  R{row} A={a!r} E={e!r} F={f!r}")
    wb2.close()

Path(r"d:\CodeUp-工资哥All项目\转换\_pn_value_cmp.txt").write_text("\n".join(out), encoding="utf-8")
print("ok")
