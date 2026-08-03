# -*- coding: utf-8 -*-
"""映射配置：从样例源账单 / PN 母版解析表头（供 Office 下拉选项）。"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bill_convert.template_rows import list_formula_example_rows as _list_formula_example_rows
from convert_mapping import find_sheet_name, resolve_convert_mapping
from xlsx_convert_utils import norm


def _header_cells(ws, header_row: int) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    seen: set[str] = set()
    for col in range(1, (ws.max_column or 0) + 1):
        raw = ws.cell(header_row, col).value
        key = norm(raw)
        if not key or key in seen:
            continue
        seen.add(key)
        label = str(raw).replace("\r\n", "\n").strip() if raw is not None else key
        out.append({"key": key, "label": label})
    return out


def inspect_source_headers(
    *,
    source_path: Path,
    engine_id: str,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    engine_id = (engine_id or "").strip()
    mapping = resolve_convert_mapping(engine_id, convert_mapping)

    if engine_id == "tw_payroll_calc":
        return _inspect_tw_source(source_path, mapping)
    if engine_id in ("china_payroll_calc", "china_hrone"):
        return _inspect_fixed_header_source(source_path, mapping, default_sheet="计算结果", default_row=1)
    if engine_id in ("hk_payroll_calc", "hk_vertical_l"):
        return _inspect_fixed_header_source(source_path, mapping, default_sheet="Hong Kong-L", default_row=7)
    if engine_id == "uae_payroll_calc":
        return _inspect_uae_source(source_path, mapping)
    if engine_id == "uk_payroll_calc":
        return _inspect_uk_vertical_source(source_path, mapping)

    return {"ok": False, "message": f"引擎「{engine_id}」暂不支持表头识别"}


def _inspect_tw_source(source_path: Path, mapping: dict[str, Any]) -> dict[str, Any]:
    from profiles.tw_payroll_calc import convert as tw_mod

    tw_mod._ACTIVE_MAPPING = mapping
    try:
        wb = load_workbook(source_path, data_only=True, read_only=True)
        sheet_names = list(wb.sheetnames)
        src_spec = mapping.get("sourceEmployeeSheet") or {}
        pc_name = find_sheet_name(sheet_names, src_spec if isinstance(src_spec, dict) else None)
        if not pc_name:
            wb.close()
            want = (src_spec.get("sheet") if isinstance(src_spec, dict) else None) or "Payroll calculation"
            return {
                "ok": False,
                "message": f"未找到工作表「{want}」",
                "sheetNames": sheet_names,
            }
        ws = wb[pc_name]
        header_row = tw_mod.find_pc_header_row(ws, sheet_label=pc_name)
        headers = _header_cells(ws, header_row)
        employees: list[dict[str, str]] = []
        try:
            for emp in tw_mod.read_pc_employees(ws, header_row):
                employees.append(
                    {
                        "cnName": str(emp.get("CN Name") or "").strip(),
                        "enName": str(emp.get("EN Name") or "").strip(),
                    }
                )
        except Exception:
            employees = []
        wb.close()
        return {
            "ok": True,
            "sheetName": pc_name,
            "headerRow": header_row,
            "headers": headers,
            "employees": employees,
        }
    except Exception as exc:
        return {"ok": False, "message": str(exc)}
    finally:
        tw_mod._ACTIVE_MAPPING = None


def _inspect_fixed_header_source(
    source_path: Path,
    mapping: dict[str, Any],
    *,
    default_sheet: str,
    default_row: int,
) -> dict[str, Any]:
    src_spec = mapping.get("sourceEmployeeSheet") or {}
    if not isinstance(src_spec, dict):
        src_spec = {}
    header_row = int(src_spec.get("headerRow") or default_row)
    name_headers = src_spec.get("nameHeaders") if isinstance(src_spec.get("nameHeaders"), list) else []
    wb = load_workbook(source_path, data_only=True, read_only=True)
    sheet_names = list(wb.sheetnames)
    name = find_sheet_name(sheet_names, src_spec) or find_sheet_name(
        sheet_names, {"sheet": default_sheet, "candidates": [default_sheet]}
    )
    if not name:
        wb.close()
        return {
            "ok": False,
            "message": f"未找到工作表「{src_spec.get('sheet') or default_sheet}」",
            "sheetNames": sheet_names,
        }
    ws = wb[name]
    headers = _header_cells(ws, header_row)
    header_map: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key and key not in header_map:
            header_map[key] = col

    employees: list[dict[str, str]] = []
    data_start = int(src_spec.get("dataStartRow") or (header_row + 1))
    name_keys = [str(x).strip() for x in name_headers if str(x).strip()] or (
        ["姓名"] if "姓名" in header_map else (
            ["Name of Employee"] if "Name of Employee" in header_map else []
        )
    )
    if name_keys:
        primary = next((k for k in name_keys if k in header_map), None)
        secondary = next((k for k in name_keys if k != primary and k in header_map), None)
        if primary:
            pcol = header_map[primary]
            scol = header_map.get(secondary) if secondary else None
            for row in range(data_start, (ws.max_row or data_start) + 1):
                cn = ws.cell(row, pcol).value
                cn_s = str(cn).strip() if cn is not None else ""
                if not cn_s:
                    continue
                en_s = ""
                if scol:
                    en = ws.cell(row, scol).value
                    en_s = str(en).strip() if en is not None else ""
                employees.append({"cnName": cn_s, "enName": en_s})

    wb.close()
    return {
        "ok": True,
        "sheetName": name,
        "headerRow": header_row,
        "headers": headers,
        "employees": employees,
    }


def _inspect_uae_source(source_path: Path, mapping: dict[str, Any]) -> dict[str, Any]:
    """
    UAE 映射样例：
    1) 已是 UAE-L → 读 UAE-L 表头（引擎输入）
    2) 供应商 Payroll Draft → 读**原始**供应商表头（供「列名对照：供应商→UAE-L」下拉）
    """
    source_path = Path(source_path)
    try:
        wb = load_workbook(source_path, data_only=True, read_only=True)
        sheet_names = list(wb.sheetnames)
        wb.close()
    except Exception as exc:
        return {"ok": False, "message": f"无法打开源表: {exc}"}

    src_spec = mapping.get("sourceEmployeeSheet") if isinstance(mapping.get("sourceEmployeeSheet"), dict) else {}
    if find_sheet_name(sheet_names, src_spec) or find_sheet_name(
        sheet_names, {"sheet": "UAE-L", "candidates": ["UAE-L"]}
    ):
        return _inspect_fixed_header_source(
            source_path, mapping, default_sheet="UAE-L", default_row=2
        )

    # 供应商原始表：扫描表头行，供列名对照左侧选项
    try:
        from pdf_ingest.profiles import auxilium_uae as aux

        wb = load_workbook(source_path, data_only=True)
        try:
            ws, header_row, headers_raw = aux._pick_payroll_draft_sheet(wb)
            headers = _header_cells(ws, header_row)
            employees: list[dict[str, str]] = []
            name_keys = {"Employee Name", "EE Name", "Name", "Staff Name", "Full Name"}
            name_col = next(
                (i for i, h in enumerate(headers_raw, start=1) if norm(h) in name_keys or "name" in norm(h).lower()),
                None,
            )
            id_col = next(
                (
                    i
                    for i, h in enumerate(headers_raw, start=1)
                    if aux._is_id_header(h)
                ),
                None,
            )
            data_start = header_row + 1
            for row in range(data_start, min((ws.max_row or data_start), data_start + 40) + 1):
                name_s = ""
                id_s = ""
                if name_col:
                    v = ws.cell(row, name_col).value
                    name_s = str(v).strip() if v is not None else ""
                if id_col:
                    v = ws.cell(row, id_col).value
                    id_s = str(v).strip() if v is not None else ""
                if not name_s and not id_s:
                    continue
                if "TOTAL" in f"{name_s} {id_s}".upper():
                    continue
                employees.append({"cnName": "", "enName": name_s or id_s})
            return {
                "ok": True,
                "sheetName": ws.title,
                "headerRow": header_row,
                "headers": headers,
                "employees": employees,
                "sourceKind": "vendor_payroll_draft",
                "hint": "已识别供应商原始表头，请在「列名对照」中配置 供应商列 → UAE-L 列",
            }
        finally:
            wb.close()
    except Exception as exc:
        return {
            "ok": False,
            "message": (
                "未找到工作表「UAE-L」，且无法识别供应商 Payroll Draft 表头。"
                f"详情: {exc}"
            ),
            "sheetNames": sheet_names,
        }


def list_formula_example_rows(
    ws,
    data_start_row: int,
    *,
    marker_col: int = 2,
    max_slots: int = 9,
) -> list[dict[str, Any]]:
    return _list_formula_example_rows(
        ws, data_start_row, marker_col=marker_col, max_slots=max_slots
    )


def inspect_pn_headers(
    *,
    template_path: Path,
    engine_id: str,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    engine_id = (engine_id or "").strip()
    mapping = resolve_convert_mapping(engine_id, convert_mapping)
    target = mapping.get("targetL") if isinstance(mapping.get("targetL"), dict) else {}
    sheet_want = str(target.get("sheet") or "TW-L")
    candidates = target.get("candidates")
    spec: dict[str, Any] = {"sheet": sheet_want}
    if isinstance(candidates, list):
        spec["candidates"] = candidates
    else:
        spec["candidates"] = [sheet_want, "China-L", "Hong Kong-L", "UK-L", "UAE-L", "TW-L"]

    if not template_path.is_file():
        return {"ok": False, "message": f"母版文件不存在: {template_path}"}

    if engine_id == "tw_payroll_calc":
        from profiles.tw_payroll_calc import convert as tw_mod

        tw_mod._ACTIVE_MAPPING = mapping
        try:
            wb = load_workbook(template_path, data_only=True, read_only=True)
            sheet_names = list(wb.sheetnames)
            try:
                name = tw_mod.resolve_tw_l_sheet_name(sheet_names)
            except ValueError as exc:
                wb.close()
                return {"ok": False, "message": str(exc), "sheetNames": sheet_names}
            ws = wb[name]
            layout = tw_mod.resolve_tw_l_layout(ws, sheet_label=name)
            headers = _header_cells(ws, layout["header_row"])
            wb.close()
            wb_f = load_workbook(template_path, data_only=False, read_only=False)
            formula_rows = tw_mod.resolve_tw_formula_rows_layout(wb_f, layout["data_start_row"])
            tw_start = formula_rows["tw_data_start"]
            ee_start = formula_rows["tw_ee_data_start"]
            tw_examples: list[dict[str, Any]] = []
            ee_examples: list[dict[str, Any]] = []
            if tw_mod.TW_SHEET in wb_f.sheetnames:
                tw_examples = list_formula_example_rows(
                    wb_f[tw_mod.TW_SHEET],
                    tw_start,
                    marker_col=2,
                )
            if tw_mod.TW_EE_SHEET in wb_f.sheetnames:
                ee_examples = list_formula_example_rows(
                    wb_f[tw_mod.TW_EE_SHEET],
                    ee_start,
                    marker_col=5,
                )
            wb_f.close()
            ft = mapping.get("formulaTemplates") if isinstance(mapping.get("formulaTemplates"), dict) else {}
            tw_tpl = ft.get("TW") if isinstance(ft.get("TW"), dict) else {}
            ee_tpl = ft.get("TW EE") if isinstance(ft.get("TW EE"), dict) else {}
            return {
                "ok": True,
                "sheetName": name,
                "headerRow": layout["header_row"],
                "dataStartRow": layout["data_start_row"],
                "autoDetectedLayout": tw_mod._target_l_auto_detect_layout(),
                "headers": headers,
                "formulaExampleRows": {
                    "TW": tw_examples,
                    "TW EE": ee_examples,
                    "twDataStartRow": tw_start,
                    "twEeDataStartRow": ee_start,
                    "defaultTwRow": int(tw_tpl.get("defaultExampleRow") or tw_start),
                    "defaultTwEeRow": int(ee_tpl.get("defaultExampleRow") or ee_start),
                },
            }
        except Exception as exc:
            return {"ok": False, "message": str(exc)}
        finally:
            tw_mod._ACTIVE_MAPPING = None

    if engine_id in ("china_payroll_calc", "china_hrone"):
        from profiles.china_payroll_calc import convert as cn_mod

        cn_mod._ACTIVE_MAPPING = mapping
        try:
            wb = load_workbook(template_path, data_only=True, read_only=True)
            sheet_names = list(wb.sheetnames)
            name = find_sheet_name(sheet_names, spec)
            if not name:
                wb.close()
                return {
                    "ok": False,
                    "message": f"母版中未找到 sheet「{sheet_want}」",
                    "sheetNames": sheet_names,
                }
            header_row = int(target.get("headerRow") or 1)
            data_start = int(target.get("dataStartRow") or cn_mod.CHINA_L_DATA_START_ROW)
            headers = _header_cells(wb[name], header_row)
            wb.close()

            wb_f = load_workbook(template_path, data_only=False, read_only=False)
            china_start = cn_mod.CHINA_DATA_START_ROW
            ee_start = cn_mod.CHINA_EE_DATA_START_ROW
            china_examples: list[dict[str, Any]] = []
            ee_examples: list[dict[str, Any]] = []
            if cn_mod.CHINA_SHEET in wb_f.sheetnames:
                china_examples = list_formula_example_rows(
                    wb_f[cn_mod.CHINA_SHEET],
                    china_start,
                    marker_col=3,
                )
            if cn_mod.CHINA_EE_SHEET in wb_f.sheetnames:
                ee_examples = list_formula_example_rows(
                    wb_f[cn_mod.CHINA_EE_SHEET],
                    ee_start,
                    marker_col=4,
                )
            wb_f.close()
            ft = mapping.get("formulaTemplates") if isinstance(mapping.get("formulaTemplates"), dict) else {}
            china_tpl = ft.get("China") if isinstance(ft.get("China"), dict) else {}
            ee_tpl = ft.get("China EE") if isinstance(ft.get("China EE"), dict) else {}
            return {
                "ok": True,
                "sheetName": name,
                "headerRow": header_row,
                "dataStartRow": data_start,
                "autoDetectedLayout": False,
                "headers": headers,
                "formulaExampleRows": {
                    "China": china_examples,
                    "China EE": ee_examples,
                    "chinaDataStartRow": china_start,
                    "chinaEeDataStartRow": ee_start,
                    "defaultChinaRow": int(china_tpl.get("defaultExampleRow") or china_start),
                    "defaultChinaEeRow": int(ee_tpl.get("defaultExampleRow") or ee_start),
                },
            }
        except Exception as exc:
            return {"ok": False, "message": str(exc)}
        finally:
            cn_mod._ACTIVE_MAPPING = None

    if engine_id in ("hk_payroll_calc", "hk_vertical_l"):
        from profiles.hk_payroll_calc import convert as hk_mod

        hk_mod._ACTIVE_MAPPING = mapping
        try:
            wb = load_workbook(template_path, data_only=True, read_only=True)
            sheet_names = list(wb.sheetnames)
            name = find_sheet_name(sheet_names, spec)
            if not name:
                wb.close()
                return {
                    "ok": False,
                    "message": f"母版中未找到 sheet「{sheet_want}」",
                    "sheetNames": sheet_names,
                }
            header_row = int(target.get("headerRow") or hk_mod.HK_L_HEADER_ROW)
            data_start = int(target.get("dataStartRow") or hk_mod.HK_L_DATA_START_ROW)
            headers = _header_cells(wb[name], header_row)
            wb.close()

            wb_f = load_workbook(template_path, data_only=False, read_only=False)
            hk_start = hk_mod.HK_DATA_START_ROW
            ee_start = hk_mod.HK_EE_DATA_START_ROW
            hk_examples: list[dict[str, Any]] = []
            ee_examples: list[dict[str, Any]] = []
            if hk_mod.HK_SHEET in wb_f.sheetnames:
                hk_examples = list_formula_example_rows(
                    wb_f[hk_mod.HK_SHEET],
                    hk_start,
                    marker_col=2,
                )
            if hk_mod.HK_EE_SHEET in wb_f.sheetnames:
                ee_examples = list_formula_example_rows(
                    wb_f[hk_mod.HK_EE_SHEET],
                    ee_start,
                    marker_col=4,
                )
            wb_f.close()
            ft = mapping.get("formulaTemplates") if isinstance(mapping.get("formulaTemplates"), dict) else {}
            hk_tpl = ft.get("Hong Kong") if isinstance(ft.get("Hong Kong"), dict) else {}
            ee_tpl = ft.get("Hong Kong EE") if isinstance(ft.get("Hong Kong EE"), dict) else {}
            return {
                "ok": True,
                "sheetName": name,
                "headerRow": header_row,
                "dataStartRow": data_start,
                "autoDetectedLayout": False,
                "headers": headers,
                "formulaExampleRows": {
                    "Hong Kong": hk_examples,
                    "Hong Kong EE": ee_examples,
                    "hkDataStartRow": hk_start,
                    "hkEeDataStartRow": ee_start,
                    "defaultHkRow": int(hk_tpl.get("defaultExampleRow") or hk_start),
                    "defaultHkEeRow": int(ee_tpl.get("defaultExampleRow") or ee_start),
                },
            }
        except Exception as exc:
            return {"ok": False, "message": str(exc)}
        finally:
            hk_mod._ACTIVE_MAPPING = None

    if engine_id == "uae_payroll_calc":
        return _inspect_uae_pn(template_path, mapping, spec, sheet_want)

    if engine_id == "uk_payroll_calc":
        return _inspect_uk_pn(template_path, mapping, spec, sheet_want)

    header_row = int(target.get("headerRow") or 7)
    wb = load_workbook(template_path, data_only=True, read_only=True)
    sheet_names = list(wb.sheetnames)
    name = find_sheet_name(sheet_names, spec)
    if not name:
        wb.close()
        return {
            "ok": False,
            "message": f"母版中未找到 sheet「{sheet_want}」",
            "sheetNames": sheet_names,
        }
    ws = wb[name]
    headers = _header_cells(ws, header_row)
    wb.close()
    return {
        "ok": True,
        "sheetName": name,
        "headerRow": header_row,
        "headers": headers,
    }


def _inspect_uk_vertical_source(source_path: Path, mapping: dict[str, Any]) -> dict[str, Any]:
    """UK-L 竖表：A 列标签视为「表头」，便于列名对照。"""
    src_spec = mapping.get("sourceEmployeeSheet") if isinstance(mapping.get("sourceEmployeeSheet"), dict) else {}
    try:
        wb = load_workbook(source_path, data_only=True, read_only=True)
    except Exception as exc:
        return {"ok": False, "message": f"无法打开源表: {exc}"}
    try:
        sheet_names = list(wb.sheetnames)
        name = find_sheet_name(sheet_names, src_spec if isinstance(src_spec, dict) else None)
        if not name:
            # 供应商原始表也可能不是 UK-L：取首张
            name = sheet_names[0] if sheet_names else None
        if not name:
            return {"ok": False, "message": "工作簿无工作表", "sheetNames": sheet_names}
        ws = wb[name]
        label_col = int(src_spec.get("labelColumn") or 1)
        headers: list[dict[str, str]] = []
        seen: set[str] = set()
        for row in range(1, min((ws.max_row or 1), 80) + 1):
            raw = ws.cell(row, label_col).value
            key = norm(raw)
            if not key or key in seen:
                continue
            # 跳过明显非金额标签的标题行
            if key.lower() in {"details", "amount in gbp"}:
                continue
            seen.add(key)
            headers.append({"key": key, "label": key})
        employees: list[dict[str, str]] = []
        # 标题行常含员工名
        title = norm(ws.cell(3, 1).value)
        if title and "salary calculation" in title.lower():
            en = title.split("-")[0].strip() if "-" in title else title
            if en:
                employees.append({"cnName": "", "enName": en})
        return {
            "ok": True,
            "sheetName": name,
            "headerRow": 0,
            "layout": "vertical_label_amount",
            "headers": headers,
            "sampleEmployees": employees,
        }
    except Exception as exc:
        return {"ok": False, "message": str(exc)}
    finally:
        wb.close()


def _inspect_uae_pn(
    template_path: Path,
    mapping: dict[str, Any],
    spec: dict[str, Any],
    sheet_want: str,
) -> dict[str, Any]:
    from profiles.uae_payroll_calc import convert as uae_mod

    target = mapping.get("targetL") if isinstance(mapping.get("targetL"), dict) else {}
    try:
        wb = load_workbook(template_path, data_only=True, read_only=True)
        sheet_names = list(wb.sheetnames)
        name = find_sheet_name(sheet_names, spec)
        if not name:
            wb.close()
            return {
                "ok": False,
                "message": f"母版中未找到 sheet「{sheet_want}」",
                "sheetNames": sheet_names,
            }
        header_row = int(target.get("headerRow") or uae_mod.UAE_L_HEADER_ROW)
        data_start = int(target.get("dataStartRow") or uae_mod.UAE_L_DATA_START)
        headers = _header_cells(wb[name], header_row)
        wb.close()

        wb_f = load_workbook(template_path, data_only=False, read_only=False)
        uae_start = uae_mod.UAE_DATA_START
        ee_start = uae_mod.UAE_EE_DATA_START
        uae_examples: list[dict[str, Any]] = []
        ee_examples: list[dict[str, Any]] = []
        if uae_mod.UAE_SHEET in wb_f.sheetnames:
            uae_examples = list_formula_example_rows(
                wb_f[uae_mod.UAE_SHEET],
                uae_start,
                marker_col=2,
            )
        if uae_mod.UAE_EE_SHEET in wb_f.sheetnames:
            ee_examples = list_formula_example_rows(
                wb_f[uae_mod.UAE_EE_SHEET],
                ee_start,
                marker_col=5,
            )
        wb_f.close()
        ft = mapping.get("formulaTemplates") if isinstance(mapping.get("formulaTemplates"), dict) else {}
        uae_tpl = ft.get("UAE") if isinstance(ft.get("UAE"), dict) else {}
        ee_tpl = ft.get("UAE EE") if isinstance(ft.get("UAE EE"), dict) else {}
        return {
            "ok": True,
            "sheetName": name,
            "headerRow": header_row,
            "dataStartRow": data_start,
            "autoDetectedLayout": False,
            "headers": headers,
            "formulaExampleRows": {
                "UAE": uae_examples,
                "UAE EE": ee_examples,
                "uaeDataStartRow": uae_start,
                "uaeEeDataStartRow": ee_start,
                "defaultUaeRow": int(uae_tpl.get("defaultExampleRow") or uae_start),
                "defaultUaeEeRow": int(ee_tpl.get("defaultExampleRow") or ee_start),
            },
        }
    except Exception as exc:
        return {"ok": False, "message": str(exc)}


def _inspect_uk_pn(
    template_path: Path,
    mapping: dict[str, Any],
    spec: dict[str, Any],
    sheet_want: str,
) -> dict[str, Any]:
    """UK-L 竖表母版：返回 A 列标签 + UK/UK EE 示例行。"""
    target = mapping.get("targetL") if isinstance(mapping.get("targetL"), dict) else {}
    try:
        wb = load_workbook(template_path, data_only=False, read_only=False)
        sheet_names = list(wb.sheetnames)
        name = find_sheet_name(sheet_names, spec)
        if not name:
            wb.close()
            return {
                "ok": False,
                "message": f"母版中未找到 sheet「{sheet_want}」",
                "sheetNames": sheet_names,
            }
        ws = wb[name]
        label_col = int(target.get("labelColumn") or 1)
        headers: list[dict[str, str]] = []
        seen: set[str] = set()
        for row in range(1, min((ws.max_row or 1), 80) + 1):
            raw = ws.cell(row, label_col).value
            key = norm(raw)
            if not key or key in seen:
                continue
            if key.lower() in {"details", "amount in gbp"}:
                continue
            seen.add(key)
            headers.append({"key": key, "label": key, "row": str(row)})

        uk_start = 9
        ee_start = 9
        uk_examples: list[dict[str, Any]] = []
        ee_examples: list[dict[str, Any]] = []
        if "UK" in wb.sheetnames:
            uk_examples = list_formula_example_rows(wb["UK"], uk_start, marker_col=2)
        if "UK EE" in wb.sheetnames:
            ee_examples = list_formula_example_rows(wb["UK EE"], ee_start, marker_col=5)
        wb.close()
        ft = mapping.get("formulaTemplates") if isinstance(mapping.get("formulaTemplates"), dict) else {}
        uk_tpl = ft.get("UK") if isinstance(ft.get("UK"), dict) else {}
        ee_tpl = ft.get("UK EE") if isinstance(ft.get("UK EE"), dict) else {}
        return {
            "ok": True,
            "sheetName": name,
            "headerRow": 0,
            "dataStartRow": 1,
            "layout": "vertical_label_amount",
            "autoDetectedLayout": False,
            "headers": headers,
            "formulaExampleRows": {
                "UK": uk_examples,
                "UK EE": ee_examples,
                "ukDataStartRow": uk_start,
                "ukEeDataStartRow": ee_start,
                "defaultUkRow": int(uk_tpl.get("defaultExampleRow") or uk_start),
                "defaultUkEeRow": int(ee_tpl.get("defaultExampleRow") or ee_start),
            },
        }
    except Exception as exc:
        return {"ok": False, "message": str(exc)}
