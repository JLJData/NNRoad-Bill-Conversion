# -*- coding: utf-8 -*-
"""PN 页外部提供的账单元数据（B8/B9/B10/F9/F10/F11）。"""
from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

PN_SHEET = "PN"

# 与各地区 convert 里 _DATE_FMT 一致，写入 PN 日期元数据时强制带上
PN_DATE_NUMBER_FORMAT = "yyyy/m/d"

# PN 页固定单元格（各地区母版一致）
PN_CELLS = {
    "customer_name": "B8",
    "customer_id": "B9",
    "billing_address": "B10",
    "invoice_number": "F9",
    "invoice_date": "F10",
    "due_date": "F11",
}


@dataclass
class PnMeta:
    customer_name: str
    customer_id: str
    billing_address: str
    invoice_date: date | None = None
    due_date: date | None = None
    invoice_number: str | None = None  # 为空则按规则自动生成

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        for key in ("invoice_date", "due_date"):
            val = d[key]
            if isinstance(val, date):
                d[key] = val.isoformat()
        return d

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> PnMeta:
        return cls(
            customer_name=str(data["customer_name"]).strip(),
            customer_id=str(data["customer_id"]).strip(),
            billing_address=str(data["billing_address"]).strip(),
            invoice_date=parse_date(data.get("invoice_date")),
            due_date=parse_date(data.get("due_date")),
            invoice_number=(str(data["invoice_number"]).strip() or None)
            if data.get("invoice_number")
            else None,
        )


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"无法解析日期: {value!r}，请使用 YYYY-MM-DD")


def format_date_for_excel(value: date) -> datetime:
    """openpyxl 写入 date 时用 datetime 避免显示异常。"""
    return datetime(value.year, value.month, value.day)


def generate_invoice_number(customer_id: str, bill_date: date, sequence: int) -> str:
    """
    规则: PN-{客户ID}-{MMDDYYYY}{当天第几单}
    例: PN-CUS15253-031820261
    """
    cid = customer_id.strip()
    if not cid:
        raise ValueError("客户 ID 不能为空，无法生成账单编号")
    if sequence < 1:
        raise ValueError("账单序号须 ≥ 1")
    return f"PN-{cid}-{bill_date.strftime('%m%d%Y')}{sequence}"


def _registry_path(base_dir: Path) -> Path:
    return base_dir / ".pn_invoice_seq.json"


def _load_registry(path: Path) -> dict[str, dict[str, int]]:
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return data
    except (json.JSONDecodeError, OSError):
        pass
    return {}


def _save_registry(path: Path, data: dict[str, dict[str, int]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def peek_invoice_sequence(customer_id: str, bill_date: date | None = None, *, registry_dir: Path) -> int:
    """预览下一个序号（不递增）。"""
    bill_date = bill_date or date.today()
    reg = _load_registry(_registry_path(registry_dir))
    day_key = bill_date.isoformat()
    return int(reg.get(customer_id.strip(), {}).get(day_key, 0)) + 1


def allocate_invoice_sequence(customer_id: str, bill_date: date | None = None, *, registry_dir: Path) -> int:
    """分配并持久化当天序号（转换成功时调用）。"""
    bill_date = bill_date or date.today()
    cid = customer_id.strip()
    path = _registry_path(registry_dir)
    reg = _load_registry(path)
    day_key = bill_date.isoformat()
    per_customer = reg.setdefault(cid, {})
    seq = int(per_customer.get(day_key, 0)) + 1
    per_customer[day_key] = seq
    _save_registry(path, reg)
    return seq


def build_invoice_number(
    customer_id: str,
    bill_date: date | None = None,
    *,
    registry_dir: Path,
    reserve: bool = False,
) -> str:
    bill_date = bill_date or date.today()
    if reserve:
        seq = allocate_invoice_sequence(customer_id, bill_date, registry_dir=registry_dir)
    else:
        seq = peek_invoice_sequence(customer_id, bill_date, registry_dir=registry_dir)
    return generate_invoice_number(customer_id, bill_date, seq)


def _write_pn_date_cell(ws, coord: str, value: date) -> None:
    cell = ws[coord]
    cell.value = format_date_for_excel(value)
    cell.number_format = PN_DATE_NUMBER_FORMAT


def apply_pn_meta(
    wb: Workbook,
    meta: PnMeta | dict[str, Any],
    *,
    registry_dir: Path,
    reserve_invoice_number: bool = True,
) -> PnMeta:
    """将外部 PN 元数据写入 workbook 的 PN sheet。"""
    if PN_SHEET not in wb.sheetnames:
        raise ValueError(f"母版缺少 sheet「{PN_SHEET}」")
    pn = meta if isinstance(meta, PnMeta) else PnMeta.from_dict(meta)

    if not pn.customer_name:
        raise ValueError("客户名称 (PN!B8) 不能为空")
    if not pn.customer_id:
        raise ValueError("客户 ID (PN!B9) 不能为空")
    if not pn.billing_address:
        raise ValueError("账单地址 (PN!B10) 不能为空")
    if pn.due_date is None:
        raise ValueError("Due date (PN!F11) 不能为空")

    invoice_date = pn.invoice_date or date.today()
    invoice_number = pn.invoice_number
    if not invoice_number:
        invoice_number = build_invoice_number(
            pn.customer_id,
            invoice_date,
            registry_dir=registry_dir,
            reserve=reserve_invoice_number,
        )

    ws = wb[PN_SHEET]
    ws[PN_CELLS["customer_name"]] = pn.customer_name
    ws[PN_CELLS["customer_id"]] = pn.customer_id
    ws[PN_CELLS["billing_address"]] = pn.billing_address
    ws[PN_CELLS["invoice_number"]] = invoice_number
    _write_pn_date_cell(ws, PN_CELLS["invoice_date"], invoice_date)
    _write_pn_date_cell(ws, PN_CELLS["due_date"], pn.due_date)

    return PnMeta(
        customer_name=pn.customer_name,
        customer_id=pn.customer_id,
        billing_address=pn.billing_address,
        invoice_date=invoice_date,
        due_date=pn.due_date,
        invoice_number=invoice_number,
    )
