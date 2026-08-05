# -*- coding: utf-8 -*-
"""
公式兼容修复（与地区无关）：写入 xlsx 前规范化，供 HyperFormula / LuckySheet / Excel 共用。

- PN 母版误写 `="- "&+"Expense...` → 合法拼接
- A1 用 =MID(CELL("filename",A1),...) 取表名 → 静态表名（避免自引用）
- `=+'Sheet'!A1` 一元加号、EOMONTH(…)/TODAY() → HyperFormula 可算写法
"""
from __future__ import annotations

import calendar
import re
from datetime import date, timedelta

from openpyxl.worksheet.worksheet import Worksheet

_AMP_PLUS_RE = re.compile(r"&\s*\+")
_DASH_AMP_QUOTE_RE = re.compile(r'="-\s*"&"([^"]*)"')
_SHEET_TITLE_SELF_REF_RE = re.compile(
    r'CELL\s*\(\s*"filename"\s*,\s*\$?A\$?1\s*\)',
    re.IGNORECASE,
)
_UNARY_PLUS_RE = re.compile(r"^=\s*\+")
_EOMONTH_TODAY_RE = re.compile(
    r"EOMONTH\s*\(\s*TODAY\s*\(\s*\)\s*,\s*(-?\d+)\s*\)(?:\s*\+\s*(\d+))?",
    re.IGNORECASE,
)
_EOMONTH_HEAD_RE = re.compile(r"EOMONTH\s*\(", re.IGNORECASE)
_TODAY_RE = re.compile(r"TODAY\s*\(\s*\)", re.IGNORECASE)
# Excel 数组写法 SUMPRODUCT(($A$1:$B$1="Y")*($A2:$B2)) → SUMIF；HF/LuckySheet 不算数组乘
_SUMPRODUCT_YN_ARRAY_RE = re.compile(
    r"SUMPRODUCT\s*\(\s*\(\s*(\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)\s*=\s*\"Y\"\s*\)\s*"
    r"\*\s*\(\s*(\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)\s*\)\s*\)",
    re.IGNORECASE,
)


def fix_pn_illegal_concat_formulas(ws: Worksheet) -> int:
    """母版误写 `="- "&+"Expense...`：Excel 能算，LuckySheet/部分引擎会 #VALUE!。"""
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=") and _AMP_PLUS_RE.search(v)):
                continue
            nv = _AMP_PLUS_RE.sub("&", v)
            nv = _DASH_AMP_QUOTE_RE.sub(r'="- \1"', nv)
            if nv != v:
                cell.value = nv
                n += 1
    return n


def flatten_sheet_title_self_refs(wb) -> int:
    """
    母版 A1 用 =MID(CELL(\"filename\",A1),...) 取表名。
    Excel 允许这种自引用；LuckySheet 重算时会弹
    「公式不可引用其本身的单元格，会导致计算结果不准确」。
    转换时写成静态表名即可（缓存值本就是表名）。
    """
    n = 0
    for name in wb.sheetnames:
        cell = wb[name].cell(1, 1)
        v = cell.value
        if isinstance(v, str) and v.startswith("=") and _SHEET_TITLE_SELF_REF_RE.search(v):
            cell.value = name
            n += 1
    return n


def _eomonth(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    last = calendar.monthrange(y, m)[1]
    return date(y, m, last)


def _replace_eomonth_today(match: re.Match) -> str:
    months = int(match.group(1))
    plus = int(match.group(2) or 0)
    end = _eomonth(date.today(), months)
    if plus:
        end = end + timedelta(days=plus)
    return f"DATE({end.year},{end.month},{end.day})"


def _rewrite_eomonth_to_date(formula: str) -> str:
    """EOMONTH(date, n) → DATE(YEAR(date), MONTH(date)+n+1, 0)；HF 无 EOMONTH。"""
    if not _EOMONTH_HEAD_RE.search(formula):
        return formula
    f = formula
    for _ in range(20):
        m = _EOMONTH_HEAD_RE.search(f)
        if not m:
            break
        start = m.start()
        open_idx = m.end() - 1
        depth = 0
        end = -1
        in_str = False
        for i in range(open_idx, len(f)):
            ch = f[i]
            if ch == '"':
                in_str = not in_str
                continue
            if in_str:
                continue
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    end = i
                    break
        if end < 0:
            break
        inner = f[open_idx + 1 : end]
        comma = -1
        d2 = 0
        s2 = False
        for i, ch in enumerate(inner):
            if ch == '"':
                s2 = not s2
                continue
            if s2:
                continue
            if ch == "(":
                d2 += 1
            elif ch == ")":
                d2 -= 1
            elif ch == "," and d2 == 0:
                comma = i
                break
        if comma < 0:
            break
        date_expr = inner[:comma].strip()
        months_expr = inner[comma + 1 :].strip()
        if not date_expr or not months_expr:
            break
        replacement = f"DATE(YEAR({date_expr}),MONTH({date_expr})+({months_expr})+1,0)"
        f = f[:start] + replacement + f[end + 1 :]
    return f


def normalize_formula_for_lucky(formula: str) -> str:
    """单条公式改写，供转换写盘前 / 扫描修复复用。"""
    if not (isinstance(formula, str) and formula.startswith("=")):
        return formula
    f = formula
    if _UNARY_PLUS_RE.match(f):
        f = _UNARY_PLUS_RE.sub("=", f, count=1)
    if _AMP_PLUS_RE.search(f):
        f = _AMP_PLUS_RE.sub("&", f)
        f = _DASH_AMP_QUOTE_RE.sub(r'="- \1"', f)
    if re.search(r"EOMONTH", f, re.I) and re.search(r"TODAY\s*\(", f, re.I):
        f = _EOMONTH_TODAY_RE.sub(_replace_eomonth_today, f)
    if _EOMONTH_HEAD_RE.search(f):
        f = _rewrite_eomonth_to_date(f)
    if _TODAY_RE.search(f) and not _EOMONTH_HEAD_RE.search(f):
        today = date.today()
        f = _TODAY_RE.sub(f"DATE({today.year},{today.month},{today.day})", f)
    return f


def fix_workbook_lucky_formulas(wb) -> int:
    """整本扫描：一元加号 / &+ / EOMONTH(TODAY()) / TODAY()。"""
    n = 0
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                nv = normalize_formula_for_lucky(v)
                if nv != v:
                    cell.value = nv
                    n += 1
    return n


def fix_sumproduct_yn_array_formulas(ws: Worksheet) -> int:
    """
    Italy-L Fee 等：SUMPRODUCT(($K$8:$W$8=\"Y\")*($K11:$W11))
    Excel 数组乘法；HyperFormula/LuckySheet 会「计算失败」。
    语义等价改写为 SUMIF(criteria_range,\"Y\",sum_range)。
    """
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=") and "SUMPRODUCT" in v.upper()):
                continue
            nv = _SUMPRODUCT_YN_ARRAY_RE.sub(r'SUMIF(\1,"Y",\2)', v)
            if nv != v:
                cell.value = nv
                n += 1
    return n


def apply_luckysheet_compat(wb, *, pn_sheet: str | None = "PN") -> dict[str, int]:
    """转换写盘前统一兼容修复。"""
    stats = {
        "amp_plus": 0,
        "sheet_title": flatten_sheet_title_self_refs(wb),
        "formula_normalize": 0,
        "table_refs": 0,
        "ref_errors": 0,
        "sumproduct_yn": 0,
    }
    if pn_sheet and pn_sheet in wb.sheetnames:
        stats["amp_plus"] = fix_pn_illegal_concat_formulas(wb[pn_sheet])
    # Italy-L / 同类 Fee 公式
    for sheet_name in wb.sheetnames:
        if sheet_name.endswith("-L") or sheet_name in ("Italy-L",):
            stats["sumproduct_yn"] += fix_sumproduct_yn_array_formulas(wb[sheet_name])
    # 再扫一遍全表（含 China!BH 的 =+、China!B2 EOMONTH、PN 残留 &+）
    stats["formula_normalize"] = fix_workbook_lucky_formulas(wb)
    return stats


_TABLE_THIS_ROW_RE = re.compile(
    r"Table(\d+)\[\[#This Row\],\[([^\]]+)\]\]",
    re.IGNORECASE,
)
_TABLE_HEADERS_RE = re.compile(
    r"Table(\d+)\[\[#Headers\],\[([^\]]+)\]\]",
    re.IGNORECASE,
)
_TABLE_COL_RE = re.compile(
    r"Table(\d+)\[([^\]]+)\]",
    re.IGNORECASE,
)


def _col_letter(col: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(col)


def _formula_text(value) -> str | None:
    from openpyxl.worksheet.formula import ArrayFormula

    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _table_header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        h = ws.cell(header_row, col).value
        if h is None:
            continue
        key = str(h).replace("\n", " ").strip()
        if key and key not in out:
            out[key] = col
    return out


def flatten_excel_table_formulas(
    wb,
    *,
    table_sheet: str = "UAE-L",
    table_name: str = "Table1",
    data_start_row: int = 3,
    header_row: int = 2,
    region_sheet: str = "UAE",
    region_data_start: int = 9,
    ee_sheet: str = "UAE EE",
    ee_data_start: int = 10,
) -> dict[str, int]:
    """
    将 Excel 表结构化引用 / ArrayFormula 展平为普通 A1 公式。

    LuckySheet / HyperFormula 不支持 Table1[...] 与动态数组，预览会一直转圈。
    """
    from openpyxl.worksheet.formula import ArrayFormula

    stats = {"table_refs": 0, "array_formulas": 0, "ref_errors": 0}
    if table_sheet not in wb.sheetnames:
        return stats

    l_ws = wb[table_sheet]
    headers = _table_header_map(l_ws, header_row)
    if not headers:
        return stats

    # Table1 → table_sheet 实际表名（母版固定 Table1）
    def resolve_l_row(sheet_name: str, excel_row: int) -> int:
        if sheet_name == region_sheet:
            # 账期等表头行（B2/C2）指向首名员工 L 行，勿用 2-9 算出负行号
            if excel_row < region_data_start:
                return data_start_row
            return data_start_row + (excel_row - region_data_start)
        if sheet_name == ee_sheet:
            if excel_row < ee_data_start:
                return data_start_row
            return data_start_row + (excel_row - ee_data_start)
        if sheet_name == table_sheet:
            return excel_row
        return data_start_row

    def cell_ref(col_name: str, l_row: int, *, from_l_sheet: bool) -> str | None:
        col = headers.get(col_name)
        if not col:
            # 容错：大小写/空白
            for k, c in headers.items():
                if k.lower() == col_name.lower():
                    col = c
                    break
        if not col:
            return None
        addr = f"{_col_letter(col)}{l_row}"
        if from_l_sheet:
            return addr
        return f"'{table_sheet}'!{addr}"

    def rewrite(formula: str, sheet_name: str, excel_row: int) -> str:
        from_l = sheet_name == table_sheet
        l_row = resolve_l_row(sheet_name, excel_row)

        def repl_this(_m: re.Match) -> str:
            # #This Row：在表内用当前行；在 UAE/EE 上用对应员工 L 行
            row_for_this = excel_row if from_l else l_row
            ref = cell_ref(_m.group(2), row_for_this, from_l_sheet=from_l)
            return ref or _m.group(0)

        def repl_headers(_m: re.Match) -> str:
            ref = cell_ref(_m.group(2), header_row, from_l_sheet=from_l)
            return ref or _m.group(0)

        def repl_col(_m: re.Match) -> str:
            ref = cell_ref(_m.group(2), l_row, from_l_sheet=from_l)
            return ref or _m.group(0)

        f = _TABLE_THIS_ROW_RE.sub(repl_this, formula)
        f = _TABLE_HEADERS_RE.sub(repl_headers, f)
        f = _TABLE_COL_RE.sub(repl_col, f)
        return f
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = min(ws.max_row or 1, 200)
        max_col = min(ws.max_column or 1, 120)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                if type(cell).__name__ == "MergedCell":
                    continue
                raw = cell.value
                text = _formula_text(raw)
                if not text:
                    continue
                is_af = isinstance(raw, ArrayFormula)
                new_text = rewrite(text, sheet_name, row)
                changed = new_text != text or is_af
                if "#REF!" in new_text:
                    cell.value = 0
                    stats["ref_errors"] += 1
                    continue
                if changed:
                    cell.value = new_text
                    if is_af:
                        stats["array_formulas"] += 1
                    if "Table" not in new_text and "Table" in text:
                        stats["table_refs"] += 1
                    elif new_text != text:
                        stats["table_refs"] += 1
    return stats


def apply_luckysheet_compat_uae(wb) -> dict[str, int]:
    """UAE：先展平 Table/ArrayFormula，再走通用兼容。"""
    table_stats = flatten_excel_table_formulas(wb)
    base = apply_luckysheet_compat(wb, pn_sheet="PN")
    base.update(table_stats)
    return base
