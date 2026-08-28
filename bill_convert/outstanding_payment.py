# -*- coding: utf-8 -*-
"""按标签写入 Outstanding payment（各地区 sheet 共用）。

规则：在工作簿中查找文案「Outstanding payment」，写入其右侧金额区
「靠左」那一格（两列金额时取左列）。balance 由 Office 注入 convert_mapping。
写入 Excel 时取相反数（Timeline -1 → 格子 1；1 → -1；0 仍为 0）。

注意：
- PN 汇总页标签常合并单元格，且金额多为公式引用地区 sheet，优先写地区 sheet
- 标签右侧可能有空列，不能简单 label_col+1
"""
from __future__ import annotations

import re
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

LABEL = "outstanding payment"
# 汇总页 / 明细页：优先跳过，避免写到合并格或公式格
_SKIP_SHEET_NAMES = {"pn"}
_SKIP_SHEET_SUFFIXES = (" ee", "-l", "_l", " detail")


def apply_outstanding_payment(wb: Workbook, mapping: dict[str, Any] | None) -> dict[str, Any] | None:
    """
    若 mapping 含 outstandingPayment.balance（含 0），则写入标签右侧左金额格。
    格子写入 Timeline 金额的相反数；找不到标签或无 balance 时不改表。
    返回诊断信息（写入位置等），供日志/响应头使用。
    """
    if wb is None or not isinstance(mapping, dict):
        return None
    timeline_amount = _resolve_balance(mapping)
    if timeline_amount is None:
        return None
    amount = _negate_amount(timeline_amount)

    candidates = _find_label_cells(wb)
    if not candidates:
        return {
            "written": False,
            "reason": "label_not_found",
            "balance": amount,
        }

    last_skip: dict[str, Any] | None = None
    for ws, row, label_col in candidates:
        value_col = _resolve_amount_col(ws, row, label_col)
        cell = ws.cell(row, value_col)
        if isinstance(cell, MergedCell):
            # 合并格只能写左上角
            top_left = _merge_top_left(ws, row, value_col)
            if top_left is None:
                last_skip = {
                    "written": False,
                    "reason": "merged_readonly",
                    "sheet": ws.title,
                    "row": row,
                    "col": value_col,
                    "balance": amount,
                }
                continue
            cell = top_left
            value_col = cell.column

        if isinstance(cell.value, str) and cell.value.startswith("="):
            last_skip = {
                "written": False,
                "reason": "formula_preserved",
                "sheet": ws.title,
                "row": row,
                "col": value_col,
                "balance": amount,
            }
            continue

        cell.value = amount
        return {
            "written": True,
            "sheet": ws.title,
            "row": row,
            "col": value_col,
            "balance": amount,
            "timelineBalance": timeline_amount,
            "negated": True,
            "foundPeriod": _nested_str(mapping, "outstandingPayment", "foundPeriod"),
            "requestMonth": _nested_str(mapping, "outstandingPayment", "requestMonth"),
        }

    return last_skip or {
        "written": False,
        "reason": "no_writable_cell",
        "balance": amount,
    }


def _resolve_balance(mapping: dict[str, Any]) -> float | int | None:
    block = mapping.get("outstandingPayment")
    raw = None
    if isinstance(block, dict):
        raw = block.get("balance")
    if raw is None:
        raw = mapping.get("outstandingPaymentBalance")
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return raw
    text = str(raw).strip().replace(",", "").replace("\xa0", "")
    if not text:
        return None
    try:
        num = float(text)
    except ValueError:
        return None
    if num.is_integer():
        return int(num)
    return num


def _negate_amount(amount: float | int) -> float | int:
    """Excel 落格用相反数；0 保持 0，避免出现 -0.0。"""
    if amount == 0:
        return 0 if isinstance(amount, int) else 0.0
    negated = -amount
    if isinstance(amount, int) and not isinstance(amount, bool):
        return int(negated)
    if isinstance(negated, float) and negated.is_integer():
        return int(negated)
    return negated


def _find_label_cells(wb: Workbook) -> list[tuple[Worksheet, int, int]]:
    """地区 sheet 优先，PN 等汇总页靠后。"""
    preferred: list[tuple[Worksheet, int, int]] = []
    fallback: list[tuple[Worksheet, int, int]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.sheet_state and str(ws.sheet_state).lower() == "hidden":
            continue
        hits = _scan_sheet_for_label(ws)
        if not hits:
            continue
        if _is_fallback_sheet(name):
            fallback.extend(hits)
        else:
            preferred.extend(hits)
    return preferred + fallback


def _is_fallback_sheet(name: str) -> bool:
    n = (name or "").strip().lower()
    if n in _SKIP_SHEET_NAMES:
        return True
    return any(n.endswith(suf) for suf in _SKIP_SHEET_SUFFIXES)


def _scan_sheet_for_label(ws: Worksheet) -> list[tuple[Worksheet, int, int]]:
    hits: list[tuple[Worksheet, int, int]] = []
    max_row = min(ws.max_row or 0, 250)
    max_col = min(ws.max_column or 0, 60)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if _norm_label(ws.cell(row, col).value) == LABEL:
                hits.append((ws, row, col))
    return hits


def _resolve_amount_col(ws: Worksheet, row: int, label_col: int) -> int:
    """标签（含合并）右侧金额区：取两列金额中靠左的那一列。"""
    start = label_col + 1
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= label_col <= m.max_col:
            start = max(start, m.max_col + 1)
            break

    scored: list[tuple[int, int]] = []
    for col in range(start, start + 10):
        score = _amount_col_score(ws, row, col)
        if score > 0:
            scored.append((col, score))
    if scored:
        # 分数最高的若干列里取最左（通常即左金额列）
        max_score = max(s for _, s in scored)
        top = [c for c, s in scored if s >= max(1, max_score - 1)]
        return min(top)

    # 回退：跳过空 spacer，落到第一个可写格
    for col in range(start, start + 10):
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            continue
        return col
    return start


def _amount_col_score(ws: Worksheet, row: int, col: int) -> int:
    score = 0
    for r in range(max(1, row - 6), row + 7):
        cell = ws.cell(r, col)
        if isinstance(cell, MergedCell):
            continue
        v = cell.value
        if isinstance(v, (int, float)):
            score += 2
        elif isinstance(v, str) and v.startswith("="):
            score += 2
    return score


def _merge_top_left(ws: Worksheet, row: int, col: int):
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            return ws.cell(m.min_row, m.min_col)
    return None


def _norm_label(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def _nested_str(mapping: dict[str, Any], *keys: str) -> str | None:
    cur: Any = mapping
    for k in keys:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(k)
    if cur is None or cur == "":
        return None
    return str(cur).strip()
