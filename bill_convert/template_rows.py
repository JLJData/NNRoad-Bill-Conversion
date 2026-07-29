# -*- coding: utf-8 -*-
from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet


def clear_row_values(ws: Worksheet, row: int) -> None:
    for col in range(1, (ws.max_column or 0) + 1):
        ws.cell(row, col).value = None


def count_data_slots(ws: Worksheet, data_start_row: int, *, marker_col: int = 2, max_slots: int = 9) -> int:
    n = 0
    for row in range(data_start_row, data_start_row + max_slots):
        if ws.cell(row, marker_col).value is not None:
            n += 1
        else:
            break
    return max(n, 1)


def list_formula_example_rows(
    ws: Worksheet,
    data_start_row: int,
    *,
    marker_col: int = 2,
    max_slots: int = 9,
) -> list[dict[str, Any]]:
    """列出可作为公式参照的示例行。

    不再在「中间空行」处提前结束：母版常在第 9 行默认公式、第 10 行第二种公式，
    中间若短暂无 marker 也应继续扫描。
    """
    out: list[dict[str, Any]] = []
    last_hit = -1
    for offset in range(max_slots):
        row = data_start_row + offset
        marker = ws.cell(row, marker_col).value
        has_formula = False
        for col in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(row, col)
            if cell.data_type == "f" and cell.value:
                has_formula = True
                break
        if marker is None and not has_formula:
            if offset == 0:
                # 默认起始行即使暂无公式也保留选项
                out.append({"row": row, "label": f"第 {row} 行（默认）"})
                last_hit = offset
            continue
        label = str(marker).strip() if marker is not None else ""
        if not label:
            label = f"第 {row} 行"
        else:
            label = f"第 {row} 行 · {label[:48]}"
        out.append({"row": row, "label": label})
        last_hit = offset
    if not out:
        out.append({"row": data_start_row, "label": f"第 {data_start_row} 行（默认）"})
    # 去掉尾部连续空占位：若只因 offset==0 塞了默认且后面全无，保留默认即可
    _ = last_hit
    return out
