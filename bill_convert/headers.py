# -*- coding: utf-8 -*-
"""表头：单行 / 父子两级（重复子列名消歧）。"""
from __future__ import annotations

from collections import Counter
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from xlsx_convert_utils import norm


def _display_text(raw: Any, fallback: str) -> str:
    if raw is None:
        return fallback
    return str(raw).replace("\r\n", "\n").strip() or fallback


def _parents_by_col(ws: Worksheet, parent_row: int, max_col: int) -> dict[int, str]:
    """
    父表头按列取值：
    - 本格有值 → 用本格
    - 否则若落在合并区域内 → 用合并区左上角的值
    - 不再做「向右盲目前向填充」（会把无父级的 Total 等误挂到左侧分组）
    """
    out: dict[int, str] = {c: "" for c in range(1, max_col + 1)}
    for col in range(1, max_col + 1):
        v = norm(ws.cell(parent_row, col).value)
        if v:
            out[col] = v

    try:
        ranges = list(ws.merged_cells.ranges)
    except Exception:
        ranges = []

    for mr in ranges:
        try:
            if parent_row < mr.min_row or parent_row > mr.max_row:
                continue
            val = norm(ws.cell(mr.min_row, mr.min_col).value)
            if not val:
                continue
            for col in range(mr.min_col, min(int(mr.max_col), max_col) + 1):
                out[col] = val
        except Exception:
            continue
    return out


def list_qualified_header_cells(
    ws: Worksheet,
    header_row: int,
    *,
    parent_row: int | None = None,
) -> list[dict[str, Any]]:
    """
    解析表头行，生成唯一 key（供列名对照 / 源表读取）。

    规则：
    - 有真实父级（本格或合并区）且不同于子列名 → key =「父/子」，下拉展示「父 › 子」
    - 无父级 → key = 子列名
    - 仍重复 → 追加 #2、#3…
    """
    max_col = int(ws.max_column or 0)
    if max_col <= 0:
        return []

    if parent_row is None and header_row > 1:
        parent_row = header_row - 1
    parent_by_col: dict[int, str] = {}
    if parent_row is not None and parent_row >= 1:
        parent_by_col = _parents_by_col(ws, parent_row, max_col)

    rows: list[dict[str, Any]] = []
    for col in range(1, max_col + 1):
        raw = ws.cell(header_row, col).value
        child = norm(raw)
        parent = parent_by_col.get(col, "")
        if not child:
            # 子格为空时用上一行分组名（India-L 的 IIT / Business Tax / Deduction）
            if not parent:
                continue
            child = parent.replace("\n", " ").replace("\r", " ").strip()
            parent = ""
            if parent_row is not None and parent_row >= 1:
                raw = ws.cell(parent_row, col).value
        rows.append(
            {
                "col": col,
                "child": child,
                "parent": parent,
                "childLabel": _display_text(raw, child),
            }
        )

    prelim: list[dict[str, Any]] = []
    for r in rows:
        child = r["child"]
        parent = r["parent"]
        child_label = r["childLabel"]
        # 有真实父级就拼接（不再要求子名重复）；父级识别已改为合并区，不会盲填
        if parent and parent != child:
            base_key = f"{parent}/{child}"
            label = f"{parent} › {child_label}"
        else:
            base_key = child
            label = child_label
            parent = ""
        prelim.append(
            {
                "col": r["col"],
                "child": child,
                "parent": parent,
                "base_key": base_key,
                "label": label,
            }
        )

    base_counts = Counter(p["base_key"] for p in prelim)
    seen: dict[str, int] = {}
    out: list[dict[str, Any]] = []
    for p in prelim:
        base = p["base_key"]
        if base_counts[base] > 1:
            n = seen.get(base, 0) + 1
            seen[base] = n
            key = base if n == 1 else f"{base}#{n}"
            label = p["label"] if n == 1 else f"{p['label']} #{n}"
        else:
            key = base
            label = p["label"]
        out.append(
            {
                "key": key,
                "label": label,
                "child": p["child"],
                "parent": p["parent"] or "",
                "col": p["col"],
            }
        )
    return out


def build_qualified_header_map(
    ws: Worksheet,
    header_row: int,
    *,
    parent_row: int | None = None,
) -> dict[str, int]:
    """资格化表头 → 列号（每 key 唯一）。"""
    return {
        str(h["key"]): int(h["col"])
        for h in list_qualified_header_cells(ws, header_row, parent_row=parent_row)
        if h.get("key") and h.get("col")
    }


def build_qualified_header_cols(
    ws: Worksheet,
    header_row: int,
    *,
    parent_row: int | None = None,
) -> dict[str, list[int]]:
    """资格化表头 → 列号列表。"""
    mapping: dict[str, list[int]] = {}
    for h in list_qualified_header_cells(ws, header_row, parent_row=parent_row):
        key = str(h.get("key") or "")
        col = h.get("col")
        if key and col:
            mapping.setdefault(key, []).append(int(col))
    return mapping


def build_header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    """单行表头（目标母版等）；同名只保留最左列。"""
    mapping: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def build_header_cols(ws: Worksheet, header_row: int) -> dict[str, list[int]]:
    """单行表头 → 同名多列列表（TW-L 等）。"""
    mapping: dict[str, list[int]] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key:
            mapping.setdefault(key, []).append(col)
    return mapping


def resolve_target_col(cols: list[int], *, dup_min_col: int = 14) -> int | None:
    if not cols:
        return None
    if len(cols) == 1:
        return cols[0]
    payroll = [c for c in cols if c >= dup_min_col]
    return max(payroll) if payroll else max(cols)


def header_lookup_keys(header_key: str) -> list[str]:
    """
    用配置里的列名反查资格化 key 时的候选：
    完整 key → 去 #n → 子段（/ 右侧）。
    """
    k = norm(header_key)
    if not k:
        return []
    out: list[str] = [k]
    base = k.split("#", 1)[0]
    if base and base not in out:
        out.append(base)
    if "/" in base:
        child = base.rsplit("/", 1)[-1]
        if child and child not in out:
            out.append(child)
    return out


def resolve_header_cols(
    header_cols: dict[str, list[int]],
    header_key: str,
) -> list[int]:
    """按资格化 key / 子名 解析列号列表。"""
    for cand in header_lookup_keys(header_key):
        cols = header_cols.get(cand)
        if cols:
            return cols
    # 再尝试：header_cols 的 key 以 /child 或 child 结尾
    want = norm(header_key)
    if not want:
        return []
    child = want.split("#", 1)[0].rsplit("/", 1)[-1]
    hits: list[int] = []
    for k, cols in header_cols.items():
        kk = norm(k)
        base = kk.split("#", 1)[0]
        if base == want or base == child or base.endswith("/" + child):
            hits.extend(cols)
    return hits
