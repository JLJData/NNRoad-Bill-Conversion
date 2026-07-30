# -*- coding: utf-8 -*-
"""UK PN 多人 UK-L 布局：UK-L / UK-L (2)… 与 UK/UK EE/PN 行扩。"""
from __future__ import annotations

import re
from copy import copy
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from bill_convert.formula_copy import shift_row_formula

UK_L_SHEET = "UK-L"
UK_SHEET = "UK"
UK_EE_SHEET = "UK EE"
PN_SHEET = "PN"


def uk_l_sheet_name(index: int) -> str:
    if index <= 0:
        return UK_L_SHEET
    return f"{UK_L_SHEET} ({index + 1})"


def list_uk_l_sheets(sheetnames: list[str]) -> list[str]:
    """按 UK-L、UK-L (2)、UK-L (3)… 顺序返回。"""
    found: dict[int, str] = {}
    for name in sheetnames:
        if name == UK_L_SHEET:
            found[0] = name
            continue
        m = re.match(rf"^{re.escape(UK_L_SHEET)}\s*\((\d+)\)$", name, re.I)
        if m:
            found[int(m.group(1)) - 1] = name
    return [found[i] for i in sorted(found)]


def _copy_cell(src, dest) -> None:
    dest.value = src.value
    if src.has_style:
        dest.font = copy(src.font)
        dest.border = copy(src.border)
        dest.fill = copy(src.fill)
        dest.number_format = src.number_format
        dest.protection = copy(src.protection)
        dest.alignment = copy(src.alignment)


def _clone_sheet_as(wb, source_name: str, new_name: str):
    if new_name in wb.sheetnames:
        return wb[new_name]
    if source_name not in wb.sheetnames:
        raise ValueError(f"缺少 sheet「{source_name}」")
    src = wb[source_name]
    dest = wb.copy_worksheet(src)
    dest.title = new_name
    return dest


def ensure_uk_l_count(wb, count: int) -> list[str]:
    if count < 1:
        raise ValueError("至少需要 1 名员工")
    if UK_L_SHEET not in wb.sheetnames:
        raise ValueError(f"母版缺少 sheet「{UK_L_SHEET}」")
    names: list[str] = []
    for i in range(count):
        name = uk_l_sheet_name(i)
        if name not in wb.sheetnames:
            _clone_sheet_as(wb, UK_L_SHEET, name)
        names.append(name)
    return names


def _retarget_uk_l_formula(formula: str, from_sheet: str, to_sheet: str) -> str:
    if not (isinstance(formula, str) and formula.startswith("=")):
        return formula
    pat = re.compile(
        rf"(?:'{re.escape(from_sheet)}'|{re.escape(from_sheet)})!",
        re.I,
    )
    return pat.sub(f"'{to_sheet}'!", formula)


def _copy_row_cells(ws: Worksheet, src_row: int, dest_row: int, max_col: int = 30) -> None:
    for c in range(1, max_col + 1):
        dest = ws.cell(dest_row, c)
        # 合并格只读，跳过（避免 MergedCell value is read-only）
        if type(dest).__name__ == "MergedCell":
            continue
        src = ws.cell(src_row, c)
        if type(src).__name__ == "MergedCell":
            continue
        _copy_cell(src, dest)
        if isinstance(dest.value, str) and dest.value.startswith("="):
            # 只平移本表相对行（如 I9→I10、UK!B9→UK!B10）；
            # 'UK-L'!B9 等跨表竖表引用保持单元格不变，稍后只改 sheet 名。
            dest.value = shift_row_formula(
                dest.value,
                src_row,
                dest_row,
                target_l_from=-1,
                target_l_to=-1,
                target_l_sheet=UK_L_SHEET,
            )


def expand_uk_employee_rows(wb, employee_count: int, sheet_names: list[str]) -> None:
    if employee_count < 1 or UK_SHEET not in wb.sheetnames:
        return
    uk = wb[UK_SHEET]
    example_row = 9
    for i in range(employee_count):
        row = example_row + i
        if i > 0:
            _copy_row_cells(uk, example_row, row)
        l_name = sheet_names[i] if i < len(sheet_names) else uk_l_sheet_name(i)
        for c in range(1, (uk.max_column or 20) + 1):
            cell = uk.cell(row, c)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = _retarget_uk_l_formula(cell.value, UK_L_SHEET, l_name)

    if UK_EE_SHEET in wb.sheetnames:
        ee = wb[UK_EE_SHEET]
        ee_example = 9
        for i in range(employee_count):
            row = ee_example + i
            if i > 0:
                _copy_row_cells(ee, ee_example, row)
            l_name = sheet_names[i] if i < len(sheet_names) else uk_l_sheet_name(i)
            for c in range(1, (ee.max_column or 20) + 1):
                cell = ee.cell(row, c)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = _retarget_uk_l_formula(cell.value, UK_L_SHEET, l_name)

    if PN_SHEET in wb.sheetnames and employee_count > 1:
        pn = wb[PN_SHEET]
        # 避免写进合并区：仅改未合并的公式/文本格
        if employee_count >= 2:
            try:
                _copy_row_cells(pn, 16, 17)
                for c in range(1, 8):
                    cell = pn.cell(17, c)
                    if type(cell).__name__ == "MergedCell":
                        continue
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cell.value = (
                            cell.value.replace("UK!B9", "UK!B10")
                            .replace("UK!J9", "UK!J10")
                            .replace("UK!N9", "UK!N10")
                            .replace("UK!Q9", "UK!Q10")
                            .replace("UK!S9", "UK!S10")
                        )
                    elif isinstance(cell.value, str) and "UK!B9" in cell.value:
                        cell.value = cell.value.replace("UK!B9", "UK!B10")
                _copy_row_cells(pn, 18, 19)
                for c in range(1, 8):
                    cell = pn.cell(19, c)
                    if type(cell).__name__ == "MergedCell":
                        continue
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cell.value = cell.value.replace("UK!B9", "UK!B10").replace(
                            "UK!S9", "UK!S10"
                        )
                    elif isinstance(cell.value, str) and "UK!B9" in cell.value:
                        cell.value = cell.value.replace("UK!B9", "UK!B10")
            except Exception:
                # PN 行扩失败不阻断主流程（UK-L / UK 已就位）
                pass
