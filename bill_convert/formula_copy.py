# -*- coding: utf-8 -*-
from __future__ import annotations

import re
from copy import copy
from typing import Any, Iterable

from openpyxl.worksheet.worksheet import Worksheet


def shift_row_formula(
    formula: str,
    from_row: int,
    to_row: int,
    *,
    target_l_from: int,
    target_l_to: int,
    target_l_sheet: str = "TW-L",
) -> str:
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    placeholders: list[str] = []

    def stash_external(match: re.Match[str]) -> str:
        placeholders.append(match.group(0))
        return f"__EXT{len(placeholders) - 1}__"

    s = re.sub(r"'[^']+'!\$?[A-Z]{1,3}\$?\d+", stash_external, formula)
    s = re.sub(
        rf"(?<!\$)(?<![A-Z])([A-Z]{{1,3}}){from_row}(?!\d)",
        lambda m: f"{m.group(1)}{to_row}",
        s,
    )
    # 只改「示例员工」对应的 L 行；账期等元数据行（如 C2/E2）必须保留
    pat = re.compile(rf"'{re.escape(target_l_sheet)}'!([A-Z]{{1,3}})(\d+)")
    for idx, ref in enumerate(placeholders):
        def _retarget_l(m: re.Match[str], _from: int = target_l_from, _to: int = target_l_to) -> str:
            if int(m.group(2)) == _from:
                return f"'{target_l_sheet}'!{m.group(1)}{_to}"
            return m.group(0)

        ref = pat.sub(_retarget_l, ref)
        s = s.replace(f"__EXT{idx}__", ref)
    return s


def snapshot_row_cells(ws: Worksheet, row: int) -> list[dict[str, Any]]:
    """拷贝一行公式/值与样式，避免后续覆盖源行后丢失示例。"""
    out: list[dict[str, Any]] = []
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        src = ws.cell(row, col)
        item: dict[str, Any] = {
            "col": col,
            "value": src.value,
            "data_type": src.data_type,
        }
        if src.has_style:
            item["font"] = copy(src.font)
            item["border"] = copy(src.border)
            item["fill"] = copy(src.fill)
            item["number_format"] = src.number_format
            item["protection"] = copy(src.protection)
            item["alignment"] = copy(src.alignment)
        out.append(item)
    return out


def copy_row_formulas_from_snapshot(
    snapshot: list[dict[str, Any]],
    ws: Worksheet,
    from_row: int,
    to_row: int,
    target_l_from: int,
    target_l_to: int,
    *,
    target_l_sheet: str = "TW-L",
) -> None:
    for item in snapshot:
        col = int(item["col"])
        dst = ws.cell(to_row, col)
        if "font" in item:
            dst.font = item["font"]
            dst.border = item["border"]
            dst.fill = item["fill"]
            dst.number_format = item["number_format"]
            dst.protection = item["protection"]
            dst.alignment = item["alignment"]
        value = item.get("value")
        data_type = item.get("data_type")
        if data_type == "f" and isinstance(value, str):
            dst.value = shift_row_formula(
                value,
                from_row,
                to_row,
                target_l_from=target_l_from,
                target_l_to=target_l_to,
                target_l_sheet=target_l_sheet,
            )
        elif value is not None and data_type != "f":
            dst.value = copy(value)


def copy_row_formulas(
    ws: Worksheet,
    from_row: int,
    to_row: int,
    target_l_from: int,
    target_l_to: int,
    *,
    target_l_sheet: str = "TW-L",
) -> None:
    copy_row_formulas_from_snapshot(
        snapshot_row_cells(ws, from_row),
        ws,
        from_row,
        to_row,
        target_l_from,
        target_l_to,
        target_l_sheet=target_l_sheet,
    )


def fix_sheet_cross_refs(
    ws: Worksheet,
    dst_row: int,
    *,
    other_sheet: str,
    other_row: int,
    quoted: bool,
    data_row_min: int | None = None,
    data_row_max: int | None = None,
) -> None:
    """修正同行公式里对其它 sheet 的「相对」行引用为配对行。

    - 尊重 $ 绝对行（如 $B$2 账期），不改
    - 可选只改落在员工数据行区间内的引用，避免误伤表头/元数据行
    """
    prefix = f"'{other_sheet}'!" if quoted else f"{other_sheet}!"
    # 组1=列前是否有 $；组2=列；组3=行前是否有 $；组4=行号
    pat = re.compile(rf"{re.escape(prefix)}(\$?)([A-Z]+)(\$?)(\d+)(?!\d)")
    repl_prefix = prefix
    lo = int(data_row_min) if data_row_min is not None else None
    hi = int(data_row_max) if data_row_max is not None else None

    def _repl(m: re.Match[str]) -> str:
        col_abs, col, row_abs, row_s = m.group(1), m.group(2), m.group(3), m.group(4)
        if row_abs == "$":
            return m.group(0)
        row_n = int(row_s)
        if lo is not None and row_n < lo:
            return m.group(0)
        if hi is not None and row_n > hi:
            return m.group(0)
        return f"{repl_prefix}{col_abs}{col}{other_row}"

    for col in range(1, (ws.max_column or 0) + 1):
        cell = ws.cell(dst_row, col)
        if cell.data_type == "f" and isinstance(cell.value, str):
            cell.value = pat.sub(_repl, cell.value)


def retarget_pn_fx_b_column_refs(
    wb,
    fx_row: int,
    *,
    from_rows: Iterable[int] | None = None,
    pn_sheet: str = "PN",
) -> int:
    """把指向「旧汇率行」的 PN!B / 本表 $B$ 改到新 fx_row。

    只改 from_rows（默认常见母版汇率行），勿动 Client Code/Name（PN!B8/B9 等）。
    返回改写单元格数。
    """
    fx = max(int(fx_row), 1)
    candidates = [int(r) for r in (from_rows if from_rows is not None else (28, 29, 30, 31, 32, 33))]
    olds = sorted({r for r in candidates if r > 0 and r != fx})
    if not olds:
        return 0
    old_alt = "|".join(str(r) for r in olds)
    pat_pn = re.compile(rf"PN!\$?B\$?(?:{old_alt})(?!\d)", re.IGNORECASE)
    pat_local = re.compile(rf"(?<!!)\$B\$(?:{old_alt})(?!\d)")
    changed = 0
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                nv = v
                if "PN!" in v.upper() and pat_pn.search(v):
                    nv = pat_pn.sub(f"PN!$B${fx}", nv)
                elif name == pn_sheet and pat_local.search(nv):
                    nv = pat_local.sub(f"$B${fx}", nv)
                if nv != v:
                    cell.value = nv
                    changed += 1
    return changed


def fix_tw_row_tw_ee_refs(ws_tw: Worksheet, dst_row: int, ee_row: int) -> None:
    fix_sheet_cross_refs(ws_tw, dst_row, other_sheet="TW EE", other_row=ee_row, quoted=True)


def fix_ee_row_tw_refs(ws_ee: Worksheet, dst_row: int, tw_row: int) -> None:
    fix_sheet_cross_refs(ws_ee, dst_row, other_sheet="TW", other_row=tw_row, quoted=False)


def fix_china_row_china_ee_refs(ws_china: Worksheet, dst_row: int, ee_row: int) -> None:
    fix_sheet_cross_refs(ws_china, dst_row, other_sheet="China EE", other_row=ee_row, quoted=True)


def fix_ee_row_china_refs(ws_ee: Worksheet, dst_row: int, china_row: int) -> None:
    fix_sheet_cross_refs(ws_ee, dst_row, other_sheet="China", other_row=china_row, quoted=False)


def fix_hk_row_hk_ee_refs(ws_hk: Worksheet, dst_row: int, ee_row: int) -> None:
    fix_sheet_cross_refs(ws_hk, dst_row, other_sheet="Hong Kong EE", other_row=ee_row, quoted=True)


def fix_ee_row_hk_refs(ws_ee: Worksheet, dst_row: int, hk_row: int) -> None:
    fix_sheet_cross_refs(ws_ee, dst_row, other_sheet="Hong Kong", other_row=hk_row, quoted=False)


def fix_pakistan_row_pakistan_ee_refs(ws_pk: Worksheet, dst_row: int, ee_row: int) -> None:
    fix_sheet_cross_refs(ws_pk, dst_row, other_sheet="Pakistan EE", other_row=ee_row, quoted=True)


def fix_ee_row_pakistan_refs(ws_ee: Worksheet, dst_row: int, pk_row: int) -> None:
    fix_sheet_cross_refs(ws_ee, dst_row, other_sheet="Pakistan", other_row=pk_row, quoted=True)
