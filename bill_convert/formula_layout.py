# -*- coding: utf-8 -*-
from __future__ import annotations

import re
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from bill_convert.formula_copy import (
    copy_row_formulas_from_snapshot,
    fix_ee_row_tw_refs,
    fix_tw_row_tw_ee_refs,
    snapshot_row_cells,
)
from bill_convert.mapping_spec import mapping_section
from bill_convert.person import (
    bill_employee_like_entry,
    norm_person_name,
    score_person_name_match,
)


def _norm_code(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def _codes_soft_equal(a: str, b: str) -> bool:
    if not a or not b:
        return False
    if a == b:
        return True
    # 兼容 CUS1525-0002 ↔ 0002
    a_tail = a.split("-")[-1] if "-" in a else a
    b_tail = b.split("-")[-1] if "-" in b else b
    return a == b_tail or b == a_tail or a_tail == b_tail


def _lookup_directory_row_for_emp(
    emp: dict[str, Any],
    directory: list[dict[str, Any]] | None,
) -> dict[str, Any] | None:
    if not directory:
        return None

    # 账单有工号时优先按工号对齐员工库（库名常与供应商姓名不一致）
    bill_code = _bill_employee_code(emp)
    if bill_code:
        soft: list[dict[str, Any]] = []
        for row in directory:
            if not isinstance(row, dict):
                continue
            got = _norm_code(row.get("employee_code") or row.get("employeeCode"))
            if got and _codes_soft_equal(bill_code, got):
                soft.append(row)
        if len(soft) == 1:
            return soft[0]
        if len(soft) > 1:
            exact = [
                r
                for r in soft
                if _norm_code(r.get("employee_code") or r.get("employeeCode")) == bill_code
            ]
            if len(exact) == 1:
                return exact[0]
            return soft[0]

    bill_labels = [
        str(emp.get("CN Name") or ""),
        str(emp.get("EN Name") or ""),
        str(emp.get("姓名") or ""),
        str(emp.get("Name of Employee") or ""),
        str(emp.get("EE Name") or ""),
        str(emp.get("Name") or ""),
    ]
    best: dict[str, Any] | None = None
    best_score = 0
    for row in directory:
        if not isinstance(row, dict):
            continue
        dir_labels = [
            str(row.get("employee_name") or row.get("employeeName") or ""),
            str(row.get("employee_name_en") or row.get("employeeNameEn") or ""),
        ]
        pair_best = 0
        for a in bill_labels:
            if not a.strip():
                continue
            for b in dir_labels:
                if not b.strip():
                    continue
                pair_best = max(pair_best, score_person_name_match(a, b))
        if pair_best > best_score:
            best_score = pair_best
            best = row
    return best if best_score >= 70 else None


def _formula_templates_spec(mapping: dict[str, Any]) -> dict[str, Any]:
    raw = mapping.get("formulaTemplates")
    return raw if isinstance(raw, dict) else {}


def _formula_sheet_block(mapping: dict[str, Any], sheet_key: str) -> dict[str, Any]:
    block = _formula_templates_spec(mapping).get(sheet_key)
    return block if isinstance(block, dict) else {}


def _formula_detect_strategy(block: dict[str, Any]) -> str:
    if block.get("fixedLayout") is True:
        return "fixed"
    if block.get("autoDetectLayout") is False:
        return "fixed"
    raw = str(block.get("detectStrategy") or "alignTwL").strip().lower()
    if raw in ("fixed", "scan"):
        return raw
    return "alignTwL"


def _scan_first_row_with_pattern(ws: Worksheet, pattern: re.Pattern[str], *, max_row: int = 35) -> int | None:
    for row in range(1, max_row + 1):
        for col in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(row, col)
            if cell.data_type == "f" and isinstance(cell.value, str) and pattern.search(cell.value):
                return row
    return None


def _default_example_row(mapping: dict[str, Any], sheet_key: str, data_start_fallback: int) -> int:
    block = _formula_sheet_block(mapping, sheet_key)
    if block.get("defaultExampleRow") is not None:
        return int(block["defaultExampleRow"])
    return data_start_fallback


def tw_l_row_for_data_row(
    data_row: int,
    *,
    data_start: int,
    target_l_data_start: int,
) -> int:
    return target_l_data_start + (data_row - data_start)


def resolve_formula_rows_layout(
    wb,
    mapping: dict[str, Any],
    target_l_data_start: int,
    *,
    tw_sheet: str = "TW",
    tw_ee_sheet: str = "TW EE",
    fallback_tw_data_start: int = 9,
    fallback_tw_ee_data_start: int = 10,
    target_l_sheet: str = "TW-L",
) -> dict[str, int]:
    tw_block = _formula_sheet_block(mapping, "TW")
    ee_block = _formula_sheet_block(mapping, "TW EE")
    tw_strategy = _formula_detect_strategy(tw_block)
    ee_strategy = _formula_detect_strategy(ee_block)

    if tw_strategy == "fixed":
        tw_start = int(tw_block.get("dataStartRow") or fallback_tw_data_start)
    elif tw_strategy == "scan" and tw_sheet in wb.sheetnames:
        pat = re.compile(rf"'{re.escape(target_l_sheet)}'!\$?[A-Z]{{1,3}}\$?\d+", re.I)
        tw_start = _scan_first_row_with_pattern(wb[tw_sheet], pat) or target_l_data_start
    else:
        tw_start = target_l_data_start

    ee_offset = int(ee_block.get("dataStartOffset") or 1)
    if ee_strategy == "fixed":
        ee_start = int(ee_block.get("dataStartRow") or fallback_tw_ee_data_start)
    elif ee_strategy == "scan" and tw_ee_sheet in wb.sheetnames:
        pat = re.compile(r"TW!\$?[A-Z]{1,3}\$?\d+", re.I)
        ee_start = _scan_first_row_with_pattern(wb[tw_ee_sheet], pat) or (tw_start + ee_offset)
    else:
        ee_start = tw_start + ee_offset

    return {
        "tw_l_data_start": target_l_data_start,
        "tw_data_start": tw_start,
        "tw_ee_data_start": ee_start,
    }


def _pick_int_field(entry: dict[str, Any], *keys: str) -> int | None:
    for key in keys:
        if entry.get(key) is None:
            continue
        try:
            return int(entry[key])
        except (TypeError, ValueError):
            continue
    return None


def _employee_formula_styles(mapping: dict[str, Any]) -> list[dict[str, Any]]:
    raw = mapping.get("employeeFormulaStyles")
    if not isinstance(raw, list):
        return []
    return [x for x in raw if isinstance(x, dict)]


def _directory_row_by_id(
    directory: list[dict[str, Any]] | None,
    employee_id: int,
) -> dict[str, Any] | None:
    if not directory:
        return None
    for row in directory:
        if not isinstance(row, dict):
            continue
        raw = row.get("employee_id") if row.get("employee_id") is not None else row.get("employeeId")
        try:
            if raw is not None and int(raw) == int(employee_id):
                return row
        except (TypeError, ValueError):
            continue
    return None


def _bill_employee_code(emp: dict[str, Any]) -> str:
    """从账单员工行提取工号（兼容多种表头）。"""
    preferred = (
        "工号",
        "员工编号",
        "雇员编号",
        "人员编号",
        "EE Code",
        "Employee Code",
        "Employee ID",
        "employee_code",
        "employeeCode",
        # HK Vertical-L：工号在「No. of EE」（样例值为 CUS1503-0001）
        "No. of EE",
        "No of EE",
        "EE No",
        "EE No.",
    )
    for key in preferred:
        got = _norm_code(emp.get(key))
        if got:
            return got
    for key, val in emp.items():
        ks = str(key)
        if "工号" in ks or re.search(
            r"employee\s*code|ee\s*code|no\.?\s*of\s*ee|ee\s*no",
            ks,
            re.I,
        ):
            got = _norm_code(val)
            if got:
                return got
    return ""


def _directory_row_by_names(
    directory: list[dict[str, Any]] | None,
    *,
    cn_name: Any = None,
    en_name: Any = None,
) -> dict[str, Any] | None:
    if not directory:
        return None
    labels = [str(x) for x in (cn_name, en_name) if x is not None and str(x).strip()]
    if not labels:
        return None
    best: dict[str, Any] | None = None
    best_score = 0
    for row in directory:
        if not isinstance(row, dict):
            continue
        dir_labels = [
            str(row.get("employee_name") or row.get("employeeName") or ""),
            str(row.get("employee_name_en") or row.get("employeeNameEn") or ""),
        ]
        pair_best = 0
        for a in labels:
            for b in dir_labels:
                if not b.strip():
                    continue
                pair_best = max(pair_best, score_person_name_match(a, b))
        if pair_best > best_score:
            best_score = pair_best
            best = row
    return best if best_score >= 70 else None


def _resolve_style_directory_person(
    entry: dict[str, Any],
    directory: list[dict[str, Any]] | None,
) -> dict[str, Any] | None:
    """映射条目 → 员工库行：employeeId / employeeCode / cnName|enName。"""
    sid = entry.get("employeeId") if entry.get("employeeId") is not None else entry.get("employee_id")
    if sid is not None:
        try:
            hit = _directory_row_by_id(directory, int(sid))
            if hit is not None:
                return hit
        except (TypeError, ValueError):
            pass
    code = _norm_code(entry.get("employeeCode") or entry.get("employee_code"))
    if code and directory:
        soft: list[dict[str, Any]] = []
        for row in directory:
            if not isinstance(row, dict):
                continue
            got = _norm_code(row.get("employee_code") or row.get("employeeCode"))
            if got and _codes_soft_equal(code, got):
                soft.append(row)
        if len(soft) == 1:
            return soft[0]
        if len(soft) > 1:
            exact = [
                r
                for r in soft
                if _norm_code(r.get("employee_code") or r.get("employeeCode")) == code
            ]
            return exact[0] if exact else soft[0]
    return _directory_row_by_names(
        directory,
        cn_name=entry.get("cnName"),
        en_name=entry.get("enName"),
    )


def _emp_matches_directory_person(emp: dict[str, Any], dir_row: dict[str, Any]) -> bool:
    bill_code = _bill_employee_code(emp)
    dir_code = _norm_code(dir_row.get("employee_code") or dir_row.get("employeeCode"))
    if bill_code and dir_code and _codes_soft_equal(bill_code, dir_code):
        return True
    return bill_employee_like_entry(
        emp,
        {
            "cnName": dir_row.get("employee_name") or dir_row.get("employeeName"),
            "enName": dir_row.get("employee_name_en") or dir_row.get("employeeNameEn"),
        },
        min_score=70,
    )


def _style_entry_matches_employee(entry: dict[str, Any], emp: dict[str, Any]) -> bool:
    return bill_employee_like_entry(emp, entry, min_score=70)


def _style_for_employee(
    mapping: dict[str, Any],
    emp: dict[str, Any],
    *,
    employee_directory: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    styles = _employee_formula_styles(mapping)
    if not styles:
        return {}

    bill_code = _bill_employee_code(emp)

    # 核心：映射人 → 员工库 → 用工号对齐账单（库名可以 ≠ 供应商姓名）
    for entry in styles:
        dir_row = _resolve_style_directory_person(entry, employee_directory)
        if dir_row is not None and _emp_matches_directory_person(emp, dir_row):
            return entry
        # 映射带 employeeId 且已解析到库行：再用库工号对账单工号（防止 entry.employeeCode 为空）
        if dir_row is not None and bill_code:
            dir_code = _norm_code(dir_row.get("employee_code") or dir_row.get("employeeCode"))
            if dir_code and _codes_soft_equal(bill_code, dir_code):
                return entry

    # 账单工号 ↔ 映射 employeeCode
    if bill_code:
        for entry in styles:
            got = _norm_code(entry.get("employeeCode") or entry.get("employee_code"))
            if got and _codes_soft_equal(bill_code, got):
                return entry

    # 账单 → 库 → employeeId
    dir_row = _lookup_directory_row_for_emp(emp, employee_directory)
    if dir_row is not None:
        eid = dir_row.get("employee_id") or dir_row.get("employeeId")
        if eid is not None:
            try:
                want = int(eid)
                for entry in styles:
                    sid = entry.get("employeeId") if entry.get("employeeId") is not None else entry.get("employee_id")
                    if sid is not None and int(sid) == want:
                        return entry
            except (TypeError, ValueError):
                pass

    # 姓名直接糊配（最后手段）：映射 cnName ↔ 账单姓名
    for entry in styles:
        if _style_entry_matches_employee(entry, emp):
            return entry
    return {}


def needed_example_rows_for_styles(
    mapping: dict[str, Any],
    employees: list[dict[str, Any]],
    *,
    main_template_key: str,
    ee_template_key: str,
    main_example_field: str,
    ee_example_field: str,
    main_data_start: int,
    ee_data_start: int,
    employee_directory: list[dict[str, Any]] | None = None,
) -> tuple[set[int], set[int]]:
    """扩行前需要保护/快照的示例行（默认行 + 映射里配置的所有示例行 + 每人配对行）。"""
    main_def = _default_example_row(mapping, main_template_key, main_data_start)
    ee_def = _default_example_row(mapping, ee_template_key, ee_data_start)
    main_needed: set[int] = {main_def}
    ee_needed: set[int] = {ee_def}
    # 先收录映射里声明的示例行，避免配对暂失败时扩行把第二种公式盖掉
    for entry in _employee_formula_styles(mapping):
        main_r = _pick_int_field(
            entry,
            main_example_field,
            "mainExampleRow",
            "chinaExampleRow",
            "twExampleRow",
            "hkExampleRow",
            "uaeExampleRow",
            "pakistanExampleRow",
            "ukExampleRow",
        )
        ee_r = _pick_int_field(
            entry,
            ee_example_field,
            "eeExampleRow",
            "chinaEeExampleRow",
            "twEeExampleRow",
            "hkEeExampleRow",
            "uaeEeExampleRow",
            "pakistanEeExampleRow",
            "ukEeExampleRow",
        )
        if main_r is not None:
            main_needed.add(main_r)
        if ee_r is not None:
            ee_needed.add(ee_r)
    for emp in employees:
        entry = _style_for_employee(mapping, emp, employee_directory=employee_directory)
        main_r = _pick_int_field(
            entry,
            main_example_field,
            "mainExampleRow",
            "chinaExampleRow",
            "twExampleRow",
            "hkExampleRow",
            "uaeExampleRow",
            "pakistanExampleRow",
            "ukExampleRow",
        )
        ee_r = _pick_int_field(
            entry,
            ee_example_field,
            "eeExampleRow",
            "chinaEeExampleRow",
            "twEeExampleRow",
            "hkEeExampleRow",
            "uaeEeExampleRow",
            "pakistanEeExampleRow",
            "ukEeExampleRow",
        )
        if main_r is not None:
            main_needed.add(main_r)
        if ee_r is not None:
            ee_needed.add(ee_r)
    return main_needed, ee_needed


def _apply_default_formula_template_to_all(mapping: dict[str, Any]) -> bool:
    spec = _formula_templates_spec(mapping)
    if "applyDefaultToAllEmployees" in spec:
        return bool(spec["applyDefaultToAllEmployees"])
    return True


def apply_employee_formula_styles(
    wb,
    employees: list[dict[str, Any]],
    mapping: dict[str, Any],
    *,
    formula_rows: dict[str, int],
    employee_directory: list[dict[str, Any]] | None = None,
    # 默认参数保持 TW 兼容；China 调用时传入 China / China EE / China-L 与 chinaExampleRow
    main_sheet: str = "TW",
    ee_sheet: str = "TW EE",
    target_l_sheet: str = "TW-L",
    main_template_key: str = "TW",
    ee_template_key: str = "TW EE",
    main_example_field: str = "twExampleRow",
    ee_example_field: str = "twEeExampleRow",
    fix_main_ee_refs=fix_tw_row_tw_ee_refs,
    fix_ee_main_refs=fix_ee_row_tw_refs,
    # 扩行前已拍好的快照（避免 fit 盖掉第二种公式示例行）
    main_snapshots: dict[int, list[dict[str, Any]]] | None = None,
    ee_snapshots: dict[int, list[dict[str, Any]]] | None = None,
    # 旧参数名（TW 调用仍可用）
    tw_sheet: str | None = None,
    tw_ee_sheet: str | None = None,
) -> None:
    if tw_sheet is not None:
        main_sheet = tw_sheet
    if tw_ee_sheet is not None:
        ee_sheet = tw_ee_sheet

    apply_all = _apply_default_formula_template_to_all(mapping)
    styles = _employee_formula_styles(mapping)
    if not apply_all and not styles:
        return []

    # 布局行号：通用键优先，其次兼容 TW 历史键名
    l_data_start = int(
        formula_rows.get("l_data_start")
        if formula_rows.get("l_data_start") is not None
        else formula_rows["tw_l_data_start"]
    )
    main_data_start = int(
        formula_rows.get("main_data_start")
        if formula_rows.get("main_data_start") is not None
        else formula_rows["tw_data_start"]
    )
    ee_data_start = int(
        formula_rows.get("ee_data_start")
        if formula_rows.get("ee_data_start") is not None
        else formula_rows["tw_ee_data_start"]
    )

    ws_main = wb[main_sheet]
    ws_ee = wb[ee_sheet]
    main_def = _default_example_row(mapping, main_template_key, main_data_start)
    ee_def = _default_example_row(mapping, ee_template_key, ee_data_start)

    # 先解析每人源行，再快照示例行，避免源行被先写坏或相互覆盖。
    # 未配对必须套用 defaultExampleRow，不能保留母版落位行（否则会误继承第二种公式）。
    plans: list[tuple[int, int, int]] = []
    main_src_needed: set[int] = set()
    ee_src_needed: set[int] = set()
    for i, emp in enumerate(employees):
        entry = _style_for_employee(mapping, emp, employee_directory=employee_directory)
        # 兼容 china/tw/hk ExampleRow / mainExampleRow 等别名，避免 UI 字段与引擎字段不一致
        main_over = _pick_int_field(
            entry,
            main_example_field,
            "mainExampleRow",
            "chinaExampleRow",
            "twExampleRow",
            "hkExampleRow",
            "uaeExampleRow",
            "pakistanExampleRow",
            "ukExampleRow",
        )
        ee_over = _pick_int_field(
            entry,
            ee_example_field,
            "eeExampleRow",
            "chinaEeExampleRow",
            "twEeExampleRow",
            "hkEeExampleRow",
            "uaeEeExampleRow",
            "pakistanEeExampleRow",
            "ukEeExampleRow",
        )
        if main_over is not None or ee_over is not None:
            src_main = main_over if main_over is not None else main_def
            src_ee = ee_over if ee_over is not None else ee_def
        else:
            src_main = main_def
            src_ee = ee_def
        plans.append((i, src_main, src_ee))
        main_src_needed.add(src_main)
        ee_src_needed.add(src_ee)

    main_snaps = dict(main_snapshots or {})
    ee_snaps = dict(ee_snapshots or {})
    for r in main_src_needed:
        if r not in main_snaps:
            main_snaps[r] = snapshot_row_cells(ws_main, r)
    for r in ee_src_needed:
        if r not in ee_snaps:
            ee_snaps[r] = snapshot_row_cells(ws_ee, r)

    for i, src_main, src_ee in plans:
        dst_main = main_data_start + i
        dst_ee = ee_data_start + i
        src_main_l = tw_l_row_for_data_row(
            src_main, data_start=main_data_start, target_l_data_start=l_data_start
        )
        dst_l = l_data_start + i
        copy_row_formulas_from_snapshot(
            main_snaps[src_main],
            ws_main,
            src_main,
            dst_main,
            src_main_l,
            dst_l,
            target_l_sheet=target_l_sheet,
        )
        fix_main_ee_refs(ws_main, dst_main, dst_ee)
        src_ee_l = tw_l_row_for_data_row(
            src_ee, data_start=ee_data_start, target_l_data_start=l_data_start
        )
        copy_row_formulas_from_snapshot(
            ee_snaps[src_ee],
            ws_ee,
            src_ee,
            dst_ee,
            src_ee_l,
            dst_l,
            target_l_sheet=target_l_sheet,
        )
        fix_ee_main_refs(ws_ee, dst_ee, dst_main)

    return [
        {
            "index": i + 1,
            "mainExampleRow": src_main,
            "eeExampleRow": src_ee,
        }
        for i, src_main, src_ee in plans
    ]
