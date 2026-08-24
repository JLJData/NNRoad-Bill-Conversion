# -*- coding: utf-8 -*-
"""把 mapping.fixedValueWrites 落到工作簿（格子约定只在 convert_mapping 里配）。"""
from __future__ import annotations

from typing import Any

from openpyxl.utils import column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from convert_mapping import find_sheet_name


def apply_fixed_value_writes(
    wb: Workbook, employees: list[dict[str, Any]], mapping: dict[str, Any] | None
) -> list[dict[str, Any]]:
    """mapping 里对应 valueKey 有值才写；缺省不改母版公式/数值。

    返回实际写入的格子列表（Excel 1-based），供 cellProvenance 使用。
    """
    if wb is None or not employees or not isinstance(mapping, dict):
        return []
    writes = mapping.get("fixedValueWrites")
    if not isinstance(writes, list):
        return []
    n = len(employees)
    cells: list[dict[str, Any]] = []
    for spec in writes:
        if isinstance(spec, dict):
            cells.extend(_apply_one(wb, spec, mapping, n))
    return cells


def _apply_one(
    wb: Workbook, spec: dict[str, Any], mapping: dict[str, Any], employee_count: int
) -> list[dict[str, Any]]:
    value_key = str(spec.get("valueKey") or "").strip()
    if not value_key:
        return []
    amount = _as_float(mapping.get(value_key))
    if amount is None:
        return []
    sheet_name = _resolve_sheet(wb, spec)
    if not sheet_name or sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    col = _resolve_column(ws, spec, mapping)
    if not col:
        return []
    start = _int_or_none(spec.get("dataStartRow"))
    if start is None and _uses_header_names(spec):
        start = _layout_int(mapping, "dataStartRow")
    if start is None or start < 1 or employee_count < 1:
        return []
    fee = round(float(amount), 6)
    label = str(spec.get("label") or value_key).strip() or value_key
    cells: list[dict[str, Any]] = []
    for i in range(employee_count):
        row = start + i
        ws.cell(row, col).value = fee
        cells.append(
            {
                "kind": "fixedValueWrites",
                "sheet": sheet_name,
                "row": row,
                "col": col,
                "sourceType": "mapping",
                "source": f"mapping.fixedValueWrites:{value_key}",
                "label": label,
                "value": fee,
                "detail": {"valueKey": value_key, "editGroup": "fixedValue"},
            }
        )
    return cells


def _resolve_sheet(wb: Workbook, spec: dict[str, Any]) -> str | None:
    sheet = str(spec.get("sheet") or "").strip()
    if not sheet:
        return None
    hit = find_sheet_name(list(wb.sheetnames), {"sheet": sheet, "candidates": spec.get("candidates") or []})
    return hit or sheet


def _uses_header_names(spec: dict[str, Any]) -> bool:
    names = spec.get("headerNames")
    return isinstance(names, list) and any(str(x).strip() for x in names)


def _resolve_column(ws: Worksheet, spec: dict[str, Any], mapping: dict[str, Any]) -> int | None:
    col = _int_or_none(spec.get("column"))
    if col and col > 0:
        return col
    letter = str(spec.get("columnLetter") or "").strip()
    if letter:
        try:
            return int(column_index_from_string(letter))
        except (ValueError, TypeError):
            return None
    names = spec.get("headerNames")
    if not isinstance(names, list):
        return None
    header_row = _int_or_none(spec.get("headerRow")) or _layout_int(mapping, "headerRow")
    if not header_row:
        return None
    want = {_norm_header(x) for x in names if str(x).strip()}
    if not want:
        return None
    max_col = min(ws.max_column or 1, 80)
    for c in range(1, max_col + 1):
        if _norm_header(ws.cell(header_row, c).value) in want:
            return c
    return None


def _layout_int(mapping: dict[str, Any], key: str) -> int | None:
    target = mapping.get("targetL") if isinstance(mapping.get("targetL"), dict) else {}
    return _int_or_none(target.get(key))


def _int_or_none(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        n = int(value)
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("\xa0", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\r", " ").replace(" ", "").replace(".", "").strip().lower()
