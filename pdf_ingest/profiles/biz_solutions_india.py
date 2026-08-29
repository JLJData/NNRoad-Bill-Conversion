# -*- coding: utf-8 -*-
"""
Biz Solutions（India）Tax Invoice PDF → India-L（profile: biz_solutions_india）

当前策略：
  - 从 PDF 提取：员工名、账期、Monthly CTC 总额、CGST+SGST（Business Tax）
  - Business Tax 取整见 mapping.indiaBusinessTaxRoundMode / indiaBusinessTaxRoundDigits（默认 ROUND 到整数）
  - 票面多出对不上 CTC/GST/合计的金额行则中止（避免 Expense Claim / Deduction 被写成 0）；
    mapping.ignoreUnknownInvoiceAmounts=true 时可忽略并继续（写入 warnings）
  - 薪资六项 + PT/IIT 走 mapping.indiaSalarySplit；Bonus 固定 0
  - 未配置拆分时：CTC 整笔进 Basic（其余 0）并 warning

用法:
  python -m pdf_ingest.profiles.biz_solutions_india <源.pdf> [-o 输出.xlsx]
"""
from __future__ import annotations

import argparse
import calendar
import re
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pdf_ingest.text_extract import extract_pdf_text
from pn_meta import PnMeta
from region_templates import get_region_template
from bill_convert.india_business_tax import (
    apply_india_business_tax_round,
    assert_no_unknown_invoice_amounts,
)

INDIA_L_SHEET = "India-L"

_MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _parse_indian_money(text: str) -> float | None:
    """2,78,805.00 / 278805 / ₹3,52,019.00 → float。"""
    raw = (text or "").strip()
    if not raw:
        return None
    raw = raw.replace("₹", "").replace(",", "").replace(" ", "").replace("\xa0", "")
    raw = re.sub(r"[^\d.\-]", "", raw)
    if not raw or raw in {".", "-", "-."}:
        return None
    try:
        return round(float(raw), 2)
    except ValueError:
        return None


def _period_bounds(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    return start, end


def parse_biz_solutions_pdf(
    pdf_path: Path,
    *,
    mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    path = Path(pdf_path).resolve()
    text = extract_pdf_text(path)
    if not text.strip():
        raise ValueError(f"PDF 无文本层或抽取为空: {path.name}")

    warnings: list[str] = []
    invoice_no = None
    m = re.search(r"Invoice\s*No\.?\s*#?\s*:?\s*(BIZ/[0-9\-]+/\d+)", text, re.I)
    if m:
        invoice_no = m.group(1).strip()

    invoice_date = None
    m = re.search(r"Invoice\s*Date\s*:?\s*(\d{1,2})-(\d{1,2})-(\d{4})", text, re.I)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        invoice_date = date(y, mo, d)

    # 员工名：Subject / Monthly CTC of Name / Project/P.O.
    employee_name = None
    m = re.search(r"Monthly CTC of\s+([A-Za-z][A-Za-z .']+?)\s*\]", text, re.I)
    if m:
        employee_name = _norm(m.group(1))
    if not employee_name:
        m = re.search(
            r"Project/P\.O\.#?\s*:?\s*[^|\n]*\|\s*([A-Za-z][A-Za-z .']+?)\s*\[",
            text,
            re.I,
        )
        if m:
            employee_name = _norm(m.group(1))
    if not employee_name:
        m = re.search(r"Subject\s*:?\s*\n?\s*([A-Za-z][A-Za-z .']+?)\s*\[", text, re.I)
        if m:
            employee_name = _norm(m.group(1))
    if not employee_name:
        raise ValueError(f"未能解析员工姓名: {path.name}")

    # 账期：month of June-2026 / for June-2026
    period_label = None
    year = month = None
    m = re.search(
        r"(?:month of|for)\s+"
        r"(January|February|March|April|May|June|July|August|September|October|November|December)"
        r"\s*[\-–]?\s*(\d{4})",
        text,
        re.I,
    )
    if m:
        month = _MONTHS[m.group(1).lower()]
        year = int(m.group(2))
        period_label = f"{m.group(1)}-{m.group(2)}"
    if not year or not month:
        warnings.append("未解析到账期，India-L 账期将留空")
        start = end = None
    else:
        start, end = _period_bounds(year, month)

    # Monthly CTC 金额：优先 [Monthly CTC ...] 段落后的金额行
    ctc = None
    m = re.search(
        r"Monthly CTC of[^\n]*\][^\n]*\n(?:.*\n){0,6}?.*?([0-9]{1,3}(?:,[0-9]{2,3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)\s*$",
        text,
        re.I | re.M,
    )
    # 更稳：Outsource payroll ... 后找第一个大金额；或直接找 998224 段
    amounts = re.findall(
        r"([0-9]{1,2}(?:,[0-9]{2,3}){2,}(?:\.\d{2})?|[0-9]{4,}(?:\.\d{2})?)",
        text,
    )
    parsed_amounts = [a for a in (_parse_indian_money(x) for x in amounts) if a is not None]

    m = re.search(
        r"Outsource payroll services.*?([0-9]{1,2}(?:,[0-9]{2,3})+(?:\.\d+)?)",
        text,
        re.I | re.S,
    )
    if m:
        ctc = _parse_indian_money(m.group(1))
    if ctc is None:
        # 次选：文中接近 278805 量级、且小于 Total 的应税拆分
        for a in parsed_amounts:
            if 50_000 <= a <= 5_000_000:
                # 跳过明显是价税合计
                if a > 300_000 and any(abs(a - x) < 1 for x in parsed_amounts if x > a):
                    continue
                ctc = a
                break
    if ctc is None:
        raise ValueError(f"未能解析 Monthly CTC: {path.name}")

    # CGST + SGST
    cgst = sgst = None
    m = re.search(r"CGST\s*9?\s*\(?\s*9%\s*\)?\s*([0-9,]+\.\d{2})", text, re.I)
    if m:
        cgst = _parse_indian_money(m.group(1))
    m = re.search(r"SGST\s*9?\s*\(?\s*9%\s*\)?\s*([0-9,]+\.\d{2})", text, re.I)
    if m:
        sgst = _parse_indian_money(m.group(1))
    if cgst is None:
        m = re.search(r"CGST9?\s*\(\s*9%\s*\)\s*([0-9,]+\.\d{2})", text, re.I)
        if m:
            cgst = _parse_indian_money(m.group(1))
    if sgst is None:
        m = re.search(r"SGST9?\s*\(\s*9%\s*\)\s*([0-9,]+\.\d{2})", text, re.I)
        if m:
            sgst = _parse_indian_money(m.group(1))
    # 汇总表备选
    if cgst is None or sgst is None:
        m = re.search(
            r"998224\s+([0-9,]+\.\d{2})\s+9%\s+([0-9,]+\.\d{2})\s+9%\s+([0-9,]+\.\d{2})",
            text,
            re.I,
        )
        if m:
            cgst = cgst or _parse_indian_money(m.group(2))
            sgst = sgst or _parse_indian_money(m.group(3))

    business_tax = None
    business_tax_formula = None
    if cgst is not None and sgst is not None:
        applied = apply_india_business_tax_round(cgst, sgst, mapping)
        business_tax = applied["value"]
        business_tax_formula = applied["formula"]
    elif cgst is not None:
        warnings.append("仅解析到 CGST，Business Tax 按 ×2 暂估")
        raise ValueError(
            "Biz Solutions PDF 仅解析到 CGST、缺少 SGST，版式可能已变更；已中止写出以免税额偏错"
        )
    else:
        warnings.append("未解析到 GST，Business Tax 置 0")
        raise ValueError(
            "Biz Solutions PDF 未解析到 GST，版式可能已变更；已中止写出以免税额被置 0"
        )

    ignore_unknown = False
    if isinstance(mapping, dict):
        ignore_unknown = bool(mapping.get("ignoreUnknownInvoiceAmounts"))

    unknown_warns = assert_no_unknown_invoice_amounts(
        text, ctc=ctc, cgst=cgst, sgst=sgst, ignore=ignore_unknown
    )
    warnings.extend(unknown_warns)

    return {
        "employee_name": employee_name,
        "invoice_no": invoice_no,
        "invoice_date": invoice_date,
        "period_label": period_label,
        "period_from": start,
        "period_to": end,
        "ctc": float(ctc),
        "cgst": cgst,
        "sgst": sgst,
        "business_tax": float(business_tax or 0),
        "business_tax_formula": business_tax_formula,
        "warnings": warnings,
        "source_file": path.name,
    }


def parsed_to_employee(
    parsed: dict[str, Any],
    *,
    mapping: dict[str, Any] | None = None,
    warnings: list[str] | None = None,
) -> dict[str, Any]:
    """PDF 字段 → India-L 员工 dict；薪资拆分+PT/IIT 优先走 mapping.indiaSalarySplit。"""
    from bill_convert.india_salary_split import apply_salary_split_to_employee

    emp: dict[str, Any] = {
        "Employee Name": parsed["employee_name"],
        "_period_from": parsed.get("period_from"),
        "_period_to": parsed.get("period_to"),
        "From": parsed.get("period_from"),
        "To": parsed.get("period_to"),
        "Business Tax": parsed.get("business_tax_formula") or parsed.get("business_tax") or 0,
        "_cgst": parsed.get("cgst"),
        "_sgst": parsed.get("sgst"),
        "_business_tax_formula": parsed.get("business_tax_formula"),
        "Expense Claim": 0,
        "Professional tax": 0,
        "Deduction": 0,
        "IIT": 0,
        "_invoice_no": parsed.get("invoice_no"),
        "_invoice_date": parsed.get("invoice_date"),
        "_period_label": parsed.get("period_label"),
        "_ctc": parsed.get("ctc"),
    }
    apply_salary_split_to_employee(
        emp,
        mapping=mapping,
        ctc=parsed.get("ctc"),
        warnings=warnings,
    )
    return emp


def convert_pdf(
    pdf_path: Path,
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    del pn_meta, registry_dir, fill_fx
    from convert_mapping import resolve_convert_mapping
    from profiles.india_payroll_calc.convert import write_india_l

    mapping_in = dict(convert_mapping) if isinstance(convert_mapping, dict) else {}
    mapping_in.setdefault("pdfProfileId", "biz_solutions_india")
    mapping = resolve_convert_mapping("india_payroll_calc", mapping_in)

    pdf_path = Path(pdf_path).resolve()
    output_path = Path(output_path).resolve()
    if not pdf_path.is_file():
        raise FileNotFoundError(f"PDF 不存在: {pdf_path}")

    parsed = parse_biz_solutions_pdf(pdf_path, mapping=mapping)
    warnings = list(parsed.get("warnings") or [])
    employee = parsed_to_employee(parsed, mapping=mapping, warnings=warnings)

    tpl = (template_path or get_region_template("India")).resolve()
    if not tpl.is_file():
        raise FileNotFoundError(f"India 母版不存在: {tpl}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tpl, output_path)
    wb = load_workbook(output_path)
    try:
        if INDIA_L_SHEET not in wb.sheetnames:
            raise ValueError(f"母版缺少 {INDIA_L_SHEET}")
        import profiles.india_payroll_calc.convert as india_mod

        prev = india_mod._ACTIVE_MAPPING
        india_mod._ACTIVE_MAPPING = mapping
        try:
            write_india_l(wb[INDIA_L_SHEET], [employee])
            from bill_convert.india_salary_split import (
                INDIA_SPLIT_PROVENANCE_FIELDS,
                build_india_business_tax_cell_writes,
                build_india_salary_split_cell_writes,
                resolve_field_columns_from_headers,
            )

            header_row, l_data_start = india_mod._india_l_layout(target=True)
            india_l_headers = india_mod._header_map(wb[INDIA_L_SHEET], header_row)
            split_field_cols = resolve_field_columns_from_headers(
                india_l_headers,
                fields=INDIA_SPLIT_PROVENANCE_FIELDS,
            )
            cell_writes = build_india_salary_split_cell_writes(
                [employee],
                sheet=INDIA_L_SHEET,
                data_start=l_data_start,
                field_cols=split_field_cols,
            )
            cell_writes.extend(
                build_india_business_tax_cell_writes(
                    [employee],
                    sheet=INDIA_L_SHEET,
                    data_start=l_data_start,
                    field_cols=resolve_field_columns_from_headers(
                        india_l_headers,
                        fields=("Business Tax",),
                    ),
                )
            )
        finally:
            india_mod._ACTIVE_MAPPING = prev
        wb.save(output_path)
    finally:
        wb.close()

    return {
        "ok": True,
        "profile_id": "biz_solutions_india",
        "region": "India",
        "source_kind": "pdf",
        "output": str(output_path),
        "employee_count": 1,
        "parsed": [
            {
                "employee_name": parsed.get("employee_name"),
                "invoice_no": parsed.get("invoice_no"),
                "period_label": parsed.get("period_label"),
                "ctc": parsed.get("ctc"),
                "business_tax": parsed.get("business_tax"),
            }
        ],
        "warnings": warnings,
        "fx_rate": None,
        "pn_meta": None,
        "cell_writes": cell_writes,
    }


def convert_pdfs(
    pdf_paths: list[Path],
    output_path: Path,
    **kwargs: Any,
) -> dict[str, Any]:
    paths = [Path(p).resolve() for p in pdf_paths]
    if not paths:
        raise ValueError("未提供 PDF")
    if len(paths) > 1:
        # 暂定单人单票；多人以后再合并
        raise ValueError("biz_solutions_india 暂只支持单份 PDF（一人一票）")
    return convert_pdf(paths[0], output_path, **kwargs)


def convert_sources(
    source_paths: list[Path],
    output_path: Path,
    **kwargs: Any,
) -> dict[str, Any]:
    paths = [Path(p).resolve() for p in source_paths]
    pdfs = [p for p in paths if p.suffix.lower() == ".pdf"]
    other = [p for p in paths if p not in pdfs]
    if other:
        raise ValueError(f"biz_solutions_india 主源为 PDF；不支持: {[p.name for p in other]}")
    if not pdfs:
        raise ValueError("请上传 Biz Solutions Tax Invoice PDF")
    return convert_pdfs(pdfs, output_path, **kwargs)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Biz Solutions India PDF → India-L")
    parser.add_argument("source", type=Path)
    parser.add_argument("-o", "--output", type=Path, default=None)
    parser.add_argument("-t", "--template", type=Path, default=None)
    args = parser.parse_args(argv)
    source = args.source.resolve()
    output = (args.output or source.with_name(f"IndiaL_{source.stem}.xlsx")).resolve()
    result = convert_pdf(source, output, template_path=args.template)
    print(result)
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
