# -*- coding: utf-8 -*-
"""
EOR Services Limited（UK）文字发票 PDF → 填 UK PN 母版的 UK-L。

样例字段（Invoice #6108）:
  Sarah Jane Walker-Monthly  3,634.70
    Monthly Salary- £3120, Empr Taxes £421.10, Empr Contributions £93.60
  Service Fee  200.00
  Invoice Total  £3,834.70

UK-L 竖表填入:
  Gross Salary / ER' NIC / ER' Pension
EE 侧 PAYE/EE NIC/EE Pension 发票通常没有 → 置 0 并 warning
Service Fee 与 PN Management Fee(450) 不是同一项 → 只进抽取结果，不瞎填 H 列
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from fx_rate import fetch_usd_rates, get_uk_gbp_per_usd
from pdf_ingest.text_extract import extract_pdf_text
from pn_meta import PnMeta, apply_pn_meta
from region_templates import get_region_template

UK_L_SHEET = "UK-L"
UK_SHEET = "UK"

# UK-L 标签 → 单元格（金额在 B 列）
_LABEL_CELLS: dict[str, str] = {
    "Gross Salary": "B7",
    "Holiday Pay": "B8",
    "Car Allowance": "B9",
    "ER' NIC": "B10",
    "ER' Pension (Auto Enrolment)": "B11",
    "PAYE (Estimated)": "B14",
    "EE'NIC": "B15",
    "EE' Pension (Auto Enrolment)": "B16",
    "App Levy": "B22",
    "Payment Fees": "B26",
}


@dataclass
class EorUkParsed:
    supplier: str = "EOR Services Limited"
    invoice_no: str | None = None
    invoice_date: date | None = None
    employee_name: str | None = None
    gross_salary: float | None = None
    er_nic: float | None = None
    er_pension: float | None = None
    service_fee: float | None = None
    invoice_total: float | None = None
    currency: str = "GBP"
    raw_line_net: float | None = None  # 工资行 Net Amt
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        if self.invoice_date is not None:
            d["invoice_date"] = self.invoice_date.isoformat()
        return d


def _money(s: str) -> float | None:
    s = (s or "").strip()
    if not s:
        return None
    s = s.replace(",", "").replace("£", "").replace("￡", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _uk_tax_year_label(d: date) -> str:
    """英国财年标签：4 月 6 日起算，如 2026-05-05 → 26-27。"""
    if d.month > 4 or (d.month == 4 and d.day >= 6):
        start = d.year
    else:
        start = d.year - 1
    return f"{str(start)[-2:]}-{str(start + 1)[-2:]}"


def parse_eor_uk_text(text: str) -> EorUkParsed:
    out = EorUkParsed()
    t = text.replace("\r", "\n")
    compact = re.sub(r"[ \t]+", " ", t)

    m = re.search(r"Invoice\s*No\s*[\n\r\s]*([0-9]+)", compact, re.I)
    if m:
        out.invoice_no = m.group(1).strip()

    m = re.search(
        r"Invoice\s*Date\s*[\n\r\s]*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})",
        compact,
        re.I,
    )
    if m:
        raw = m.group(1).replace("-", "/")
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y"):
            try:
                out.invoice_date = datetime.strptime(raw, fmt).date()
                break
            except ValueError:
                continue
        if out.invoice_date is None:
            out.warnings.append(f"无法解析发票日期: {raw}")

    # 员工名：描述行「Name-Monthly」
    m = re.search(
        r"([A-Za-z][A-Za-z .'-]{1,80}?)-Monthly",
        compact,
        re.I,
    )
    if m:
        out.employee_name = re.sub(r"\s+", " ", m.group(1)).strip(" -")

    # Salary / Empr Taxes / Contributions
    m = re.search(
        r"Monthly\s+Salary\s*[-–—]?\s*£?\s*([0-9,]+\.?[0-9]*)\s*,?\s*"
        r"Empr\s+Taxes?\s*£?\s*([0-9,]+\.?[0-9]*)\s*,?\s*"
        r"Empr\s+Contributions?\s*£?\s*([0-9,]+\.?[0-9]*)",
        compact,
        re.I,
    )
    if m:
        out.gross_salary = _money(m.group(1))
        out.er_nic = _money(m.group(2))
        out.er_pension = _money(m.group(3))
    else:
        out.warnings.append(
            "未匹配到「Monthly Salary / Empr Taxes / Empr Contributions」描述，版式可能已变更"
        )

    m = re.search(
        r"Service\s+Fee\s+([0-9,]+\.?[0-9]*)",
        compact,
        re.I,
    )
    if m:
        out.service_fee = _money(m.group(1))

    m = re.search(
        r"Invoice\s+Total\s*£?\s*([0-9,]+\.?[0-9]*)",
        compact,
        re.I,
    )
    if m:
        out.invoice_total = _money(m.group(1))

    # 工资行净额（描述后第一个金额块，作勾稽）
    m = re.search(
        r"-Monthly\s+([0-9,]+\.?[0-9]*)",
        compact,
        re.I,
    )
    if m:
        out.raw_line_net = _money(m.group(1))

    # 勾稽：Salary+NIC+Pension ≈ 工资行；+ServiceFee ≈ Total
    if (
        out.gross_salary is not None
        and out.er_nic is not None
        and out.er_pension is not None
        and out.raw_line_net is not None
    ):
        labor = round(out.gross_salary + out.er_nic + out.er_pension, 2)
        if abs(labor - out.raw_line_net) > 0.05:
            out.warnings.append(
                f"工资构成合计 {labor} 与行净额 {out.raw_line_net} 不一致"
            )
    if (
        out.raw_line_net is not None
        and out.service_fee is not None
        and out.invoice_total is not None
    ):
        expect = round(out.raw_line_net + out.service_fee, 2)
        if abs(expect - out.invoice_total) > 0.05:
            out.warnings.append(
                f"行净额+ServiceFee={expect} 与 Invoice Total {out.invoice_total} 不一致"
            )

    if not out.employee_name:
        out.warnings.append("未解析到员工姓名")
    if out.gross_salary is None:
        out.warnings.append("未解析到 Gross Salary")

    return out


def parse_eor_uk_pdf(pdf_path: Path) -> EorUkParsed:
    text = extract_pdf_text(pdf_path)
    parsed = parse_eor_uk_text(text)
    # 版式指纹
    low = text.lower()
    if "eor services limited" not in low and "eorservices.co.uk" not in low:
        parsed.warnings.append(
            "正文未出现 EOR Services 关键字，可能不是本 profile 对应的发票"
        )
    return parsed


def _set_amount_by_label(ws, label: str, value: float | None) -> None:
    addr = _LABEL_CELLS.get(label)
    if not addr or value is None:
        return
    ws[addr] = float(value)


def apply_to_workbook(
    wb,
    parsed: EorUkParsed,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
) -> tuple[list[str], float | None, PnMeta | None]:
    warnings = list(parsed.warnings)
    if UK_L_SHEET not in wb.sheetnames:
        raise ValueError(f"母版缺少 sheet「{UK_L_SHEET}」，现有: {wb.sheetnames}")

    ws = wb[UK_L_SHEET]
    name = parsed.employee_name or "Employee"
    fy = _uk_tax_year_label(parsed.invoice_date) if parsed.invoice_date else "YY-YY"
    ws["A3"] = f"{name}- Salary Calculation for FY {fy}"

    _set_amount_by_label(ws, "Gross Salary", parsed.gross_salary)
    _set_amount_by_label(ws, "Holiday Pay", 0.0)
    _set_amount_by_label(ws, "Car Allowance", 0.0)
    _set_amount_by_label(ws, "ER' NIC", parsed.er_nic)
    _set_amount_by_label(ws, "ER' Pension (Auto Enrolment)", parsed.er_pension)
    # 发票通常无 EE 侧明细
    _set_amount_by_label(ws, "PAYE (Estimated)", 0.0)
    _set_amount_by_label(ws, "EE'NIC", 0.0)
    _set_amount_by_label(ws, "EE' Pension (Auto Enrolment)", 0.0)
    _set_amount_by_label(ws, "App Levy", 0.0)
    warnings.append(
        "EE 侧 PAYE / EE NIC / EE Pension 发票未提供，已置 0（需人工或其它来源）"
    )
    if parsed.service_fee is not None:
        warnings.append(
            f"PDF Service Fee={parsed.service_fee}（与 PN Management Fee 不同项，未自动写入 UK!H）"
        )

    fx_rate = None
    if fill_fx:
        try:
            rates = fetch_usd_rates()
            fx_rate = get_uk_gbp_per_usd(rates)
            ws["D24"] = fx_rate
        except Exception as exc:
            warnings.append(f"写入 UK-L!D24 汇率失败: {exc}")

    if UK_SHEET in wb.sheetnames:
        wb[UK_SHEET]["B9"] = name

    applied_pn = None
    if pn_meta is not None:
        applied_pn = apply_pn_meta(
            wb,
            pn_meta,
            registry_dir=registry_dir,
            reserve_invoice_number=True,
        )

    return warnings, fx_rate, applied_pn


def convert_pdf(
    pdf_path: Path,
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    pdf_path = pdf_path.resolve()
    output_path = output_path.resolve()
    if not pdf_path.is_file():
        raise FileNotFoundError(f"PDF 不存在: {pdf_path}")

    tpl = (template_path or get_region_template("UK")).resolve()
    if not tpl.is_file():
        raise FileNotFoundError(f"UK 母版不存在: {tpl}")

    parsed = parse_eor_uk_pdf(pdf_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tpl, output_path)

    wb = load_workbook(output_path)
    warnings, fx_rate, applied_pn = apply_to_workbook(
        wb,
        parsed,
        pn_meta=pn_meta,
        registry_dir=registry_dir or output_path.parent,
        fill_fx=fill_fx,
    )
    wb.save(output_path)
    wb.close()

    return {
        "ok": True,
        "profile_id": "eor_uk",
        "region": "UK",
        "output": str(output_path),
        "parsed": parsed.to_dict(),
        "warnings": warnings,
        "fx_rate": fx_rate,
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
    }

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="EOR UK 发票 PDF → UK-L PN 源表")
    parser.add_argument("pdf", type=Path, help="供应商 PDF 路径")
    parser.add_argument("-o", "--output", type=Path, help="输出 xlsx")
    parser.add_argument("-t", "--template", type=Path, help="UK PN 母版")
    args = parser.parse_args(argv)

    pdf = args.pdf.resolve()
    out = (
        args.output.resolve()
        if args.output
        else pdf.parent / f"UK_L_from_pdf_{pdf.stem}.xlsx"
    )
    try:
        result = convert_pdf(pdf, out, template_path=args.template)
    except Exception as exc:
        print(f"失败: {exc}", file=sys.stderr)
        return 1

    p = result["parsed"]
    print("完成")
    print(f"  输出: {result['output']}")
    print(f"  发票号: {p.get('invoice_no')}  日期: {p.get('invoice_date')}")
    print(f"  员工: {p.get('employee_name')}")
    print(
        f"  Gross={p.get('gross_salary')}  ER NIC={p.get('er_nic')}  "
        f"ER Pension={p.get('er_pension')}  ServiceFee={p.get('service_fee')}  "
        f"Total={p.get('invoice_total')}"
    )
    for w in result.get("warnings") or []:
        print(f"  ! {w}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
