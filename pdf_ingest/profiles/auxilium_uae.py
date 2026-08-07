# -*- coding: utf-8 -*-
"""
Auxilium UAE Payroll Draft Excel → UAE-L 源表（profile: auxilium_uae）

识别：表头含 AX ID / Payroll Draft / NNRoad (UAE)。
多员工：跳过 TOTALS 行；Payroll Days 用工作日（不硬抄源 30）。
Admin Fee 写入 UAE-L「Admin Fees」。Recurring Fee 默认跟母版；若映射配置了
uaeRecurringFeeFixed 则写入该固定值（如 Omal）。
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pn_meta import PnMeta
from region_templates import get_region_template

UAE_L_SHEET = "UAE-L"
HEADER_ROW = 4
DATA_START = 5

# Payroll Draft 列号（1-based；同名列按位置区分）
COL_AX_ID = 2
COL_NAME = 3
COL_DESIGNATION = 4
COL_EC_BASIC = 5
COL_EC_HOUSING = 6
COL_EC_TRANSPORT = 7
COL_EC_OTHER = 8
COL_EC_GROSS = 9
COL_LOCALIZATION = 10
COL_GRATUITY_ACCRUAL = 11
COL_ADMIN_FEE = 12
COL_MONTH_DAYS = 13
COL_PAYROLL_DAYS = 14
COL_OT1_HRS = 15
COL_OT2_HRS = 16
COL_ACT_BASIC = 17
COL_ACT_HOUSING = 18
COL_ACT_TRANSPORT = 19
COL_ACT_SCHOOL = 20
COL_ACT_FOOD = 21
COL_ACT_MOBILE = 22
COL_ACT_OTHER = 23
COL_ACT_OT1 = 24
COL_ACT_OT2 = 25
COL_EXPENSES = 26
COL_DEDUCTIONS = 28
COL_TOTAL_PAY = 29
COL_VISA = 33
COL_MEDICAL = 34
COL_GRATUITY = 36
COL_AIRFARE = 38
COL_LEAVE = 39
COL_AJEER = 42
COL_LMRA = 43
COL_LOCALIZATION_GOV = 44
COL_GOSI = 45
COL_WORKMEN = 47
COL_ZAKAT = 48
COL_FAMILY_VISA = 49
COL_PAYOUT = 51
COL_INSTALLMENTS = 52
COL_ACCRUALS = 53
COL_GOV_FEES = 54
COL_ADMIN_FEE_INV = 55
COL_INVOICE_TOTAL = 57


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def _header_key(value: Any) -> str:
    """宽松表头键：Emp ID / EmpID / emp_id 等价。"""
    return re.sub(r"[^a-z0-9]+", "", _norm(value).lower())


_ID_HEADER_KEYS = frozenset(
    {"axid", "empid", "employeeid", "eeid", "employeeidno", "employeecode", "eecode"}
)
_NAME_HEADER_KEYS = frozenset({"employeename", "eename", "name", "fullname", "staffname"})


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("AED", "").replace("\xa0", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _as_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if value is None:
        return None
    text = _norm(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt)
        except ValueError:
            continue
    return None


def weekdays_inclusive(d0: datetime, d1: datetime) -> int:
    if d1 < d0:
        d0, d1 = d1, d0
    n = 0
    cur = d0.date() if isinstance(d0, datetime) else d0
    end = d1.date() if isinstance(d1, datetime) else d1
    while cur <= end:
        if cur.weekday() < 5:
            n += 1
        cur += timedelta(days=1)
    return n


def looks_like_auxilium_payroll_draft(path: Path) -> bool:
    path = Path(path)
    try:
        wb = load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return False
    try:
        for ws in wb.worksheets:
            max_c = min((ws.max_column or 1), 30)
            max_r = min((ws.max_row or 1), 16)
            head = " ".join(
                _norm(ws.cell(r, c).value).lower()
                for r in range(1, max_r + 1)
                for c in range(1, max_c + 1)
            )
            compact = re.sub(r"[^a-z0-9]+", "", head)
            if (
                "axid" in compact
                or "empid" in compact
                or "employeeid" in compact
                or "employeename" in compact
                or "payrolldraft" in compact
                or "nnroaduae" in compact
                or ("basic" in compact and "housing" in compact)
            ):
                return True
        return False
    finally:
        wb.close()


def looks_like_uae_l_workbook(path: Path) -> bool:
    path = Path(path)
    try:
        wb = load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return False
    try:
        if UAE_L_SHEET not in wb.sheetnames:
            return False
        ws = wb[UAE_L_SHEET]
        headers = {_norm(ws.cell(2, c).value) for c in range(1, 12)}
        return "Emp ID" in headers and "Employee Name" in headers
    finally:
        wb.close()


def parse_period_from_sheet(ws: Worksheet) -> tuple[datetime | None, datetime | None]:
    for r in range(1, 4):
        text = _norm(ws.cell(r, 1).value)
        m = re.search(
            r"Period:\s*(\d{4}-\d{2}-\d{2})\s*to\s*(\d{4}-\d{2}-\d{2})",
            text,
            re.I,
        )
        if m:
            return _as_date(m.group(1)), _as_date(m.group(2))
        m2 = re.search(r"(\d{4}-\d{2}-\d{2}).{1,8}(\d{4}-\d{2}-\d{2})", text)
        if m2 and "period" in text.lower():
            return _as_date(m2.group(1)), _as_date(m2.group(2))
    return None, None


def parse_period_from_filename(path: Path) -> tuple[datetime | None, datetime | None]:
    m = re.search(r"(20\d{2})-(\d{2})", path.name)
    if not m:
        return None, None
    y, mo = int(m.group(1)), int(m.group(2))
    start = datetime(y, mo, 1)
    if mo == 12:
        end = datetime(y, 12, 31)
    else:
        end = datetime(y, mo + 1, 1) - timedelta(days=1)
    return start, end


def _cell(ws: Worksheet, row: int, col: int) -> Any:
    return ws.cell(row, col).value


def _strip_currency_noise(key: str) -> str:
    """去掉 aed/usd 等币种尾巴，便于 Basic(AED) ↔ Basic。"""
    return re.sub(r"(aed|usd|sar|eur|gbp)$", "", key)


def _header_alias_keys(header: str) -> set[str]:
    """表头可匹配集合：完整 key、去 #n、以及「父/子」的子段。"""
    raw = str(header or "").strip()
    if not raw:
        return set()
    out: set[str] = set()
    k = _header_key(raw)
    if k:
        out.add(k)
    base = raw.split("#", 1)[0]
    bk = _header_key(base)
    if bk:
        out.add(bk)
    if "/" in base:
        child = _header_key(base.rsplit("/", 1)[-1])
        if child:
            out.add(child)
    return out


def _find_header_col(
    headers: list[str],
    *aliases: str,
    occurrence: int = 0,
    used: set[int] | None = None,
) -> int | None:
    """
    按别名找列（1-based），必须同名：仅 `_header_key` 规范化后全等
    （忽略大小写/空格/标点，如 Emp ID ≡ EmpID；不做包含、去币种等模糊匹配）。
    occurrence: 同名第几次（0=左起第一次；用于 EC Basic vs Actual Basic）。
    used: 已占用列会跳过，继续找下一个同名候选（避免 Overtime 2 占住后 EC - OT2 落空）。
    支持资格化表头「父/子」「子#2」：别名既可匹配完整 key，也可匹配子列名。
    """
    want_keys = [_header_key(a) for a in aliases if a and str(a).strip()]
    want_keys = [k for k in want_keys if k]
    if not want_keys:
        return None
    want_set = set(want_keys)
    used = used if used is not None else set()

    matches = [
        i
        for i, h in enumerate(headers, start=1)
        if want_set & _header_alias_keys(h)
    ]
    if not matches:
        return None
    # 优先取第 occurrence 个同名列（未占用时）
    if occurrence < len(matches) and matches[occurrence] not in used:
        return matches[occurrence]
    # occurrence=0：别名集合命中多列时，跳过已占用，取下一个（如 Overtime 2 已被工时占用 → EC - OT2）
    if occurrence == 0:
        for col in matches:
            if col not in used:
                return col
    return None


def _resolve_col(
    headers: list[str],
    aliases: tuple[str, ...],
    *,
    fallback: int | None = None,
    occurrence: int = 0,
    used: set[int] | None = None,
) -> int | None:
    col = _find_header_col(headers, *aliases, occurrence=occurrence, used=used)
    if col is not None:
        if used is not None:
            used.add(col)
        return col
    if fallback is not None and (used is None or fallback not in used):
        if used is not None:
            used.add(fallback)
        return fallback
    return None


def _is_id_header(text: Any) -> bool:
    key = _header_key(text)
    if not key:
        return False
    if key in _ID_HEADER_KEYS:
        return True
    return ("id" in key) and ("emp" in key or "ax" in key or "ee" in key)


def _is_name_header(text: Any) -> bool:
    key = _header_key(text)
    return key in _NAME_HEADER_KEYS or key in {"employeefullname", "stafffullname"}


def _row_looks_like_payroll_header(headers: list[str]) -> bool:
    """弱识别：有姓名/工号，或有多列薪酬特征，即视为表头（不强制 AX ID）。"""
    keys = [_header_key(h) for h in headers if h]
    if not keys:
        return False
    has_id = any(_is_id_header(h) for h in headers)
    has_name = any(_is_name_header(h) for h in headers)
    moneyish = 0
    markers = (
        "basic",
        "housing",
        "transport",
        "adminfee",
        "admin",
        "payrolldays",
        "invoicetotal",
        "gratuity",
        "localization",
        "employeepayout",
        "gross",
    )
    for k in keys:
        plain = _strip_currency_noise(k)
        if any(m in plain for m in markers):
            moneyish += 1
    if has_id or has_name:
        return moneyish >= 1 or has_id and has_name
    return moneyish >= 3


def _detect_header_row(
    ws: Worksheet,
    *,
    scan_max: int = 30,
    marker_keys: frozenset[str] | None = None,
) -> tuple[int, list[str]] | None:
    """扫描前若干行定位表头；配置了标志列时优先按标志列定位。"""
    max_col = min(max(ws.max_column or 1, 1), 100)
    if marker_keys and len(marker_keys) >= 2:
        try:
            from bill_convert.header_scan import find_header_row_by_markers

            row = find_header_row_by_markers(
                ws,
                marker_keys=marker_keys,
                max_scan=scan_max,
                sheet_label=ws.title or "Payroll Draft",
            )
            headers = [_norm(_cell(ws, row, c)) for c in range(1, max_col + 1)]
            return row, headers
        except ValueError:
            return None
    for row in range(1, min((ws.max_row or 1), scan_max) + 1):
        headers = [_norm(_cell(ws, row, c)) for c in range(1, max_col + 1)]
        if _row_looks_like_payroll_header(headers):
            return row, headers
    return None


def _pick_payroll_draft_sheet(
    wb,
    *,
    marker_keys: frozenset[str] | None = None,
    scan_max: int = 30,
) -> tuple[Any, int, list[str]]:
    """优先活动表，否则扫全部 sheet。"""
    candidates = []
    active = wb.active
    ordered = [active] + [wb[n] for n in wb.sheetnames if wb[n] is not active]
    last_sample: list[str] = []
    last_title = ""
    for ws in ordered:
        found = _detect_header_row(ws, scan_max=scan_max, marker_keys=marker_keys)
        if found is not None:
            header_row, headers = found
            candidates.append((ws, header_row, headers))
            if ws is active:
                return ws, header_row, headers
        else:
            max_col = min(max(ws.max_column or 1, 1), 40)
            row = min(4, max(ws.max_row or 1, 1))
            last_sample = [h for h in (_norm(_cell(ws, row, c)) for c in range(1, max_col + 1)) if h][:12]
            last_title = ws.title
    if candidates:
        return candidates[0]
    raise ValueError(
        f"未识别为 Auxilium/UAE Payroll Draft（各 sheet 前 {scan_max} 行未见员工/薪酬表头）。"
        f"sheet「{last_title}」样例: {last_sample or '（空）'}"
    )


def _vendor_names_for_targets(column_rename: dict[str, str] | None, *targets: str) -> list[str]:
    """
    columnRename 为 供应商列→UAE-L列；取映射到指定 UAE-L 列的供应商列名。

    右侧目标可能是资格化 key（母版第 1 行父级 + 第 2 行子列），例如：
    「Standard Salary Rates/EC - Basic Salary」「c/Emp ID」。
    须与引擎字段裸名「EC - Basic Salary」「Emp ID」按完整 key / 子段均可命中。
    """
    if not column_rename:
        return []
    want: set[str] = set()
    for t in targets:
        if not t or not str(t).strip():
            continue
        want |= _header_alias_keys(str(t).strip())
    if not want:
        return []
    out: list[str] = []
    seen: set[str] = set()
    for src, tgt in column_rename.items():
        if not src or not tgt:
            continue
        tgt_s = str(tgt).strip()
        if tgt_s in targets or (want & _header_alias_keys(tgt_s)):
            key = str(src).strip()
            if key and key not in seen:
                seen.add(key)
                out.append(key)
    return out


def _build_colmap(
    headers: list[str],
    column_rename: dict[str, str] | None = None,
) -> dict[str, int | None]:
    """
    逻辑字段 → 列号。只认两类来源，无内置别名、无固定列号兜底：
    1) columnRename 供应商列 → UAE-L 列
    2) 表头与目标列同名（_header_key 规范化后全等）
    同名重复列：occurrence 取第几次（0=左起第一次；用于 EC Rate vs Actual）。
    """
    used: set[int] = set()
    m: dict[str, int | None] = {}
    rename = column_rename if isinstance(column_rename, dict) else {}

    def put(field: str, targets: tuple[str, ...], *, occurrence: int = 0):
        from_map = _vendor_names_for_targets(rename, *targets)
        names = tuple(from_map + list(targets))
        m[field] = _resolve_col(
            headers, names, fallback=None, occurrence=occurrence, used=used
        )

    put("id", ("Emp ID",))
    put("name", ("Employee Name",))
    put("designation", ("Designation",))

    put("ec_basic", ("EC - Basic Salary",))
    put("ec_housing", ("EC - Housing Allowance",))
    put("ec_transport", ("EC - Transport Allowance",))
    put("ec_other", ("EC - Other allowance",))
    put("ec_gross", ("EC - Gross",))

    put("localization", ("EC - Localization", "Localization"))
    put("gratuity_accrual", ("EC - Gratuity Accrual",))
    put("admin_fee", ("Admin Fees",))
    put("month_days", ("Month Days",))
    put("payroll_days", ("Payroll Days",))
    put("ot1_hrs", ("Overtime 1",))
    put("ot2_hrs", ("Overtime 2",))

    put("act_basic", ("Basic Salary",), occurrence=1)
    put("act_housing", ("Housing Allowance",), occurrence=1)
    put("act_transport", ("Transport Allowance",), occurrence=1)
    put("act_school", ("School Allowance", "EC - School Allowance"))
    put("act_food", ("Food Allowance", "EC - Food Allowance"))
    put("act_mobile", ("Mobile Allowance", "EC - Mobile Allowance"))
    put("act_other", ("Other Allowance",), occurrence=1)
    put("act_ot1", ("EC - OT1", "OT1"))
    put("act_ot2", ("EC - OT2", "OT2"))

    put("expenses", ("Other Adjustments",))
    put("deductions", ("Deduction",))
    put("total_pay", ("Total Payout",))
    put("visa", ("Visa",))
    put("medical", ("Medical Insurance",))
    put("gratuity", ("Gratuity",))
    put("airfare", ("Airfare",))
    put("leave", ("Annual Leave",))
    put("ajeer", ("Ajeer",))
    put("lmra", ("LMRA",))
    put("localization_gov", ("Localization",), occurrence=1)
    put("gosi", ("GOSI", "EC - GOSI"))
    put("workmen", ("Workmen Comp", "EC - Workmen Comp"))
    put("zakat", ("Zakat",))
    put("family_visa", ("Family Visa Levy",))
    put("payout", ("Employee payout",))
    put("installments", ("Installments", "Total Installments"))
    put("accruals", ("Accruals", "Total Accruals"))
    put("gov_fees", ("Government Fees", "Total Government Fees"))
    put("admin_fee_inv", ("Admin Fees",), occurrence=1)
    put("invoice_total", ("Invoice total",))
    return m


def _col_val(ws: Worksheet, row: int, colmap: dict[str, int | None], field: str) -> Any:
    col = colmap.get(field)
    if not col:
        return None
    return _cell(ws, row, col)


def parse_auxilium_payroll_draft(
    excel_path: Path,
    *,
    column_rename: dict[str, str] | None = None,
    source_spec: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    path = Path(excel_path).resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Excel 不存在: {path}")
    from bill_convert.header_scan import marker_keys_from_spec

    spec = source_spec if isinstance(source_spec, dict) else {}
    markers = marker_keys_from_spec(spec)
    scan_max = int(spec.get("headerScanMaxRow") or 30)
    wb = load_workbook(path, data_only=True)
    try:
        ws, header_row, _headers_raw = _pick_payroll_draft_sheet(
            wb,
            marker_keys=markers or None,
            scan_max=scan_max,
        )
        from bill_convert.headers import list_qualified_header_cells

        qualified = list_qualified_header_cells(ws, header_row)
        max_col = max((int(h["col"]) for h in qualified), default=0)
        headers = [""] * max_col
        for h in qualified:
            headers[int(h["col"]) - 1] = str(h["key"])
        data_start = header_row + 1
        colmap = _build_colmap(headers, column_rename=column_rename)
        id_col = colmap.get("id")
        name_col = colmap.get("name")
        if not id_col and not name_col:
            raise ValueError(
                "未识别员工工号/姓名列：请在列名对照中配置到 Emp ID / Employee Name，"
                "或保证表头与这两列同名"
            )
        if not id_col:
            id_col = name_col
        if not name_col:
            name_col = id_col

        period_from, period_to = parse_period_from_sheet(ws)
        if period_from is None or period_to is None:
            period_from, period_to = parse_period_from_filename(path)

        employees: list[dict[str, Any]] = []
        for row in range(data_start, (ws.max_row or data_start) + 1):
            ax_id = _norm(_cell(ws, row, id_col))
            name = _norm(_cell(ws, row, name_col))
            if not ax_id and not name:
                continue
            blob = f"{ax_id} {name}".upper()
            if "TOTAL" in blob:
                continue

            d_from = period_from
            d_to = period_to
            if d_from and d_to:
                payroll_days = weekdays_inclusive(d_from, d_to)
            else:
                payroll_days = _as_float(_col_val(ws, row, colmap, "payroll_days"))

            month_days = _as_float(_col_val(ws, row, colmap, "month_days"))
            admin_fee = _as_float(_col_val(ws, row, colmap, "admin_fee"))
            if admin_fee is None:
                admin_fee = _as_float(_col_val(ws, row, colmap, "admin_fee_inv"))

            localization = _as_float(_col_val(ws, row, colmap, "localization"))
            if localization is None:
                localization = _as_float(_col_val(ws, row, colmap, "localization_gov"))

            def f(field: str) -> float | None:
                return _as_float(_col_val(ws, row, colmap, field))

            def f0(field: str) -> float:
                v = f(field)
                return v if v is not None else 0.0

            emp: dict[str, Any] = {
                "Emp ID": ax_id or None,
                "Employee Name": name.upper() if name else None,
                "Designation": _norm(_col_val(ws, row, colmap, "designation")) or None,
                "Payroll Currency": "AED",
                "Month Days": month_days,
                "From": d_from,
                "To": d_to,
                "Payroll Days": payroll_days,
                "EC - Basic Salary": f("ec_basic"),
                "EC - Housing Allowance": f("ec_housing"),
                "EC - Transport Allowance": f("ec_transport"),
                "EC - School Allowance": f0("act_school"),
                "EC - Other allowance": f("ec_other"),
                "EC - Mobile Allowance": f0("act_mobile"),
                "EC - Food Allowance": f0("act_food"),
                "OT Eligible": "No",
                "EC - OT1": f0("act_ot1"),
                "EC - OT2": f0("act_ot2"),
                "EC - Workmen Comp": f0("workmen"),
                "EC - Localization": localization,
                "EC - Gratuity Accrual": f("gratuity_accrual"),
                "EC - Health Insurance Installment": 0,
                "EC - GOSI": f0("gosi"),
                "EC - Admin Fees": None,
                "Absence Days": 0,
                "Days w/o Absence": payroll_days,
                "Overtime 1": f0("ot1_hrs"),
                "Overtime 2": f0("ot2_hrs"),
                "Basic Salary": f("act_basic"),
                "Housing Allowance": f("act_housing"),
                "Transport Allowance": f("act_transport"),
                "School Allowance": f0("act_school"),
                "Mobile Allowance": f0("act_mobile"),
                "Food Allowance": f0("act_food"),
                "Other Allowance": f("act_other"),
                "OT1": f0("act_ot1"),
                "OT2": f0("act_ot2"),
                "Other Adjustments": f0("expenses"),
                "Deduction": f0("deductions"),
                "Total Payout": f("total_pay") or f("payout"),
                "Visa": f0("visa"),
                "Medical Insurance": f0("medical"),
                "Total Installments": f0("installments"),
                "Gratuity": f("gratuity") or f("gratuity_accrual"),
                "Airfare": f0("airfare"),
                "Annual Leave": f0("leave"),
                "Other Accruals": 0,
                "Total Accruals": f("accruals"),
                "Ajeer": f0("ajeer"),
                "LMRA": f0("lmra"),
                "Localization": localization,
                "GOSI": f0("gosi"),
                "Workmen Comp": f0("workmen"),
                "Zakat": f0("zakat"),
                "Family Visa Levy": f0("family_visa"),
                "Total Government Fees": f("gov_fees") if f("gov_fees") is not None else localization,
                "Employee payout": f("payout") or f("total_pay"),
                "Installments": f0("installments"),
                "Accruals": f("accruals"),
                "Government Fees": f("gov_fees") if f("gov_fees") is not None else localization,
                "Admin Fees": admin_fee,
                "Invoice total": f("invoice_total"),
                "_admin_fee": admin_fee,
            }
            employees.append(emp)
        if not employees:
            raise ValueError(f"未解析到员工行: {path.name}")
        return employees
    finally:
        wb.close()


def convert_excels(
    excel_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Auxilium Payroll Draft（可多文件合并员工）→ 含 UAE-L 的源表。"""
    del pn_meta, registry_dir, fill_fx  # 仅产出 UAE-L，PN 由引擎填写
    from convert_mapping import resolve_convert_mapping

    mapping_in = dict(convert_mapping) if isinstance(convert_mapping, dict) else {}
    mapping_in.setdefault("pdfProfileId", "auxilium_uae")
    mapping = resolve_convert_mapping("uae_payroll_calc", mapping_in)
    rename = mapping.get("columnRename") if isinstance(mapping.get("columnRename"), dict) else {}
    source_spec = (
        mapping.get("sourceEmployeeSheet")
        if isinstance(mapping.get("sourceEmployeeSheet"), dict)
        else {}
    )

    paths = [Path(p).resolve() for p in excel_paths]
    if not paths:
        raise ValueError("未提供 Excel")
    for p in paths:
        if not p.is_file():
            raise FileNotFoundError(f"Excel 不存在: {p}")

    output_path = Path(output_path).resolve()
    uae_l_flags = [looks_like_uae_l_workbook(p) for p in paths]
    if all(uae_l_flags):
        if len(paths) > 1:
            raise ValueError("多份已是 UAE-L 的 Excel 无法自动合并，请只传一份")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(paths[0], output_path)
        return {
            "ok": True,
            "profile_id": "auxilium_uae",
            "region": "UAE",
            "source_kind": "excel_uae_l",
            "output": str(output_path),
            "employee_count": None,
            "parsed": [],
            "warnings": ["源表已是 UAE-L，已原样用作转换输入"],
            "fx_rate": None,
            "pn_meta": None,
        }

    employees: list[dict[str, Any]] = []
    warnings: list[str] = []
    for p in paths:
        if looks_like_uae_l_workbook(p):
            raise ValueError("请不要混传已成型 UAE-L 与 Payroll Draft")
        if not looks_like_auxilium_payroll_draft(p):
            warnings.append(f"文件可能不是 Auxilium Payroll Draft，仍尝试解析: {p.name}")
        employees.extend(
            parse_auxilium_payroll_draft(p, column_rename=rename, source_spec=source_spec)
        )

    tpl = (template_path or get_region_template("UAE")).resolve()
    if not tpl.is_file():
        raise FileNotFoundError(f"UAE 母版不存在: {tpl}")

    from profiles.uae_payroll_calc.convert import write_uae_l

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tpl, output_path)
    wb = load_workbook(output_path)
    try:
        if UAE_L_SHEET not in wb.sheetnames:
            raise ValueError(f"母版缺少 {UAE_L_SHEET}")
        write_uae_l(wb[UAE_L_SHEET], employees)
        wb.save(output_path)
    finally:
        wb.close()

    return {
        "ok": True,
        "profile_id": "auxilium_uae",
        "region": "UAE",
        "source_kind": "excel",
        "output": str(output_path),
        "employee_count": len(employees),
        "parsed": [
            {
                "emp_id": e.get("Emp ID"),
                "employee_name": e.get("Employee Name"),
                "admin_fee": e.get("Admin Fees"),
                "payroll_days": e.get("Payroll Days"),
                "invoice_total": e.get("Invoice total"),
            }
            for e in employees
        ],
        "warnings": warnings,
        "fx_rate": None,
        "pn_meta": None,
    }


def convert_sources(
    source_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """
    Payroll Draft Excel → UAE-L；Admin Fee 发票 PDF 作为旁路事实（Total VAT），不进员工表。
    """
    from bill_convert.vendor_plugins.runtime import parse_artifact_facts, split_main_and_artifacts

    paths = [Path(p).resolve() for p in source_paths]
    if not paths:
        raise ValueError("未提供源文件")
    main_paths, artifact_paths, split_warnings = split_main_and_artifacts(
        paths, pdf_profile_id="auxilium_uae"
    )
    excels = [p for p in main_paths if p.suffix.lower() in (".xlsx", ".xlsm", ".xls")]
    leftover_pdfs = [p for p in main_paths if p.suffix.lower() == ".pdf"]
    other = [p for p in main_paths if p not in excels and p not in leftover_pdfs]
    if other:
        raise ValueError(f"不支持的文件类型: {[p.name for p in other]}")
    if leftover_pdfs:
        raise ValueError(
            "auxilium_uae 主源仅支持 Excel Payroll Draft；"
            f"无法识别的 PDF: {[p.name for p in leftover_pdfs]}（Admin Fee 发票应能自动识别）"
        )
    if not excels:
        raise ValueError("请至少上传一份 Auxilium Payroll Draft Excel")

    artifact_facts, artifact_warnings = parse_artifact_facts(
        artifact_paths, pdf_profile_id="auxilium_uae"
    )
    # 同批解析到的 VAT 同时作为 latest，供引擎当 curr（即使以后不同批也走 latest）
    if isinstance(artifact_facts, dict) and "auxilium.admin_fee.total_vat" in artifact_facts:
        artifact_facts["auxilium.admin_fee.latest_vat"] = artifact_facts["auxilium.admin_fee.total_vat"]
    result = convert_excels(
        excels,
        output_path,
        template_path=template_path,
        pn_meta=pn_meta,
        registry_dir=registry_dir,
        fill_fx=fill_fx,
        convert_mapping=convert_mapping,
    )
    warnings = list(result.get("warnings") or [])
    warnings.extend(split_warnings)
    warnings.extend(artifact_warnings)
    if artifact_paths and not artifact_facts:
        warnings.append("已识别 Admin Fee PDF 但未解析到事实")
    elif artifact_facts:
        warnings.append(
            "已解析 Admin Fee PDF Total VAT="
            f"{artifact_facts.get('auxilium.admin_fee.total_vat')} "
            f"({artifact_facts.get('auxilium.admin_fee.source_file')})"
        )
    result["warnings"] = warnings
    result["artifact_facts"] = artifact_facts
    return result


def convert_pdf(
    pdf_path: Path,
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    raise ValueError("auxilium_uae 暂不支持 PDF，请上传 Payroll Draft Excel")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Auxilium UAE Payroll Draft → UAE-L")
    parser.add_argument("sources", nargs="+", type=Path)
    parser.add_argument("-o", "--output", type=Path)
    parser.add_argument("-t", "--template", type=Path)
    args = parser.parse_args(argv)
    sources = [p.resolve() for p in args.sources]
    out = (
        args.output.resolve()
        if args.output
        else sources[0].parent / f"UAE_L_from_auxilium_{sources[0].stem}.xlsx"
    )
    try:
        result = convert_sources(sources, out, template_path=args.template)
    except Exception as exc:
        print(f"失败: {exc}", file=sys.stderr)
        return 1
    print("完成", result.get("output"), "人数", result.get("employee_count"))
    for w in result.get("warnings") or []:
        print(" !", w)
    return 0


if __name__ == "__main__":
    sys.exit(main())
