# -*- coding: utf-8 -*-
"""
SafeGuard (SGWI) Italy Payroll Excel → Italy-L（profile: safeguard_italy）

源 sheet「Calculation」：
  姓名取供应商账单姓名列（通常 A 列，如 Matteo Cupi）
  列映射与 Auxilium 相同：有 columnRename 用对照，没有才表头同名；
  不按列号、无内置 Accruals→Leave。Fee Min 走 mapping.italyFeeMin。

用法:
  python -m pdf_ingest.profiles.safeguard_italy <源.xlsx> [-o 输出.xlsx]
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pn_meta import PnMeta
from region_templates import get_region_template

ITALY_L_SHEET = "Italy-L"
SRC_SHEET_CANDIDATES = ("Calculation", "calculation", "Italy-L")

# 列名对照不再内置：须在 Office「转换映射」columnRename 配置。
# 未配置时仅「源列名与 Italy-L 列名同名」自动匹配；SGWI→Fee 等改名必须显式配置。


def _norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").replace("\xa0", " ").replace("\uFEFF", "")
    return re.sub(r"\s+", " ", text).strip()


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("\xa0", "")
    if not text or text == "-":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _money(value: Any) -> float | None:
    """金额统一两位小数，避免源表浮点尾巴把 PN USD 合计顶到差 0.01。"""
    num = _as_float(value)
    if num is None:
        return None
    return round(num, 2)


def _header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    from bill_convert.headers import build_qualified_header_map

    return build_qualified_header_map(ws, header_row)


def _col_by_label(headers: dict[str, int], *labels: str) -> int | None:
    want = {_norm(x).lower() for x in labels if x}
    if not want:
        return None
    for key, col in headers.items():
        base = str(key).split("#", 1)[0]
        child = base.rsplit("/", 1)[-1]
        cands = {_norm(key).lower(), _norm(base).lower(), _norm(child).lower()}
        if cands & want:
            return col
    return None


def _header_lookup_keys(header: str) -> list[str]:
    """完整 key / 去 #n / 子段，供 columnRename 命中（含小写便于大小写不敏感）。"""
    k = _norm(header)
    if not k:
        return []
    out: list[str] = [k]
    base = k.split("#", 1)[0]
    if base and base not in out:
        out.append(base)
    if "/" in base:
        child = base.rsplit("/", 1)[-1]
        if child and child not in out:
            out.append(child)
    # 小写副本，匹配时不区分大小写
    for x in list(out):
        low = x.lower()
        if low not in out:
            out.append(low)
    return out


def _strip_target_label(tgt: str) -> str:
    """资格化目标取子段，写 Italy-L 用裸列名。"""
    raw = _norm(tgt)
    if not raw:
        return ""
    base = raw.split("#", 1)[0]
    return base.rsplit("/", 1)[-1] or raw


def _rename_target_for_source(src_h: str, column_rename: dict[str, str] | None) -> str | None:
    """columnRename：供应商列 → Italy-L 列；支持资格化「父/子」与子段命中。"""
    if not isinstance(column_rename, dict) or not column_rename:
        return None
    src_keys = set(_header_lookup_keys(src_h))
    for src, tgt in column_rename.items():
        if not src or not tgt:
            continue
        if _norm(src) in src_keys or _norm(src).lower() in src_keys:
            return _norm(tgt)
    for src, tgt in column_rename.items():
        if not src or not tgt:
            continue
        if set(_header_lookup_keys(src)) & src_keys:
            return _norm(tgt)
    return None


def _explicit_rename_targets(column_rename: dict[str, str] | None) -> set[str]:
    """已被列名对照占用的 Italy-L 列（规范化裸名）。保留供外部诊断。"""
    out: set[str] = set()
    if not isinstance(column_rename, dict):
        return out
    for v in column_rename.values():
        t = _strip_target_label(str(v))
        if t:
            out.add(t)
            out.add(t.lower())
    return out


def _resolve_target_for_source(
    src_h: str,
    column_rename: dict[str, str] | None = None,
    *,
    claimed: set[str] | None = None,
) -> str | None:
    """有对照用对照；没有才同名。对照已占用的目标，同名不再写。"""
    renamed = _rename_target_for_source(src_h, column_rename)
    if renamed:
        return _strip_target_label(renamed) or None

    keys = _header_lookup_keys(src_h)
    if not keys:
        return None
    same = keys[-1] if "/" in keys[0] else keys[0]
    tgt = _norm(same) or None
    if not tgt:
        return None
    if claimed and (tgt in claimed or tgt.lower() in claimed):
        return None
    return tgt


def _find_source_col(headers: dict[str, int], src_name: str) -> int | None:
    """按对照左侧列名反查源列（支持父/子资格化）。"""
    return _col_by_label(headers, src_name)


def _put_cell_value(emp: dict[str, Any], tgt: str, val: Any) -> None:
    if not tgt or val is None or val == "":
        return
    num = _money(val)
    tgt_l = tgt.lower()
    if tgt_l in (
        "po number",
        "currency",
        "sgwi minimum currency",
        "fee minimum currency",
        "applied sgwi minimum currency",
    ):
        emp[tgt] = val
    elif num is not None:
        emp[tgt] = num
    else:
        emp[tgt] = val


def _find_header_row_fallback(ws: Worksheet) -> int:
    for r in range(1, min(20, (ws.max_row or 1) + 1)):
        labels = {_norm(ws.cell(r, c).value).lower() for c in range(1, min(15, (ws.max_column or 1) + 1))}
        if "employee id" in labels or "employee name" in labels:
            return r
        if "po number" in labels and ("currency" in labels or "sgwi min" in labels):
            return r
    raise ValueError("未找到 SafeGuard 员工表头行（Employee ID / Employee Name）")


def _find_header_row(
    ws: Worksheet,
    *,
    source_spec: dict[str, Any] | None = None,
) -> int:
    from bill_convert.header_scan import resolve_header_row

    return resolve_header_row(
        ws,
        spec=source_spec,
        sheet_label=str((source_spec or {}).get("sheet") or ws.title or "SafeGuard"),
        fallback=lambda: _find_header_row_fallback(ws),
    )


def _find_sheet(wb) -> Worksheet:
    for name in SRC_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    # 兜底：第一张
    return wb[wb.sheetnames[0]]


def looks_like_safeguard_italy(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return False
    try:
        ws = _find_sheet(wb)
        blob = []
        for r in range(1, min(12, (ws.max_row or 1) + 1)):
            for c in range(1, min(12, (ws.max_column or 1) + 1)):
                v = _norm(ws.cell(r, c).value).lower()
                if v:
                    blob.append(v)
        text = " ".join(blob)
        return ("italy" in text or "animal equality" in text) and (
            "sgwi" in text or "safeguard" in text or "pay period" in text
        )
    finally:
        wb.close()


def looks_like_italy_l_workbook(path: Path) -> bool:
    from profiles.italy_payroll_calc.convert import looks_like_italy_l_workbook as _fn

    return _fn(path)


def _read_meta(ws: Worksheet, header_row: int) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    for r in range(1, header_row):
        label = _norm(ws.cell(r, 2).value).lower()
        val = ws.cell(r, 3).value
        if label == "customer":
            meta["_customer"] = _norm(val)
        elif label == "location":
            meta["_location"] = _norm(val)
        elif label == "pay period":
            meta["_pay_period"] = val
            meta["Pay Period"] = val
    # Invoice ID: 常见 R2 K 标签 / R3 K 号码
    if _norm(ws.cell(2, 11).value).lower().startswith("invoice"):
        meta["_invoice_id"] = ws.cell(3, 11).value
    return meta


def _salary_src_header(headers: dict[str, int]) -> str | None:
    for h in headers:
        if re.search(r"\bsalary\b", h, flags=re.I):
            return h
    return None


def parse_safeguard_italy_excel(
    excel_path: Path,
    *,
    column_rename: dict[str, str] | None = None,
    source_spec: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    path = Path(excel_path).resolve()
    wb = load_workbook(path, data_only=True)
    # 无缓存公式值时 data_only 为空；再开一份取单元格常量
    wb_raw = load_workbook(path, data_only=False)
    try:
        ws = _find_sheet(wb)
        ws_raw = _find_sheet(wb_raw)
        header_row = _find_header_row(ws, source_spec=source_spec)
        headers = _header_map(ws, header_row)
        meta = _read_meta(ws, header_row)
        rename = column_rename if isinstance(column_rename, dict) else {}

        def _cell_val(row: int, col: int) -> Any:
            v = ws.cell(row, col).value
            if v is not None and v != "":
                return v
            raw = ws_raw.cell(row, col).value
            if isinstance(raw, (int, float)):
                return raw
            return v

        # 姓名：优先 A 列非空；否则 Employee Name。
        # 不用 Employee ID 顶姓名——源表缺姓名时留给母版自带名（write_italy_l 保留）。
        name_col_a = 1
        name_header_col = _col_by_label(headers, "Employee Name")
        id_col = _col_by_label(headers, "Employee ID")

        salary_h = _salary_src_header(headers)

        employees: list[dict[str, Any]] = []
        data_start = header_row + 1
        for row in range(data_start, (ws.max_row or data_start) + 1):
            name = _norm(_cell_val(row, name_col_a))
            if not name and name_header_col:
                name = _norm(_cell_val(row, name_header_col))
            eid_raw = _cell_val(row, id_col) if id_col else None
            eid = _norm(eid_raw)
            # 工号或纯数字不当姓名（源表常把 ID 写在 Name 列 / A 列空只剩 ID）
            if name and eid and name == eid:
                name = ""
            if name and name.isdigit():
                if not eid:
                    eid = name
                    eid_raw = name
                name = ""
            if not name and not eid:
                continue
            low = (name or eid).lower()
            if "invoice total" in low or low.startswith("sgwi"):
                continue

            emp: dict[str, Any] = dict(meta)
            if name:
                emp["Employee Name"] = name
            if eid_raw not in (None, ""):
                emp["_employee_id"] = eid_raw

            claimed = _explicit_rename_targets(rename)
            applied_rename = 0
            mapped_cols: set[int] = set()
            for src_h, col in headers.items():
                tgt = _resolve_target_for_source(src_h, rename, claimed=claimed)
                if not tgt:
                    continue
                if _rename_target_for_source(src_h, rename):
                    applied_rename += 1
                _put_cell_value(emp, tgt, _cell_val(row, col))
                mapped_cols.add(col)
            # 对照 key 与表头字面不完全一致时再反查一次
            for src, tgt_raw in rename.items():
                if not src or not tgt_raw:
                    continue
                col = _find_source_col(headers, str(src))
                if not col or col in mapped_cols:
                    continue
                tgt = _strip_target_label(str(tgt_raw))
                if not tgt:
                    continue
                _put_cell_value(emp, tgt, _cell_val(row, col))
                applied_rename += 1
                mapped_cols.add(col)
            emp["_rename_applied"] = applied_rename

            # 薪资列未配对照：Italy-L 表头会被改成账期全称（Aug → August），补一格同值
            if salary_h and not _rename_target_for_source(salary_h, rename):
                from profiles.italy_payroll_calc.convert import salary_header_for_period

                period_title = _norm(salary_header_for_period(meta.get("_pay_period")) or salary_h)
                if period_title and period_title != _norm(salary_h):
                    _put_cell_value(emp, period_title, _cell_val(row, headers[salary_h]))

            # Fee Min 留给 mapping；源 SGWI Min 仅作参考字段
            if "SGWI Min" in headers:
                emp["_source_fee_min"] = _money(_cell_val(row, headers["SGWI Min"]))
            if "Fee Min" in headers:
                emp["_source_fee_min"] = _money(_cell_val(row, headers["Fee Min"]))

            employees.append(emp)

        if not employees:
            raise ValueError(f"未解析到员工行: {path.name}")
        return employees
    finally:
        wb.close()
        wb_raw.close()


def convert_excels(
    excel_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """SafeGuard Calculation Excel → 含 Italy-L 的源表。"""
    del pn_meta, registry_dir, fill_fx
    from convert_mapping import resolve_convert_mapping
    from profiles.italy_payroll_calc.convert import set_fee_min, write_italy_l

    mapping_in = dict(convert_mapping) if isinstance(convert_mapping, dict) else {}
    mapping_in.setdefault("pdfProfileId", "safeguard_italy")
    mapping = resolve_convert_mapping("italy_payroll_calc", mapping_in)
    rename = mapping.get("columnRename") if isinstance(mapping.get("columnRename"), dict) else {}
    rename = {str(k): str(v) for k, v in rename.items() if k and v and str(k).strip() != str(v).strip()}
    source_spec = (
        mapping.get("sourceEmployeeSheet")
        if isinstance(mapping.get("sourceEmployeeSheet"), dict)
        else {}
    )
    print(
        f"[italy-rename] vendor-to-source entries={len(rename)} "
        f"keys={list(rename.keys())[:8]} "
        f"markers={source_spec.get('headerMarkerKeys')}"
    )

    paths = [Path(p).resolve() for p in excel_paths]
    if not paths:
        raise ValueError("未提供 Excel")
    for p in paths:
        if not p.is_file():
            raise FileNotFoundError(f"Excel 不存在: {p}")

    output_path = Path(output_path).resolve()
    italy_l_flags = [looks_like_italy_l_workbook(p) for p in paths]
    # 已是 Italy-L 且无需列名对照时才原样拷贝；有 columnRename 必须重解析，否则对照不生效
    if all(italy_l_flags) and not rename:
        if len(paths) > 1:
            raise ValueError("多份已是 Italy-L 的 Excel 无法自动合并，请只传一份")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(paths[0], output_path)
        return {
            "ok": True,
            "profile_id": "safeguard_italy",
            "region": "Italy",
            "source_kind": "excel_italy_l",
            "output": str(output_path),
            "employee_count": None,
            "parsed": [],
            "warnings": ["源表已是 Italy-L，已原样用作转换输入"],
            "fx_rate": None,
            "pn_meta": None,
        }

    employees: list[dict[str, Any]] = []
    warnings: list[str] = []
    if not rename:
        warnings.append("columnRename 为空：仅同名列匹配。请确认已在映射里配置并点「保存映射」。")
    else:
        warnings.append(f"columnRename 已加载 {len(rename)} 条对照")
    for p in paths:
        if looks_like_italy_l_workbook(p) and not rename:
            raise ValueError("请不要混传已成型 Italy-L 与 SafeGuard 源账单")
        if not looks_like_safeguard_italy(p) and not looks_like_italy_l_workbook(p):
            warnings.append(f"文件可能不是 SafeGuard Italy 账单，仍尝试解析: {p.name}")
        elif looks_like_italy_l_workbook(p) and rename:
            warnings.append(f"源表含 Italy-L 且配置了列名对照，已按对照重解析: {p.name}")
        employees.extend(
            parse_safeguard_italy_excel(p, column_rename=rename, source_spec=source_spec)
        )
    if employees and rename:
        applied = int(employees[0].get("_rename_applied") or 0)
        if applied <= 0:
            warnings.append(
                "columnRename 已配置但未命中任何源列，请核对供应商列名是否与示例表头一致"
            )
        else:
            warnings.append(f"columnRename 本批命中写入 {applied} 次（按源列计）")
    for e in employees:
        e.pop("_rename_applied", None)

    tpl = (template_path or get_region_template("Italy")).resolve()
    if not tpl.is_file():
        raise FileNotFoundError(f"Italy 母版不存在: {tpl}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tpl, output_path)
    wb = load_workbook(output_path)
    try:
        if ITALY_L_SHEET not in wb.sheetnames:
            raise ValueError(f"母版缺少 {ITALY_L_SHEET}")
        # 临时激活 mapping，供 set_fee_min 读取
        import profiles.italy_payroll_calc.convert as italy_mod

        prev = italy_mod._ACTIVE_MAPPING
        italy_mod._ACTIVE_MAPPING = mapping
        try:
            write_italy_l(wb[ITALY_L_SHEET], employees)
            set_fee_min(wb, employees)
        finally:
            italy_mod._ACTIVE_MAPPING = prev
        wb.save(output_path)
    finally:
        wb.close()

    return {
        "ok": True,
        "profile_id": "safeguard_italy",
        "region": "Italy",
        "source_kind": "excel",
        "output": str(output_path),
        "employee_count": len(employees),
        "parsed": [
            {
                "employee_name": e.get("Employee Name"),
                "employee_id": e.get("_employee_id"),
                "pay_period": e.get("Pay Period") or e.get("_pay_period"),
                "invoice_id": e.get("_invoice_id"),
            }
            for e in employees
        ],
        "warnings": warnings,
        "fx_rate": None,
        "pn_meta": None,
    }


def convert_sources(
    source_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    paths = [Path(p).resolve() for p in source_paths]
    if not paths:
        raise ValueError("未提供源文件")
    excels = [p for p in paths if p.suffix.lower() in (".xlsx", ".xlsm", ".xls")]
    other = [p for p in paths if p not in excels]
    if other:
        raise ValueError(
            "safeguard_italy 目前仅支持 Excel 源账单；"
            f"不支持: {[p.name for p in other]}"
        )
    return convert_excels(
        excels,
        output_path,
        template_path=template_path,
        pn_meta=pn_meta,
        registry_dir=registry_dir,
        fill_fx=fill_fx,
        convert_mapping=convert_mapping,
    )


def convert_pdf(*_args, **_kwargs):
    raise RuntimeError("safeguard_italy 主源为 Excel，不支持 PDF；请上传 SGWI Payroll xlsx")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="SafeGuard Italy Excel → Italy-L")
    parser.add_argument("source", type=Path)
    parser.add_argument("-o", "--output", type=Path, default=None)
    parser.add_argument("-t", "--template", type=Path, default=None)
    args = parser.parse_args(argv)
    source = args.source.resolve()
    output = (args.output or source.with_name(f"ItalyL_{source.stem}.xlsx")).resolve()
    result = convert_excels([source], output, template_path=args.template)
    print(result)
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
