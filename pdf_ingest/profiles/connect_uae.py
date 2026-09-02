# -*- coding: utf-8 -*-
"""
Connect Resources（UAE）税票 PDF → UAE-L。

- 以 PDF 为准抽取：员工月薪、Commission、报销、EOSB、Emiratization、账期、发票号
- Basic / Housing / Transport：mapping.connectSalarySplit 按姓名维护；缺省整笔进 Basic
- Agency Fees 不写入 Recurring Fee（母版公式保留）
"""
from __future__ import annotations

import argparse
import calendar
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from convert_mapping import resolve_convert_mapping
from pdf_ingest.text_extract import extract_pdf_text
from pn_meta import PnMeta

UAE_L_SHEET = "UAE-L"
HEADER_ROW = 7
DATA_START = 8

CONNECT_HEADERS = [
    "S.No",
    "English Name",
    "Client",
    "From",
    "To",
    "No of days",
    "Basic Salary",
    "Housing Allowance",
    "Transport",
    "Other allowances",
    "Project Allowance",
    "Commission /bonus",
    "Additional Expenses",
    "Deduction",
    "Net salary",
    "Sevice Fees",
    "EOSB Accrual",
    "Emiratization fee",
    "Invoice Before VAT",
    "Total Invoice Value in AED",
]

_INV_NO_RE = re.compile(r"Invoice#\s*(CR\d+)", re.I)
_INV_DATE_RE = re.compile(r"Invoice\s+Date\s*:?\s*(\d{1,2}\s+\w+\s+\d{4})", re.I)
_SUBJECT_RE = re.compile(r"Subject\s*:?\s*\n?\s*([^\n]+)", re.I)
_PERIOD_RE = re.compile(
    r"(January|February|March|April|May|June|July|August|September|October|November|December)"
    r"['’]?\s*(\d{4})",
    re.I,
)
_NAME_AED_RE = re.compile(
    r"([A-Za-z][A-Za-z .']*?)\s*[-–—]\s*AED\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?|\d+(?:\.\d+)?)",
    re.I,
)
# CR76056 等新版：• Kevin Willmaser - Aug 2026: 18418.00 AED
_NAME_PERIOD_AED_RE = re.compile(
    r"(?:[•\u2022]\s*)?"
    r"([A-Za-z][A-Za-z .']*?)\s*[-–—]\s*"
    r"(?:January|February|March|April|May|June|July|August|September|October|November|December|"
    r"Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)"
    r"['’]?\s*\d{4}\s*:\s*"
    r"([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?|\d+(?:\.\d+)?)\s*AED",
    re.I,
)
# CR77401：• Mohamad Fiazul Huq -  41800.00 AED（金额在 AED 前，无月份）
_NAME_AMOUNT_AED_RE = re.compile(
    r"(?:[•\u2022]\s*)?"
    r"([A-Za-z][A-Za-z .']*?)\s*[-–—]\s*"
    r"([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?|\d+(?:\.\d+)?)\s*AED",
    re.I,
)
_NAME_GBP_RE = re.compile(
    r"([A-Za-z][A-Za-z .']*?)\s*[-–—]\s*[£￡]\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?|\d+(?:\.\d+)?)",
    re.I,
)
_MONEY_RE = re.compile(
    r"(?<![A-Za-z])([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?|\d+(?:\.\d+)?)(?!\s*%)"
)
# 「Commission - Kevin Willmaser 1.00 59,342.94」行首姓名
_LINE_KIND_NAME_RE = re.compile(
    r"^\s*\d+\s+"
    r"(?:commission|reimbursement|expense(?:s)?|payroll|eosb|end\s+of\s+service|"
    r"emiratisation|emiritisation|emiratization)\s*[-–—:]\s*"
    r"([A-Za-z][A-Za-z .,'']+?)(?=\s+\d|\s*$)",
    re.I | re.M,
)
_MONTH_MAP = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def _parse_money(text: str | None) -> float | None:
    if text is None:
        return None
    s = str(text).replace(",", "").replace("\xa0", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _norm_name(value: Any) -> str:
    s = str(value or "")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _name_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", _norm_name(value).lower())


def _parse_uk_date(text: str) -> date | None:
    raw = (text or "").strip()
    for fmt in ("%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _period_bounds(month_name: str, year: int) -> tuple[date, date]:
    m = _MONTH_MAP[month_name.lower()]
    last = calendar.monthrange(year, m)[1]
    return date(year, m, 1), date(year, m, last)


def _ensure_emp(bucket: dict[str, dict[str, Any]], name: str) -> dict[str, Any]:
    key = _name_key(name)
    if key not in bucket:
        bucket[key] = {
            "English Name": _norm_name(name),
            "payroll": 0.0,
            "commission": 0.0,
            "expenses": 0.0,
            "eosb": 0.0,
            "emiratization": 0.0,
        }
    else:
        # 保留第一次见到的较完整姓名
        if len(_norm_name(name)) > len(bucket[key]["English Name"]):
            bucket[key]["English Name"] = _norm_name(name)
    return bucket[key]


def _match_split(splits: dict[str, Any], name: str) -> dict[str, Any] | None:
    if not isinstance(splits, dict) or not name:
        return None
    want = _name_key(name)
    for k, v in splits.items():
        if _name_key(k) == want and isinstance(v, dict):
            return v
    # 宽松：包含关系
    for k, v in splits.items():
        if not isinstance(v, dict):
            continue
        kk = _name_key(k)
        if kk and (kk in want or want in kk):
            return v
    return None


def looks_like_connect_invoice(path: Path, text: str | None = None) -> bool:
    name = path.name.lower()
    if name.startswith("cr") and name.endswith(".pdf"):
        return True
    body = (text if text is not None else extract_pdf_text(path)).lower()
    return "connect resources" in body and ("tax invoice" in body or "invoice#" in body)


def parse_connect_invoice(
    path: Path,
    *,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    text = extract_pdf_text(path)
    if not looks_like_connect_invoice(path, text):
        raise ValueError(f"不是 Connect Resources 税票: {path.name}")

    mapping_in = dict(convert_mapping) if isinstance(convert_mapping, dict) else {}
    mapping_in["pdfProfileId"] = "connect_uae"
    mapping = resolve_convert_mapping("uae_payroll_calc", mapping_in)

    inv_no = None
    m = _INV_NO_RE.search(text)
    if m:
        inv_no = m.group(1).strip()

    inv_date = None
    m = _INV_DATE_RE.search(text)
    if m:
        inv_date = _parse_uk_date(m.group(1))

    subject = None
    m = _SUBJECT_RE.search(text)
    if m:
        subject = _norm_name(m.group(1))
        if subject.lower() in {"bill to", "nn road", "# description"}:
            subject = None

    client_label = _norm_name(mapping.get("connectClientLabel") or subject or "")

    period_from = period_to = None
    m = _PERIOD_RE.search(text)
    if m:
        period_from, period_to = _period_bounds(m.group(1), int(m.group(2)))
    elif inv_date:
        period_from, period_to = _period_bounds(
            calendar.month_name[inv_date.month], inv_date.year
        )

    # 按描述块归类金额
    low = text.replace("\xa0", " ")
    blocks = re.split(r"\n(?=\d+\s+)", low)
    emps: dict[str, dict[str, Any]] = {}
    emiratization_pool: list[float] = []
    warnings: list[str] = []

    for block in blocks:
        bl = block.lower()
        if "agency fee" in bl:
            continue  # 不进 Recurring Fee
        if "bank charge" in bl:
            continue
        kind = None
        if "emiratization" in bl or "emiratisation" in bl or "emiritisation" in bl:
            kind = "emiratization"
        elif "eosb" in bl or "end of service" in bl:
            kind = "eosb"
        elif "commission" in bl:
            kind = "commission"
        elif "reimbursement" in bl or "expense" in bl:
            kind = "expenses"
        elif "payroll" in bl or "outsourcing" in bl:
            kind = "payroll"
        else:
            continue

        name_hits = list(_NAME_AED_RE.finditer(block))
        if not name_hits:
            name_hits = list(_NAME_PERIOD_AED_RE.finditer(block))
        if not name_hits:
            name_hits = list(_NAME_AMOUNT_AED_RE.finditer(block))

        if kind == "commission" and not name_hits:
            # Kevin - £9144 后跟 AED 行
            gbp_hits = list(_NAME_GBP_RE.finditer(block))
            moneys = [_parse_money(x) for x in _MONEY_RE.findall(block)]
            moneys = [x for x in moneys if x is not None and x >= 100]
            for i, gh in enumerate(gbp_hits):
                emp = _ensure_emp(emps, gh.group(1))
                # 取块内最大 AED（通常为折算后金额）
                aed = max(moneys) if moneys else _parse_money(gh.group(2))
                if aed is not None:
                    emp["commission"] = round(emp["commission"] + aed, 6)
            if gbp_hits:
                continue
            # 「Commission - Kevin Willmaser 1.00 59,342.94」
            m_line = _LINE_KIND_NAME_RE.search(block)
            if m_line and moneys:
                emp = _ensure_emp(emps, m_line.group(1))
                emp["commission"] = round(emp["commission"] + max(moneys), 6)
            continue

        if kind == "emiratization" and not name_hits:
            # "Emiratization Fee 2.00 660.00" / "Emiritisation Fee - Per Employee ... 1.00 1,320.00"
            m_fee = re.search(
                r"(?:emir[ia]tisation|emiratization)\s+fee\b[^\n]*?"
                r"(?<![0-9,])(\d+(?:\.\d+)?)\s+([0-9]{1,3}(?:,[0-9]{3})*\.\d{2}|\d+\.\d{2})(?!\d)",
                block,
                re.I,
            )
            added = False
            if m_fee:
                qty = int(float(m_fee.group(1)))
                rate = _parse_money(m_fee.group(2))
                # qty=1 且 rate 为总额时进池一次；qty>1 时按人头复制单价
                if qty > 0 and rate is not None and rate >= 10:
                    if qty == 1:
                        emiratization_pool.append(rate)
                    else:
                        for _ in range(qty):
                            emiratization_pool.append(rate)
                    added = True
            if added:
                continue
            nums = [_parse_money(x) for x in _MONEY_RE.findall(block)]
            nums = [x for x in nums if x is not None]
            qty, rate = None, None
            for i, n in enumerate(nums):
                # 跳过行号；单价通常 >= 10
                if n in (1.0, 2.0, 3.0, 4.0, 5.0) and i + 1 < len(nums) and nums[i + 1] >= 10:
                    qty, rate = n, nums[i + 1]
                    break
            if qty and rate:
                if int(qty) == 1:
                    emiratization_pool.append(rate)
                else:
                    for _ in range(int(qty)):
                        emiratization_pool.append(rate)
            elif nums:
                big = [x for x in nums if x >= 10]
                if big:
                    emiratization_pool.append(big[-1])
            continue

        if kind == "expenses" and not name_hits:
            # 「Reimbursement - Kevin Willmaser, Zineb Messaoud 1.00 8,617.00」
            m_line = _LINE_KIND_NAME_RE.search(block)
            moneys = [_parse_money(x) for x in _MONEY_RE.findall(block)]
            moneys = [x for x in moneys if x is not None and x >= 1]
            if m_line and moneys:
                names = [n.strip() for n in re.split(r"\s*,\s*", m_line.group(1)) if n.strip()]
                # 去掉尾部误吃的数字碎片
                names = [re.sub(r"\s+\d[\d,.]*$", "", n).strip() for n in names]
                names = [n for n in names if n and re.search(r"[A-Za-z]", n)]
                total = max(moneys)
                if names:
                    each = round(total / len(names), 6)
                    for n in names:
                        emp = _ensure_emp(emps, n)
                        emp["expenses"] = round(float(emp.get("expenses") or 0) + each, 6)
            continue

        for hit in name_hits:
            emp = _ensure_emp(emps, hit.group(1))
            amt = _parse_money(hit.group(2))
            if amt is None:
                continue
            emp[kind] = round(float(emp.get(kind) or 0) + amt, 6)

    if not emps:
        raise ValueError(f"Connect 税票未解析到员工行: {path.name}")

    # Emiratization 未写到人名时，按人头均分票面池
    if emiratization_pool:
        ordered = list(emps.values())
        if len(emiratization_pool) == len(ordered):
            for emp, amt in zip(ordered, emiratization_pool):
                emp["emiratization"] = round(amt, 6)
        else:
            each = round(sum(emiratization_pool) / max(len(ordered), 1), 6)
            for emp in ordered:
                if not emp.get("emiratization"):
                    emp["emiratization"] = each

    splits = mapping.get("connectSalarySplit")
    if not isinstance(splits, dict):
        splits = {}

    employees: list[dict[str, Any]] = []
    for idx, emp in enumerate(emps.values(), start=1):
        name = emp["English Name"]
        payroll = float(emp.get("payroll") or 0)
        split = _match_split(splits, name) or {}
        basic = _parse_money(split.get("basic"))
        housing = _parse_money(split.get("housing"))
        transport = _parse_money(split.get("transport"))
        if basic is None and housing is None and transport is None:
            basic, housing, transport = payroll, 0.0, 0.0
            if payroll:
                warnings.append(
                    f"{name}：映射未配置 Basic/Housing/Transport，已将 PDF 月薪 {payroll} 全部计入 Basic"
                )
        else:
            basic = basic if basic is not None else 0.0
            housing = housing if housing is not None else 0.0
            transport = transport if transport is not None else 0.0
            split_sum = round(basic + housing + transport, 2)
            if payroll and abs(split_sum - round(payroll, 2)) > 0.05:
                warnings.append(
                    f"{name}：映射拆分合计 {split_sum} ≠ PDF 月薪 {payroll}"
                )
                raise ValueError(
                    f"Connect UAE：员工「{name}」映射拆分合计 {split_sum} ≠ PDF 月薪 {payroll}，已中止写出"
                )

        row = {
            "S.No": idx,
            "English Name": name,
            "Employee Name": name,
            "Client": client_label or None,
            "From": period_from,
            "To": period_to,
            "Basic Salary": basic,
            "Housing Allowance": housing,
            "Transport": transport,
            "Other allowances": None,
            "Project Allowance": 0,
            "Commission /bonus": emp.get("commission") or 0,
            "Additional Expenses": emp.get("expenses") or 0,
            "Deduction": None,
            "Sevice Fees": None,
            "EOSB Accrual": emp.get("eosb") or 0,
            "Emiratization fee": emp.get("emiratization") or 0,
            "_client": client_label or None,
            "_invoice_no": inv_no,
            "_invoice_date": inv_date.isoformat() if inv_date else None,
            "_payroll": payroll,
        }
        employees.append(row)

    return {
        "employees": employees,
        "warnings": warnings,
        "invoice_no": inv_no,
        "invoice_date": inv_date.isoformat() if inv_date else None,
        "client_label": client_label,
        "period_from": period_from.isoformat() if period_from else None,
        "period_to": period_to.isoformat() if period_to else None,
        "source_file": path.name,
    }


def _write_source_workbook(employees: list[dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = UAE_L_SHEET
    ws["A1"] = "Company name:"
    client = employees[0].get("Client") or employees[0].get("_client")
    if client:
        ws["C1"] = client
    ws["A2"] = "Payroll period:"
    from_dt = employees[0].get("From")
    if from_dt is not None:
        ws["C2"] = from_dt
        ws["C2"].number_format = "yyyy/m/d"
    ws["D2"] = "to"
    to_dt = employees[0].get("To")
    if to_dt is not None:
        ws["E2"] = to_dt
        ws["E2"].number_format = "yyyy/m/d"
    ws["A3"] = "Currency: "
    ws["C3"] = "AED"

    for col, h in enumerate(CONNECT_HEADERS, start=1):
        ws.cell(HEADER_ROW, col).value = h
    for i, emp in enumerate(employees):
        row = DATA_START + i
        for col, h in enumerate(CONNECT_HEADERS, start=1):
            if h in emp and not str(h).startswith("_"):
                ws.cell(row, col).value = emp.get(h)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def convert_pdf(
    pdf_path: Path,
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    pdf_path = Path(pdf_path).resolve()
    output_path = Path(output_path).resolve()
    parsed = parse_connect_invoice(pdf_path, convert_mapping=convert_mapping)
    employees = parsed["employees"]
    _write_source_workbook(employees, output_path)
    return {
        "ok": True,
        "profile_id": "connect_uae",
        "region": "UAE",
        "source_kind": "pdf",
        "output": str(output_path),
        "employee_count": len(employees),
        "parsed": [
            {
                "employee_name": e.get("English Name"),
                "basic": e.get("Basic Salary"),
                "housing": e.get("Housing Allowance"),
                "transport": e.get("Transport"),
                "commission": e.get("Commission /bonus"),
                "expenses": e.get("Additional Expenses"),
                "eosb": e.get("EOSB Accrual"),
                "emiratization": e.get("Emiratization fee"),
            }
            for e in employees
        ],
        "warnings": parsed.get("warnings") or [],
        "fx_rate": None,
        "pn_meta": None,
        "invoice_no": parsed.get("invoice_no"),
        "client_label": parsed.get("client_label"),
    }


def convert_sources(
    source_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    paths = [Path(p).resolve() for p in source_paths]
    pdfs = [p for p in paths if p.suffix.lower() == ".pdf"]
    if not pdfs:
        raise ValueError("connect_uae 请上传 Connect Tax Invoice PDF")
    if len(pdfs) > 1:
        # 取第一份税票；其余忽略并警告
        pass
    result = convert_pdf(
        pdfs[0],
        output_path,
        template_path=template_path,
        pn_meta=pn_meta,
        registry_dir=registry_dir,
        fill_fx=fill_fx,
        convert_mapping=convert_mapping,
    )
    if len(pdfs) > 1:
        warnings = list(result.get("warnings") or [])
        warnings.append(f"本批 {len(pdfs)} 份 PDF，已采用 {pdfs[0].name}")
        result["warnings"] = warnings
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Connect UAE Tax Invoice → UAE-L")
    parser.add_argument("sources", nargs="+", type=Path)
    parser.add_argument("-o", "--output", type=Path)
    args = parser.parse_args(argv)
    sources = [p.resolve() for p in args.sources]
    out = (
        args.output.resolve()
        if args.output
        else sources[0].parent / f"UAE_L_from_connect_{sources[0].stem}.xlsx"
    )
    try:
        result = convert_sources(sources, out)
    except Exception as exc:
        print(f"失败: {exc}")
        return 1
    print("完成", result.get("output"), "人数", result.get("employee_count"))
    for w in result.get("warnings") or []:
        print(" !", w)
    for row in result.get("parsed") or []:
        print(" -", row)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
