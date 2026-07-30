# -*- coding: utf-8 -*-
"""
TopSource Worldwide (UK) 文字发票 PDF → 填 UK PN 母版。

样例（Invoice_TGS25266767 / 768）:
  一人一张票；金额为 USD 打包价，无 GBP 明细行。
  Description: Gross Salary + ER NIC + ER Pension + App Levy for {Name} - Feb - 2026
  Service Charge / Total USD

能自动填: 员工名、财年标题、汇率 D24、（多人）UK-L / UK-L (2)…
需人工补: Gross/Holiday/Car/ER NIC/Pension/PAYE/EE…（同事用截图补）
Service Charge 与 PN Management Fee 公式不同 → 只进抽取结果 + warning
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bill_convert.uk_layout import (
    UK_L_SHEET,
    UK_SHEET,
    ensure_uk_l_count,
    expand_uk_employee_rows,
)
from fx_rate import fetch_usd_rates, get_uk_gbp_per_usd
from pdf_ingest.text_extract import extract_pdf_text
from pn_meta import PnMeta, apply_pn_meta
from region_templates import get_region_template

UK_EE_SHEET = "UK EE"
PN_SHEET = "PN"

_MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


@dataclass
class TopSourceUkParsed:
    supplier: str = "TopSource Worldwide (UK) Limited"
    invoice_no: str | None = None
    invoice_date: date | None = None
    due_date: date | None = None
    employee_name: str | None = None
    period_label: str | None = None  # e.g. Feb-2026
    title: str | None = None  # Excel A3 原标题
    labor_usd: float | None = None  # Gross+ERNIC+Pension+AppLevy 打包
    service_charge_usd: float | None = None
    invoice_total_usd: float | None = None
    currency: str = "USD"
    account_number: str | None = None
    # Excel 源：标签 → GBP 金额（写入 UK-L）
    amounts: dict[str, float] = field(default_factory=dict)
    # 发票 Exchange Rate（USD per GBP）；写入 D24 时用 1/rate
    fx_usd_per_gbp: float | None = None
    source_kind: str = "pdf"  # pdf | excel
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        for key in ("invoice_date", "due_date"):
            val = d.get(key)
            if isinstance(val, date):
                d[key] = val.isoformat()
        return d


# Excel 竖表标签别名 → UK-L 标准标签
_LABEL_ALIASES: dict[str, str] = {
    "gross salary": "Gross Salary",
    "holiday pay": "Holiday Pay",
    "holiday": "Holiday Pay",
    "car allowance": "Car Allowance",
    "er' nic": "ER' NIC",
    "er nic": "ER' NIC",
    "er' pension (auto enrolment)": "ER' Pension (Auto Enrolment)",
    "er' pension": "ER' Pension (Auto Enrolment)",
    "paye (estimated)": "PAYE (Estimated)",
    "paye": "PAYE (Estimated)",
    "ee'nic": "EE'NIC",
    "ee' nic": "EE'NIC",
    "ee nic": "EE'NIC",
    "ee' pension (auto enrolment)": "EE' Pension (Auto Enrolment)",
    "ee' pension": "EE' Pension (Auto Enrolment)",
    "app levy": "App Levy",
    "payment fees": "Payment Fees",
    "setup fees": "Setup Fees",
}

_UK_L_AMOUNT_CELLS: dict[str, str] = {
    "Gross Salary": "B7",
    "Holiday Pay": "B8",
    "Car Allowance": "B9",
    "ER' NIC": "B10",
    "ER' Pension (Auto Enrolment)": "B11",
    "PAYE (Estimated)": "B14",
    "EE'NIC": "B15",
    "EE' Pension (Auto Enrolment)": "B16",
    "App Levy": "B22",
    "Setup Fees": "B25",
    "Payment Fees": "B26",
}

_TITLE_RE = re.compile(
    r"^\s*(.+?)\s*-?\s*Salary Calculation\s+for\s+FY\s+(.+?)\s*$",
    re.IGNORECASE,
)

def _money(s: str) -> float | None:
    s = (s or "").strip()
    if not s:
        return None
    s = s.replace(",", "").replace("$", "").replace("£", "").replace("\xa0", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _parse_uk_date(raw: str) -> date | None:
    raw = (raw or "").strip().replace("-", "/")
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _uk_tax_year_label(d: date) -> str:
    if d.month > 4 or (d.month == 4 and d.day >= 6):
        start = d.year
    else:
        start = d.year - 1
    return f"{str(start)[-2:]}-{str(start + 1)[-2:]}"


def parse_topsource_uk_text(text: str) -> TopSourceUkParsed:
    out = TopSourceUkParsed()
    t = text.replace("\r", "\n").replace("\xa0", " ")
    compact = re.sub(r"[ \t]+", " ", t)

    m = re.search(r"Invoice\s+Number\s*\n?\s*(TGS\d+)", compact, re.I)
    if m:
        out.invoice_no = m.group(1).strip()

    m = re.search(r"Invoice\s+Date\s*\n?\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})", compact, re.I)
    if m:
        out.invoice_date = _parse_uk_date(m.group(1))
        if out.invoice_date is None:
            out.warnings.append(f"无法解析发票日期: {m.group(1)}")

    m = re.search(r"Due\s+Date\s*:?\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})", compact, re.I)
    if m:
        out.due_date = _parse_uk_date(m.group(1))

    m = re.search(r"Account\s+Number\s*\n?\s*(C\d+)", compact, re.I)
    if m:
        out.account_number = m.group(1).strip()

    # Reference: 24640-Adrian Grace-Feb-26
    m = re.search(
        r"Reference\s*\n?\s*\d+\s*-\s*([A-Za-z][A-Za-z .'-]{1,80}?)\s*-\s*([A-Za-z]{3,9})\s*-?\s*(\d{2,4})",
        compact,
        re.I,
    )
    if m:
        out.employee_name = re.sub(r"\s+", " ", m.group(1)).strip(" -")
        mon = _MONTHS.get(m.group(2).lower())
        yy = int(m.group(3))
        year = yy if yy > 100 else 2000 + yy
        if mon:
            out.period_label = f"{m.group(2).title()}-{year}"

    # Description 行里的姓名（兜底）
    if not out.employee_name:
        m = re.search(
            r"App\s+Levy\s+for\s*\n?\s*([A-Za-z][A-Za-z .'-]{1,80}?)\s*-\s*([A-Za-z]{3,9})",
            compact,
            re.I,
        )
        if m:
            out.employee_name = re.sub(r"\s+", " ", m.group(1)).strip(" -")

    # 打包人工成本 USD（描述后的金额）
    m = re.search(
        r"App\s+Levy\s+for[\s\S]{0,120}?1\s+\$?([0-9,]+\.?[0-9]*)\s+0%\s+\$?([0-9,]+\.?[0-9]*)",
        compact,
        re.I,
    )
    if m:
        out.labor_usd = _money(m.group(2)) or _money(m.group(1))
    else:
        m = re.search(
            r"Gross\s+Salary\s*\+\s*ER\s+NIC[\s\S]{0,200}?\$([0-9,]+\.?[0-9]*)",
            compact,
            re.I,
        )
        if m:
            out.labor_usd = _money(m.group(1))

    m = re.search(
        r"Service\s+Charge\s+1\s+\$?([0-9,]+\.?[0-9]*)\s+0%\s+\$?([0-9,]+\.?[0-9]*)",
        compact,
        re.I,
    )
    if m:
        out.service_charge_usd = _money(m.group(2)) or _money(m.group(1))
    else:
        m = re.search(r"Service\s+Charge\s+.*?\$([0-9,]+\.?[0-9]*)", compact, re.I)
        if m:
            out.service_charge_usd = _money(m.group(1))

    m = re.search(r"Total\s*USD\s*\$?([0-9,]+\.?[0-9]*)", compact, re.I)
    if m:
        out.invoice_total_usd = _money(m.group(1))

    if not out.employee_name:
        out.warnings.append("未解析到员工姓名")
    if out.labor_usd is None:
        out.warnings.append("未解析到人工成本 USD 打包金额")
    out.warnings.append(
        "TopSource PDF 仅为 USD 打包价，无 GBP 明细；Gross/Holiday/PAYE 等请按截图人工补齐"
    )
    if out.labor_usd is not None:
        out.warnings.append(f"发票人工打包 USD={out.labor_usd}（供核对，未写入 UK-L 明细）")
    if out.service_charge_usd is not None:
        out.warnings.append(
            f"发票 Service Charge USD={out.service_charge_usd}（与 PN Management Fee 公式不同，未自动写入）"
        )
    return out


def parse_topsource_uk_pdf(pdf_path: Path) -> TopSourceUkParsed:
    text = extract_pdf_text(pdf_path)
    parsed = parse_topsource_uk_text(text)
    parsed.source_kind = "pdf"
    low = text.lower()
    if "topsource" not in low:
        parsed.warnings.append("正文未出现 TopSource 关键字，可能不是本 profile 对应的发票")
    return parsed


def _norm_label(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\xa0", " ").strip()


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("£", "").replace("$", "").replace("\xa0", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _canonical_amount_label(label: str) -> str | None:
    key = _norm_label(label).lower()
    if not key:
        return None
    if key in _LABEL_ALIASES:
        return _LABEL_ALIASES[key]
    # One-time Setup Fee: GBP 200 per client
    if key.startswith("one-time setup fee") or key.startswith("setup fee"):
        return "Setup Fees"
    return None


def _pick_employee_sheet(wb) -> Any:
    skip = {"exchange rate", "excange rate", "fx", "rates"}
    for name in wb.sheetnames:
        if name.strip().lower() in skip:
            continue
        return wb[name]
    return wb[wb.sheetnames[0]] if wb.sheetnames else None


def _looks_like_uk_l_workbook(path: Path) -> bool:
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return any(str(n).strip().upper().startswith("UK-L") for n in wb.sheetnames)
    finally:
        wb.close()


def _looks_like_topsource_invoice_excel(path: Path) -> bool:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _pick_employee_sheet(wb)
        if ws is None:
            return False
        head = " ".join(
            _norm_label(ws.cell(r, 1).value).lower() for r in range(1, 5)
        )
        return "topsource" in head or "salary calculation" in head
    finally:
        wb.close()


def parse_topsource_uk_excel(excel_path: Path) -> TopSourceUkParsed:
    """解析 TopSource NN Road 一人一表 Excel 发票。"""
    out = TopSourceUkParsed(source_kind="excel", currency="GBP")
    wb = load_workbook(excel_path, data_only=True)
    try:
        ws = _pick_employee_sheet(wb)
        if ws is None:
            out.warnings.append("Excel 无可用工作表")
            return out

        a1 = _norm_label(ws["A1"].value).lower()
        if "topsource" not in a1 and "topsource" not in _norm_label(ws["A2"].value).lower():
            out.warnings.append("表头未出现 TopSource，可能不是本版式 Excel")

        title = _norm_label(ws["A3"].value)
        out.title = title or None
        m = _TITLE_RE.match(title) if title else None
        if m:
            out.employee_name = m.group(1).strip(" -")
        period = ws["B3"].value
        if period is not None:
            out.period_label = _norm_label(period)

        max_row = min(ws.max_row or 40, 80)
        max_col = min(ws.max_column or 8, 10)
        for r in range(1, max_row + 1):
            label = _norm_label(ws.cell(r, 1).value)
            canon = _canonical_amount_label(label)
            if canon:
                num = _as_float(ws.cell(r, 2).value)
                if num is not None:
                    out.amounts[canon] = num
            # Exchange Rate 表头常见于 C/D
            for c in range(1, max_col + 1):
                cell_l = _norm_label(ws.cell(r, c).value).lower()
                if cell_l == "exchange rate":
                    for rr in range(r, min(r + 3, max_row + 1)):
                        fx = _as_float(ws.cell(rr, c).value)
                        if fx is None:
                            fx = _as_float(ws.cell(rr, c + 1).value) if c + 1 <= max_col else None
                        if fx is not None and fx > 0:
                            out.fx_usd_per_gbp = fx
                            break

            low = label.lower()
            if "ts margin" in low or "service charge" in low:
                margin = _as_float(ws.cell(r, 2).value)
                if margin is not None:
                    out.warnings.append(
                        f"发票含 TS Margin/服务费 GBP={margin}（与 PN Management Fee 不同，未自动写入）"
                    )

        if not out.employee_name:
            # 文件名：NN Road - Adrian Grace - June 2026 Invoice.xlsx
            stem = excel_path.stem
            m = re.search(r"NN\s*Road\s*-\s*(.+?)\s*-\s*", stem, re.I)
            if m:
                out.employee_name = m.group(1).strip()
            else:
                out.employee_name = ws.title.strip() or None
                out.warnings.append("未从 A3 解析到姓名，已用 sheet/文件名兜底")

        if not out.amounts:
            out.warnings.append("未从 Excel 解析到任何 GBP 明细金额")
    finally:
        wb.close()
    return out


def fill_uk_l_employee(ws, parsed: TopSourceUkParsed, *, clear_amounts: bool = True) -> None:
    if parsed.title:
        ws["A3"] = parsed.title
    else:
        name = parsed.employee_name or "Employee"
        fy = _uk_tax_year_label(parsed.invoice_date) if parsed.invoice_date else "YY-YY"
        ws["A3"] = f"{name} Salary Calculation for FY {fy}"

    if clear_amounts:
        for addr in _UK_L_AMOUNT_CELLS.values():
            ws[addr] = 0.0

    for label, addr in _UK_L_AMOUNT_CELLS.items():
        if label in parsed.amounts:
            ws[addr] = float(parsed.amounts[label])


def apply_parsed_list(
    wb,
    parsed_list: list[TopSourceUkParsed],
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
) -> tuple[list[str], float | None, PnMeta | None]:
    warnings: list[str] = []
    for p in parsed_list:
        warnings.extend(p.warnings)

    n = len(parsed_list)
    sheet_names = ensure_uk_l_count(wb, n)
    expand_uk_employee_rows(wb, n, sheet_names)

    for i, parsed in enumerate(parsed_list):
        fill_uk_l_employee(wb[sheet_names[i]], parsed, clear_amounts=True)
        if UK_SHEET in wb.sheetnames:
            wb[UK_SHEET].cell(9 + i, 2).value = parsed.employee_name or f"Employee {i + 1}"

    fx_rate = None
    if fill_fx:
        try:
            src_fx = next((p.fx_usd_per_gbp for p in parsed_list if p.fx_usd_per_gbp), None)
            if src_fx and src_fx > 0:
                fx_rate = 1.0 / float(src_fx)
            else:
                rates = fetch_usd_rates()
                fx_rate = get_uk_gbp_per_usd(rates)
            for name in sheet_names:
                wb[name]["D24"] = fx_rate
        except Exception as exc:
            warnings.append(f"写入 UK-L!D24 汇率失败: {exc}")

    applied_pn = None
    if pn_meta is not None:
        applied_pn = apply_pn_meta(
            wb,
            pn_meta,
            registry_dir=registry_dir,
            reserve_invoice_number=True,
        )
    return warnings, fx_rate, applied_pn


def convert_pdfs(
    pdf_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    paths = [Path(p).resolve() for p in pdf_paths]
    if not paths:
        raise ValueError("未提供 PDF")
    for p in paths:
        if not p.is_file():
            raise FileNotFoundError(f"PDF 不存在: {p}")

    output_path = Path(output_path).resolve()
    tpl = (template_path or get_region_template("UK")).resolve()
    if not tpl.is_file():
        raise FileNotFoundError(f"UK 母版不存在: {tpl}")

    parsed_list = [parse_topsource_uk_pdf(p) for p in paths]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tpl, output_path)

    wb = load_workbook(output_path)
    warnings, fx_rate, applied_pn = apply_parsed_list(
        wb,
        parsed_list,
        pn_meta=pn_meta,
        registry_dir=registry_dir or output_path.parent,
        fill_fx=fill_fx,
    )
    wb.save(output_path)
    wb.close()

    return {
        "ok": True,
        "profile_id": "topsource_uk",
        "region": "UK",
        "source_kind": "pdf",
        "output": str(output_path),
        "employee_count": len(parsed_list),
        "parsed": [p.to_dict() for p in parsed_list],
        "warnings": warnings,
        "fx_rate": fx_rate,
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
    }


def convert_excels(
    excel_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    """多份 TopSource 一人一表 Excel → 一份含 UK-L / UK-L(2)… 的源表。"""
    paths = [Path(p).resolve() for p in excel_paths]
    if not paths:
        raise ValueError("未提供 Excel")
    for p in paths:
        if not p.is_file():
            raise FileNotFoundError(f"Excel 不存在: {p}")

    output_path = Path(output_path).resolve()
    # 已是 UK-L 源表：单文件直接拷贝；多文件暂不支持合并已成型 PN
    uk_l_flags = [_looks_like_uk_l_workbook(p) for p in paths]
    if all(uk_l_flags):
        if len(paths) > 1:
            raise ValueError("多份已是 UK-L 的 Excel 无法自动合并，请只传一份或传供应商原始发票")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(paths[0], output_path)
        return {
            "ok": True,
            "profile_id": "topsource_uk",
            "region": "UK",
            "source_kind": "excel_uk_l",
            "output": str(output_path),
            "employee_count": None,
            "parsed": [],
            "warnings": ["源表已是 UK-L，已原样用作转换输入"],
            "fx_rate": None,
            "pn_meta": None,
        }

    tpl = (template_path or get_region_template("UK")).resolve()
    if not tpl.is_file():
        raise FileNotFoundError(f"UK 母版不存在: {tpl}")

    parsed_list = [parse_topsource_uk_excel(p) for p in paths]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tpl, output_path)

    wb = load_workbook(output_path)
    warnings, fx_rate, applied_pn = apply_parsed_list(
        wb,
        parsed_list,
        pn_meta=pn_meta,
        registry_dir=registry_dir or output_path.parent,
        fill_fx=fill_fx,
    )
    wb.save(output_path)
    wb.close()

    return {
        "ok": True,
        "profile_id": "topsource_uk",
        "region": "UK",
        "source_kind": "excel",
        "output": str(output_path),
        "employee_count": len(parsed_list),
        "parsed": [p.to_dict() for p in parsed_list],
        "warnings": warnings,
        "fx_rate": fx_rate,
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
    }


def convert_sources(
    source_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    """按扩展名自动分流：PDF → convert_pdfs；Excel → convert_excels。"""
    paths = [Path(p).resolve() for p in source_paths]
    if not paths:
        raise ValueError("未提供源文件")
    pdfs = [p for p in paths if p.suffix.lower() == ".pdf"]
    excels = [p for p in paths if p.suffix.lower() in (".xlsx", ".xlsm", ".xls")]
    other = [p for p in paths if p not in pdfs and p not in excels]
    if other:
        raise ValueError(f"不支持的文件类型: {[p.name for p in other]}")
    if pdfs and excels:
        raise ValueError("同一批次请不要混传 PDF 与 Excel，请只传其中一类")
    if pdfs:
        return convert_pdfs(
            pdfs,
            output_path,
            template_path=template_path,
            pn_meta=pn_meta,
            registry_dir=registry_dir,
            fill_fx=fill_fx,
        )
    return convert_excels(
        excels,
        output_path,
        template_path=template_path,
        pn_meta=pn_meta,
        registry_dir=registry_dir,
        fill_fx=fill_fx,
    )


def convert_pdf(
    pdf_path: Path,
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    return convert_pdfs(
        [pdf_path],
        output_path,
        template_path=template_path,
        pn_meta=pn_meta,
        registry_dir=registry_dir,
        fill_fx=fill_fx,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="TopSource UK 发票 PDF/Excel → UK PN 源表")
    parser.add_argument("sources", nargs="+", type=Path, help="一份或多份 PDF/Excel（一人一张）")
    parser.add_argument("-o", "--output", type=Path, help="输出 xlsx")
    parser.add_argument("-t", "--template", type=Path, help="UK PN 母版")
    args = parser.parse_args(argv)

    sources = [p.resolve() for p in args.sources]
    out = (
        args.output.resolve()
        if args.output
        else sources[0].parent / f"UK_from_topsource_{sources[0].stem}.xlsx"
    )
    try:
        result = convert_sources(sources, out, template_path=args.template)
    except Exception as exc:
        print(f"失败: {exc}", file=sys.stderr)
        return 1

    print("完成")
    print(f"  输出: {result['output']}")
    print(f"  类型: {result.get('source_kind')}")
    print(f"  人数: {result.get('employee_count')}")
    print(f"  汇率D24: {result.get('fx_rate')}")
    for i, p in enumerate(result.get("parsed") or []):
        am = p.get("amounts") or {}
        er_nic = am.get("ER' NIC")
        print(
            f"  [{i + 1}] {p.get('employee_name')}  gross={am.get('Gross Salary')}  "
            f"erNIC={er_nic}  appLevy={am.get('App Levy')}"
        )
    for w in result.get("warnings") or []:
        print(f"  ! {w}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
