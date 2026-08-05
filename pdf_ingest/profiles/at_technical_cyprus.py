# -*- coding: utf-8 -*-
"""
A&T Professional Technical Management Services（Cyprus）
Invoice PDF + Payroll Calculation PDF → Cyprus-L（profile: at_technical_cyprus）

两份 PDF 一起上传：
  - Invoice：姓名（First Last）、账期、Gross、ER Contributions、Public Liability、Administration Fee
  - Payroll Calculation：同人明细（Last First）、EE Social Ins / Tax / N.H.S.、ER Contributions 校验

Recurring Fee 不写死：走 mapping.cyprusRecurringFee（引擎写 Cyprus!I）。
"""
from __future__ import annotations

import argparse
import calendar
import re
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bill_convert.person import compact_person_name, score_person_name_match
from pdf_ingest.text_extract import extract_pdf_text
from pn_meta import PnMeta
from region_templates import get_region_template

CYPRUS_L_SHEET = "Cyprus-L"

_MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "")
    if not text:
        return None
    if re.search(r",\d{2}$", text) and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(",") == 1 and "." not in text:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _period_bounds(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    return start, end


def _westernize_name(name: str) -> str:
    """Sidorov Anatoly → Anatoly Sidorov（两词时）；已是 First Last 则保持。"""
    parts = [p for p in re.split(r"\s+", (name or "").strip()) if p]
    if len(parts) == 2:
        # Invoice 用 First Last；Payroll 用 Last First — 无法绝对判断，合并时以 invoice 为准
        return f"{parts[0]} {parts[1]}"
    return " ".join(parts)


def classify_at_pdf(text: str, path: Path | None = None) -> str:
    """返回 'invoice' | 'payroll' | 'unknown'。"""
    low = (text or "").lower()
    name = (path.name if path else "").lower()
    if "payroll calculation" in low or "payroll type:" in low or "empl.id:" in low:
        return "payroll"
    if "invoice" in low or "total due" in low or "services fee calculation" in low:
        return "invoice"
    if "invoice" in name:
        return "invoice"
    if "payroll" in name:
        return "payroll"
    return "unknown"


def parse_at_invoice_pdf(pdf_path: Path) -> dict[str, Any]:
    path = Path(pdf_path).resolve()
    text = extract_pdf_text(path)
    warnings: list[str] = []

    inv_m = re.search(r"Invoice\s+(AT\d+)", text, flags=re.I)
    invoice_no = inv_m.group(1).strip() if inv_m else None

    date_m = re.search(r"DATE\s+(\d{1,2}/\d{1,2}/\d{4})", text, flags=re.I)
    invoice_date = None
    if date_m:
        try:
            invoice_date = datetime.strptime(date_m.group(1), "%d/%m/%Y").date().isoformat()
        except ValueError:
            warnings.append(f"无法解析发票日期: {date_m.group(1)}")

    period_label = None
    year = month = None
    pm = re.search(
        r"Monthly cost:\s*(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})",
        text,
        flags=re.I,
    )
    if pm:
        month = _MONTHS.get(pm.group(1).lower())
        year = int(pm.group(2))
        period_label = f"{pm.group(1)}-{pm.group(2)}"

    employees: list[dict[str, Any]] = []
    block_re = re.compile(
        r"Monthly cost:\s*(?P<month>January|February|March|April|May|June|July|August|September|"
        r"October|November|December)\s+(?P<year>\d{4})\s*-\s*(?P<name>[^\n]+?)\s+"
        r"Gross Salary\s+(?P<gross>[\d,]+\.?\d*)\s*"
        r"Employer's Contributions\s+(?P<er>[\d,]+\.?\d*)\s*"
        r"Employer's\s*&\s*Public Liability\s+(?P<liab>[\d,]+\.?\d*)\s*"
        r"Administration Fee\s+(?P<admin>[\d,]+\.?\d*)",
        flags=re.I | re.S,
    )
    for m in block_re.finditer(text):
        name = re.sub(r"\s+", " ", m.group("name")).strip()
        name = re.sub(r"\s+Gross Salary.*$", "", name, flags=re.I).strip()
        y = int(m.group("year"))
        mo = _MONTHS.get(m.group("month").lower())
        if year is None:
            year, month = y, mo
            period_label = f"{m.group('month')}-{y}"
        employees.append(
            {
                "Employee Name": name,
                "Name of Employee": name,
                "Base salary": _as_float(m.group("gross")),
                "Employer's contributions": _as_float(m.group("er")),
                "Employer's & Public Liability": _as_float(m.group("liab")),
                "_admin_fee": _as_float(m.group("admin")),
                "_source": "invoice",
            }
        )

    if not employees:
        raise ValueError(f"A&T Invoice 未解析到员工块: {path.name}")

    start = end = None
    if year and month:
        start, end = _period_bounds(year, month)
        for e in employees:
            e["_period_from"] = start
            e["_period_to"] = end
            e["_period_label"] = period_label
            e["From"] = start
            e["To"] = end

    return {
        "kind": "invoice",
        "employees": employees,
        "invoice_no": invoice_no,
        "invoice_date": invoice_date,
        "period_label": period_label,
        "period_from": start,
        "period_to": end,
        "warnings": warnings,
        "source_file": path.name,
        "text": text,
    }


def parse_at_payroll_pdf(pdf_path: Path) -> dict[str, Any]:
    path = Path(pdf_path).resolve()
    text = extract_pdf_text(path)
    warnings: list[str] = []

    year = month = None
    # Post Month: 6/2026
    pm = re.search(r"(\d{1,2})\s*/\s*(\d{4})\s*Post Month", text, flags=re.I)
    if not pm:
        pm = re.search(r"Post Month:\s*(\d{1,2})\s*/\s*(\d{4})", text, flags=re.I)
    if not pm:
        # text order sometimes "6/2026Post Month:"
        pm = re.search(r"(\d{1,2})/(\d{4})\s*Post Month", text, flags=re.I)
    if pm:
        month, year = int(pm.group(1)), int(pm.group(2))

    employees: list[dict[str, Any]] = []
    # Split by employee blocks
    parts = re.split(r"Empl\.ID:\s*", text, flags=re.I)
    for part in parts[1:]:
        head = re.match(
            r"(?P<eid>\d+)\s*-\s*(?P<name>[^\n]+?)(?:\s{2,}|\s+Empl\.Date:)",
            part,
            flags=re.I,
        )
        if not head:
            warnings.append("Payroll 某员工块无法解析姓名")
            continue
        raw_name = re.sub(r"\s+", " ", head.group("name")).strip()
        # Last First → keep for matching; westernize later when merging with invoice
        tokens = raw_name.split()
        if len(tokens) == 2:
            display = f"{tokens[1]} {tokens[0]}"
        else:
            display = raw_name

        basic_m = re.search(r"Basic Salary:\s*([\d.]+,\d{2}|\d[\d,]*\.?\d*)", part, flags=re.I)
        base = _as_float(basic_m.group(1)) if basic_m else None

        # Deductions：Social Ins / Tax-1 / N.H.S.-SI（Notice 行上的 EE NHS，勿取 Contributions 侧）
        si_m = re.search(r"(?:^|\n)[^\n]*?\bBasic\b[^\n]*?\bSocial Ins\s+([\d.]+,\d{2}|\d[\d.,]*)", part, flags=re.I)
        if not si_m:
            si_m = re.search(r"Social Ins\s+([\d.]+,\d{2}|\d[\d.,]*)", part, flags=re.I)
        tax_m = re.search(r"Tax-1\s+([\d.]+,\d{2}|\d[\d.,]*)", part, flags=re.I)
        nhs_m = re.search(
            r"Notice\s+[\d.,]+\s+[\d.,]+\s+N\.H\.S\.-SI\s+([\d.]+,\d{2}|\d[\d.,]*)",
            part,
            flags=re.I,
        )
        if not nhs_m:
            # 回退：取较小的那个 NHS（EE 通常小于 ER）
            nhs_vals = [_as_float(x) for x in re.findall(r"N\.H\.S\.-SI\s+([\d.]+,\d{2}|\d[\d.,]*)", part, flags=re.I)]
            nhs_vals = [x for x in nhs_vals if x is not None]
            nhs = min(nhs_vals) if nhs_vals else None
        else:
            nhs = _as_float(nhs_m.group(1))

        # ER contributions：Ear.+Con. 块末行第三个数
        er = None
        er_m = re.search(
            r"Ear\.\+Con\.\s*[\d.,]+\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)",
            part,
            flags=re.I | re.S,
        )
        if er_m:
            er = _as_float(er_m.group(3))
        else:
            triples = re.findall(r"([\d.]+,\d{2})\s+([\d.]+,\d{2})\s+([\d.]+,\d{2})", part)
            if triples:
                er = _as_float(triples[-1][2])

        ee_si = _as_float(si_m.group(1)) if si_m else None
        ee_tax = _as_float(tax_m.group(1)) if tax_m else None

        employees.append(
            {
                "Employee Name": display,
                "Name of Employee": display,
                "_payroll_name": raw_name,
                "_empl_id": head.group("eid"),
                "Base salary": base,
                "Employer's contributions": er,
                "Employee's Social Insurance": -abs(ee_si) if ee_si is not None else None,
                "Employee's tax": -abs(ee_tax) if ee_tax is not None else None,
                "Employee - N.H.S.-SI": -abs(nhs) if nhs is not None else None,
                "_source": "payroll",
            }
        )

    if not employees:
        raise ValueError(f"A&T Payroll 未解析到员工: {path.name}")

    start = end = None
    period_label = None
    if year and month:
        start, end = _period_bounds(year, month)
        period_label = f"{year}-{month:02d}"
        for e in employees:
            e["_period_from"] = start
            e["_period_to"] = end
            e["_period_label"] = period_label
            e["From"] = start
            e["To"] = end

    return {
        "kind": "payroll",
        "employees": employees,
        "period_label": period_label,
        "period_from": start,
        "period_to": end,
        "warnings": warnings,
        "source_file": path.name,
        "text": text,
    }


def _match_emp(name: str, pool: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not name or not pool:
        return None
    best = None
    best_score = 0
    compact = compact_person_name(name)
    for e in pool:
        candidates = [
            e.get("Employee Name"),
            e.get("Name of Employee"),
            e.get("_payroll_name"),
        ]
        for c in candidates:
            if not c:
                continue
            if compact_person_name(c) == compact:
                return e
            score = score_person_name_match(name, str(c))
            if score > best_score:
                best_score = score
                best = e
    return best if best_score >= 70 else None


def merge_invoice_and_payroll(
    invoice: dict[str, Any] | None,
    payroll: dict[str, Any] | None,
) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if not invoice and not payroll:
        raise ValueError("未提供可用的 Invoice / Payroll PDF 解析结果")

    if invoice and not payroll:
        warnings.append("仅有 Invoice：缺少 EE Social Ins / Tax / N.H.S.，对应列置空")
        return list(invoice["employees"]), warnings
    if payroll and not invoice:
        warnings.append("仅有 Payroll：缺少 Public Liability，对应列置 0")
        out = []
        for e in payroll["employees"]:
            row = dict(e)
            row.setdefault("Employer's & Public Liability", 0.0)
            out.append(row)
        return out, warnings

    assert invoice is not None and payroll is not None
    inv_emps = list(invoice["employees"])
    pay_emps = list(payroll["employees"])
    used: set[int] = set()
    merged: list[dict[str, Any]] = []

    for inv in inv_emps:
        name = str(inv.get("Employee Name") or "")
        hit = _match_emp(name, pay_emps)
        row = dict(inv)
        if hit:
            used.add(id(hit))
            for k in (
                "Employee's Social Insurance",
                "Employee's tax",
                "Employee - N.H.S.-SI",
                "_empl_id",
                "_payroll_name",
            ):
                if hit.get(k) is not None:
                    row[k] = hit[k]
            # ER contrib：两边都有则优先 invoice（与税票一致），差异告警
            inv_er = inv.get("Employer's contributions")
            pay_er = hit.get("Employer's contributions")
            if inv_er is not None and pay_er is not None and abs(float(inv_er) - float(pay_er)) > 0.05:
                warnings.append(
                    f"{name}：Invoice ER Contributions {inv_er} ≠ Payroll {pay_er}，已用 Invoice"
                )
        else:
            warnings.append(f"{name}：Payroll 中未匹配到同名员工，EE 扣款列为空")
        # period：优先 invoice
        if invoice.get("period_from"):
            row["_period_from"] = invoice["period_from"]
            row["_period_to"] = invoice["period_to"]
            row["From"] = invoice["period_from"]
            row["To"] = invoice["period_to"]
            row["_period_label"] = invoice.get("period_label")
        merged.append(row)

    for pay in pay_emps:
        if id(pay) in used:
            continue
        warnings.append(
            f"{pay.get('Employee Name')}：仅出现在 Payroll，已追加（Public Liability=0）"
        )
        row = dict(pay)
        row.setdefault("Employer's & Public Liability", 0.0)
        if invoice.get("period_from"):
            row["_period_from"] = invoice["period_from"]
            row["_period_to"] = invoice["period_to"]
            row["From"] = invoice["period_from"]
            row["To"] = invoice["period_to"]
        merged.append(row)

    warnings.extend(invoice.get("warnings") or [])
    warnings.extend(payroll.get("warnings") or [])
    return merged, warnings


def convert_sources(
    source_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    del pn_meta, registry_dir, fill_fx
    from convert_mapping import resolve_convert_mapping
    from profiles.cyprus_payroll_calc.convert import set_recurring_fees, write_cyprus_l

    mapping_in = dict(convert_mapping) if isinstance(convert_mapping, dict) else {}
    mapping_in.setdefault("pdfProfileId", "at_technical_cyprus")
    mapping = resolve_convert_mapping("cyprus_payroll_calc", mapping_in)

    paths = [Path(p).resolve() for p in source_paths]
    if not paths:
        raise ValueError("未提供 PDF")
    pdfs = [p for p in paths if p.suffix.lower() == ".pdf"]
    if not pdfs:
        raise ValueError("at_technical_cyprus 需要 PDF（Invoice + Payroll Calculation）")

    invoice = payroll = None
    warnings: list[str] = []
    for p in pdfs:
        text = extract_pdf_text(p)
        kind = classify_at_pdf(text, p)
        if kind == "invoice":
            if invoice:
                warnings.append(f"重复 Invoice，忽略: {p.name}")
                continue
            invoice = parse_at_invoice_pdf(p)
        elif kind == "payroll":
            if payroll:
                warnings.append(f"重复 Payroll，忽略: {p.name}")
                continue
            payroll = parse_at_payroll_pdf(p)
        else:
            warnings.append(f"无法识别 PDF 类型，尝试按内容解析: {p.name}")
            if "empl.id:" in text.lower():
                payroll = parse_at_payroll_pdf(p)
            else:
                invoice = parse_at_invoice_pdf(p)

    if not invoice and not payroll:
        raise ValueError("未能解析任何 A&T Invoice / Payroll PDF")
    if not invoice or not payroll:
        warnings.append(
            "建议同时上传 Invoice 与 Payroll Calculation；当前缺一份，已尽力合并"
        )

    employees, merge_warnings = merge_invoice_and_payroll(invoice, payroll)
    warnings.extend(merge_warnings)

    tpl = (template_path or get_region_template("Cyprus")).resolve()
    if not tpl.is_file():
        raise FileNotFoundError(f"Cyprus 母版不存在: {tpl}")

    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tpl, output_path)
    wb = load_workbook(output_path)
    try:
        if CYPRUS_L_SHEET not in wb.sheetnames:
            raise ValueError(f"母版缺少 {CYPRUS_L_SHEET}")
        import profiles.cyprus_payroll_calc.convert as cyprus_mod

        prev = cyprus_mod._ACTIVE_MAPPING
        cyprus_mod._ACTIVE_MAPPING = mapping
        try:
            write_cyprus_l(wb[CYPRUS_L_SHEET], employees)
            set_recurring_fees(wb, employees)
        finally:
            cyprus_mod._ACTIVE_MAPPING = prev
        wb.save(output_path)
    finally:
        wb.close()

    return {
        "ok": True,
        "profile_id": "at_technical_cyprus",
        "region": "Cyprus",
        "source_kind": "pdf",
        "output": str(output_path),
        "employee_count": len(employees),
        "parsed": [
            {
                "employee_name": e.get("Employee Name"),
                "base_salary": e.get("Base salary"),
                "er_contributions": e.get("Employer's contributions"),
                "liability": e.get("Employer's & Public Liability"),
            }
            for e in employees
        ],
        "warnings": warnings,
        "fx_rate": None,
        "pn_meta": None,
        "invoice_no": (invoice or {}).get("invoice_no"),
        "period_label": (invoice or payroll or {}).get("period_label"),
    }


def convert_pdfs(
    pdf_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return convert_sources(
        pdf_paths,
        output_path,
        template_path=template_path,
        pn_meta=pn_meta,
        registry_dir=registry_dir,
        fill_fx=fill_fx,
        convert_mapping=convert_mapping,
    )


def convert_pdf(
    pdf_path: Path,
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return convert_pdfs(
        [pdf_path],
        output_path,
        template_path=template_path,
        pn_meta=pn_meta,
        registry_dir=registry_dir,
        fill_fx=fill_fx,
        convert_mapping=convert_mapping,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="A&T Cyprus PDF → Cyprus-L")
    parser.add_argument("sources", nargs="+", type=Path)
    parser.add_argument("-o", "--output", type=Path, required=True)
    parser.add_argument("-t", "--template", type=Path, default=None)
    args = parser.parse_args(argv)
    result = convert_sources(args.sources, args.output, template_path=args.template)
    print(result)
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
