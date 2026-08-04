# -*- coding: utf-8 -*-
"""
Panda Work Global（Pakistan）季度发票 PDF → Pakistan-L。

一人一票；账期为季度（如 Q3-2026 Jul–Sep）。
合并多 PDF 后按人拆成 3 个月行（Base = 季度 PKR / 3）。
"""
from __future__ import annotations

import argparse
import calendar
import re
import shutil
import sys
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from bill_convert.pakistan_business_tax import (
    BT_FEDERAL_HEADER,
    BT_SINDH_HEADER,
    apply_derived_coeffs_to_rows,
    parse_pakistan_business_tax_cfg,
)
from bill_convert.pakistan_employee_fees import apply_fees_to_employee_rows, parse_pakistan_employee_fees
from convert_mapping import resolve_convert_mapping
from pdf_ingest.text_extract import extract_pdf_text
from pn_meta import PnMeta

PK_L_SHEET = "Pakistan-L"
HEADER_ROW = 7
DATA_START = 8

_MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}

_HEADERS = [
    "No. of EE",
    "Name of Employee",
    "Base Salary",
    "Salary Adjusment",
    "OT",
    "Monthly Bonus",
    "Other Bonus",
    "Leave Payment",
    "Other ",
    "Gross Salary",
    "Expense Reimbursment",
    "E.O.B.I",
    "IT",
    "Net Salary",
    "Social Benefit",
    "Medical Insurance",
    "Bank Charges",
    "Total Cost",
]


@dataclass
class PandaWorkPkParsed:
    supplier: str = "Panda Work Global (Private) Limited"
    invoice_no: str | None = None
    invoice_date: date | None = None
    employee_name: str | None = None
    client_label: str | None = None
    period_from: date | None = None
    period_to: date | None = None
    salary_usd: float | None = None
    salary_pkr: float | None = None
    management_fee_usd: float | None = None
    management_fee_pkr: float | None = None
    # 季度总额（发票原值，非月）
    bank_fee_usd: float | None = None
    federal_it_usd: float | None = None
    sindh_sales_tax_usd: float | None = None
    fx_rate: float | None = None
    currency: str = "PKR"
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        for key in ("invoice_date", "period_from", "period_to"):
            val = getattr(self, key)
            if isinstance(val, date):
                d[key] = val.isoformat()
        return d


def looks_like_panda_work_pk(path: Path, text: str | None = None) -> bool:
    raw = (text if text is not None else extract_pdf_text(path)).lower()
    return any(
        k in raw
        for k in (
            "panda work global",
            "pandaworkglobal.com",
            "panda work global (private) limited",
        )
    )


def _money(s: str) -> float | None:
    s = (s or "").strip().replace(",", "").replace("\xa0", "").replace("$", "").replace("PKR", "")
    s = s.replace("USD", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_invoice_date(raw: str) -> date | None:
    raw = (raw or "").strip()
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_quarter_period(text: str) -> tuple[date | None, date | None]:
    # Q3-2026 - (July 2026-September 2026)
    m = re.search(
        r"Q\d\s*-?\s*(\d{4})\s*[-–]\s*\(\s*([A-Za-z]+)\s+(\d{4})\s*[-–]\s*([A-Za-z]+)\s+(\d{4})\s*\)",
        text,
        re.I,
    )
    if not m:
        m = re.search(
            r"\(\s*([A-Za-z]+)\s+(\d{4})\s*[-–]\s*([A-Za-z]+)\s+(\d{4})\s*\)",
            text,
            re.I,
        )
        if not m:
            return None, None
        m1, y1, m2, y2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    else:
        m1, y1, m2, y2 = m.group(2), int(m.group(3)), m.group(4), int(m.group(5))
    mo1 = _MONTHS.get(m1.lower())
    mo2 = _MONTHS.get(m2.lower())
    if not mo1 or not mo2:
        return None, None
    last = calendar.monthrange(y2, mo2)[1]
    return date(y1, mo1, 1), date(y2, mo2, last)


def parse_panda_work_pk_text(text: str) -> PandaWorkPkParsed:
    out = PandaWorkPkParsed()
    t = text.replace("\r", "\n").replace("\xa0", " ")

    m = re.search(r"INVOICE\s*NUMBER\s*:\s*([A-Z0-9\-]+)", t, re.I)
    if m:
        out.invoice_no = m.group(1).strip()

    m = re.search(r"INVOICE\s*DATE\s*:\s*([A-Za-z0-9,\s/]+?)(?:\s+INVOICE\s*DUE|\s+ADDRESS|\n)", t, re.I)
    if m:
        out.invoice_date = _parse_invoice_date(m.group(1).strip())
        if out.invoice_date is None:
            out.warnings.append(f"无法解析发票日期: {m.group(1).strip()}")

    m = re.search(
        r"(?:Consultancy\s+Fee|Salaries)\s*[-–]\s*(?:Mr\.|Ms\.|Mrs\.)?\s*([A-Za-z][A-Za-z. ]+)",
        t,
        re.I,
    )
    if m:
        name = re.sub(r"\s+", " ", m.group(1)).strip(" .-")
        # 防止跨行吞进供应商名
        for stop in ("PANDA", "WORK", "GLOBAL", "PRIVATE", "LIMITED", "INVOICE"):
            idx = name.upper().find(stop)
            if idx > 0:
                name = name[:idx].strip(" .-")
                break
        out.employee_name = name or None

    m = re.search(r"\(\s*([^)]+FZCo[^)]*)\)", t, re.I)
    if m:
        out.client_label = re.sub(r"\s+", " ", m.group(1)).strip()
    else:
        m = re.search(r"\(\s*(Danfoss[^)]*)\)", t, re.I)
        if m:
            out.client_label = re.sub(r"\s+", " ", m.group(1)).strip()

    out.period_from, out.period_to = _parse_quarter_period(t)
    if out.period_from is None:
        out.warnings.append("未解析到季度账期")

    # 首行大额 USD+PKR 为工资（Consultancy/Salaries）
    money_pairs = re.findall(
        r"([0-9,]+\.?[0-9]*)\s*\$\s+([0-9,]+\.?[0-9]*)\s*PKR",
        t,
        re.I,
    )
    if money_pairs:
        usd0 = _money(money_pairs[0][0])
        pkr0 = _money(money_pairs[0][1])
        out.salary_usd = usd0
        out.salary_pkr = pkr0
        if usd0 and pkr0 and usd0 > 0:
            out.fx_rate = round(pkr0 / usd0, 6)

    m = re.search(
        r"PWG\s+Management\s+Fee\s+([0-9,]+\.?[0-9]*)\s*\$\s+([0-9,]+\.?[0-9]*)\s*PKR",
        t,
        re.I,
    )
    if m:
        out.management_fee_usd = _money(m.group(1))
        out.management_fee_pkr = _money(m.group(2))
        # 管理费之后金额行（文本与标签常拆列）：Subtotal / Bank Fee / Federal IT / Sindh / Fee TOTAL
        tail = t[m.end() :]
        cut = re.search(
            r"Bank\s+Details|Bank\s+Name|Amount in words|DESCRIPTION",
            tail,
            re.I,
        )
        if cut:
            tail = tail[: cut.start()]
        fee_pairs = re.findall(
            r"([0-9,]+\.?[0-9]*)\s*\$\s+([0-9,]+\.?[0-9]*)\s*PKR",
            tail,
            re.I,
        )
        fee_usds = [u for u in (_money(a) for a, _ in fee_pairs) if u is not None]
        # 期望：0=小计, 1=Bank Fee, 2=Federal IT, 3=Sindh Sales Tax, …
        if len(fee_usds) >= 4:
            out.bank_fee_usd = fee_usds[1]
            out.federal_it_usd = fee_usds[2]
            out.sindh_sales_tax_usd = fee_usds[3]
        elif len(fee_usds) >= 3:
            # 偶发无小计：Bank / Federal / Sindh
            out.bank_fee_usd = fee_usds[0]
            out.federal_it_usd = fee_usds[1]
            out.sindh_sales_tax_usd = fee_usds[2]
        else:
            out.warnings.append("未从 PDF 解析到 Federal IT / Sindh Sales Tax 金额行")

    if not out.employee_name:
        out.warnings.append("未解析到员工姓名")
    if out.salary_pkr is None:
        out.warnings.append("未解析到季度薪资 PKR")
    return out


def parse_panda_work_pk_pdf(path: Path) -> PandaWorkPkParsed:
    text = extract_pdf_text(path)
    if not looks_like_panda_work_pk(path, text):
        raise ValueError(f"不是 Panda Work Global Pakistan 发票: {path.name}")
    parsed = parse_panda_work_pk_text(text)
    return parsed


def _split_months(n: int = 3) -> int:
    return max(int(n or 3), 1)


def employees_from_parsed(
    parsed_list: list[PandaWorkPkParsed],
    *,
    split_months: int = 3,
) -> tuple[list[dict[str, Any]], list[str], float | None]:
    warnings: list[str] = []
    months = _split_months(split_months)
    rows: list[dict[str, Any]] = []
    fx: float | None = None
    for idx, p in enumerate(parsed_list):
        warnings.extend(p.warnings)
        if p.fx_rate and (fx is None):
            fx = p.fx_rate
        name = (p.employee_name or "").strip() or f"Employee {idx + 1}"
        total = float(p.salary_pkr or 0)
        monthly = round(total / months, 6) if total else None
        if monthly is None:
            warnings.append(f"{name}：无季度薪资，未写入 Base Salary")
        for _ in range(months):
            rows.append(
                {
                    "No. of EE": idx + 1,
                    "Name of Employee": name,
                    "Base Salary": monthly,
                    "From": p.period_from,
                    "To": p.period_to,
                    "Client": p.client_label,
                    "_invoice_no": p.invoice_no,
                    "_salary_usd": p.salary_usd,
                    "_management_fee_usd": p.management_fee_usd,
                    "_federal_it_usd": p.federal_it_usd,
                    "_sindh_sales_tax_usd": p.sindh_sales_tax_usd,
                    "_bank_fee_usd": p.bank_fee_usd,
                }
            )
    return rows, warnings, fx


def _write_pakistan_l_workbook(
    output_path: Path,
    employees: list[dict[str, Any]],
    *,
    fx_rate: float | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = PK_L_SHEET

    ws["A1"] = "Company name:"
    client = ""
    for emp in employees:
        client = str(emp.get("Client") or "").strip()
        if client:
            break
    ws["C1"] = client or None

    ws["A2"] = "Payroll period:"
    ws["D2"] = "to"
    if employees:
        ws["C2"] = employees[0].get("From")
        ws["E2"] = employees[0].get("To")
        if hasattr(ws["C2"].value, "strftime"):
            ws["C2"].number_format = "yyyy/m/d"
        if hasattr(ws["E2"].value, "strftime"):
            ws["E2"].number_format = "yyyy/m/d"

    ws["A3"] = "Currency: "
    ws["C3"] = "PKR"
    if fx_rate is not None:
        ws["E3"] = "FX USD:PKR"
        ws["F3"] = float(fx_rate)

    for col, h in enumerate(_HEADERS, start=1):
        ws.cell(HEADER_ROW, col).value = h
    # 可选：Business Tax 月度 USD 系数（Danfoss 发票推导）
    extra_headers: list[str] = []
    for h in (BT_SINDH_HEADER, BT_FEDERAL_HEADER):
        if any(emp.get(h) is not None or emp.get("_bt_sindh_usd" if "Sindh" in h else "_bt_federal_usd") is not None for emp in employees):
            extra_headers.append(h)
    for i, h in enumerate(extra_headers):
        ws.cell(HEADER_ROW, len(_HEADERS) + 1 + i).value = h

    for i, emp in enumerate(employees):
        row = DATA_START + i
        ws.cell(row, 1).value = emp.get("No. of EE")
        ws.cell(row, 2).value = emp.get("Name of Employee")
        base = emp.get("Base Salary")
        if base is not None:
            ws.cell(row, 3).value = float(base)
        eobi = emp.get("E.O.B.I")
        if eobi is not None:
            try:
                ws.cell(row, 12).value = float(eobi)
            except (TypeError, ValueError):
                ws.cell(row, 12).value = eobi
        it = emp.get("IT")
        if it is not None:
            try:
                ws.cell(row, 13).value = float(it)
            except (TypeError, ValueError):
                ws.cell(row, 13).value = it
        for j, h in enumerate(extra_headers):
            val = emp.get(h)
            if val is None and h == BT_SINDH_HEADER:
                val = emp.get("_bt_sindh_usd")
            if val is None and h == BT_FEDERAL_HEADER:
                val = emp.get("_bt_federal_usd")
            if val is not None:
                try:
                    ws.cell(row, len(_HEADERS) + 1 + j).value = float(val)
                except (TypeError, ValueError):
                    ws.cell(row, len(_HEADERS) + 1 + j).value = val
        # Gross = SUM(C:I)
        ws.cell(row, 10).value = f"=SUM(C{row}:I{row})"
        # Net = Gross + Expense - EOBI - IT
        ws.cell(row, 14).value = f"=J{row}+K{row}-L{row}-M{row}"
        ws.cell(row, 18).value = f"=J{row}+K{row}+O{row}+P{row}+Q{row}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def convert_pdfs(
    pdf_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    del template_path, pn_meta, registry_dir, fill_fx
    mapping_in = dict(convert_mapping) if isinstance(convert_mapping, dict) else {}
    mapping_in.setdefault("pdfProfileId", "panda_work_pk")
    mapping = resolve_convert_mapping("pakistan_payroll_calc", mapping_in)
    split = 3
    overlay_like = mapping.get("quarterSplitMonths")
    if overlay_like is not None:
        try:
            split = int(overlay_like)
        except (TypeError, ValueError):
            split = 3

    paths = [Path(p).resolve() for p in pdf_paths]
    if not paths:
        raise ValueError("未提供 PDF")
    parsed_list: list[PandaWorkPkParsed] = []
    for p in paths:
        if not p.is_file():
            raise FileNotFoundError(f"PDF 不存在: {p}")
        parsed_list.append(parse_panda_work_pk_pdf(p))

    # mapping 客户名优先
    client_label = str(mapping.get("connectClientLabel") or mapping.get("clientLabel") or "").strip()
    if client_label:
        for p in parsed_list:
            p.client_label = client_label

    employees, warnings, fx = employees_from_parsed(parsed_list, split_months=split)
    if not employees:
        raise ValueError("未解析到员工")

    bt_cfg = parse_pakistan_business_tax_cfg(mapping)
    if bt_cfg:
        warnings.extend(apply_derived_coeffs_to_rows(employees, bt_cfg, split_months=split))

    fees = parse_pakistan_employee_fees(mapping)
    if fees:
        warnings.extend(apply_fees_to_employee_rows(employees, fees))
    else:
        seen: set[str] = set()
        for emp in employees:
            name = str(emp.get("Name of Employee") or "").strip()
            key = name.lower()
            if name and key not in seen:
                seen.add(key)
                warnings.append(f"{name}：E.O.B.I / IT 若映射未配置则 Pakistan-L 对应列留空")

    output_path = Path(output_path).resolve()
    _write_pakistan_l_workbook(output_path, employees, fx_rate=fx)

    return {
        "ok": True,
        "profile_id": "panda_work_pk",
        "region": "Pakistan",
        "output": str(output_path),
        "employee_count": len(employees),
        "person_count": len(parsed_list),
        "fx_rate": fx,
        "parsed": [p.to_dict() for p in parsed_list],
        "warnings": warnings,
    }


def convert_pdf(
    pdf_path: Path,
    output_path: Path,
    *,
    template_path: Path | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    fill_fx: bool = True,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return convert_pdfs(
        [pdf_path],
        output_path,
        template_path=template_path,
        pn_meta=pn_meta,
        registry_dir=registry_dir,
        fill_fx=fill_fx,
        convert_mapping=convert_mapping,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Panda Work PK 季度发票 → Pakistan-L")
    parser.add_argument("pdfs", nargs="+", type=Path)
    parser.add_argument("-o", "--output", type=Path, required=True)
    args = parser.parse_args(argv)
    result = convert_pdfs(args.pdfs, args.output)
    print(result)
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
