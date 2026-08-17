# -*- coding: utf-8 -*-
"""
Italy-L 横向源账单 → Italy PN（引擎 italy_payroll_calc）

用法:
  python -m profiles.italy_payroll_calc.convert <源.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

源账单: sheet「Italy-L」第 10 行表头、第 11 行起员工（可由 safeguard_italy ingest 产出）。
默认母版: templates/italy/template.xlsx

原则：PN / Italy / Italy EE 以母版公式为准；只写 Italy-L 数据与必要元数据。
Fee Min：mapping.italyFeeMin 有值才写；格子见 mapping.fixedValueWrites。
"""
from __future__ import annotations

import argparse
import calendar
import re
import shutil
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from bill_convert.fixed_value_writes import apply_fixed_value_writes
from bill_convert.formula_copy import shift_row_formula
from convert_mapping import find_sheet_name, resolve_convert_mapping
from fx_rate import fetch_usd_rates, get_italy_pn_fx_rate
from pn_meta import PnMeta, apply_pn_meta
from profiles.tw_payroll_calc.convert import match_ee_code
from region_templates import get_region_template
from xlsx_convert_utils import coerce_datetime_for_excel, is_date_column_header
from xlsx_luckysheet_compat import apply_luckysheet_compat
from xlsx_postprocess import postprocess_converted_xlsx

DEFAULT_TEMPLATE = get_region_template("Italy")

ITALY_L_SHEET = "Italy-L"
ITALY_SHEET = "Italy"
ITALY_EE_SHEET = "Italy EE"
PN_SHEET = "PN"

ITALY_L_HEADER_ROW = 10
ITALY_L_DATA_START = 11
ITALY_DATA_START = 9
ITALY_EE_DATA_START = 10
MAX_EMPLOYEES = 20

_DATE_FMT = "yyyy/m/d"

_ACTIVE_MAPPING: dict[str, Any] | None = None

_MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def _active_mapping() -> dict[str, Any]:
    return (
        _ACTIVE_MAPPING
        if isinstance(_ACTIVE_MAPPING, dict)
        else resolve_convert_mapping("italy_payroll_calc", None)
    )


def _italy_l_layout(*, target: bool = False) -> tuple[int, int, list[str]]:
    mapping = _active_mapping()
    key = "targetL" if target else "sourceEmployeeSheet"
    spec = mapping.get(key) if isinstance(mapping.get(key), dict) else {}
    if not spec and target:
        spec = (
            mapping.get("sourceEmployeeSheet")
            if isinstance(mapping.get("sourceEmployeeSheet"), dict)
            else {}
        )
    header = int(spec.get("headerRow") or ITALY_L_HEADER_ROW)
    data_start = int(spec.get("dataStartRow") or ITALY_L_DATA_START)
    names = spec.get("nameHeaders") if isinstance(spec.get("nameHeaders"), list) else None
    name_headers = [str(x).strip() for x in (names or ["Employee Name"]) if str(x).strip()]
    if not name_headers:
        name_headers = ["Employee Name"]
    return header, data_start, name_headers


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        h = _norm(ws.cell(header_row, col).value)
        if h and h not in out:
            out[h] = col
    return out


def _cell_formula_text(value: Any) -> str | None:
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _set_cell_value(cell, value: Any) -> None:
    if isinstance(value, float):
        cell.value = round(value, 6) if abs(value) >= 1e-9 else 0
    else:
        cell.value = value


def _italy_l_formula_cols(ws: Worksheet, data_start: int) -> dict[int, str]:
    out: dict[int, str] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        text = _cell_formula_text(ws.cell(data_start, col).value)
        if text:
            out[col] = text
    return out


def _emp_display_name(emp: dict[str, Any]) -> str:
    for key in ("Employee Name", "Name"):
        name = _norm(emp.get(key))
        if name:
            return name
    return ""


def parse_pay_period(label: Any) -> tuple[int, int] | None:
    """May'26 / May-2026 / May 2026 → (year, month)。"""
    text = _norm(label)
    if not text:
        return None
    m = re.search(
        r"(jan|january|feb|february|mar|march|apr|april|may|jun|june|jul|july|"
        r"aug|august|sep|sept|september|oct|october|nov|november|dec|december)"
        r"[\s\-']*(\d{2,4})",
        text,
        flags=re.I,
    )
    if not m:
        return None
    month = _MONTHS.get(m.group(1).lower())
    if not month:
        return None
    year = int(m.group(2))
    if year < 100:
        year += 2000
    return year, month


def salary_header_for_period(label: Any) -> str | None:
    parsed = parse_pay_period(label)
    if not parsed:
        return None
    year, month = parsed
    names = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]
    return f"{names[month - 1]} {year} Salary"


def period_bounds(label: Any) -> tuple[date, date] | None:
    parsed = parse_pay_period(label)
    if not parsed:
        return None
    year, month = parsed
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    return start, end


def looks_like_italy_l_workbook(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return False
    try:
        return ITALY_L_SHEET in wb.sheetnames
    finally:
        wb.close()


def parse_italy_l_employees(
    ws: Worksheet,
    *,
    column_rename: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    header_row, data_start, name_headers = _italy_l_layout(target=False)
    headers = _header_map(ws, header_row)
    if not headers:
        raise ValueError(f"「{ITALY_L_SHEET}」第 {header_row} 行表头为空")

    name_cols = [headers[h] for h in name_headers if h in headers]
    if not name_cols and "Employee Name" in headers:
        name_cols = [headers["Employee Name"]]
    if not name_cols:
        raise ValueError(f"Italy-L 未找到姓名列（期望: {name_headers}）")

    # sheet meta
    meta: dict[str, Any] = {}
    for r in range(1, header_row):
        label = _norm(ws.cell(r, 2).value).lower()
        val = ws.cell(r, 3).value
        if label == "customer":
            meta["_customer"] = _norm(val)
        elif label == "location":
            meta["_location"] = _norm(val)
        elif label == "pay period":
            meta["_pay_period"] = val
            meta["Pay Period"] = val

    rename = column_rename if isinstance(column_rename, dict) else {}
    # 复用 SafeGuard 对照逻辑：显式 columnRename 覆盖同名
    from pdf_ingest.profiles.safeguard_italy import (
        _explicit_rename_targets,
        _find_source_col,
        _put_cell_value,
        _resolve_target_for_source,
        _strip_target_label,
    )

    claimed = _explicit_rename_targets(rename)

    employees: list[dict[str, Any]] = []
    max_row = max(ws.max_row or data_start, data_start)
    for row in range(data_start, max_row + 1):
        name = ""
        for c in name_cols:
            name = _norm(ws.cell(row, c).value)
            if name:
                break
        if not name:
            continue
        low = name.lower()
        if "invoice total" in low or "fee invoice" in low:
            continue
        emp: dict[str, Any] = dict(meta)
        emp["Employee Name"] = name

        def _cell_val(col: int) -> Any:
            return ws.cell(row, col).value

        # 1) 同名（跳过公式格、已被对照占用的目标）
        for h, col in headers.items():
            if col in name_cols:
                continue
            val = _cell_val(col)
            if _cell_formula_text(val):
                continue
            if val is None or val == "":
                continue
            tgt = _resolve_target_for_source(h, None) or h
            if tgt in claimed or str(tgt).lower() in claimed:
                continue
            _put_cell_value(emp, tgt, val)

        # 2) 显式 columnRename 覆盖
        for src, tgt_raw in rename.items():
            if not src or not tgt_raw:
                continue
            col = _find_source_col(headers, str(src))
            if not col:
                continue
            val = _cell_val(col)
            if _cell_formula_text(val):
                continue
            tgt = _strip_target_label(str(tgt_raw))
            if not tgt:
                continue
            _put_cell_value(emp, tgt, val)

        employees.append(emp)
    return employees


def write_italy_l_sheet_meta(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    if not employees:
        return
    emp0 = employees[0]
    customer = _norm(emp0.get("_customer") or emp0.get("Customer"))
    location = _norm(emp0.get("_location") or emp0.get("Location")) or "Italy"
    pay_period = emp0.get("_pay_period") or emp0.get("Pay Period")

    if _norm(ws.cell(2, 2).value).lower() == "customer" and customer:
        ws.cell(2, 3).value = customer
    if _norm(ws.cell(3, 2).value).lower() == "location":
        ws.cell(3, 3).value = location
    if _norm(ws.cell(5, 2).value).lower() == "pay period" and pay_period is not None:
        ws.cell(5, 3).value = pay_period

    # 动态薪资列表头（Italy / Italy EE 引用 Italy-L!K10）
    header_row, _, _ = _italy_l_layout(target=True)
    salary_title = salary_header_for_period(pay_period)
    if salary_title:
        headers = _header_map(ws, header_row)
        for h, col in list(headers.items()):
            if re.search(r"salary", h, flags=re.I):
                ws.cell(header_row, col).value = salary_title
                break


def write_italy_l(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    """只覆盖数据列，母版公式列保留/扩行复制。"""
    header_row, data_start, _ = _italy_l_layout(target=True)
    headers = _header_map(ws, header_row)
    if not headers:
        raise ValueError(f"「{ITALY_L_SHEET}」第 {header_row} 行表头为空")
    write_italy_l_sheet_meta(ws, employees)
    # meta 可能改了 Salary 表头，重读
    headers = _header_map(ws, header_row)

    n = len(employees)
    formula_by_col = _italy_l_formula_cols(ws, data_start)
    max_col = max(max(headers.values(), default=1), max(formula_by_col.keys(), default=1), 1)

    for i in range(1, n):
        _copy_row_style_and_formula(
            ws,
            data_start,
            data_start + i,
            max_col=max_col,
            l_from=data_start,
            l_to=data_start + i,
        )

    last_keep = data_start + max(n, 1) - 1
    max_row = max(ws.max_row or data_start, data_start + MAX_EMPLOYEES)
    for row in range(data_start, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ == "MergedCell":
                continue
            if row <= last_keep and col in formula_by_col:
                continue
            cell.value = None

    for i in range(n):
        row = data_start + i
        for col, formula in formula_by_col.items():
            cur = ws.cell(row, col).value
            if _cell_formula_text(cur):
                continue
            ws.cell(row, col).value = shift_row_formula(
                formula,
                data_start,
                row,
                target_l_from=data_start,
                target_l_to=row,
                target_l_sheet=ITALY_L_SHEET,
            )

    for idx, emp in enumerate(employees):
        row = data_start + idx
        for h, col in headers.items():
            if col in formula_by_col:
                continue
            if h not in emp:
                continue
            val = emp[h]
            if val is None:
                continue
            cell = ws.cell(row, col)
            if is_date_column_header(h) or h in ("From", "To"):
                dt = coerce_datetime_for_excel(val)
                if dt is not None:
                    cell.value = dt
                    cell.number_format = _DATE_FMT
                    continue
            if isinstance(val, datetime):
                cell.value = val
                cell.number_format = _DATE_FMT
            else:
                _set_cell_value(cell, val)


def _copy_row_style_and_formula(
    ws: Worksheet,
    src_row: int,
    dest_row: int,
    max_col: int = 40,
    *,
    l_from: int | None = None,
    l_to: int | None = None,
) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dest = ws.cell(dest_row, c)
        if type(dest).__name__ == "MergedCell" or type(src).__name__ == "MergedCell":
            continue
        if src.has_style:
            dest.font = copy(src.font)
            dest.border = copy(src.border)
            dest.fill = copy(src.fill)
            dest.number_format = src.number_format
            dest.protection = copy(src.protection)
            dest.alignment = copy(src.alignment)
        text = _cell_formula_text(src.value)
        if text:
            dest.value = shift_row_formula(
                text,
                src_row,
                dest_row,
                target_l_from=l_from if l_from is not None else -1,
                target_l_to=l_to if l_to is not None else -1,
                target_l_sheet=ITALY_L_SHEET,
            )
        elif src.value is not None and not isinstance(src.value, ArrayFormula):
            dest.value = src.value


def _retarget_ee_refs(formula: str, ee_from: int, ee_to: int) -> str:
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    return re.sub(
        rf"('Italy EE'!\$?[A-Z]{{1,3}})\$?{ee_from}(?!\d)",
        lambda m: f"{m.group(1)}{ee_to}",
        formula,
        flags=re.I,
    )


def expand_italy_employee_rows(wb, employee_count: int) -> None:
    n = max(int(employee_count), 1)
    _, l_data_start, _ = _italy_l_layout(target=True)
    if ITALY_SHEET in wb.sheetnames:
        italy = wb[ITALY_SHEET]
        for i in range(1, n):
            dest = ITALY_DATA_START + i
            l_row = l_data_start + i
            ee_row = ITALY_EE_DATA_START + i
            _copy_row_style_and_formula(
                italy,
                ITALY_DATA_START,
                dest,
                max_col=40,
                l_from=l_data_start,
                l_to=l_row,
            )
            for c in range(1, 41):
                cell = italy.cell(dest, c)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = _retarget_ee_refs(cell.value, ITALY_EE_DATA_START, ee_row)

    if ITALY_EE_SHEET in wb.sheetnames:
        ee = wb[ITALY_EE_SHEET]
        for i in range(1, n):
            dest = ITALY_EE_DATA_START + i
            l_row = l_data_start + i
            _copy_row_style_and_formula(
                ee,
                ITALY_EE_DATA_START,
                dest,
                max_col=40,
                l_from=l_data_start,
                l_to=l_row,
            )


def set_fee_min(wb, employees: list[dict[str, Any]]) -> None:
    """Italy-L Fee Min：格子见 mapping.fixedValueWrites（convert_mapping）。"""
    apply_fixed_value_writes(wb, employees, _active_mapping())


def set_period(wb, employees: list[dict[str, Any]]) -> None:
    """按 Pay Period 写 Italy!B2/C2（覆盖 TODAY 公式，否则账期会漂）。"""
    if ITALY_SHEET not in wb.sheetnames or not employees:
        return
    bounds = period_bounds(employees[0].get("_pay_period") or employees[0].get("Pay Period"))
    if not bounds:
        return
    start, end = bounds
    italy = wb[ITALY_SHEET]
    for col, dt in ((2, start), (3, end)):
        cell = italy.cell(2, col)
        cell.value = datetime(dt.year, dt.month, dt.day)
        cell.number_format = _DATE_FMT


def _find_pn_row_by_label(ws: Worksheet, keyword: str, col: int = 1) -> int | None:
    key = keyword.lower()
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and key in v.lower():
            return row
    return None


def fit_italy_pn_employees(wb, employee_count: int) -> dict[str, Any]:
    """
    人数 == 母版槽位：PN 公式不动。
    多人：暂定只扩 Italy/Italy EE（expand 已处理）；PN 多行 Labor/Fee 后续再细化。
    """
    fx_row = _find_pn_row_by_label(wb[PN_SHEET], "FX rate") if PN_SHEET in wb.sheetnames else None
    return {"fx_row": fx_row or 28, "employee_count": employee_count}


def _pn_customer_id(pn_meta: PnMeta | dict[str, Any] | None) -> str | None:
    if pn_meta is None:
        return None
    if isinstance(pn_meta, PnMeta):
        cid = (pn_meta.customer_id or "").strip()
        return cid or None
    if isinstance(pn_meta, dict):
        cid = str(pn_meta.get("customer_id") or pn_meta.get("customerId") or "").strip()
        return cid or None
    return None


def apply_italy_ee_codes(
    wb,
    employees: list[dict[str, Any]],
    *,
    employee_directory: list[dict[str, Any]] | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
) -> list[str]:
    if ITALY_EE_SHEET not in wb.sheetnames:
        return []
    ws = wb[ITALY_EE_SHEET]
    client_code = _pn_customer_id(pn_meta)
    directory = list(employee_directory or [])
    warnings: list[str] = []
    for i, emp in enumerate(employees):
        row = ITALY_EE_DATA_START + i
        if client_code:
            ws.cell(row, 2).value = client_code
        excel_names = [_emp_display_name(emp)]
        code, warn = match_ee_code([n for n in excel_names if n], directory)
        ws.cell(row, 4).value = code
        if warn:
            warnings.append(f"Italy EE 第{i + 1}人：{warn}")
    return warnings


def apply_fx(wb, *, fill_fx: bool = True, fx_row: int | None = None) -> float | None:
    if not fill_fx or PN_SHEET not in wb.sheetnames:
        return None
    rates = fetch_usd_rates()
    fx = get_italy_pn_fx_rate(rates)
    row = fx_row or _find_pn_row_by_label(wb[PN_SHEET], "FX rate") or 28
    cell = wb[PN_SHEET].cell(row, 2)
    if _cell_formula_text(cell.value):
        return fx
    cell.value = fx
    return fx


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    registry_dir: Path | None = None,
    convert_mapping: dict[str, Any] | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    global _ACTIVE_MAPPING
    _ACTIVE_MAPPING = resolve_convert_mapping("italy_payroll_calc", convert_mapping)
    try:
        source_path = Path(source_path).resolve()
        output_path = Path(output_path).resolve()
        template_path = Path(template_path).resolve()
        if not source_path.is_file():
            raise FileNotFoundError(f"源文件不存在: {source_path}")
        if not template_path.is_file():
            raise FileNotFoundError(f"母版不存在: {template_path}")

        src_wb = load_workbook(source_path, data_only=False)
        try:
            l_name = find_sheet_name(list(src_wb.sheetnames), _active_mapping().get("sourceEmployeeSheet"))
            if not l_name or l_name not in src_wb.sheetnames:
                if ITALY_L_SHEET in src_wb.sheetnames:
                    l_name = ITALY_L_SHEET
                else:
                    raise ValueError(f"未找到 Italy-L，现有: {src_wb.sheetnames}")
            employees = parse_italy_l_employees(
                src_wb[l_name],
                column_rename=_active_mapping().get("columnRename")
                if isinstance(_active_mapping().get("columnRename"), dict)
                else None,
            )
        finally:
            src_wb.close()

        if not employees:
            raise ValueError("Italy-L 未解析到员工行")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)
        warnings: list[str] = []
        fx = None
        applied_pn = None
        try:
            if ITALY_L_SHEET not in wb.sheetnames:
                raise ValueError(f"母版缺少 {ITALY_L_SHEET}")
            write_italy_l(wb[ITALY_L_SHEET], employees)
            set_fee_min(wb, employees)
            set_period(wb, employees)
            expand_italy_employee_rows(wb, len(employees))
            pn_layout = fit_italy_pn_employees(wb, len(employees))
            if len(employees) > 1:
                warnings.append(
                    f"Italy PN 多人 Labor/Service Fee 行扩行暂定：已扩 Italy/Italy EE（{len(employees)} 人），"
                    "PN 明细行请人工核对"
                )
            try:
                fx = apply_fx(wb, fill_fx=fill_fx, fx_row=pn_layout.get("fx_row"))
            except Exception as exc:
                warnings.append(f"写入 PN 汇率失败: {exc}")

            if pn_meta is not None:
                applied_pn = apply_pn_meta(
                    wb,
                    pn_meta,
                    registry_dir=registry_dir or output_path.parent,
                    reserve_invoice_number=True,
                )
            warnings.extend(
                apply_italy_ee_codes(
                    wb,
                    employees,
                    employee_directory=employee_directory,
                    pn_meta=applied_pn or pn_meta,
                )
            )
            apply_luckysheet_compat(wb, pn_sheet=PN_SHEET)
            wb.save(output_path)
        finally:
            wb.close()

        postprocess_converted_xlsx(output_path)
        return {
            "ok": True,
            "engine_id": "italy_payroll_calc",
            "region": "Italy",
            "output": str(output_path),
            "employee_count": len(employees),
            "fx_rate": fx,
            "warnings": warnings,
            "pn_meta": applied_pn.to_dict() if applied_pn else None,
        }
    finally:
        _ACTIVE_MAPPING = None


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Italy-L → Italy PN")
    parser.add_argument("source", type=Path)
    parser.add_argument("-o", "--output", type=Path, default=None)
    parser.add_argument("-t", "--template", type=Path, default=None)
    parser.add_argument("--no-fx", action="store_true")
    args = parser.parse_args(argv)
    source = args.source.resolve()
    output = (args.output or source.with_name(f"PN_Italy_{source.stem}.xlsx")).resolve()
    template = (args.template or DEFAULT_TEMPLATE).resolve()
    result = convert(source, output, template, fill_fx=not args.no_fx)
    print(result)
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
