# -*- coding: utf-8 -*-
"""
香港 Hong Kong-L 源账单 → Hong Kong PN（引擎 hk_payroll_calc）

用法:
  python -m profiles.hk_payroll_calc.convert <原始 T-N 账单.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

原始账单: sheet「Hong Kong-L」第 7 行表头、第 8 行起员工数据（按表头名匹配）
默认母版: templates/hongkong/template.xlsx
Office 映射: convert_mapping（列对照 / 员工公式配对 hkExampleRow）
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from bill_convert.convert_checks import check_column_rename_hits

from bill_convert.formula_copy import (
    copy_row_formulas as shared_copy_row_formulas,
    fix_ee_row_hk_refs,
    fix_hk_row_hk_ee_refs,
    retarget_pn_fx_b_column_refs,
    snapshot_row_cells,
)
from bill_convert.formula_layout import (
    apply_employee_formula_styles,
    needed_example_rows_for_styles,
    tw_l_row_for_data_row,
)
from bill_convert.formula_layout import _default_example_row as default_example_row_for_mapping
from bill_convert.headers import list_qualified_header_cells
from convert_mapping import find_sheet_name, resolve_convert_mapping
from fx_rate import fetch_usd_rates, get_hk_pn_fx_rate
from pn_meta import PnMeta, apply_pn_meta
from region_templates import get_region_template
from xlsx_convert_utils import (
    clean_value,
    coerce_datetime_for_excel,
    is_date_column_header,
    norm,
)
from xlsx_postprocess import postprocess_converted_xlsx

DEFAULT_TEMPLATE = get_region_template("Hong Kong")

HK_L_SHEET = "Hong Kong-L"
HK_SHEET = "Hong Kong"
HK_EE_SHEET = "Hong Kong EE"
PN_SHEET = "PN"

HK_L_HEADER_ROW = 7
HK_L_DATA_START_ROW = 8
HK_DATA_START_ROW = 9
HK_EE_DATA_START_ROW = 10
_DATE_FMT = "yyyy/m/d"
MAX_EMPLOYEES = 10

NAME_HEADERS = ("Name of Employee", "EE Name", "Name")

_ACTIVE_MAPPING: dict[str, Any] | None = None


def _active_mapping() -> dict[str, Any]:
    return _ACTIVE_MAPPING if isinstance(_ACTIVE_MAPPING, dict) else resolve_convert_mapping("hk_payroll_calc", None)


def _hk_l_layout() -> tuple[int, int]:
    target = _active_mapping().get("targetL") if isinstance(_active_mapping().get("targetL"), dict) else {}
    header = int(target.get("headerRow") or HK_L_HEADER_ROW)
    data_start = int(target.get("dataStartRow") or HK_L_DATA_START_ROW)
    return header, data_start


def _hk_formula_rows() -> dict[str, int]:
    _, l_start = _hk_l_layout()
    return {
        "l_data_start": l_start,
        "main_data_start": HK_DATA_START_ROW,
        "ee_data_start": HK_EE_DATA_START_ROW,
        # 兼容 formula_layout 旧键
        "tw_l_data_start": l_start,
        "tw_data_start": HK_DATA_START_ROW,
        "tw_ee_data_start": HK_EE_DATA_START_ROW,
    }


def _column_rename() -> dict[str, str]:
    raw = _active_mapping().get("columnRename") or {}
    if not isinstance(raw, dict):
        return {}
    return {norm(str(k)): str(v).strip() for k, v in raw.items() if k and v}


def _skip_source_headers() -> set[str]:
    raw = _active_mapping().get("skipSourceHeaders") or []
    if not isinstance(raw, list):
        return set()
    return {norm(x) for x in raw if x}


def map_source_header(source_header: str) -> str | None:
    h = norm(source_header)
    if not h or h in _skip_source_headers():
        return None
    rename = _column_rename()
    if h in rename:
        return rename[h]
    claimed = {norm(v) for v in rename.values()}
    for v in list(claimed):
        base = v.split("#", 1)[0]
        claimed.add(base)
        claimed.add(base.rsplit("/", 1)[-1])
    if h in claimed:
        return None
    base = h.split("#", 1)[0]
    child = base.rsplit("/", 1)[-1]
    if base in claimed or child in claimed:
        return None
    return h


def build_header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def format_payroll_date(value: Any) -> datetime | None:
    return coerce_datetime_for_excel(value)


def normalize_company_name(value: Any) -> Any:
    if value is None:
        return None
    s = norm(value)
    m = re.search(r"\(([^)]+)\)", s)
    if m:
        return m.group(1).strip()
    return s


def read_hk_l_employees(ws: Worksheet) -> list[dict[str, Any]]:
    header_row, data_start = _hk_l_layout()
    qualified = list_qualified_header_cells(ws, header_row)
    source_headers = {str(h["key"]): int(h["col"]) for h in qualified}
    child_to_keys: dict[str, list[str]] = {}
    for h in qualified:
        child_to_keys.setdefault(str(h["child"]), []).append(str(h["key"]))
    name_col = None
    src_spec = _active_mapping().get("sourceEmployeeSheet") if isinstance(_active_mapping().get("sourceEmployeeSheet"), dict) else {}
    name_headers = src_spec.get("nameHeaders") if isinstance(src_spec.get("nameHeaders"), list) else list(NAME_HEADERS)
    for nh in name_headers:
        key = norm(nh)
        if key in source_headers:
            name_col = source_headers[key]
            break
        for qk in child_to_keys.get(key, []):
            if qk in source_headers:
                name_col = source_headers[qk]
                break
        if name_col is not None:
            break
    if name_col is None:
        raise ValueError(f"「Hong Kong-L」第 {header_row} 行须包含员工姓名表头（如 Name of Employee）")

    employees: list[dict[str, Any]] = []
    for row in range(data_start, (ws.max_row or 0) + 1):
        name = clean_value(ws.cell(row, name_col).value)
        if name is None:
            continue
        record: dict[str, Any] = {}
        rename = _column_rename()
        if rename:
            check_column_rename_hits(rename, source_headers, strict_if_configured=True)
        ordered = list(source_headers.items())
        explicit_first = [(s, c) for s, c in ordered if norm(s) in rename]
        auto_rest = [(s, c) for s, c in ordered if norm(s) not in rename]
        for src_hdr, col in explicit_first + auto_rest:
            target_hdr = map_source_header(src_hdr)
            if target_hdr is None:
                continue
            val = clean_value(ws.cell(row, col).value)
            if val is not None:
                if is_date_column_header(target_hdr):
                    dt = coerce_datetime_for_excel(val)
                    if dt is not None:
                        val = dt
                if target_hdr in record and norm(src_hdr) not in rename:
                    continue
                record[target_hdr] = val
        employees.append(record)

    if not employees:
        raise ValueError("「Hong Kong-L」中未找到有效员工行")
    if len(employees) > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {len(employees)} 超过模板上限 {MAX_EMPLOYEES}")
    return employees


def read_hk_l_meta(ws: Worksheet) -> dict[str, Any]:
    return {
        "company_name": normalize_company_name(ws.cell(1, 3).value),
        "period_from": format_payroll_date(ws.cell(2, 3).value),
        "period_to": format_payroll_date(ws.cell(2, 5).value),
        "currency": ws.cell(3, 3).value,
    }


def ensure_hk_period_date_formats(wb) -> None:
    """Hong Kong!B2/C2 引用账期；强制日期格式，避免显示成金额。"""
    try:
        hkl = wb[HK_L_SHEET]
        for row, col in ((2, 3), (2, 5)):
            hkl.cell(row, col).number_format = _DATE_FMT
    except KeyError:
        pass
    try:
        hk = wb[HK_SHEET]
        for row, col in ((2, 2), (2, 3)):
            hk.cell(row, col).number_format = _DATE_FMT
    except KeyError:
        pass


def clear_hk_l_data(ws: Worksheet, from_row: int | None = None) -> None:
    if from_row is None:
        _, from_row = _hk_l_layout()
    max_row = max(ws.max_row or from_row, from_row + MAX_EMPLOYEES)
    max_col = ws.max_column or 1
    for row in range(from_row, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def write_hk_l(ws: Worksheet, employees: list[dict[str, Any]], meta: dict[str, Any]) -> None:
    header_row, data_start = _hk_l_layout()
    target_headers = build_header_map(ws, header_row)
    if not target_headers:
        raise ValueError(f"「Hong Kong-L」第 {header_row} 行表头为空")

    if meta.get("company_name") is not None:
        ws.cell(1, 3).value = meta["company_name"]
    if meta.get("period_from") is not None:
        cell = ws.cell(2, 3)
        cell.value = meta["period_from"]
        cell.number_format = _DATE_FMT
    if meta.get("period_to") is not None:
        cell = ws.cell(2, 5)
        cell.value = meta["period_to"]
        cell.number_format = _DATE_FMT
    if meta.get("currency") is not None:
        ws.cell(3, 3).value = meta["currency"]

    clear_hk_l_data(ws, data_start)
    for idx, emp in enumerate(employees):
        row = data_start + idx
        for hdr, val in emp.items():
            col = target_headers.get(hdr)
            if col is not None:
                out_val = val
                if is_date_column_header(hdr):
                    dt = coerce_datetime_for_excel(val)
                    if dt is not None:
                        out_val = dt
                cell = ws.cell(row, col)
                cell.value = out_val
                if is_date_column_header(hdr) and isinstance(out_val, datetime):
                    cell.number_format = _DATE_FMT


def clear_employee_row(ws: Worksheet, row: int) -> None:
    for col in range(1, (ws.max_column or 0) + 1):
        ws.cell(row, col).value = None


def count_template_employee_slots(ws: Worksheet, data_start_row: int, marker_col: int) -> int:
    """连续占位：marker 非空，或该行存在公式。"""
    n = 0
    for row in range(data_start_row, data_start_row + MAX_EMPLOYEES):
        marker = ws.cell(row, marker_col).value
        has_formula = False
        if marker is None:
            for col in range(1, (ws.max_column or 0) + 1):
                cell = ws.cell(row, col)
                if cell.data_type == "f" and cell.value:
                    has_formula = True
                    break
        if marker is not None or has_formula:
            n += 1
            continue
        break
    return max(n, 1)


def fit_hk_formula_sheets(
    wb,
    employee_count: int,
    *,
    clear_excess: bool = True,
    protected_hk_rows: set[int] | None = None,
    protected_ee_rows: set[int] | None = None,
) -> None:
    """
    扩/缩 Hong Kong / Hong Kong EE 公式行。
    配对前须 clear_excess=False，并先快照示例行。
    """
    n = max(int(employee_count), 1)
    mapping = _active_mapping()
    _, l_start = _hk_l_layout()
    hk = wb[HK_SHEET]
    ee = wb[HK_EE_SHEET]
    hk_slots = count_template_employee_slots(hk, HK_DATA_START_ROW, marker_col=2)
    ee_slots = count_template_employee_slots(ee, HK_EE_DATA_START_ROW, marker_col=4)
    hk_tpl = default_example_row_for_mapping(mapping, "Hong Kong", HK_DATA_START_ROW)
    ee_tpl = default_example_row_for_mapping(mapping, "Hong Kong EE", HK_EE_DATA_START_ROW)
    prot_h = set(protected_hk_rows or ())
    prot_e = set(protected_ee_rows or ())
    prot_h.add(hk_tpl)
    prot_e.add(ee_tpl)
    for r in prot_h:
        hk_slots = max(hk_slots, r - HK_DATA_START_ROW + 1)
    for r in prot_e:
        ee_slots = max(ee_slots, r - HK_EE_DATA_START_ROW + 1)
    src_hk_l = tw_l_row_for_data_row(hk_tpl, data_start=HK_DATA_START_ROW, target_l_data_start=l_start)
    src_ee_l = tw_l_row_for_data_row(ee_tpl, data_start=HK_EE_DATA_START_ROW, target_l_data_start=l_start)

    if clear_excess:
        for i in range(n, hk_slots):
            clear_employee_row(hk, HK_DATA_START_ROW + i)
        for i in range(n, ee_slots):
            clear_employee_row(ee, HK_EE_DATA_START_ROW + i)

    if n > hk_slots:
        for i in range(hk_slots, n):
            dst_row = HK_DATA_START_ROW + i
            dst_l = l_start + i
            ee_row = HK_EE_DATA_START_ROW + i
            shared_copy_row_formulas(
                hk, hk_tpl, dst_row, src_hk_l, dst_l, target_l_sheet=HK_L_SHEET
            )
            fix_hk_row_hk_ee_refs(hk, dst_row, ee_row)

    if n > ee_slots:
        for i in range(ee_slots, n):
            dst_row = HK_EE_DATA_START_ROW + i
            dst_l = l_start + i
            hk_row = HK_DATA_START_ROW + i
            shared_copy_row_formulas(
                ee, ee_tpl, dst_row, src_ee_l, dst_l, target_l_sheet=HK_L_SHEET
            )
            fix_ee_row_hk_refs(ee, dst_row, hk_row)


def clear_excess_hk_formula_rows(wb, employee_count: int) -> None:
    n = max(int(employee_count), 0)
    hk = wb[HK_SHEET]
    ee = wb[HK_EE_SHEET]
    hk_slots = count_template_employee_slots(hk, HK_DATA_START_ROW, marker_col=2)
    ee_slots = count_template_employee_slots(ee, HK_EE_DATA_START_ROW, marker_col=4)
    for i in range(n, hk_slots):
        clear_employee_row(hk, HK_DATA_START_ROW + i)
    for i in range(n, ee_slots):
        clear_employee_row(ee, HK_EE_DATA_START_ROW + i)


# ---------- PN 人员扩减（母版默认 1 人：Labor16 / Expense17 / Office18 / Service20 / Mgmt21 / FX B28）----------

_PN_EOR_ROW = 15
_PN_LABOR_START_ROW = 16
_PN_HKD_FMT = "#,##0.00"
_PN_USD_FMT = '$#,##0.00'


def _find_pn_row_by_label(ws: Worksheet, keyword: str, col: int = 1) -> int | None:
    key = keyword.lower()
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and key in v.lower():
            return row
    return None


def _find_pn_fx_row(ws: Worksheet) -> int:
    return _find_pn_row_by_label(ws, "FX rate") or 28


def _count_pn_labor_slots(ws: Worksheet) -> int:
    n = 0
    for row in range(_PN_LABOR_START_ROW, _PN_LABOR_START_ROW + MAX_EMPLOYEES + 2):
        v = ws.cell(row, 1).value
        if isinstance(v, str) and "Labor cost" in v:
            n += 1
            continue
        break
    return max(n, 1)


def _copy_pn_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    max_col = max(ws.max_column or 6, 6)
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _ensure_merge_a_c(ws: Worksheet, row: int) -> None:
    rng = f"A{row}:C{row}"
    for m in list(ws.merged_cells.ranges):
        if m.min_row <= row <= m.max_row and m.min_col <= 1 and m.max_col >= 3:
            if m.min_row == row and m.max_row == row and m.min_col == 1 and m.max_col == 3:
                return
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    try:
        ws.merge_cells(rng)
    except ValueError:
        pass


def _collect_merges_from(ws: Worksheet, start_row: int) -> list[tuple[int, int, int, int]]:
    kept: list[tuple[int, int, int, int]] = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row < start_row:
            continue
        kept.append((m.min_row, m.min_col, m.max_row, m.max_col))
        try:
            ws.unmerge_cells(str(m))
        except ValueError:
            pass
    return kept


def _restore_merges(
    ws: Worksheet,
    merges: list[tuple[int, int, int, int]],
    row_shift: int = 0,
) -> None:
    for min_r, min_c, max_r, max_c in merges:
        nr1, nr2 = min_r + row_shift, max_r + row_shift
        if nr1 < 1 or nr2 < nr1:
            continue
        try:
            ws.merge_cells(
                start_row=nr1,
                start_column=min_c,
                end_row=nr2,
                end_column=max_c,
            )
        except ValueError:
            pass


def _shift_row_heights(ws: Worksheet, start_row: int, amount: int) -> None:
    if amount == 0:
        return
    captured: dict[int, float] = {}
    max_r = max(ws.max_row or start_row, start_row)
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            captured[r] = h
        max_r = max(max_r, r)
    for r in range(start_row, max_r + abs(amount) + 2):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    if amount > 0:
        for r in sorted(captured.keys(), reverse=True):
            ws.row_dimensions[r + amount].height = captured[r]
    else:
        for r in sorted(captured.keys()):
            new_r = r + amount
            if new_r >= start_row:
                ws.row_dimensions[new_r].height = captured[r]


def _capture_row_heights_from(ws: Worksheet, start_row: int) -> dict[int, float]:
    out: dict[int, float] = {}
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            out[r] = h
    return out


def _apply_row_heights(ws: Worksheet, heights: dict[int, float], start_row: int) -> None:
    max_r = max([start_row, *heights.keys()], default=start_row)
    for r in range(start_row, max_r + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r, h in heights.items():
        if r >= start_row:
            ws.row_dimensions[r].height = h


def _pn_insert_rows(ws: Worksheet, idx: int, amount: int, fill_style_row: int | None = None) -> None:
    if amount <= 0:
        return
    merges = _collect_merges_from(ws, idx)
    _shift_row_heights(ws, idx, amount)
    ws.insert_rows(idx, amount)
    _restore_merges(ws, merges, row_shift=amount)
    if fill_style_row is not None:
        style_row = fill_style_row + amount if fill_style_row >= idx else fill_style_row
        for r in range(idx, idx + amount):
            _copy_pn_row_style(ws, style_row, r)
            _ensure_merge_a_c(ws, r)


def _pn_delete_rows(ws: Worksheet, idx: int, amount: int) -> None:
    if amount <= 0:
        return
    merges = _collect_merges_from(ws, idx + amount)
    heights = _capture_row_heights_from(ws, idx)
    ws.delete_rows(idx, amount)
    new_heights: dict[int, float] = {}
    for r, h in heights.items():
        if r < idx or r < idx + amount:
            continue
        new_heights[r - amount] = h
    _apply_row_heights(ws, new_heights, idx)
    _restore_merges(ws, merges, row_shift=-amount)


def _rewrite_pn_settlement_block(ws: Worksheet, *, fx_row: int, total_row: int) -> None:
    """
    香港母版 FX 下方：
      F{fx}=F{total}；F{fx+1..4}=Hong Kong!E22..E25；Sub/Tax/Total 相对锚点。
    """
    f_total_usd = ws.cell(fx_row, 6)
    f_total_usd.value = f"=F{total_row}"
    f_total_usd.number_format = _PN_USD_FMT
    for i, hk_e in enumerate((22, 23, 24, 25), start=1):
        cell = ws.cell(fx_row + i, 6)
        cell.value = f"='Hong Kong'!E{hk_e}"
    sub_row = fx_row + 5
    tax_row = fx_row + 6
    grand_row = fx_row + 7
    ws.cell(sub_row, 6).value = f"=SUM(F{fx_row}:F{fx_row + 4})"
    ws.cell(tax_row, 6).value = "='Hong Kong'!F6"
    ws.cell(grand_row, 6).value = f"=SUM(F{sub_row}:F{tax_row})"
    for r in (sub_row, tax_row, grand_row):
        ws.cell(r, 6).number_format = _PN_USD_FMT


def fit_pn_employees(ws: Worksheet, employee_count: int) -> dict[str, int]:
    """按人数扩/缩 PN：Labor×n + Expense×n + Office Rental×1，并重写合计/服务费/汇率区。"""
    n = max(int(employee_count), 1)
    if n > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {n} 超过模板上限 {MAX_EMPLOYEES}")

    old_n = _count_pn_labor_slots(ws)
    labor_start = _PN_LABOR_START_ROW
    expense_start = labor_start + old_n
    office_row = _find_pn_row_by_label(ws, "Office Rental") or (expense_start + old_n)
    delta = n - old_n

    if delta > 0:
        # 先扩 Labor（插在首个 Expense 前），再扩 Expense（插在 Office 前）
        _pn_insert_rows(ws, expense_start, delta, fill_style_row=labor_start)
        office_row = _find_pn_row_by_label(ws, "Office Rental") or (office_row + delta)
        expense_start = labor_start + n
        _pn_insert_rows(ws, office_row, delta, fill_style_row=expense_start)
    elif delta < 0:
        # 先删多余 Expense，再删多余 Labor
        expense_start = labor_start + old_n
        _pn_delete_rows(ws, expense_start + n, -delta)
        _pn_delete_rows(ws, labor_start + n, -delta)

    expense_start = labor_start + n
    office_row = _find_pn_row_by_label(ws, "Office Rental") or (expense_start + n)
    eor_row = _find_pn_row_by_label(ws, "EOR/PEO Cost") or _PN_EOR_ROW
    fx_row = _find_pn_fx_row(ws)

    for i in range(n):
        hk_row = HK_DATA_START_ROW + i
        labor_row = labor_start + i
        expense_row = expense_start + i
        if i > 0:
            _copy_pn_row_style(ws, labor_start, labor_row)
            _copy_pn_row_style(ws, expense_start, expense_row)
        _ensure_merge_a_c(ws, labor_row)
        _ensure_merge_a_c(ws, expense_row)

        ws.cell(labor_row, 1).value = (
            f'="- Labor cost for "&\'Hong Kong\'!B{hk_row}&"  -  "&MONTH(\'Hong Kong\'!B2)&"-"&YEAR(\'Hong Kong\'!C2)'
        )
        e_labor = ws.cell(labor_row, 5)
        e_labor.value = f"='Hong Kong'!I{hk_row}+'Hong Kong'!R{hk_row}-'Hong Kong'!P{hk_row}"
        e_labor.number_format = _PN_HKD_FMT
        f_labor = ws.cell(labor_row, 6)
        f_labor.value = f"=E{labor_row}/$B${fx_row}"
        f_labor.number_format = _PN_USD_FMT

        ws.cell(expense_row, 1).value = (
            f'="- "&+"Expense claim for "&\'Hong Kong\'!B{hk_row}&"  -  "&MONTH(\'Hong Kong\'!B2)&"-"&YEAR(\'Hong Kong\'!C2)'
        )
        e_exp = ws.cell(expense_row, 5)
        e_exp.value = f"='Hong Kong'!V{hk_row}"
        e_exp.number_format = _PN_HKD_FMT
        f_exp = ws.cell(expense_row, 6)
        f_exp.value = f"=E{expense_row}/$B${fx_row}"
        f_exp.number_format = _PN_USD_FMT

    # Office Rental：共用一行（Hong Kong!P6 合计）
    _ensure_merge_a_c(ws, office_row)
    ws.cell(office_row, 1).value = '="- Office Rental"'
    e_off = ws.cell(office_row, 5)
    e_off.value = "='Hong Kong'!P6"
    e_off.number_format = _PN_HKD_FMT
    f_off = ws.cell(office_row, 6)
    f_off.value = f"=E{office_row}/$B${fx_row}"
    f_off.number_format = _PN_USD_FMT

    e_eor = ws.cell(eor_row, 5)
    e_eor.value = f"=SUM(E{labor_start}:E{office_row})"
    e_eor.number_format = _PN_HKD_FMT
    f_eor = ws.cell(eor_row, 6)
    f_eor.value = f"=SUM(F{labor_start}:F{office_row})"
    f_eor.number_format = _PN_USD_FMT

    svc = _find_pn_row_by_label(ws, "Service Fee")
    mgmt = _find_pn_row_by_label(ws, "Management Fee")
    if svc is None or mgmt is None:
        svc = office_row + 2
        mgmt = svc + 1
    e_svc = ws.cell(svc, 5)
    e_svc.value = f"=SUM(E{mgmt}:E{mgmt + 1})"
    e_svc.number_format = _PN_HKD_FMT
    f_svc = ws.cell(svc, 6)
    f_svc.value = f"=SUM(F{mgmt}:F{mgmt + 1})"
    f_svc.number_format = _PN_USD_FMT
    # Management：用 Hong Kong!G6 汇总（多人合计；单人亦等于 G9）
    ws.cell(mgmt, 1).value = (
        '="- "&+"Management Fee - "&MONTH(\'Hong Kong\'!B2)&"-"&YEAR(\'Hong Kong\'!B2)'
    )
    e_mgmt = ws.cell(mgmt, 5)
    e_mgmt.value = "='Hong Kong'!G6"
    e_mgmt.number_format = _PN_HKD_FMT
    f_mgmt = ws.cell(mgmt, 6)
    f_mgmt.value = f"=E{mgmt}/$B${fx_row}"
    f_mgmt.number_format = _PN_USD_FMT
    _ensure_merge_a_c(ws, svc)
    _ensure_merge_a_c(ws, mgmt)

    total_row = _find_pn_row_by_label(ws, "EOR/PEO Service Cost")
    if total_row is None or total_row <= eor_row:
        total_row = svc + 5
    e_total = ws.cell(total_row, 5)
    e_total.value = f"=E{eor_row}+E{svc}"
    e_total.number_format = _PN_HKD_FMT
    f_total = ws.cell(total_row, 6)
    f_total.value = f"=F{eor_row}+F{svc}"
    f_total.number_format = _PN_USD_FMT

    fx_row = _find_pn_fx_row(ws)
    _rewrite_pn_settlement_block(ws, fx_row=fx_row, total_row=total_row)
    for row in range(labor_start, fx_row + 8):
        fcell = ws.cell(row, 6)
        if isinstance(fcell.value, str) and "$B$" in fcell.value:
            fcell.value = re.sub(r"\$B\$\d+", f"$B${fx_row}", fcell.value)

    return {
        "n": n,
        "eor_row": eor_row,
        "labor_start": labor_start,
        "expense_start": expense_start,
        "office_row": office_row,
        "service_row": svc,
        "mgmt_row": mgmt,
        "total_row": total_row,
        "fx_row": fx_row,
    }


def retarget_pn_fx_refs(wb, fx_row: int, *, from_rows: list[int] | None = None) -> None:
    """插删 PN 行后，把「旧汇率行」引用改到新行。

    只改指向旧 FX 行的 PN!Bxx / $B$xx；勿动 Client Code/Name（PN!B9 / PN!B8）。
    """
    retarget_pn_fx_b_column_refs(
        wb,
        fx_row,
        from_rows=from_rows,
        pn_sheet=PN_SHEET,
    )


def apply_hk_employee_formula_styles(
    wb,
    employees: list[dict[str, Any]],
    *,
    formula_rows: dict[str, int] | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    main_snapshots: dict[int, list[dict[str, Any]]] | None = None,
    ee_snapshots: dict[int, list[dict[str, Any]]] | None = None,
) -> list[dict[str, Any]]:
    """Hong Kong / Hong Kong EE 按人盖公式：配对用 hkExampleRow，未配对用默认第一行。"""
    return apply_employee_formula_styles(
        wb,
        employees,
        _active_mapping(),
        formula_rows=formula_rows or _hk_formula_rows(),
        employee_directory=employee_directory,
        main_sheet=HK_SHEET,
        ee_sheet=HK_EE_SHEET,
        target_l_sheet=HK_L_SHEET,
        main_template_key="Hong Kong",
        ee_template_key="Hong Kong EE",
        main_example_field="hkExampleRow",
        ee_example_field="hkEeExampleRow",
        fix_main_ee_refs=fix_hk_row_hk_ee_refs,
        fix_ee_main_refs=fix_ee_row_hk_refs,
        main_snapshots=main_snapshots,
        ee_snapshots=ee_snapshots,
    )


def parse_source(source_path: Path) -> dict[str, Any]:
    mapping = _active_mapping()
    src_spec = mapping.get("sourceEmployeeSheet") if isinstance(mapping.get("sourceEmployeeSheet"), dict) else {}
    wb = load_workbook(source_path, data_only=True, read_only=True)
    name = find_sheet_name(list(wb.sheetnames), src_spec) or (
        HK_L_SHEET if HK_L_SHEET in wb.sheetnames else None
    )
    if not name:
        names = wb.sheetnames
        wb.close()
        raise ValueError(f"未找到 sheet「{src_spec.get('sheet') or HK_L_SHEET}」，现有: {names}")

    ws = wb[name]
    employees = read_hk_l_employees(ws)
    meta = read_hk_l_meta(ws)
    wb.close()
    return {"employees": employees, "meta": meta}


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    global _ACTIVE_MAPPING
    _ACTIVE_MAPPING = resolve_convert_mapping("hk_payroll_calc", convert_mapping)
    try:
        return _convert_impl(
            source_path,
            output_path,
            template_path,
            pn_meta=pn_meta,
            registry_dir=registry_dir,
            employee_directory=employee_directory,
        )
    finally:
        _ACTIVE_MAPPING = None


def _convert_impl(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if not template_path.is_file():
        raise FileNotFoundError(f"母版不存在: {template_path}")
    if not source_path.is_file():
        raise FileNotFoundError(f"原始账单不存在: {source_path}")

    parsed = parse_source(source_path)
    employees = parsed["employees"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, rich_text=True)
    for name in (HK_L_SHEET, HK_SHEET, HK_EE_SHEET, PN_SHEET):
        if name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"母版缺少 sheet「{name}」，现有: {wb.sheetnames}")

    applied_pn: PnMeta | None = None
    if pn_meta is not None:
        applied_pn = apply_pn_meta(
            wb,
            pn_meta,
            registry_dir=registry_dir or output_path.parent,
            reserve_invoice_number=True,
        )

    write_hk_l(wb[HK_L_SHEET], employees, parsed["meta"])

    mapping = _active_mapping()
    formula_rows = _hk_formula_rows()
    need_hk, need_ee = needed_example_rows_for_styles(
        mapping,
        employees,
        main_template_key="Hong Kong",
        ee_template_key="Hong Kong EE",
        main_example_field="hkExampleRow",
        ee_example_field="hkEeExampleRow",
        main_data_start=HK_DATA_START_ROW,
        ee_data_start=HK_EE_DATA_START_ROW,
        employee_directory=employee_directory,
    )
    main_snaps = {r: snapshot_row_cells(wb[HK_SHEET], r) for r in need_hk}
    ee_snaps = {r: snapshot_row_cells(wb[HK_EE_SHEET], r) for r in need_ee}
    fit_hk_formula_sheets(
        wb,
        len(employees),
        clear_excess=False,
        protected_hk_rows=need_hk,
        protected_ee_rows=need_ee,
    )
    formula_plan = apply_hk_employee_formula_styles(
        wb,
        employees,
        formula_rows=formula_rows,
        employee_directory=employee_directory,
        main_snapshots=main_snaps,
        ee_snapshots=ee_snaps,
    )
    clear_excess_hk_formula_rows(wb, len(employees))

    fx_row_before = _find_pn_fx_row(wb[PN_SHEET]) if PN_SHEET in wb.sheetnames else 28
    pn_layout = fit_pn_employees(wb[PN_SHEET], len(employees))
    fx_row = int(pn_layout.get("fx_row") or 28)
    fx_rate = None
    fx_source = None
    # 优先供应商源表 Hong Kong!B3（Exchange rate HKD per USD）
    try:
        src_wb = load_workbook(source_path, data_only=True)
        try:
            hk_name = None
            for n in src_wb.sheetnames:
                if n.strip().lower() in ("hong kong", "hongkong"):
                    hk_name = n
                    break
            if hk_name:
                raw = src_wb[hk_name]["B3"].value
                if isinstance(raw, (int, float)) and float(raw) > 0:
                    fx_rate = float(raw)
                    fx_source = "vendor:Hong Kong!B3"
        finally:
            src_wb.close()
    except Exception:
        pass
    if fx_rate is None:
        try:
            # 公式未缓存时再读非 data_only
            src_wb = load_workbook(source_path, data_only=False)
            try:
                hk_name = None
                for n in src_wb.sheetnames:
                    if n.strip().lower() in ("hong kong", "hongkong"):
                        hk_name = n
                        break
                if hk_name:
                    cell = src_wb[hk_name]["B3"].value
                    text = str(cell or "").strip()
                    if text.startswith("="):
                        # 常见 =7.81*0.97：取乘积若可解析，否则回退 API
                        body = text[1:].replace(" ", "")
                        if "*" in body:
                            parts = body.split("*")
                            try:
                                fx_rate = float(parts[0]) * float(parts[1])
                                fx_source = "vendor:Hong Kong!B3(formula)"
                            except ValueError:
                                fx_rate = None
                    elif isinstance(cell, (int, float)) and float(cell) > 0:
                        fx_rate = float(cell)
                        fx_source = "vendor:Hong Kong!B3"
            finally:
                src_wb.close()
        except Exception:
            pass
    if fx_rate is None:
        rates = fetch_usd_rates()
        fx_rate = get_hk_pn_fx_rate(rates)
        fx_source = "api:HKD*0.97"
    wb[PN_SHEET].cell(fx_row, 2).value = fx_rate
    retarget_pn_fx_refs(wb, fx_row, from_rows=[int(fx_row_before), 28, 29, 30, 31])
    from fx_policy import make_pn_fx_provenance

    write_source = "api" if str(fx_source or "").startswith("api:") or str(fx_source or "").startswith("vendor:") else "mapping"
    pn_fx_write = make_pn_fx_provenance(
        PN_SHEET,
        fx_row,
        2,
        mapping,
        float(fx_rate),
        write_source=write_source,
        fx_source=str(fx_source or ""),
    )

    ensure_hk_period_date_formats(wb)
    wb.save(output_path)
    wb.close()
    postprocess_converted_xlsx(output_path)

    styles = mapping.get("employeeFormulaStyles") if isinstance(mapping.get("employeeFormulaStyles"), list) else []
    warnings: list[str] = [f"映射员工公式样式条数: {len(styles)}"]
    for p in formula_plan or []:
        warnings.append(
            f"公式配对：第{p.get('index')}人 → Hong Kong第{p.get('mainExampleRow')}行 / Hong Kong EE第{p.get('eeExampleRow')}行"
        )

    return {
        "employee_count": len(employees),
        "employee_names": [
            e.get("Name of Employee") or e.get("EE Name") or e.get("Name")
            for e in employees
        ],
        "company_name": parsed["meta"].get("company_name"),
        "period": (parsed["meta"].get("period_from"), parsed["meta"].get("period_to")),
        "fx_rate": fx_rate,
        "fx_source": fx_source,
        "output": str(output_path),
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
        "warnings": warnings,
        "mapping_style_count": len(styles),
        "formula_main_rows": ",".join(
            str(p.get("mainExampleRow")) for p in (formula_plan or []) if p.get("mainExampleRow") is not None
        ),
        "pn_fx_write": pn_fx_write,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Hong Kong T-N 账单 → PN N-C 自动转换")
    parser.add_argument("source", type=Path, help="原始 T-N Excel 路径")
    parser.add_argument("-o", "--output", type=Path, help="输出 PN 路径")
    parser.add_argument("-t", "--template", type=Path, default=DEFAULT_TEMPLATE, help="PN 母版路径")
    args = parser.parse_args(argv)

    source = args.source.resolve()
    output = args.output.resolve() if args.output else source.parent / f"PN_auto_{source.stem}.xlsx"

    try:
        result = convert(source, output, args.template.resolve())
    except Exception as exc:
        print(f"转换失败: {exc}", file=sys.stderr)
        return 1

    print("转换完成")
    print(f"  输出: {result['output']}")
    print(f"  公司: {result['company_name']}")
    print(f"  账期: {result['period'][0]} ~ {result['period'][1]}")
    print(f"  员工: {result['employee_count']} 人 → {result['employee_names']}")
    print(f"  汇率 PN!B28: {result['fx_rate']} ({result.get('fx_source')})")
    if result.get("formula_main_rows"):
        print(f"  公式行: {result['formula_main_rows']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
