# -*- coding: utf-8 -*-
"""分析 hk_payroll_calc 样例账单字段结构，输出 structure.json。"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[2]
PROFILE_DIR = Path(__file__).resolve().parent
SRC = ROOT / r"账单\now\Hong Kong\PN_NNRoad_UECorp_T-N 20260318-v1.xlsx"
OUT = ROOT / r"账单\now\Hong Kong\PN_UECorp_N-C 20260318 revised.xlsx"
OUT_FILE = PROFILE_DIR / "structure.json"


def norm(v):
    if v is None:
        return None
    return str(v).strip()


def scan_hk_l(ws, data_only=False):
    pairs = []
    for c in range(1, (ws.max_column or 0) + 1):
        k = ws.cell(1, c).value
        if k is not None:
            pairs.append({
                "col": get_column_letter(c),
                "field": norm(k),
                "row2": ws.cell(2, c).value,
                "row3": ws.cell(3, c).value if ws.max_row >= 3 else None,
            })
    return pairs


def scan_row_formulas(ws, row, max_col=80):
    cells = []
    for c in range(1, min(max_col, (ws.max_column or 0) + 1)):
        cell = ws.cell(row, c)
        v = cell.value
        if v is not None:
            item = {"col": get_column_letter(c), "v": str(v)[:150]}
            if cell.data_type == "f":
                item["formula"] = str(v)
            cells.append(item)
    return cells


def diff_hk_l(src_pairs, out_pairs):
    src_map = {p["field"]: p for p in src_pairs}
    out_map = {p["field"]: p for p in out_pairs}
    all_keys = sorted(set(src_map) | set(out_map))
    diffs = []
    for k in all_keys:
        s = src_map.get(k, {})
        o = out_map.get(k, {})
        if s.get("row2") != o.get("row2"):
            diffs.append({
                "field": k,
                "src_row2": s.get("row2"),
                "out_row2": o.get("row2"),
            })
    return diffs


def main():
    r = {}

    wb_src = load_workbook(SRC, data_only=True)
    wb_out = load_workbook(OUT, data_only=True)
    wb_src_f = load_workbook(SRC, data_only=False)
    wb_out_f = load_workbook(OUT, data_only=False)

    r["src_hk_l"] = scan_hk_l(wb_src["Hong Kong-L"])
    r["out_hk_l"] = scan_hk_l(wb_out["Hong Kong-L"])
    r["hk_l_diffs"] = diff_hk_l(r["src_hk_l"], r["out_hk_l"])

    # employee count in HK-L
    hk_l = wb_out["Hong Kong-L"]
    ee_rows = []
    for row in range(2, min(15, (hk_l.max_row or 0) + 1)):
        vals = [hk_l.cell(row, c).value for c in range(1, 6)]
        if any(v for v in vals):
            ee_rows.append({"row": row, "preview": vals})
    r["out_hk_l_data_rows"] = ee_rows

    # Hong Kong sheet row 9 formulas
    r["src_hk_row9"] = scan_row_formulas(wb_src_f["Hong Kong"], 9)
    r["out_hk_row9"] = scan_row_formulas(wb_out_f["Hong Kong"], 9)

    # Hong Kong EE row 10
    r["src_hk_ee_row10"] = scan_row_formulas(wb_src_f["Hong Kong EE"], 10)
    r["out_hk_ee_row10"] = scan_row_formulas(wb_out_f["Hong Kong EE"], 10)

    # PN key cells
    for addr in ["B8", "B29", "C49"]:
        r[f"src_pn_{addr}"] = wb_src_f["PN"][addr].value
        r[f"out_pn_{addr}"] = wb_out_f["PN"][addr].value

    # Compare Hong Kong row9 col by col values (data_only)
    row9_diff = []
    ws_s = wb_src["Hong Kong"]
    ws_o = wb_out["Hong Kong"]
    for c in range(1, 60):
        vs, vo = ws_s.cell(9, c).value, ws_o.cell(9, c).value
        if vs != vo and (vs or vo):
            h = ws_o.cell(4, c).value or ws_o.cell(3, c).value
            row9_diff.append({"col": get_column_letter(c), "header": norm(h), "src": vs, "out": vo})
    r["hong_kong_row9_value_diffs"] = row9_diff[:30]

    wb_src.close()
    wb_out.close()
    wb_src_f.close()
    wb_out_f.close()

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(r, f, ensure_ascii=False, indent=2, default=str)
    print("written", OUT_FILE)
    print("HK-L field count src/out:", len(r["src_hk_l"]), len(r["out_hk_l"]))
    print("HK-L diffs count:", len(r["hk_l_diffs"]))
    print("Employee data rows:", r["out_hk_l_data_rows"])


if __name__ == "__main__":
    main()
