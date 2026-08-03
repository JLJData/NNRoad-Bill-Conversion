# -*- coding: utf-8 -*-
"""
UAE-L 横向源账单 → UAE PN（引擎 uae_payroll_calc）

用法:
  python -m profiles.uae_payroll_calc.convert <源.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

源账单: sheet「UAE-L」第 2 行表头、第 3 行起员工（可由 auxilium_uae ingest 产出）。
默认母版: templates/uae/template.xlsx
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from bill_convert.formula_copy import shift_row_formula
from convert_mapping import find_sheet_name, resolve_convert_mapping
from fx_rate import fetch_usd_rates, get_uae_pn_fx_rate
from pn_meta import PnMeta, apply_pn_meta
from profiles.tw_payroll_calc.convert import match_ee_code
from region_templates import get_region_template
from xlsx_convert_utils import coerce_datetime_for_excel, is_date_column_header
from xlsx_luckysheet_compat import apply_luckysheet_compat_uae
from xlsx_postprocess import postprocess_converted_xlsx

DEFAULT_TEMPLATE = get_region_template("UAE")

UAE_L_SHEET = "UAE-L"
UAE_SHEET = "UAE"
UAE_EE_SHEET = "UAE EE"
PN_SHEET = "PN"

UAE_L_HEADER_ROW = 2
UAE_L_DATA_START = 3
UAE_DATA_START = 9
UAE_EE_DATA_START = 10
MAX_EMPLOYEES = 20

# 与 TW/HK 一致：强制日期格式，避免网页把 Excel 序列显示成 46,113.00
_DATE_FMT = "yyyy/m/d"

_ACTIVE_MAPPING: dict[str, Any] | None = None


def _active_mapping() -> dict[str, Any]:
    return (
        _ACTIVE_MAPPING
        if isinstance(_ACTIVE_MAPPING, dict)
        else resolve_convert_mapping("uae_payroll_calc", None)
    )


def _uae_l_layout(*, target: bool = False) -> tuple[int, int, list[str]]:
    """返回 (header_row, data_start, name_headers)。Connect 母版多为 7/8 + English Name。"""
    mapping = _active_mapping()
    key = "targetL" if target else "sourceEmployeeSheet"
    spec = mapping.get(key) if isinstance(mapping.get(key), dict) else {}
    if not spec and target:
        spec = mapping.get("sourceEmployeeSheet") if isinstance(mapping.get("sourceEmployeeSheet"), dict) else {}
    header = int(spec.get("headerRow") or UAE_L_HEADER_ROW)
    data_start = int(spec.get("dataStartRow") or UAE_L_DATA_START)
    names = spec.get("nameHeaders") if isinstance(spec.get("nameHeaders"), list) else None
    name_headers = [str(x).strip() for x in (names or ["Employee Name", "English Name"]) if str(x).strip()]
    if not name_headers:
        name_headers = ["Employee Name", "English Name"]
    return header, data_start, name_headers


def _emp_display_name(emp: dict[str, Any]) -> str:
    for key in ("Employee Name", "English Name"):
        name = _norm(emp.get(key))
        if name:
            return name
    return ""


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("AED", "").replace("\xa0", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        h = _norm(ws.cell(header_row, col).value)
        if h and h not in out:
            out[h] = col
    return out


def _cell_formula_text(value: Any) -> str | None:
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _set_cell_value(cell, value: Any) -> None:
    """写入值；若目标是 ArrayFormula 则改为普通公式/值。"""
    cell.value = value


def parse_uae_l_employees(ws: Worksheet) -> list[dict[str, Any]]:
    header_row, data_start, name_headers = _uae_l_layout(target=False)
    headers = _header_map(ws, header_row)
    if not headers:
        raise ValueError(f"「UAE-L」第 {header_row} 行表头为空")
    employees: list[dict[str, Any]] = []
    for row in range(data_start, (ws.max_row or data_start) + 1):
        name = None
        for nh in name_headers:
            if nh in headers:
                name = ws.cell(row, headers[nh]).value
                if _norm(name):
                    break
        emp_id = None
        if "Emp ID" in headers:
            emp_id = ws.cell(row, headers["Emp ID"]).value
        name_s = _norm(name)
        id_s = _norm(emp_id)
        if not name_s and not id_s:
            continue
        if name_s.upper().startswith("TOTAL") or id_s.upper().startswith("TOTAL"):
            continue
        row_data: dict[str, Any] = {}
        for h, col in headers.items():
            row_data[h] = ws.cell(row, col).value
        employees.append(row_data)
    return employees


def clear_uae_l_data(ws: Worksheet) -> None:
    """清空员工数据区。保留母版首行上的公式列结构由 write_uae_l 按「只清数据列」处理。"""
    _, data_start, _ = _uae_l_layout(target=True)
    max_row = max(ws.max_row or data_start, data_start + MAX_EMPLOYEES)
    max_col = min(ws.max_column or 1, 64)
    for row in range(data_start, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ == "MergedCell":
                continue
            cell.value = None


def _uae_l_formula_cols(ws: Worksheet, data_start: int) -> dict[int, str]:
    """母版首名员工行上的公式列 → 公式文本。"""
    out: dict[int, str] = {}
    for col in range(1, max((ws.max_column or 1), 1) + 1):
        text = _cell_formula_text(ws.cell(data_start, col).value)
        if text:
            out[col] = text
    return out


def ensure_uae_period_date_formats(wb) -> None:
    """UAE!B2/C2 引用 UAE-L 账期；强制日期格式，避免网页显示成 46,113.00。"""
    try:
        uae_l = wb[UAE_L_SHEET]
        if _norm(uae_l.cell(2, 1).value).lower().startswith("payroll"):
            for col in (3, 5):
                uae_l.cell(2, col).number_format = _DATE_FMT
        header_row, data_start, _ = _uae_l_layout(target=True)
        headers = _header_map(uae_l, header_row)
        for key in ("From", "To"):
            col = headers.get(key)
            if not col:
                continue
            for row in range(data_start, data_start + MAX_EMPLOYEES):
                cell = uae_l.cell(row, col)
                if type(cell).__name__ == "MergedCell":
                    continue
                if cell.value is None and row > data_start:
                    break
                cell.number_format = _DATE_FMT
    except KeyError:
        pass
    try:
        uae = wb[UAE_SHEET]
        for col in (2, 3):
            uae.cell(2, col).number_format = _DATE_FMT
    except KeyError:
        pass


def write_uae_l_sheet_meta(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    """Connect 风格 UAE-L：C1 客户名、C2/E2 账期（写 datetime，勿写文本）。"""
    if not employees:
        return
    emp0 = employees[0]
    if _norm(ws.cell(1, 1).value).lower().startswith("company"):
        client = _norm(emp0.get("Client")) or _norm(emp0.get("_client"))
        if client:
            ws.cell(1, 3).value = client
    if _norm(ws.cell(2, 1).value).lower().startswith("payroll"):
        for key, col in (("From", 3), ("To", 5)):
            dt = coerce_datetime_for_excel(emp0.get(key))
            if dt is None:
                continue
            cell = ws.cell(2, col)
            cell.value = dt
            cell.number_format = _DATE_FMT


def write_uae_l(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    """
    写入 UAE-L：只覆盖数据列，母版公式列原样保留/扩行复制。
    避免「整行清空再补公式」——从源头不破坏母版公式。
    """
    header_row, data_start, _ = _uae_l_layout(target=True)
    headers = _header_map(ws, header_row)
    if not headers:
        raise ValueError(f"「UAE-L」第 {header_row} 行表头为空")
    write_uae_l_sheet_meta(ws, employees)

    n = len(employees)
    formula_by_col = _uae_l_formula_cols(ws, data_start)
    max_col = max(max(headers.values(), default=1), max(formula_by_col.keys(), default=1), 1)

    # 多人：先按首行复制样式+公式，再只改数据格
    for i in range(1, n):
        _copy_row_style_and_formula(
            ws,
            data_start,
            data_start + i,
            max_col=max_col,
            l_from=data_start,
            l_to=data_start + i,
        )

    # 只清空「非公式」数据列；多余员工行整行清空
    last_keep = data_start + max(n, 1) - 1
    max_row = max(ws.max_row or data_start, data_start + MAX_EMPLOYEES)
    for row in range(data_start, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ == "MergedCell":
                continue
            if row <= last_keep and col in formula_by_col:
                continue
            cell.value = None

    # 公式行被多余清空后：确保前 n 行公式仍在（首行未清；扩行已复制）
    for i in range(n):
        row = data_start + i
        for col, formula in formula_by_col.items():
            cur = ws.cell(row, col).value
            if _cell_formula_text(cur):
                continue
            ws.cell(row, col).value = shift_row_formula(
                formula,
                data_start,
                row,
                target_l_from=data_start,
                target_l_to=row,
                target_l_sheet=UAE_L_SHEET,
            )

    for idx, emp in enumerate(employees):
        row = data_start + idx
        for h, col in headers.items():
            if col in formula_by_col:
                continue
            if h not in emp:
                continue
            val = emp[h]
            if val is None:
                continue
            cell = ws.cell(row, col)
            if is_date_column_header(h) or h in ("From", "To"):
                dt = coerce_datetime_for_excel(val)
                if dt is not None:
                    cell.value = dt
                    cell.number_format = _DATE_FMT
                    continue
            if isinstance(val, datetime):
                cell.value = val
                cell.number_format = _DATE_FMT
            else:
                _set_cell_value(cell, val)
        # Auxilium：母版若无 Gross 公式则生成
        gross_col = headers.get("EC - Gross Salary")
        if gross_col and gross_col not in formula_by_col:
            parts = []
            for key in (
                "EC - Basic Salary",
                "EC - Housing Allowance",
                "EC - Transport Allowance",
                "EC - School Allowance",
                "EC - Other allowance",
                "EC - Mobile Allowance",
                "EC - Food Allowance",
            ):
                c = headers.get(key)
                if c:
                    parts.append(ws.cell(row, c).coordinate)
            if parts:
                ws.cell(row, gross_col).value = "=" + "+".join(parts)


def _copy_row_style_and_formula(
    ws: Worksheet,
    src_row: int,
    dest_row: int,
    max_col: int = 40,
    *,
    l_from: int | None = None,
    l_to: int | None = None,
) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dest = ws.cell(dest_row, c)
        if type(dest).__name__ == "MergedCell" or type(src).__name__ == "MergedCell":
            continue
        if src.has_style:
            dest.font = copy(src.font)
            dest.border = copy(src.border)
            dest.fill = copy(src.fill)
            dest.number_format = src.number_format
            dest.protection = copy(src.protection)
            dest.alignment = copy(src.alignment)
        text = _cell_formula_text(src.value)
        if text:
            dest.value = shift_row_formula(
                text,
                src_row,
                dest_row,
                target_l_from=l_from if l_from is not None else -1,
                target_l_to=l_to if l_to is not None else -1,
                target_l_sheet=UAE_L_SHEET,
            )
        elif src.value is not None and not isinstance(src.value, ArrayFormula):
            dest.value = src.value


def _retarget_ee_refs(formula: str, ee_from: int, ee_to: int) -> str:
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    return re.sub(
        rf"('UAE EE'!\$?[A-Z]{{1,3}})\$?{ee_from}(?!\d)",
        lambda m: f"{m.group(1)}{ee_to}",
        formula,
        flags=re.I,
    )


def expand_uae_employee_rows(wb, employee_count: int) -> None:
    n = max(int(employee_count), 1)
    _, l_data_start, _ = _uae_l_layout(target=True)
    if UAE_SHEET in wb.sheetnames:
        uae = wb[UAE_SHEET]
        for i in range(1, n):
            dest = UAE_DATA_START + i
            l_row = l_data_start + i
            ee_row = UAE_EE_DATA_START + i
            _copy_row_style_and_formula(
                uae,
                UAE_DATA_START,
                dest,
                max_col=40,
                l_from=l_data_start,
                l_to=l_row,
            )
            for c in range(1, 41):
                cell = uae.cell(dest, c)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = _retarget_ee_refs(cell.value, UAE_EE_DATA_START, ee_row)
            # 姓名列等公式已由上面从母版首行复制并平移，勿再硬编码覆盖

    if UAE_EE_SHEET in wb.sheetnames:
        ee = wb[UAE_EE_SHEET]
        for i in range(1, n):
            dest = UAE_EE_DATA_START + i
            l_row = l_data_start + i
            _copy_row_style_and_formula(
                ee,
                UAE_EE_DATA_START,
                dest,
                max_col=40,
                l_from=l_data_start,
                l_to=l_row,
            )
            # EE 公式（含姓名引用）同样只复制平移，不硬写 E 列


# ---------- PN 多人：Labor×n + Expense×n（无 Office Rental）----------

_PN_LABOR_START = 16
_PN_AED_FMT = "#,##0.00"
_PN_USD_FMT = "$#,##0.00"


def _find_pn_row_by_label(ws: Worksheet, keyword: str, col: int = 1) -> int | None:
    key = keyword.lower()
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and key in v.lower():
            return row
    return None


def _copy_pn_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    max_col = max(ws.max_column or 6, 6)
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _ensure_merge_a_c(ws: Worksheet, row: int) -> None:
    rng = f"A{row}:C{row}"
    for m in list(ws.merged_cells.ranges):
        if m.min_row <= row <= m.max_row and m.min_col <= 1 and m.max_col >= 3:
            if m.min_row == row and m.max_row == row and m.min_col == 1 and m.max_col == 3:
                return
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    try:
        ws.merge_cells(rng)
    except ValueError:
        pass


def _collect_merges_from(ws: Worksheet, start_row: int) -> list[tuple[int, int, int, int]]:
    kept: list[tuple[int, int, int, int]] = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row < start_row:
            continue
        kept.append((m.min_row, m.min_col, m.max_row, m.max_col))
        try:
            ws.unmerge_cells(str(m))
        except ValueError:
            pass
    return kept


def _restore_merges(
    ws: Worksheet,
    merges: list[tuple[int, int, int, int]],
    row_shift: int = 0,
) -> None:
    for min_r, min_c, max_r, max_c in merges:
        nr1, nr2 = min_r + row_shift, max_r + row_shift
        if nr1 < 1 or nr2 < nr1:
            continue
        try:
            ws.merge_cells(
                start_row=nr1,
                start_column=min_c,
                end_row=nr2,
                end_column=max_c,
            )
        except ValueError:
            pass


def _shift_row_heights(ws: Worksheet, start_row: int, amount: int) -> None:
    if amount == 0:
        return
    captured: dict[int, float] = {}
    max_r = max(ws.max_row or start_row, start_row)
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            captured[r] = h
            max_r = max(max_r, r)
    for r in range(start_row, max_r + abs(amount) + 2):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    if amount > 0:
        for r in sorted(captured.keys(), reverse=True):
            ws.row_dimensions[r + amount].height = captured[r]
    else:
        for r in sorted(captured.keys()):
            new_r = r + amount
            if new_r >= start_row:
                ws.row_dimensions[new_r].height = captured[r]


def _pn_insert_rows(ws: Worksheet, idx: int, amount: int, fill_style_row: int | None = None) -> None:
    if amount <= 0:
        return
    merges = _collect_merges_from(ws, idx)
    _shift_row_heights(ws, idx, amount)
    ws.insert_rows(idx, amount)
    _restore_merges(ws, merges, row_shift=amount)
    if fill_style_row is not None:
        style_row = fill_style_row + amount if fill_style_row >= idx else fill_style_row
        for r in range(idx, idx + amount):
            _copy_pn_row_style(ws, style_row, r)
            _ensure_merge_a_c(ws, r)


def _count_pn_labor_slots(ws: Worksheet) -> int:
    n = 0
    for row in range(_PN_LABOR_START, _PN_LABOR_START + MAX_EMPLOYEES + 2):
        v = ws.cell(row, 1).value
        if isinstance(v, str) and "Labor cost" in v:
            n += 1
            continue
        break
    return max(n, 1)


def _pn_snapshot_row_formulas(ws: Worksheet, row: int, max_col: int = 6) -> dict[int, str]:
    """快照一行公式文本（col → formula）。"""
    out: dict[int, str] = {}
    for col in range(1, max_col + 1):
        text = _cell_formula_text(ws.cell(row, col).value)
        if text:
            out[col] = text
    return out


def _pn_stash_sheet_refs(formula: str) -> tuple[str, list[str]]:
    placeholders: list[str] = []

    def stash(match: re.Match[str]) -> str:
        placeholders.append(match.group(0))
        return f"__SHEETREF{len(placeholders) - 1}__"

    # 'Sheet'!A1 或 Sheet!A1
    s = re.sub(r"'(?:[^']+)'!\$?[A-Z]{1,3}\$?\d+", stash, formula)
    s = re.sub(r"(?<![A-Z])[A-Za-z][A-Za-z0-9 ]*!\$?[A-Z]{1,3}\$?\d+", stash, s)
    return s, placeholders


def _pn_restore_sheet_refs(body: str, placeholders: list[str]) -> str:
    for idx, ref in enumerate(placeholders):
        body = body.replace(f"__SHEETREF{idx}__", ref)
    return body


def _pn_shift_local_rows(formula: str, from_row: int, to_row: int) -> str:
    """只平移本表相对行引用；跨表引用原样保留。"""
    if from_row == to_row or not (isinstance(formula, str) and formula.startswith("=")):
        return formula
    body, refs = _pn_stash_sheet_refs(formula)

    def repl(m: re.Match[str]) -> str:
        return f"{m.group(1)}{to_row}"

    body = re.sub(
        rf"(?<!\$)(?<![A-Z])([A-Z]{{1,3}}){from_row}(?!\d)",
        repl,
        body,
    )
    return _pn_restore_sheet_refs(body, refs)


def _pn_retarget_uae_emp_row(formula: str, from_emp: int, to_emp: int) -> str:
    """UAE!X{from_emp} → UAE!X{to_emp}；账期/合计等其它行号不动。"""
    if from_emp == to_emp or not (isinstance(formula, str) and formula.startswith("=")):
        return formula

    def repl(m: re.Match[str]) -> str:
        return f"{m.group(1)}{m.group(2)}{to_emp}"

    return re.sub(
        rf"(UAE!|'UAE'!)(\$?[A-Z]{{1,3}}\$?){from_emp}(?!\d)",
        repl,
        formula,
    )


def _pn_retarget_fx_row(formula: str, old_fx: int, new_fx: int) -> str:
    if old_fx == new_fx or not (isinstance(formula, str) and formula.startswith("=")):
        return formula
    out = formula
    out = re.sub(rf"\$B\${old_fx}(?!\d)", f"$B${new_fx}", out)
    out = re.sub(rf"(?<!\$)B{old_fx}(?!\d)", f"B{new_fx}", out)
    out = re.sub(
        rf"('?PN'?!?)\$?B\$?{old_fx}(?!\d)",
        lambda m: f"{m.group(1)}B{new_fx}" if "!" in m.group(1) else f"$B${new_fx}",
        out,
    )
    # 'PN'!B28 / PN!B28
    out = re.sub(rf"'PN'!\$?B\$?{old_fx}(?!\d)", f"'PN'!B{new_fx}", out)
    out = re.sub(rf"(?<![A-Z])PN!\$?B\$?{old_fx}(?!\d)", f"PN!B{new_fx}", out)
    return out


def _pn_remap_local_rows(formula: str, remap: dict[int, int]) -> str:
    """按行映射表平移本表引用（扩/缩行后修正合计区公式文本）。"""
    if not remap or not (isinstance(formula, str) and formula.startswith("=")):
        return formula
    body, refs = _pn_stash_sheet_refs(formula)

    def repl(m: re.Match[str]) -> str:
        row = int(m.group(2))
        new_row = remap.get(row, row)
        return f"{m.group(1)}{new_row}"

    body = re.sub(r"(?<!\$)(?<![A-Z])([A-Z]{1,3})(\d+)", repl, body)
    # $A$12 这类绝对引用也跟行走（FX 旁合计常用）
    body = re.sub(
        r"\$([A-Z]{1,3})\$(\d+)",
        lambda m: f"${m.group(1)}${remap.get(int(m.group(2)), int(m.group(2)))}",
        body,
    )
    return _pn_restore_sheet_refs(body, refs)


def _pn_expand_row_remap(
    *,
    labor_start: int,
    expense_start_old: int,
    svc_old: int,
    delta: int,
    max_row: int = 80,
) -> dict[int, int]:
    """
    delta>0：先在 expense 前插 Labor，再在 Service 前插 Expense →
      r < expense_start_old → 不动
      expense_start_old <= r < svc_old → +delta
      r >= svc_old → +2*delta
    delta<0：对称收缩。
    """
    remap: dict[int, int] = {}
    if delta == 0:
        return remap
    for r in range(1, max_row + 1):
        if delta > 0:
            if r < expense_start_old:
                remap[r] = r
            elif r < svc_old:
                remap[r] = r + delta
            else:
                remap[r] = r + 2 * delta
        else:
            # 先删多余 Expense 再删多余 Labor（与 _pn_delete_rows 顺序一致）
            shrink = -delta
            if r < labor_start + (expense_start_old - labor_start - shrink):
                # 简化：labor 保留 n 行后的映射在 delete 后由标签重找；这里给合计区用
                pass
            if r < expense_start_old:
                # labor 区：删的是 labor_start+n .. expense-1，保留行号不变
                remap[r] = r
            elif r < svc_old:
                # 原 expense 区整体上移 shrink（删 labor 后再删 expense 各 shrink）
                remap[r] = r - shrink
            else:
                remap[r] = r - 2 * shrink
    return remap


def _pn_apply_detail_formulas(
    ws: Worksheet,
    *,
    snapshot: dict[int, str],
    template_row: int,
    dest_row: int,
    uae_from: int,
    uae_to: int,
    old_fx: int,
    new_fx: int,
) -> None:
    for col, formula in snapshot.items():
        text = _pn_shift_local_rows(formula, template_row, dest_row)
        text = _pn_retarget_uae_emp_row(text, uae_from, uae_to)
        text = _pn_retarget_fx_row(text, old_fx, new_fx)
        cell = ws.cell(dest_row, col)
        cell.value = text


def _pn_patch_eor_sum(ws: Worksheet, eor_row: int, labor_start: int, last_detail: int) -> None:
    """EOR 明细合计：若母版是 SUM，只改起止行；否则不动。"""
    for col in (5, 6):
        cell = ws.cell(eor_row, col)
        text = _cell_formula_text(cell.value)
        if not text:
            continue
        letter = "E" if col == 5 else "F"
        if re.search(rf"SUM\(\s*{letter}\d+\s*:\s*{letter}\d+\s*\)", text, re.I):
            cell.value = f"=SUM({letter}{labor_start}:{letter}{last_detail})"


def _pn_layout_dict(ws: Worksheet, n: int) -> dict[str, int]:
    labor_start = _PN_LABOR_START
    expense_start = labor_start + n
    last_eor_detail = expense_start + n - 1
    eor_row = _find_pn_row_by_label(ws, "EOR/PEO Cost") or 15
    fx_row = _find_pn_row_by_label(ws, "FX rate") or 28
    svc = _find_pn_row_by_label(ws, "Service Fee")
    mgmt = _find_pn_row_by_label(ws, "Management Fee")
    if svc is None:
        svc = last_eor_detail + 3
    if mgmt is None:
        mgmt = svc + 1
    total_row = _find_pn_row_by_label(ws, "EOR/PEO Service Cost") or (mgmt + 4)
    return {
        "labor_start": labor_start,
        "expense_start": expense_start,
        "eor_row": eor_row,
        "svc_row": svc,
        "mgmt_row": mgmt,
        "fx_row": fx_row,
        "total_row": total_row,
        "employee_count": n,
    }


def fit_uae_pn_employees(wb, employee_count: int) -> dict[str, int]:
    """
    PN 按人数扩/缩 Labor + Expense。

    原则：母版公式是唯一真相——不发明 I/S/AE/H6 等列引用。
    - 人数 == 母版槽位：整张 PN 公式不动，只返回布局。
    - 人数变化：从母版首人 Labor/Expense 公式复制并平移行号；合计区按行映射修正文本。
    """
    if PN_SHEET not in wb.sheetnames:
        return {}
    ws = wb[PN_SHEET]
    n = max(int(employee_count), 1)
    if n > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {n} 超过模板上限 {MAX_EMPLOYEES}")

    old_n = _count_pn_labor_slots(ws)
    labor_start = _PN_LABOR_START
    expense_start_old = labor_start + old_n
    svc_old = _find_pn_row_by_label(ws, "Service Fee") or 20
    fx_old = _find_pn_row_by_label(ws, "FX rate") or 28
    delta = n - old_n

    if delta == 0:
        return _pn_layout_dict(ws, n)

    labor_snap = _pn_snapshot_row_formulas(ws, labor_start)
    expense_snap = _pn_snapshot_row_formulas(ws, expense_start_old)
    # 合计/结算区：扩行前快照，扩行后按行映射写回（不换成硬编码业务公式）
    summary_rows = {
        "eor": _find_pn_row_by_label(ws, "EOR/PEO Cost"),
        "svc": svc_old,
        "mgmt": _find_pn_row_by_label(ws, "Management Fee"),
        "total": _find_pn_row_by_label(ws, "EOR/PEO Service Cost"),
        "fx": fx_old,
    }
    summary_snaps: dict[str, dict[int, str]] = {}
    for key, row in summary_rows.items():
        if row:
            summary_snaps[key] = _pn_snapshot_row_formulas(ws, row)
    settle_snaps: dict[int, dict[int, str]] = {}
    for off in range(1, 8):
        settle_snaps[off] = _pn_snapshot_row_formulas(ws, fx_old + off)

    remap = _pn_expand_row_remap(
        labor_start=labor_start,
        expense_start_old=expense_start_old,
        svc_old=svc_old,
        delta=delta,
        max_row=max(ws.max_row or 40, fx_old + 10),
    )

    if delta > 0:
        _pn_insert_rows(ws, expense_start_old, delta, fill_style_row=labor_start)
        svc_row = _find_pn_row_by_label(ws, "Service Fee") or (svc_old + delta)
        expense_start = labor_start + n
        _pn_insert_rows(ws, svc_row, delta, fill_style_row=expense_start)
    else:

        def _pn_delete_rows(wss: Worksheet, idx: int, amount: int) -> None:
            if amount <= 0:
                return
            merges = _collect_merges_from(wss, idx + amount)
            wss.delete_rows(idx, amount)
            _restore_merges(wss, merges, row_shift=-amount)

        _pn_delete_rows(ws, expense_start_old + n, -delta)
        _pn_delete_rows(ws, labor_start + n, -delta)

    layout = _pn_layout_dict(ws, n)
    expense_start = layout["expense_start"]
    last_eor_detail = expense_start + n - 1
    fx_row = layout["fx_row"]
    eor_row = layout["eor_row"]
    svc = layout["svc_row"]
    mgmt = layout["mgmt_row"]
    total_row = layout["total_row"]

    for i in range(n):
        labor_row = labor_start + i
        expense_row = expense_start + i
        uae_to = UAE_DATA_START + i
        if i > 0:
            _copy_pn_row_style(ws, labor_start, labor_row)
            _copy_pn_row_style(ws, expense_start, expense_row)
        _ensure_merge_a_c(ws, labor_row)
        _ensure_merge_a_c(ws, expense_row)
        _pn_apply_detail_formulas(
            ws,
            snapshot=labor_snap,
            template_row=labor_start,
            dest_row=labor_row,
            uae_from=UAE_DATA_START,
            uae_to=uae_to,
            old_fx=fx_old,
            new_fx=fx_row,
        )
        _pn_apply_detail_formulas(
            ws,
            snapshot=expense_snap,
            template_row=expense_start_old,
            dest_row=expense_row,
            uae_from=UAE_DATA_START,
            uae_to=uae_to,
            old_fx=fx_old,
            new_fx=fx_row,
        )

    # 合计区：母版公式 + 行映射；EOR 的 SUM 起止扩到新明细末行
    label_to_row = {
        "eor": eor_row,
        "svc": svc,
        "mgmt": mgmt,
        "total": total_row,
        "fx": fx_row,
    }
    for key, snap in summary_snaps.items():
        dest = label_to_row.get(key)
        if not dest:
            continue
        for col, formula in snap.items():
            text = _pn_remap_local_rows(formula, remap)
            text = _pn_retarget_fx_row(text, fx_old, fx_row)
            ws.cell(dest, col).value = text
    for off, snap in settle_snaps.items():
        dest = fx_row + off
        for col, formula in snap.items():
            text = _pn_remap_local_rows(formula, remap)
            text = _pn_retarget_fx_row(text, fx_old, fx_row)
            ws.cell(dest, col).value = text

    _pn_patch_eor_sum(ws, eor_row, labor_start, last_eor_detail)
    _ensure_merge_a_c(ws, svc)
    _ensure_merge_a_c(ws, mgmt)

    return layout


def set_period(wb, employees: list[dict[str, Any]]) -> None:
    if not employees or UAE_SHEET not in wb.sheetnames:
        return
    emp = employees[0]
    uae = wb[UAE_SHEET]
    _, data_start, _ = _uae_l_layout(target=True)
    for key, col in (("From", 2), ("To", 3)):
        cell = uae.cell(2, col)
        existing = cell.value
        # Connect 母版 UAE!B2 已引用 UAE-L!C2，勿强行覆盖公式格
        if isinstance(existing, str) and existing.startswith("="):
            cell.number_format = _DATE_FMT
            continue
        dt = coerce_datetime_for_excel(emp.get(key))
        if dt is not None:
            cell.value = dt
            cell.number_format = _DATE_FMT
        elif emp.get(key) is None:
            # 兜底：指向 UAE-L 首名员工账期（Auxilium F/G）
            from_col = "F" if col == 2 else "G"
            cell.value = f"='{UAE_L_SHEET}'!{from_col}{data_start}"
            cell.number_format = _DATE_FMT


def set_recurring_fees(wb, employees: list[dict[str, Any]]) -> None:
    """UAE!H = Admin Fee × 1.5（Auxilium）。Connect 无 Admin Fees，保留母版 Recurring 公式。"""
    if UAE_SHEET not in wb.sheetnames:
        return
    if _resolve_pdf_profile_id(_active_mapping()) == "connect_uae":
        return
    uae = wb[UAE_SHEET]
    for i, emp in enumerate(employees):
        admin = _as_float(emp.get("EC - Admin Fees"))
        if admin is None:
            admin = _as_float(emp.get("Admin Fees"))
        if admin is None:
            admin = _as_float(emp.get("_admin_fee"))
        if admin is None:
            continue
        fee = round(admin * 1.5, 6)
        uae.cell(UAE_DATA_START + i, 8).value = fee


def _resolve_pdf_profile_id(mapping: dict[str, Any] | None) -> str | None:
    if not isinstance(mapping, dict):
        return None
    for key in ("pdfProfileId", "_pdfProfileId"):
        val = mapping.get(key)
        if val is not None and str(val).strip():
            return str(val).strip()
    from bill_convert.fact_store import get_batch_facts

    batch = get_batch_facts(mapping)
    if any(str(k).startswith("auxilium.") for k in batch):
        return "auxilium_uae"
    return None


def _apply_vendor_plugins(wb, warnings: list[str], *, employee_count: int = 1) -> dict[str, Any]:
    """按 pdfProfile 加载供应商旁路插件（如 Auxilium Admin Fee → Business Tax）。"""
    from bill_convert.fact_store import get_batch_facts
    from bill_convert.vendor_plugins.runtime import apply_vendor_plugins

    mapping = _active_mapping()
    pdf_profile_id = _resolve_pdf_profile_id(mapping)
    if not pdf_profile_id:
        return {}
    return apply_vendor_plugins(
        wb,
        pdf_profile_id=pdf_profile_id,
        mapping=mapping,
        batch_facts=get_batch_facts(mapping),
        warnings=warnings,
        employee_count=employee_count,
    ) or {}


def _retarget_pn_fx_refs_in_formula(formula: str, fx_row: int) -> str:
    """母版公式里的 PN!B{n} / 'PN'!B{n} 统一指到当前 FX 行，保留原引用写法。"""
    if not (isinstance(formula, str) and formula.startswith("=")):
        return formula
    fx = max(int(fx_row), 1)

    def repl(m: re.Match[str]) -> str:
        return f"{m.group(1)}B{fx}"

    return re.sub(r"('PN'!|PN!)\$?B\$?\d+", repl, formula)


def normalize_uae_other_fee_block(wb, *, fx_row: int = 28) -> None:
    """
    Other Fee 区（E22:E26）：
    - 空单元格补 0；已有公式/数值不覆盖
    - E25 Bank Charges：以母版为准（如 10+49.99/PN!B29），仅重定向到当前 FX 行
    """
    if UAE_SHEET not in wb.sheetnames:
        return
    uae = wb[UAE_SHEET]
    fx = max(int(fx_row or 28), 1)
    for row, default in ((22, 0), (23, 0), (24, 0), (26, 0)):
        cell = uae.cell(row, 5)
        val = cell.value
        if val is None or val == "":
            cell.value = default
    bank = uae.cell(25, 5)
    text = _cell_formula_text(bank.value)
    if text:
        bank.value = _retarget_pn_fx_refs_in_formula(text, fx)
    # 母版是常数（如默认模板 E25=10）则保持，不发明 10+50
    e21 = uae.cell(21, 5)
    if e21.value is None or e21.value == "":
        e21.value = "=SUM(E22:E26)"
    elif isinstance(e21.value, str) and e21.value.startswith("="):
        pass
    for row in (23, 24):
        fcell = uae.cell(row, 6)
        if fcell.value == 0:
            fcell.value = None


def _pn_customer_id(pn_meta: PnMeta | dict[str, Any] | None) -> str | None:
    if pn_meta is None:
        return None
    if isinstance(pn_meta, PnMeta):
        cid = (pn_meta.customer_id or "").strip()
        return cid or None
    if isinstance(pn_meta, dict):
        cid = str(pn_meta.get("customer_id") or pn_meta.get("customerId") or "").strip()
        return cid or None
    return None


def apply_uae_ee_codes(
    wb,
    employees: list[dict[str, Any]],
    *,
    employee_directory: list[dict[str, Any]] | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
) -> list[str]:
    """
    UAE EE!B = Client Code（库客户编号，同 PN!B9）
    UAE EE!D = EE Code（按姓名匹配客户员工目录）
    """
    if UAE_EE_SHEET not in wb.sheetnames:
        return []
    ws = wb[UAE_EE_SHEET]
    client_code = _pn_customer_id(pn_meta)
    directory = list(employee_directory or [])
    warnings: list[str] = []

    for i, emp in enumerate(employees):
        row = UAE_EE_DATA_START + i
        # Client Code：仅在有客户编号时写入数值；母版公式/空单元格一律不发明引用
        if client_code:
            ws.cell(row, 2).value = client_code
        # Client Name：母版公式保留（Connect 常用 UAE-L!C1；Auxilium 常用 PN!B8），勿覆盖

        excel_names = [_emp_display_name(emp)]
        code, warn = match_ee_code([n for n in excel_names if n], directory)
        ws.cell(row, 4).value = code  # 匹配不到显式清空，避免母版残留
        if warn:
            warnings.append(f"UAE EE 第{i + 1}人：{warn}")
    return warnings


def apply_fx(wb, *, fill_fx: bool = True, fx_row: int | None = None) -> float | None:
    if not fill_fx or PN_SHEET not in wb.sheetnames:
        return None
    rates = fetch_usd_rates()
    fx = get_uae_pn_fx_rate(rates)
    row = fx_row or _find_pn_row_by_label(wb[PN_SHEET], "FX rate") or 28
    cell = wb[PN_SHEET].cell(row, 2)
    # 母版若是公式（如 =3.6725*0.97）则保留，勿覆盖成裸数值
    if _cell_formula_text(cell.value):
        return fx
    cell.value = fx
    return fx


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    registry_dir: Path | None = None,
    convert_mapping: dict[str, Any] | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    global _ACTIVE_MAPPING
    _ACTIVE_MAPPING = resolve_convert_mapping("uae_payroll_calc", convert_mapping)
    try:
        source_path = Path(source_path).resolve()
        output_path = Path(output_path).resolve()
        template_path = Path(template_path).resolve()
        if not source_path.is_file():
            raise FileNotFoundError(f"源文件不存在: {source_path}")
        if not template_path.is_file():
            raise FileNotFoundError(f"母版不存在: {template_path}")

        src_wb = load_workbook(source_path, data_only=False)
        try:
            l_name = find_sheet_name(list(src_wb.sheetnames), _active_mapping().get("sourceEmployeeSheet"))
            if not l_name or l_name not in src_wb.sheetnames:
                if UAE_L_SHEET in src_wb.sheetnames:
                    l_name = UAE_L_SHEET
                else:
                    raise ValueError(f"未找到 UAE-L，现有: {src_wb.sheetnames}")
            employees = parse_uae_l_employees(src_wb[l_name])
        finally:
            src_wb.close()

        if not employees:
            raise ValueError("UAE-L 未解析到员工行")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)
        warnings: list[str] = []
        fact_store_updates: dict[str, Any] = {}
        fx = None
        applied_pn = None
        try:
            if UAE_L_SHEET not in wb.sheetnames:
                raise ValueError(f"母版缺少 {UAE_L_SHEET}")
            write_uae_l(wb[UAE_L_SHEET], employees)
            # UAE 姓名列若母版已是公式（='UAE-L'!B…）则保留，勿用文本覆盖
            set_period(wb, employees)
            expand_uae_employee_rows(wb, len(employees))
            set_recurring_fees(wb, employees)
            fact_store_updates = _apply_vendor_plugins(wb, warnings, employee_count=len(employees))
            pn_layout = fit_uae_pn_employees(wb, len(employees))
            try:
                fx = apply_fx(wb, fill_fx=fill_fx, fx_row=pn_layout.get("fx_row"))
            except Exception as exc:
                warnings.append(f"写入 PN 汇率失败: {exc}")
            # 必须在 PN 扩行 + 写汇率之后，Bank Charges 才能指向正确的 B{fx}
            normalize_uae_other_fee_block(wb, fx_row=int(pn_layout.get("fx_row") or 28))

            if pn_meta is not None:
                applied_pn = apply_pn_meta(
                    wb,
                    pn_meta,
                    registry_dir=registry_dir or output_path.parent,
                    reserve_invoice_number=True,
                )
            warnings.extend(
                apply_uae_ee_codes(
                    wb,
                    employees,
                    employee_directory=employee_directory,
                    pn_meta=applied_pn or pn_meta,
                )
            )
            # Table1/ArrayFormula → 普通 A1，避免核对页 LuckySheet/HF 一直转圈
            apply_luckysheet_compat_uae(wb)
            ensure_uae_period_date_formats(wb)
            wb.save(output_path)
        finally:
            wb.close()

        postprocess_converted_xlsx(output_path)
        return {
            "ok": True,
            "engine_id": "uae_payroll_calc",
            "region": "UAE",
            "output": str(output_path),
            "employee_count": len(employees),
            "fx_rate": fx,
            "warnings": warnings,
            "pn_meta": applied_pn.to_dict() if applied_pn else None,
            "fact_store_updates": fact_store_updates,
        }
    finally:
        _ACTIVE_MAPPING = None


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="UAE-L → UAE PN")
    parser.add_argument("source", type=Path)
    parser.add_argument("-o", "--output", type=Path)
    parser.add_argument("-t", "--template", type=Path, default=None)
    args = parser.parse_args(argv)
    src = args.source.resolve()
    out = args.output.resolve() if args.output else src.parent / f"PN_UAE_{src.stem}.xlsx"
    tpl = (args.template or DEFAULT_TEMPLATE).resolve()
    try:
        result = convert(src, out, tpl)
    except Exception as exc:
        print(f"失败: {exc}", file=sys.stderr)
        return 1
    print("完成", result.get("output"), "人数", result.get("employee_count"), "FX", result.get("fx_rate"))
    for w in result.get("warnings") or []:
        print(" !", w)
    return 0


if __name__ == "__main__":
    sys.exit(main())
