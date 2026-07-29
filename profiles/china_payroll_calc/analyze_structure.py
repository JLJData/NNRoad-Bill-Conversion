# -*- coding: utf-8 -*-
"""分析 china_payroll_calc 样例账单字段结构，输出 structure.json。"""
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parents[2]
PROFILE_DIR = Path(__file__).resolve().parent
SRC = ROOT / r"账单\now\China\HROne HK Payment Notice-NNRoad-Hermetic Solutions-202603C.xlsx"
OUT = ROOT / r"账单\now\China\PN_Hermetic Solutions Group_N-C 20260318.xlsx"
OUT_FILE = PROFILE_DIR / "structure.json"


def read_headers_2level(ws, r1=1, r2=2, max_col=60):
    cols = []
    for c in range(1, max_col + 1):
        h1 = ws.cell(r1, c).value
        h2 = ws.cell(r2, c).value
        if h1 or h2:
            cols.append({
                "col": c,
                "letter": get_column_letter(c),
                "parent": str(h1).strip() if h1 else None,
                "header": str(h2).strip() if h2 else (str(h1).strip() if h1 else None),
            })
    return cols


def main():
    r = {}

    wb_src = load_workbook(SRC, data_only=True)
    r["source_sheets"] = wb_src.sheetnames

    # find calc sheet - last one that looks like 计算结果 or contains 计算
    calc_sheet = None
    for n in wb_src.sheetnames:
        if "计算" in n or n.endswith("结果"):
            calc_sheet = n
    if not calc_sheet:
        # fallback: sheet index for garbled name - user said 计算结果
        for n in wb_src.sheetnames:
            if wb_src.sheetnames.index(n) == 16:  # from list position
                calc_sheet = n
    r["calc_sheet_name"] = calc_sheet

    if calc_sheet:
        ws = wb_src[calc_sheet]
        headers = read_headers_2level(ws, 1, 1, 80)  # might be single row
        if len(headers) < 5:
            headers = read_headers_2level(ws, 1, 2, 80)
        r["calc_headers"] = headers
        # data from row 2+
        data_rows = []
        for row in range(2, min(ws.max_row + 1, 6)):
            row_data = {}
            for h in headers[:40]:
                v = ws.cell(row, h["col"]).value
                if v is not None:
                    row_data[h["header"]] = v
            if row_data:
                data_rows.append({"row": row, "data": row_data})
        r["calc_sample_rows"] = data_rows

    ws_other = wb_src["Other Fee"]
    r["other_fee"] = {
        "C15": ws_other["C15"].value,
        "D19": ws_other["D19"].value,
    }
    ws_pn_src = wb_src["S-Payment Notice"]
    r["payment_notice_C49"] = ws_pn_src["C49"].value

    wb_src.close()

    wb_out = load_workbook(OUT, data_only=False)
    r["output_sheets"] = wb_out.sheetnames

    china_l_name = next((n for n in wb_out.sheetnames if "China-L" in n or n == "China-L"), None)
    china_name = next((n for n in wb_out.sheetnames if n == "China"), None)
    china_ee_name = next((n for n in wb_out.sheetnames if "China EE" in n), None)
    pn_name = next((n for n in wb_out.sheetnames if n == "PN"), None)

    if china_l_name:
        ws = wb_out[china_l_name]
        cols = read_headers_2level(ws, 1, 2, 80)
        r["china_l"] = {
            "sheet": china_l_name,
            "headers": cols,
            "row3_values": {h["header"]: ws.cell(3, h["col"]).value for h in cols if ws.cell(3, h["col"]).value is not None},
        }

    if china_name:
        ws = wb_out[china_name]
        cols = read_headers_2level(ws, 1, 2, 80)
        specials = {}
        for h in cols:
            if h["header"] and ("Other" in h["header"] or "Expense" in h["header"] or "For" in (h["header"] or "")):
                cell = ws.cell(3, h["col"])
                specials[h["header"]] = {
                    "col": h["letter"],
                    "parent": h["parent"],
                    "value": cell.value,
                    "is_formula": cell.data_type == "f",
                }
        r["china"] = {"headers": cols, "specials_row3": specials}

    if pn_name:
        ws = wb_out[pn_name]
        c = ws["B29"]
        r["pn_B29"] = {"value": c.value, "formula": c.value if c.data_type == "f" else None}

    if china_ee_name:
        ws = wb_out[china_ee_name]
        ee_rows = []
        for row in range(1, ws.max_row + 1):
            vals = [ws.cell(row, c).value for c in range(1, 8)]
            if any(v for v in vals):
                ee_rows.append({"row": row, "A-G": vals})
        r["china_ee"] = {"sheet": china_ee_name, "rows": ee_rows[:15], "max_row": ws.max_row}

    wb_out.close()

    # mapping match: source calc headers vs china-l headers via vendor mapping
    calc_h = {h["header"] for h in r.get("calc_headers", []) if h.get("header")}
    china_l_h = {h["header"] for h in r.get("china_l", {}).get("headers", []) if h.get("header")}
    r["match"] = {
        "calc_count": len(calc_h),
        "china_l_count": len(china_l_h),
        "matched": sorted(calc_h & china_l_h),
        "calc_only": sorted(calc_h - china_l_h)[:20],
        "china_l_only": sorted(china_l_h - calc_h)[:20],
    }

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(r, f, ensure_ascii=False, indent=2, default=str)

    print("written", OUT_FILE)


if __name__ == "__main__":
    main()
