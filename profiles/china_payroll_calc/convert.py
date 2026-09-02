# -*- coding: utf-8 -*-
"""
大陆「计算结果」源账单 → China PN（引擎 china_payroll_calc）

用法:
  python -m profiles.china_payroll_calc.convert <原始账单.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

默认母版: templates/china/template.xlsx
（含 PN / China / China EE / China-L 公式结构）
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from pathlib import Path
from typing import Any

from pn_meta import PnMeta, apply_pn_meta

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from fx_policy import api_fx_for_currency, fx_policy
from fx_rate import fetch_usd_rates, get_china_pn_fx_rate
from region_templates import get_region_template
from xlsx_convert_utils import clean_value, norm
from xlsx_luckysheet_compat import apply_luckysheet_compat
from xlsx_postprocess import postprocess_converted_xlsx
from convert_mapping import find_sheet_name, resolve_convert_mapping
from bill_convert.formula_copy import (
    copy_row_formulas as shared_copy_row_formulas,
    fix_china_row_china_ee_refs,
    fix_ee_row_china_refs,
    retarget_pn_fx_b_column_refs,
    snapshot_row_cells,
)
from bill_convert.formula_layout import (
    apply_employee_formula_styles,
    needed_example_rows_for_styles,
    sort_employees_by_code,
    tw_l_row_for_data_row,
)
from bill_convert.formula_layout import _default_example_row as default_example_row_for_mapping
from bill_convert.headers import list_qualified_header_cells

DEFAULT_TEMPLATE = get_region_template("China")

CALC_SHEET_NAMES = ("计算结果",)
OTHER_FEE_NAMES = ("Other Fee",)
PAYMENT_NOTICE_NAMES = ("S-Payment Notice", "Payment Notice", "付款通知")
_FX_LABEL_KEYS = (
    "汇率",
    "兑换率",
    "exchange rate",
    "fx rate",
    "usd/cny",
    "usd to cny",
    "美元汇率",
)
_SIMPLE_CELL_REF_RE = re.compile(r"^\$?([A-Za-z]+)\$?(\d+)$")
_CNY_USD_MIN, _CNY_USD_MAX = 4.0, 12.0

CHINA_SHEET = "China"
CHINA_EE_SHEET = "China EE"
CHINA_L_SHEET = "China-L"
CHINA_DATA_START_ROW = 9
CHINA_EE_DATA_START_ROW = 10
CHINA_L_DATA_START_ROW = 2
MAX_EMPLOYEES = 10

_ACTIVE_MAPPING: dict[str, Any] | None = None


def _active_mapping() -> dict[str, Any]:
    return _ACTIVE_MAPPING if isinstance(_ACTIVE_MAPPING, dict) else resolve_convert_mapping("china_payroll_calc", None)


def _china_l_data_start() -> int:
    target = _active_mapping().get("targetL")
    if isinstance(target, dict) and target.get("dataStartRow") is not None:
        return int(target["dataStartRow"])
    return CHINA_L_DATA_START_ROW


def _china_formula_rows() -> dict[str, int]:
    l_start = _china_l_data_start()
    return {
        "l_data_start": l_start,
        "main_data_start": CHINA_DATA_START_ROW,
        "ee_data_start": CHINA_EE_DATA_START_ROW,
        # 兼容旧键名（若公共层回退）
        "tw_l_data_start": l_start,
        "tw_data_start": CHINA_DATA_START_ROW,
        "tw_ee_data_start": CHINA_EE_DATA_START_ROW,
    }


def find_sheet(wb, candidates: tuple[str, ...]) -> str | None:
    names = {n: n for n in wb.sheetnames}
    for c in candidates:
        if c in names:
            return c
    lower = {n.lower(): n for n in wb.sheetnames}
    for c in candidates:
        if c.lower() in lower:
            return lower[c.lower()]
    for n in wb.sheetnames:
        for c in candidates:
            if c in n:
                return n
    return None


def build_header_map(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def parse_expense_count(text: Any) -> int:
    if text is None:
        return 0
    m = re.search(r"(\d+)", str(text))
    return int(m.group(1)) if m else 0


def extract_by_label_col_a(ws: Worksheet, label: str, value_col: int) -> Any:
    target = norm(label)
    for row in range(1, (ws.max_row or 0) + 1):
        if norm(ws.cell(row, 1).value) == target:
            return clean_value(ws.cell(row, value_col).value)
    return None


def read_calc_employees(ws: Worksheet) -> list[dict[str, Any]]:
    qualified = list_qualified_header_cells(ws, 1)
    headers = {str(h["key"]): int(h["col"]) for h in qualified}
    child_to_keys: dict[str, list[str]] = {}
    for h in qualified:
        child_to_keys.setdefault(str(h["child"]), []).append(str(h["key"]))

    def _col_for(label: str) -> int | None:
        if label in headers:
            return headers[label]
        for qk in child_to_keys.get(label, []):
            if qk in headers:
                return headers[qk]
        return None

    name_col = _col_for("姓名")
    if name_col is None:
        raise ValueError("「计算结果」sheet 第 1 行须包含表头「姓名」")

    employees: list[dict[str, Any]] = []
    for row in range(2, (ws.max_row or 0) + 1):
        name = clean_value(ws.cell(row, name_col).value)
        if name is None:
            continue
        record: dict[str, Any] = {}
        for hdr, col in headers.items():
            record[hdr] = clean_value(ws.cell(row, col).value)
        employees.append(record)

    if not employees:
        raise ValueError("「计算结果」中未找到有效员工行（姓名非空）")
    if len(employees) > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {len(employees)} 超过模板上限 {MAX_EMPLOYEES}")
    return employees


def clear_china_l_data(ws: Worksheet, from_row: int = CHINA_L_DATA_START_ROW) -> None:
    max_row = max(ws.max_row or from_row, from_row + MAX_EMPLOYEES)
    max_col = ws.max_column or 1
    for row in range(from_row, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def write_china_l(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    target_headers = build_header_map(ws, 1)
    if not target_headers:
        raise ValueError("China-L 第 1 行表头为空")

    rename = _active_mapping().get("columnRename") or {}
    if not isinstance(rename, dict):
        rename = {}
    # China-L 列名 → 源账单列名
    target_to_source = {
        str(v).strip(): str(k).strip()
        for k, v in rename.items()
        if k and v and str(k).strip() != str(v).strip()
    }
    if rename:
        sample_keys: set[str] = set()
        for emp in employees[:8]:
            sample_keys.update(str(k) for k in emp.keys())
        check_column_rename_hits(rename, sample_keys, strict_if_configured=True)

    clear_china_l_data(ws)
    for idx, emp in enumerate(employees):
        row = _china_l_data_start() + idx
        for hdr, col in target_headers.items():
            val = emp.get(hdr)
            if val is None and hdr in target_to_source:
                val = emp.get(target_to_source[hdr])
            if val is not None:
                ws.cell(row, col).value = val


def shift_formula(formula: str, from_row: int, to_row: int, china_l_from: int, china_l_to: int) -> str:
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    s = re.sub(
        r"'China-L'!([A-Z]{1,3})(\d+)",
        lambda m: f"'China-L'!{m.group(1)}{china_l_to}",
        formula,
    )
    s = re.sub(
        rf"(?<!\$)(?<![A-Z])([A-Z]{{1,3}}){from_row}(?!\d)",
        lambda m: f"{m.group(1)}{to_row}",
        s,
    )
    return s


def copy_row_formulas(ws: Worksheet, from_row: int, to_row: int, china_l_from: int, china_l_to: int) -> None:
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        src = ws.cell(from_row, col)
        dst = ws.cell(to_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
        if src.data_type == "f" and isinstance(src.value, str):
            dst.value = shift_formula(src.value, from_row, to_row, china_l_from, china_l_to)
        elif src.value is not None and src.data_type != "f":
            dst.value = copy(src.value)


def clear_employee_row(ws: Worksheet, row: int) -> None:
    for col in range(1, (ws.max_column or 0) + 1):
        ws.cell(row, col).value = None


def count_template_employee_slots(ws: Worksheet, data_start_row: int, marker_col: int) -> int:
    """连续占位：marker 非空，或该行存在公式（第二种公式行 C 列可能为空）。"""
    n = 0
    for row in range(data_start_row, data_start_row + MAX_EMPLOYEES):
        marker = ws.cell(row, marker_col).value
        has_formula = False
        if marker is None:
            for col in range(1, (ws.max_column or 0) + 1):
                cell = ws.cell(row, col)
                if cell.data_type == "f" and cell.value:
                    has_formula = True
                    break
        if marker is not None or has_formula:
            n += 1
            continue
        break
    return max(n, 1)


def fit_china_formula_sheets(
    wb,
    employee_count: int,
    *,
    clear_excess: bool = True,
    protected_china_rows: set[int] | None = None,
    protected_ee_rows: set[int] | None = None,
) -> None:
    """
    母版 China / China EE 预置公式行。
    人数更多时从「默认示例行」（第一行公式）复制扩展；人数不足时可清多余行。

    若随后还要按 employeeFormulaStyles 从示例行盖公式，须先 clear_excess=False，
    再 apply_china_employee_formula_styles，最后 clear_excess_china_formula_rows。
    protected_*：扩行时不得覆盖的示例行（如第二种公式行）。
    """
    n = max(int(employee_count), 1)
    mapping = _active_mapping()
    l_start = _china_l_data_start()
    china = wb[CHINA_SHEET]
    ee = wb[CHINA_EE_SHEET]
    china_slots = count_template_employee_slots(china, CHINA_DATA_START_ROW, marker_col=3)
    ee_slots = count_template_employee_slots(ee, CHINA_EE_DATA_START_ROW, marker_col=4)
    china_tpl = default_example_row_for_mapping(mapping, "China", CHINA_DATA_START_ROW)
    ee_tpl = default_example_row_for_mapping(mapping, "China EE", CHINA_EE_DATA_START_ROW)
    # 保护第二种公式示例行：槽位至少覆盖到这些行，避免扩行时从第 1 种公式盖掉
    prot_c = set(protected_china_rows or ())
    prot_e = set(protected_ee_rows or ())
    prot_c.add(china_tpl)
    prot_e.add(ee_tpl)
    for r in prot_c:
        china_slots = max(china_slots, r - CHINA_DATA_START_ROW + 1)
    for r in prot_e:
        ee_slots = max(ee_slots, r - CHINA_EE_DATA_START_ROW + 1)
    src_china_l = tw_l_row_for_data_row(
        china_tpl, data_start=CHINA_DATA_START_ROW, target_l_data_start=l_start
    )
    src_ee_l = tw_l_row_for_data_row(
        ee_tpl, data_start=CHINA_EE_DATA_START_ROW, target_l_data_start=l_start
    )

    if clear_excess:
        for i in range(n, china_slots):
            clear_employee_row(china, CHINA_DATA_START_ROW + i)
        for i in range(n, ee_slots):
            clear_employee_row(ee, CHINA_EE_DATA_START_ROW + i)

    if n > china_slots:
        for i in range(china_slots, n):
            dst_row = CHINA_DATA_START_ROW + i
            dst_l = l_start + i
            ee_row = CHINA_EE_DATA_START_ROW + i
            shared_copy_row_formulas(
                china, china_tpl, dst_row, src_china_l, dst_l, target_l_sheet=CHINA_L_SHEET
            )
            fix_china_row_china_ee_refs(china, dst_row, ee_row)

    if n > ee_slots:
        for i in range(ee_slots, n):
            dst_row = CHINA_EE_DATA_START_ROW + i
            dst_l = l_start + i
            china_row = CHINA_DATA_START_ROW + i
            shared_copy_row_formulas(
                ee, ee_tpl, dst_row, src_ee_l, dst_l, target_l_sheet=CHINA_L_SHEET
            )
            fix_ee_row_china_refs(ee, dst_row, china_row, data_row_min=CHINA_DATA_START_ROW)


def clear_excess_china_formula_rows(wb, employee_count: int) -> None:
    n = max(int(employee_count), 0)
    china = wb[CHINA_SHEET]
    ee = wb[CHINA_EE_SHEET]
    china_slots = count_template_employee_slots(china, CHINA_DATA_START_ROW, marker_col=3)
    ee_slots = count_template_employee_slots(ee, CHINA_EE_DATA_START_ROW, marker_col=4)
    for i in range(n, china_slots):
        clear_employee_row(china, CHINA_DATA_START_ROW + i)
    for i in range(n, ee_slots):
        clear_employee_row(ee, CHINA_EE_DATA_START_ROW + i)


def apply_china_employee_formula_styles(
    wb,
    employees: list[dict[str, Any]],
    *,
    formula_rows: dict[str, int] | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    main_snapshots: dict[int, list[dict[str, Any]]] | None = None,
    ee_snapshots: dict[int, list[dict[str, Any]]] | None = None,
) -> list[dict[str, Any]]:
    """China / China EE 按人盖公式：配对用 chinaExampleRow，未配对用默认第一行公式。"""
    return apply_employee_formula_styles(
        wb,
        employees,
        _active_mapping(),
        formula_rows=formula_rows or _china_formula_rows(),
        employee_directory=employee_directory,
        main_sheet=CHINA_SHEET,
        ee_sheet=CHINA_EE_SHEET,
        target_l_sheet=CHINA_L_SHEET,
        main_template_key="China",
        ee_template_key="China EE",
        main_example_field="chinaExampleRow",
        ee_example_field="chinaEeExampleRow",
        fix_main_ee_refs=fix_china_row_china_ee_refs,
        fix_ee_main_refs=lambda ws_ee, dst_row, china_row: fix_ee_row_china_refs(
            ws_ee, dst_row, china_row, data_row_min=CHINA_DATA_START_ROW
        ),
        main_snapshots=main_snapshots,
        ee_snapshots=ee_snapshots,
    )


_CHINA_EE_NAME_COL = 2  # China!B = EE Name（China EE!E 引用 =China!B{row}）


def _directory_row_by_employee_code(
    code: Any,
    directory: list[dict[str, Any]],
) -> dict[str, Any] | None:
    want = norm(code)
    if not want:
        return None
    exact: list[dict[str, Any]] = []
    for row in directory:
        if not isinstance(row, dict):
            continue
        got = norm(row.get("employee_code") or row.get("employeeCode"))
        if got and got == want:
            exact.append(row)
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        return exact[0]
    # 兼容库里只存短工号、账单带客户前缀：CUS1525-0002 ↔ 0002
    tail = want.split("-")[-1] if "-" in want else want
    soft: list[dict[str, Any]] = []
    for row in directory:
        if not isinstance(row, dict):
            continue
        got = norm(row.get("employee_code") or row.get("employeeCode"))
        if not got:
            continue
        got_tail = got.split("-")[-1] if "-" in got else got
        if got == tail or got_tail == want or got_tail == tail:
            soft.append(row)
    if len(soft) == 1:
        return soft[0]
    return None


def apply_china_sheet_names_from_directory(
    ws_china: Worksheet,
    employees: list[dict[str, Any]],
    directory: list[dict[str, Any]] | None,
    *,
    china_data_start_row: int = CHINA_DATA_START_ROW,
) -> tuple[list[str], list[str]]:
    """
    写入 China!B（EE Name）：只用员工库名称（按工号匹配），绝不使用供应商账单「姓名」。
    China EE!E = China!B{row}，须在公式扩行之后覆盖母版占位（如 CPT）。
    未匹配或库姓名为空时清空该格并记 warning。

    Returns:
        (warnings, written_names) — written_names 与 employees 对齐，未写入则为 ""。
    """
    warnings: list[str] = []
    written: list[str] = []
    dir_list = [r for r in (directory or []) if isinstance(r, dict)]
    for i, emp in enumerate(employees):
        row = china_data_start_row + i
        code = emp.get("工号")
        cell = ws_china.cell(row, _CHINA_EE_NAME_COL)
        if not dir_list:
            cell.value = None
            written.append("")
            warnings.append(f"第{i + 1}人：未传入员工库，China!B 未写入库名称（工号 {code or '（空）'}）")
            continue
        hit = _directory_row_by_employee_code(code, dir_list)
        if hit is None:
            cell.value = None
            written.append("")
            warnings.append(f"第{i + 1}人：工号 {code or '（空）'} 未在员工库匹配，China!B 未填（不用供应商姓名）")
            continue
        lib_name = (
            str(hit.get("employee_name") or hit.get("employeeName") or "").strip()
            or str(hit.get("employee_name_en") or hit.get("employeeNameEn") or "").strip()
        )
        if not lib_name:
            cell.value = None
            written.append("")
            warnings.append(f"第{i + 1}人：工号 {code} 已匹配员工库，但库中姓名为空，China!B 未填")
            continue
        cell.value = lib_name
        written.append(lib_name)
    return warnings, written


def apply_china_specials(
    ws: Worksheet,
    employee_count: int,
    expense_count: int,
    other_amount: Any,
    *,
    fx_row: int = 29,
) -> None:
    for i in range(employee_count):
        row = CHINA_DATA_START_ROW + i
        # I 列：保留母版/配对公式（如 50*PN!$B$xx*n），只改汇率行与末尾报销笔数
        # 旧逻辑写死 =40*... 会把第二种公式的 50 盖掉
        i_val = ws.cell(row, 9).value
        if isinstance(i_val, str) and i_val.startswith("="):
            updated = re.sub(r"PN!\$?B\$?\d+", f"PN!$B${fx_row}", i_val, flags=re.IGNORECASE)
            if re.search(r"\*\d+\s*$", updated):
                updated = re.sub(r"\*\d+\s*$", f"*{int(expense_count)}", updated)
            ws.cell(row, 9).value = updated
        else:
            ws.cell(row, 9).value = f"=40*PN!$B${fx_row}*{int(expense_count)}"
        # 同步母版 H 列里的汇率绝对行号
        h = ws.cell(row, 8).value
        if isinstance(h, str) and "PN!" in h:
            ws.cell(row, 8).value = re.sub(r"PN!\$?B\$?\d+", f"PN!$B${fx_row}", h)
        if other_amount is not None:
            ws.cell(row, 10).value = other_amount


# ---------- PN 人员扩减 ----------

PN_SHEET = "PN"
_PN_EOR_ROW = 15
_PN_LABOR_START_ROW = 16
_PN_RMB_FMT = "#,##0.00"
_PN_USD_FMT = '$#,##0.00'


def _pn_layout(employee_count: int) -> dict[str, int]:
    n = max(int(employee_count), 1)
    labor_start = _PN_LABOR_START_ROW
    expense_start = labor_start + n
    blank_row = expense_start + n
    service_row = blank_row + 1
    mgmt_row = service_row + 1
    return {
        "n": n,
        "eor_row": _PN_EOR_ROW,
        "labor_start": labor_start,
        "expense_start": expense_start,
        "blank_row": blank_row,
        "service_row": service_row,
        "mgmt_row": mgmt_row,
        "sum_end": expense_start + n - 1,
    }


def _count_pn_labor_slots(ws: Worksheet) -> int:
    n = 0
    for row in range(_PN_LABOR_START_ROW, _PN_LABOR_START_ROW + MAX_EMPLOYEES + 2):
        v = ws.cell(row, 1).value
        if isinstance(v, str) and "Labor cost" in v:
            n += 1
            continue
        break
    return max(n, 1)


def _find_pn_row_by_label(ws: Worksheet, keyword: str, col: int = 1) -> int | None:
    key = keyword.lower()
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and key in v.lower():
            return row
    return None


def _find_pn_fx_row(ws: Worksheet) -> int:
    return _find_pn_row_by_label(ws, "FX rate") or 29


def _copy_pn_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    max_col = max(ws.max_column or 6, 6)
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _ensure_merge_a_c(ws: Worksheet, row: int) -> None:
    rng = f"A{row}:C{row}"
    for m in list(ws.merged_cells.ranges):
        if m.min_row <= row <= m.max_row and m.min_col <= 1 and m.max_col >= 3:
            if m.min_row == row and m.max_row == row and m.min_col == 1 and m.max_col == 3:
                return
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    try:
        ws.merge_cells(rng)
    except ValueError:
        pass


def _collect_merges_from(ws: Worksheet, start_row: int) -> list[tuple[int, int, int, int]]:
    kept: list[tuple[int, int, int, int]] = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row < start_row:
            continue
        kept.append((m.min_row, m.min_col, m.max_row, m.max_col))
        try:
            ws.unmerge_cells(str(m))
        except ValueError:
            pass
    return kept


def _restore_merges(
    ws: Worksheet,
    merges: list[tuple[int, int, int, int]],
    row_shift: int = 0,
) -> None:
    for min_r, min_c, max_r, max_c in merges:
        nr1, nr2 = min_r + row_shift, max_r + row_shift
        if nr1 < 1 or nr2 < nr1:
            continue
        try:
            ws.merge_cells(
                start_row=nr1,
                start_column=min_c,
                end_row=nr2,
                end_column=max_c,
            )
        except ValueError:
            pass


def _shift_row_heights(ws: Worksheet, start_row: int, amount: int) -> None:
    if amount == 0:
        return
    captured: dict[int, float] = {}
    max_r = max(ws.max_row or start_row, start_row)
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            captured[r] = h
        max_r = max(max_r, r)
    for r in range(start_row, max_r + abs(amount) + 2):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    if amount > 0:
        for r in sorted(captured.keys(), reverse=True):
            ws.row_dimensions[r + amount].height = captured[r]
    else:
        for r in sorted(captured.keys()):
            new_r = r + amount
            if new_r >= start_row:
                ws.row_dimensions[new_r].height = captured[r]


def _capture_row_heights_from(ws: Worksheet, start_row: int) -> dict[int, float]:
    out: dict[int, float] = {}
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            out[r] = h
    return out


def _apply_row_heights(ws: Worksheet, heights: dict[int, float], start_row: int) -> None:
    max_r = max([start_row, *heights.keys()], default=start_row)
    for r in range(start_row, max_r + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r, h in heights.items():
        if r >= start_row:
            ws.row_dimensions[r].height = h


def _pn_insert_rows(ws: Worksheet, idx: int, amount: int, fill_style_row: int | None = None) -> None:
    if amount <= 0:
        return
    merges = _collect_merges_from(ws, idx)
    _shift_row_heights(ws, idx, amount)
    ws.insert_rows(idx, amount)
    _restore_merges(ws, merges, row_shift=amount)
    if fill_style_row is not None:
        style_row = fill_style_row + amount if fill_style_row >= idx else fill_style_row
        for r in range(idx, idx + amount):
            _copy_pn_row_style(ws, style_row, r)
            _ensure_merge_a_c(ws, r)


def _pn_delete_rows(ws: Worksheet, idx: int, amount: int) -> None:
    if amount <= 0:
        return
    merges = _collect_merges_from(ws, idx + amount)
    heights = _capture_row_heights_from(ws, idx)
    ws.delete_rows(idx, amount)
    new_heights: dict[int, float] = {}
    for r, h in heights.items():
        if r < idx:
            continue
        if r < idx + amount:
            continue
        new_heights[r - amount] = h
    _apply_row_heights(ws, new_heights, idx)
    _restore_merges(ws, merges, row_shift=-amount)


def _rewrite_pn_settlement_block(ws: Worksheet, *, fx_row: int, total_row: int) -> None:
    """
    中国母版 FX 下方：
      F{fx}=F{total}；F{fx+1..4}=China!E22..E25；Sub/Tax/Total 相对锚点。
    """
    f_eor = ws.cell(fx_row, 6)
    f_eor.value = f"=F{total_row}"
    f_eor.number_format = _PN_USD_FMT
    for i, china_e in enumerate((22, 23, 24, 25), start=1):
        cell = ws.cell(fx_row + i, 6)
        cell.value = f"=China!E{china_e}"
    sub_row = fx_row + 5
    tax_row = fx_row + 6
    grand_row = fx_row + 7
    ws.cell(sub_row, 6).value = f"=SUM(F{fx_row}:F{fx_row + 4})"
    ws.cell(tax_row, 6).value = f"=China!F6/$B${fx_row}"
    ws.cell(grand_row, 6).value = f"=SUM(F{sub_row}:F{tax_row})"
    for r in (sub_row, tax_row, grand_row):
        ws.cell(r, 6).number_format = _PN_USD_FMT


def fit_pn_employees(ws: Worksheet, employee_count: int) -> dict[str, int]:
    """按人数扩/缩 PN Labor+Expense，并重写 Service/Deposit/合计/汇率区公式。"""
    n = max(int(employee_count), 1)
    if n > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {n} 超过模板上限 {MAX_EMPLOYEES}")

    old_n = _count_pn_labor_slots(ws)
    labor_start = _PN_LABOR_START_ROW
    expense_start = labor_start + old_n
    delta = n - old_n

    if delta > 0:
        _pn_insert_rows(ws, expense_start, delta, fill_style_row=labor_start)
        expense_start += delta
        _pn_insert_rows(ws, expense_start + old_n, delta, fill_style_row=expense_start)
    elif delta < 0:
        _pn_delete_rows(ws, expense_start + n, -delta)
        _pn_delete_rows(ws, labor_start + n, -delta)

    layout = _pn_layout(n)
    expense_start = layout["expense_start"]
    blank_row = layout["blank_row"]
    eor_row = layout["eor_row"]
    sum_end = layout["sum_end"]

    for m in list(ws.merged_cells.ranges):
        if m.min_row == blank_row and m.max_row == blank_row:
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    for col in range(1, 7):
        cell = ws.cell(blank_row, col)
        if type(cell).__name__ == "MergedCell":
            continue
        cell.value = None
    _ensure_merge_a_c(ws, blank_row)

    fx_row = _find_pn_fx_row(ws)

    for i in range(n):
        china_row = CHINA_DATA_START_ROW + i
        labor_row = labor_start + i
        expense_row = expense_start + i
        if i > 0:
            _copy_pn_row_style(ws, labor_start, labor_row)
            _copy_pn_row_style(ws, expense_start, expense_row)
        _ensure_merge_a_c(ws, labor_row)
        _ensure_merge_a_c(ws, expense_row)

        ws.cell(labor_row, 1).value = (
            f'="- Labor cost for "&China!B{china_row}&"  -  "&MONTH(China!$B$2)&"-"&YEAR(China!$B$2)'
        )
        e_labor = ws.cell(labor_row, 5)
        e_labor.value = f"=China!K{china_row}+China!AQ{china_row}+China!BC{china_row}"
        e_labor.number_format = _PN_RMB_FMT
        f_labor = ws.cell(labor_row, 6)
        f_labor.value = f"=E{labor_row}/$B${fx_row}"
        f_labor.number_format = _PN_USD_FMT

        ws.cell(expense_row, 1).value = f'="- Expense claim for "&China!B{china_row}'
        e_exp = ws.cell(expense_row, 5)
        e_exp.value = f"=China!BF{china_row}"
        e_exp.number_format = _PN_RMB_FMT
        f_exp = ws.cell(expense_row, 6)
        f_exp.value = f"=E{expense_row}/$B${fx_row}"
        f_exp.number_format = _PN_USD_FMT

    e_eor = ws.cell(eor_row, 5)
    e_eor.value = f"=SUM(E{labor_start}:E{sum_end})"
    e_eor.number_format = _PN_RMB_FMT
    f_eor = ws.cell(eor_row, 6)
    f_eor.value = f"=SUM(F{labor_start}:F{sum_end})"
    f_eor.number_format = _PN_USD_FMT

    svc = _find_pn_row_by_label(ws, "Service Fee") or layout["service_row"]
    mgmt = _find_pn_row_by_label(ws, "Management Fee") or layout["mgmt_row"]
    e_svc = ws.cell(svc, 5)
    e_svc.value = f"=SUM(E{mgmt}:E{mgmt + 1})"
    e_svc.number_format = _PN_RMB_FMT
    f_svc = ws.cell(svc, 6)
    f_svc.value = f"=SUM(F{mgmt}:F{mgmt + 1})"
    f_svc.number_format = _PN_USD_FMT
    ws.cell(mgmt, 1).value = '="- Management Fee - "&MONTH(China!$B$2)&"-"&YEAR(China!$B$2)'
    e_mgmt = ws.cell(mgmt, 5)
    e_mgmt.value = "=China!G6"
    e_mgmt.number_format = _PN_RMB_FMT
    f_mgmt = ws.cell(mgmt, 6)
    f_mgmt.value = f"=E{mgmt}/$B${fx_row}"
    f_mgmt.number_format = _PN_USD_FMT
    _ensure_merge_a_c(ws, svc)
    _ensure_merge_a_c(ws, mgmt)

    deposit_hdr = _find_pn_row_by_label(ws, "Recurring Deposit")
    deposit_row = None
    if deposit_hdr:
        deposit_row = deposit_hdr + 1
        e_dh = ws.cell(deposit_hdr, 5)
        e_dh.value = f"=SUM(E{deposit_row})"
        e_dh.number_format = _PN_RMB_FMT
        f_dh = ws.cell(deposit_hdr, 6)
        f_dh.value = f"=SUM(F{deposit_row})"
        f_dh.number_format = _PN_USD_FMT
        # Deposit 只有一行：始终保留完整描述公式（勿因人数>1 简化成 "=- Deposit"）
        ws.cell(deposit_row, 1).value = (
            '="- Deposit  for "&China!B9&"  -  "&MONTH(China!B2)&"-"&YEAR(China!B2)'
        )
        e_dep = ws.cell(deposit_row, 5)
        e_dep.value = "=China!BI6"
        e_dep.number_format = _PN_RMB_FMT
        f_dep = ws.cell(deposit_row, 6)
        f_dep.value = f"=E{deposit_row}/$B${fx_row}"
        f_dep.number_format = _PN_USD_FMT

    total_row = _find_pn_row_by_label(ws, "EOR/PEO Service Cost")
    if total_row is None or total_row <= eor_row:
        total_row = (deposit_hdr or svc) + 3
    e_total = ws.cell(total_row, 5)
    if deposit_hdr:
        e_total.value = f"=E{eor_row}+E{svc}+E{deposit_hdr}"
    else:
        e_total.value = f"=E{eor_row}+E{svc}"
    e_total.number_format = _PN_RMB_FMT
    f_total = ws.cell(total_row, 6)
    if deposit_hdr:
        f_total.value = f"=F{eor_row}+F{svc}+F{deposit_hdr}"
    else:
        f_total.value = f"=F{eor_row}+F{svc}"
    f_total.number_format = _PN_USD_FMT

    fx_row = _find_pn_fx_row(ws)
    _rewrite_pn_settlement_block(ws, fx_row=fx_row, total_row=total_row)
    for row in range(labor_start, fx_row + 8):
        fcell = ws.cell(row, 6)
        if isinstance(fcell.value, str) and "$B$" in fcell.value:
            fcell.value = re.sub(r"\$B\$\d+", f"$B${fx_row}", fcell.value)

    layout["service_row"] = svc
    layout["mgmt_row"] = mgmt
    layout["deposit_hdr"] = deposit_hdr or 0
    layout["deposit_row"] = deposit_row or 0
    layout["total_row"] = total_row
    layout["fx_row"] = fx_row
    return layout



def retarget_pn_fx_refs(wb, fx_row: int, *, from_rows: list[int] | None = None) -> None:
    """插删 PN 行后，仅把指向旧汇率行的引用改到新行。

    勿动 Client Name/Code（PN!B8 / PN!B9 及绝对 $B$8/$B$9）。
    """
    retarget_pn_fx_b_column_refs(
        wb,
        fx_row,
        from_rows=from_rows if from_rows is not None else (28, 29, 30, 31, 32, 33),
        pn_sheet=PN_SHEET,
    )


def _plausible_cny_per_usd(value: float) -> bool:
    return _CNY_USD_MIN <= float(value) <= _CNY_USD_MAX


def _coerce_fx_rate(value: Any) -> float | None:
    cleaned = clean_value(value)
    if isinstance(cleaned, (int, float)) and not isinstance(cleaned, bool):
        n = float(cleaned)
        return n if n > 0 else None
    text = norm(value)
    if not text:
        return None
    if text.startswith("="):
        body = text[1:].replace(" ", "").replace(",", "")
        try:
            n = float(body)
            return n if n > 0 else None
        except ValueError:
            pass
        if "*" in body and all(ch not in body for ch in "/()"):
            parts = body.split("*")
            if len(parts) == 2:
                try:
                    n = float(parts[0]) * float(parts[1])
                    return n if n > 0 else None
                except ValueError:
                    return None
        return None
    return None


def _fx_from_cell(ws: Worksheet, addr: str) -> float | None:
    try:
        raw = ws[addr].value
    except Exception:
        return None
    fx = _coerce_fx_rate(raw)
    if fx is not None:
        return fx
    text = norm(raw)
    if text.startswith("="):
        body = text[1:].replace("$", "").replace(" ", "")
        m = _SIMPLE_CELL_REF_RE.fullmatch(body)
        if m:
            try:
                return _coerce_fx_rate(ws[f"{m.group(1)}{m.group(2)}"].value)
            except Exception:
                return None
    return None


def _scan_fx_by_label(ws: Worksheet) -> tuple[float | None, str | None]:
    max_r = min(ws.max_row or 0, 80)
    max_c = min(ws.max_column or 0, 12)
    for row in range(1, max_r + 1):
        for col in range(1, max_c + 1):
            label = norm(ws.cell(row, col).value).lower()
            if not label or not any(k in label for k in _FX_LABEL_KEYS):
                continue
            for dc in (1, 2, 3):
                fx = _coerce_fx_rate(ws.cell(row, col + dc).value)
                if fx is not None and _plausible_cny_per_usd(fx):
                    return fx, f"label:{ws.cell(row, col).value}"
    return None, None


def _payment_sheet_hints() -> tuple[str, ...]:
    policy = fx_policy(_active_mapping())
    hints = policy.get("sourceSheetHints") if isinstance(policy.get("sourceSheetHints"), list) else []
    names = [str(x).strip() for x in hints if str(x).strip()]
    return tuple(names) if names else PAYMENT_NOTICE_NAMES


def _payment_fx_cell() -> str:
    policy = fx_policy(_active_mapping())
    cell = str(policy.get("sourceCell") or "").strip().upper()
    return cell or "C49"


def read_vendor_fx(source_path: Path, *, payment_name: str | None = None) -> tuple[float | None, str | None]:
    """供应商账单汇率：先读映射格（默认 S-Payment Notice!C49），公式无缓存则再读公式，再扫「汇率」标签。"""
    hints = _payment_sheet_hints()
    cell = _payment_fx_cell()
    sheet = payment_name
    fx = None
    source = None

    def resolve_sheet(wb) -> str | None:
        return find_sheet(wb, hints) if not sheet else (sheet if sheet in wb.sheetnames else find_sheet(wb, hints))

    try:
        wb = load_workbook(source_path, data_only=True, read_only=True)
        try:
            name = resolve_sheet(wb)
            if name:
                fx = _fx_from_cell(wb[name], cell)
                if fx is not None:
                    source = f"vendor:{name}!{cell}"
        finally:
            wb.close()
    except Exception:
        pass

    if fx is None:
        try:
            wb = load_workbook(source_path, data_only=False)
            try:
                name = resolve_sheet(wb)
                if name:
                    fx = _fx_from_cell(wb[name], cell)
                    if fx is not None:
                        source = f"vendor:{name}!{cell}(formula)"
                    if fx is None:
                        fx, lab = _scan_fx_by_label(wb[name])
                        if fx is not None:
                            source = f"vendor:{name}!{lab}"
            finally:
                wb.close()
        except Exception:
            pass

    if fx is not None and not _plausible_cny_per_usd(fx):
        return None, None
    return fx, source


def parse_source(source_path: Path) -> dict[str, Any]:
    mapping = _active_mapping()
    src_spec = mapping.get("sourceEmployeeSheet") if isinstance(mapping.get("sourceEmployeeSheet"), dict) else {}
    wb = load_workbook(source_path, data_only=True, read_only=True)
    calc_name = find_sheet_name(list(wb.sheetnames), src_spec) or find_sheet(wb, CALC_SHEET_NAMES)
    if not calc_name:
        names = wb.sheetnames
        wb.close()
        want = str(src_spec.get("sheet") or "计算结果")
        raise ValueError(f"未找到 sheet「{want}」，现有: {names}")

    employees = read_calc_employees(wb[calc_name])

    other_name = find_sheet(wb, OTHER_FEE_NAMES)
    payment_name = find_sheet(wb, _payment_sheet_hints())

    other_amount = None
    expense_count = 0
    if other_name:
        other_ws = wb[other_name]
        other_amount = extract_by_label_col_a(other_ws, "其他", 4)
        expense_text = extract_by_label_col_a(other_ws, "报销服务费", 2)
        expense_count = parse_expense_count(expense_text)

    wb.close()

    vendor_fx_rate, vendor_fx_src = read_vendor_fx(source_path, payment_name=payment_name)
    policy = fx_policy(mapping)
    mode = str(policy.get("mode") or "vendor_bill").strip().lower()
    fallback = str(policy.get("fallback") or "api").strip().lower()
    currency = str(policy.get("defaultCurrency") or "CNY").strip().upper() or "CNY"

    fx_rate = None
    fx_source = "none"
    if mode != "none" and vendor_fx_rate is not None:
        fx_rate = vendor_fx_rate
        fx_source = vendor_fx_src or "vendor_bill"
    elif mode == "none":
        fx_rate = None
        fx_source = "none"
    elif fallback == "api" or mode == "api":
        try:
            fx_rate = api_fx_for_currency(currency, rates=fetch_usd_rates())
            fx_source = f"api:{currency}"
        except RuntimeError:
            try:
                fx_rate = get_china_pn_fx_rate(fetch_usd_rates())
                fx_source = "api:CNY"
            except RuntimeError:
                fx_rate = None
                fx_source = "none"

    return {
        "employees": employees,
        "other_amount": other_amount,
        "expense_count": expense_count,
        "fx_rate": fx_rate,
        "fx_source": fx_source,
        "vendor_fx_rate": vendor_fx_rate,
    }


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    global _ACTIVE_MAPPING
    _ACTIVE_MAPPING = resolve_convert_mapping("china_payroll_calc", convert_mapping)
    try:
        return _convert_impl(
            source_path,
            output_path,
            template_path,
            pn_meta=pn_meta,
            registry_dir=registry_dir,
            employee_directory=employee_directory,
        )
    finally:
        _ACTIVE_MAPPING = None


def _convert_impl(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if not template_path.is_file():
        raise FileNotFoundError(f"母版不存在: {template_path}")
    if not source_path.is_file():
        raise FileNotFoundError(f"原始账单不存在: {source_path}")

    parsed = parse_source(source_path)
    employees = parsed["employees"]
    sort_employees_by_code(employees, employee_directory)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, rich_text=True)
    required = (CHINA_L_SHEET, CHINA_SHEET, CHINA_EE_SHEET, PN_SHEET)
    for name in required:
        if name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"母版缺少 sheet「{name}」，现有: {wb.sheetnames}")

    applied_pn: PnMeta | None = None
    if pn_meta is not None:
        applied_pn = apply_pn_meta(
            wb,
            pn_meta,
            registry_dir=registry_dir or output_path.parent,
            reserve_invoice_number=True,
        )

    write_china_l(wb[CHINA_L_SHEET], employees)
    # 扩行前先快照示例行（含第二种公式），否则 4 人扩行会先用第 1 种公式盖掉第 10 行
    mapping = _active_mapping()
    formula_rows = _china_formula_rows()
    need_china, need_ee = needed_example_rows_for_styles(
        mapping,
        employees,
        main_template_key="China",
        ee_template_key="China EE",
        main_example_field="chinaExampleRow",
        ee_example_field="chinaEeExampleRow",
        main_data_start=CHINA_DATA_START_ROW,
        ee_data_start=CHINA_EE_DATA_START_ROW,
        employee_directory=employee_directory,
    )
    main_snaps = {r: snapshot_row_cells(wb[CHINA_SHEET], r) for r in need_china}
    ee_snaps = {r: snapshot_row_cells(wb[CHINA_EE_SHEET], r) for r in need_ee}
    fit_china_formula_sheets(
        wb,
        len(employees),
        clear_excess=False,
        protected_china_rows=need_china,
        protected_ee_rows=need_ee,
    )
    formula_plan = apply_china_employee_formula_styles(
        wb,
        employees,
        formula_rows=formula_rows,
        employee_directory=employee_directory,
        main_snapshots=main_snaps,
        ee_snapshots=ee_snaps,
    )
    clear_excess_china_formula_rows(wb, len(employees))
    # 公式扩行可能带上母版占位名（如 CPT）；用工号查员工库覆盖 China!B（只用库名称）
    name_warnings, lib_names = apply_china_sheet_names_from_directory(
        wb[CHINA_SHEET],
        employees,
        employee_directory,
        china_data_start_row=CHINA_DATA_START_ROW,
    )
    for p in formula_plan or []:
        name_warnings.append(
            f"公式配对：第{p.get('index')}人 → China第{p.get('mainExampleRow')}行 / China EE第{p.get('eeExampleRow')}行"
        )
    styles = mapping.get("employeeFormulaStyles") if isinstance(mapping.get("employeeFormulaStyles"), list) else []
    name_warnings.insert(0, f"映射员工公式样式条数: {len(styles)}")
    fx_src = str(parsed.get("fx_source") or "")
    if parsed.get("vendor_fx_rate") is None and fx_src.startswith("api:"):
        name_warnings.append(
            "供应商账单未读到汇率（S-Payment Notice!C49 或「汇率」标签），已回退网上 CNY；"
            "请确认付款通知页有数值汇率（不要只留未计算的公式）"
        )
    elif fx_src.startswith("vendor:"):
        name_warnings.append(f"汇率已取自供应商账单 {fx_src} = {parsed.get('fx_rate')}")
    if styles:
        from bill_convert.formula_layout import _pick_int_field

        want_rows = set()
        for s in styles:
            if not isinstance(s, dict):
                continue
            r = _pick_int_field(s, "chinaExampleRow", "mainExampleRow", "twExampleRow")
            if r is not None:
                want_rows.add(r)
        used = {int(p["mainExampleRow"]) for p in (formula_plan or []) if p.get("mainExampleRow") is not None}
        if want_rows and used.isdisjoint(want_rows):
            name_warnings.append(
                "公式配对未命中：映射要求 China 示例行 "
                + ",".join(str(x) for x in sorted(want_rows))
                + "，但实际全部落在默认行；请检查员工库工号是否与账单「工号」一致"
            )
            formula_match_hint = "style-row-miss"
        elif not want_rows:
            name_warnings.append(
                "映射有员工公式样式，但未找到 chinaExampleRow/mainExampleRow 字段（可能未保存成功）"
            )
            formula_match_hint = "style-row-missing-field"
        elif want_rows and want_rows.issubset(used):
            formula_match_hint = "style-row-hit"
        else:
            formula_match_hint = "style-row-partial"
    else:
        formula_match_hint = ""

    fx_row_before = _find_pn_fx_row(wb[PN_SHEET]) if PN_SHEET in wb.sheetnames else 29
    pn_layout = fit_pn_employees(wb[PN_SHEET], len(employees))
    fx_row = int(pn_layout.get("fx_row") or _find_pn_fx_row(wb[PN_SHEET]))
    pn_fx_write = None

    if parsed["fx_rate"] is not None:
        from fx_policy import make_pn_fx_provenance

        wb[PN_SHEET].cell(fx_row, 2).value = parsed["fx_rate"]
        fx_src = str(parsed.get("fx_source") or "mapping")
        write_source = "api" if fx_src.startswith("api:") or fx_src.startswith("vendor:") else "mapping"
        pn_fx_write = make_pn_fx_provenance(
            PN_SHEET,
            fx_row,
            2,
            _ACTIVE_MAPPING,
            float(parsed["fx_rate"]),
            write_source=write_source,
            fx_source=fx_src,
        )

    retarget_pn_fx_refs(wb, fx_row, from_rows=[int(fx_row_before), 28, 29, 30, 31, 32])
    apply_china_specials(
        wb[CHINA_SHEET],
        len(employees),
        parsed["expense_count"],
        parsed["other_amount"],
        fx_row=fx_row,
    )

    apply_luckysheet_compat(wb, pn_sheet=PN_SHEET)

    wb.save(output_path)
    wb.close()
    # 主题填充 / 富文本；金额由前端 HyperFormula 按公式重算，不再注入 PN 缓存
    postprocess_converted_xlsx(output_path)

    return {
        "employee_count": len(employees),
        "employee_names": lib_names,
        "fx_rate": parsed["fx_rate"],
        "fx_source": parsed.get("fx_source"),
        "vendor_fx_rate": parsed.get("vendor_fx_rate"),
        "other_amount": parsed["other_amount"],
        "expense_count": parsed["expense_count"],
        "fx_row": fx_row,
        "output": str(output_path),
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
        "warnings": name_warnings,
        "mapping_style_count": len(styles),
        "formula_main_rows": ",".join(
            str(p.get("mainExampleRow")) for p in (formula_plan or []) if p.get("mainExampleRow") is not None
        ),
        "formula_match_hint": formula_match_hint,
        "pn_fx_write": pn_fx_write,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="China 供应商账单 → PN 自动转换")
    parser.add_argument("source", type=Path, help="原始账单 Excel 路径")
    parser.add_argument("-o", "--output", type=Path, help="输出 PN 路径（默认同目录 PN_auto_*.xlsx）")
    parser.add_argument("-t", "--template", type=Path, default=DEFAULT_TEMPLATE, help="PN 母版路径")
    args = parser.parse_args(argv)

    source = args.source.resolve()
    if args.output:
        output = args.output.resolve()
    else:
        output = source.parent / f"PN_auto_{source.stem}.xlsx"

    try:
        result = convert(source, output, args.template.resolve())
    except Exception as exc:
        print(f"转换失败: {exc}", file=sys.stderr)
        return 1

    print("转换完成")
    print(f"  输出: {result['output']}")
    print(f"  员工: {result['employee_count']} 人 → {result['employee_names']}")
    print(f"  汇率 PN!B{result.get('fx_row', 29)}: {result['fx_rate']} ({result.get('fx_source')})")
    if str(result.get("fx_source") or "").startswith("api:"):
        print("  （供应商账单无有效汇率，已回退网上 CNY）")
    print(f"  Other Fee → China!J*: {result['other_amount']}")
    print(f"  报销笔数 → For Expense: {result['expense_count']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
