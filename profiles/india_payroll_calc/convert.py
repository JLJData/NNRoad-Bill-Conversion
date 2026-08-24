# -*- coding: utf-8 -*-
"""
India-L → India PN（引擎 india_payroll_calc）

用法:
  python -m profiles.india_payroll_calc.convert <源.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

源账单: sheet「India-L」（可由 biz_solutions_india ingest 产出）。
默认母版: templates/india/template.xlsx

原则：PN / India / India EE 以母版公式为准；只写 India-L 数据与必要元数据。
Biz Solutions：PDF 仅有 CTC 总额/税/姓名/账期；七项拆分走 mapping.indiaSalarySplit。
"""
from __future__ import annotations

import argparse
import calendar
import re
import shutil
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from bill_convert.formula_copy import shift_row_formula
from bill_convert.india_salary_split import (
    INDIA_L_KNOWN_DATA_FIELDS,
    resolve_field_columns_from_headers,
)
from convert_mapping import find_sheet_name, resolve_convert_mapping
from fx_rate import get_india_pn_fx_rate, get_usd_rate
from pn_meta import PnMeta, apply_pn_meta
from profiles.tw_payroll_calc.convert import match_ee_code
from region_templates import get_region_template
from xlsx_convert_utils import coerce_datetime_for_excel, is_date_column_header
from xlsx_luckysheet_compat import apply_luckysheet_compat
from xlsx_postprocess import postprocess_converted_xlsx

DEFAULT_TEMPLATE = get_region_template("India")

INDIA_L_SHEET = "India-L"
INDIA_SHEET = "India"
INDIA_EE_SHEET = "India EE"
PN_SHEET = "PN"

INDIA_L_HEADER_ROW = 4
INDIA_L_DATA_START = 10
INDIA_L_PERIOD_ROW = 8
INDIA_DATA_START = 9
INDIA_EE_DATA_START = 9
MAX_EMPLOYEES = 20
_DATE_FMT = "yyyy/m/d"

# India-L 默认列（仅作表头未匹配时的兜底；读写优先 _header_map）
COL_NAME = 2

_ACTIVE_MAPPING: dict[str, Any] | None = None

_MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def _active_mapping() -> dict[str, Any]:
    return (
        _ACTIVE_MAPPING
        if isinstance(_ACTIVE_MAPPING, dict)
        else resolve_convert_mapping("india_payroll_calc", None)
    )


def _india_l_layout(*, target: bool = False) -> tuple[int, int]:
    mapping = _active_mapping()
    key = "targetL" if target else "sourceEmployeeSheet"
    spec = mapping.get(key) if isinstance(mapping.get(key), dict) else {}
    if not spec and target:
        spec = (
            mapping.get("sourceEmployeeSheet")
            if isinstance(mapping.get("sourceEmployeeSheet"), dict)
            else {}
        )
    header = int(spec.get("headerRow") or INDIA_L_HEADER_ROW)
    data_start = int(spec.get("dataStartRow") or INDIA_L_DATA_START)
    return header, data_start


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("\xa0", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _cell_formula_text(value: Any) -> str | None:
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _set_cell_value(cell, value: Any) -> None:
    if isinstance(value, float):
        cell.value = round(value, 6) if abs(value) >= 1e-9 else 0
    else:
        cell.value = value


def _header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        h = _norm(ws.cell(header_row, col).value)
        if h and h not in out:
            out[h] = col
    return out


def _india_l_formula_cols(ws: Worksheet, data_start: int) -> dict[int, str]:
    out: dict[int, str] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        text = _cell_formula_text(ws.cell(data_start, col).value)
        if text:
            out[col] = text
    return out


def parse_pay_period_label(label: Any) -> tuple[int, int] | None:
    text = _norm(label)
    if not text:
        return None
    m = re.search(
        r"(jan|january|feb|february|mar|march|apr|april|may|jun|june|jul|july|"
        r"aug|august|sep|sept|september|oct|october|nov|november|dec|december)"
        r"[\s\-']*(\d{2,4})",
        text,
        flags=re.I,
    )
    if not m:
        return None
    month = _MONTHS.get(m.group(1).lower())
    if not month:
        return None
    year = int(m.group(2))
    if year < 100:
        year += 2000
    return year, month


def period_bounds_from_label(label: Any) -> tuple[date, date] | None:
    parsed = parse_pay_period_label(label)
    if not parsed:
        return None
    year, month = parsed
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    return start, end


def looks_like_india_l_workbook(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return False
    try:
        return INDIA_L_SHEET in wb.sheetnames
    finally:
        wb.close()


def parse_india_l_employees(ws: Worksheet) -> list[dict[str, Any]]:
    header_row, data_start = _india_l_layout(target=False)
    headers = _header_map(ws, header_row)
    field_cols = resolve_field_columns_from_headers(headers, fields=INDIA_L_KNOWN_DATA_FIELDS)
    if not field_cols.get("Employee Name"):
        field_cols = {**field_cols, "Employee Name": COL_NAME}
    name_col = field_cols["Employee Name"]
    known_cols = set(field_cols.values())

    period_from = ws.cell(INDIA_L_PERIOD_ROW, 2).value
    period_to = ws.cell(INDIA_L_PERIOD_ROW, 3).value

    employees: list[dict[str, Any]] = []
    max_row = max(ws.max_row or data_start, data_start)
    for row in range(data_start, max_row + 1):
        name = _norm(ws.cell(row, name_col).value)
        if not name:
            continue
        emp: dict[str, Any] = {
            "Employee Name": name,
            "_period_from": period_from,
            "_period_to": period_to,
            "From": period_from,
            "To": period_to,
        }
        for key, col in field_cols.items():
            if key == "Employee Name":
                continue
            val = ws.cell(row, col).value
            if _cell_formula_text(val):
                continue
            if val is None or val == "":
                continue
            emp[key] = val

        for h, col in headers.items():
            if col in known_cols or h in emp:
                continue
            val = ws.cell(row, col).value
            if _cell_formula_text(val) or val is None or val == "":
                continue
            emp[h] = val
        employees.append(emp)
    return employees


def write_india_l_period(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    if not employees:
        return
    emp0 = employees[0]
    start = emp0.get("_period_from") or emp0.get("From")
    end = emp0.get("_period_to") or emp0.get("To")
    if isinstance(start, str):
        bounds = period_bounds_from_label(start)
        if bounds:
            start, end = bounds
    for col, val in ((2, start), (3, end)):
        if val is None:
            continue
        cell = ws.cell(INDIA_L_PERIOD_ROW, col)
        dt = coerce_datetime_for_excel(val)
        if dt is not None:
            cell.value = dt
            cell.number_format = _DATE_FMT
        elif isinstance(val, date):
            cell.value = datetime(val.year, val.month, val.day)
            cell.number_format = _DATE_FMT
        else:
            cell.value = val


def write_india_l(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    """只覆盖数据列，母版公式列保留/扩行复制。"""
    _, data_start = _india_l_layout(target=True)
    write_india_l_period(ws, employees)

    n = len(employees)
    formula_by_col = _india_l_formula_cols(ws, data_start)
    max_col = max(ws.max_column or 31, 31)

    for i in range(1, n):
        _copy_row_style_and_formula(
            ws,
            data_start,
            data_start + i,
            max_col=max_col,
            l_from=data_start,
            l_to=data_start + i,
        )

    last_keep = data_start + max(n, 1) - 1
    max_row = max(ws.max_row or data_start, data_start + MAX_EMPLOYEES)
    for row in range(data_start, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ == "MergedCell":
                continue
            if row <= last_keep and col in formula_by_col:
                continue
            cell.value = None

    for i in range(n):
        row = data_start + i
        for col, formula in formula_by_col.items():
            cur = ws.cell(row, col).value
            if _cell_formula_text(cur):
                continue
            ws.cell(row, col).value = shift_row_formula(
                formula,
                data_start,
                row,
                target_l_from=data_start,
                target_l_to=row,
                target_l_sheet=INDIA_L_SHEET,
            )

    field_cols = resolve_field_columns_from_headers(
        _header_map(ws, _india_l_layout(target=True)[0]),
    )
    if not field_cols.get("Employee Name"):
        # 姓名列：表头未匹配时仍用母版约定列（与 parse 侧一致）
        field_cols = {**field_cols, "Employee Name": COL_NAME}

    for idx, emp in enumerate(employees):
        row = data_start + idx
        for key, col in field_cols.items():
            if col in formula_by_col:
                continue
            if key not in emp:
                continue
            val = emp[key]
            if val is None:
                continue
            cell = ws.cell(row, col)
            if is_date_column_header(key):
                dt = coerce_datetime_for_excel(val)
                if dt is not None:
                    cell.value = dt
                    cell.number_format = _DATE_FMT
                    continue
            _set_cell_value(cell, val)


def _copy_row_style_and_formula(
    ws: Worksheet,
    src_row: int,
    dest_row: int,
    max_col: int = 40,
    *,
    l_from: int | None = None,
    l_to: int | None = None,
) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dest = ws.cell(dest_row, c)
        if type(dest).__name__ == "MergedCell" or type(src).__name__ == "MergedCell":
            continue
        if src.has_style:
            dest.font = copy(src.font)
            dest.border = copy(src.border)
            dest.fill = copy(src.fill)
            dest.number_format = src.number_format
            dest.protection = copy(src.protection)
            dest.alignment = copy(src.alignment)
        text = _cell_formula_text(src.value)
        if text:
            dest.value = shift_row_formula(
                text,
                src_row,
                dest_row,
                target_l_from=l_from if l_from is not None else -1,
                target_l_to=l_to if l_to is not None else -1,
                target_l_sheet=INDIA_L_SHEET,
            )
        elif src.value is not None and not isinstance(src.value, ArrayFormula):
            dest.value = src.value


def expand_india_employee_rows(wb, employee_count: int) -> None:
    n = max(int(employee_count), 1)
    _, l_data_start = _india_l_layout(target=True)
    if INDIA_SHEET in wb.sheetnames:
        india = wb[INDIA_SHEET]
        for i in range(1, n):
            dest = INDIA_DATA_START + i
            l_row = l_data_start + i
            _copy_row_style_and_formula(
                india,
                INDIA_DATA_START,
                dest,
                max_col=40,
                l_from=l_data_start,
                l_to=l_row,
            )
    if INDIA_EE_SHEET in wb.sheetnames:
        ee = wb[INDIA_EE_SHEET]
        for i in range(1, n):
            dest = INDIA_EE_DATA_START + i
            l_row = l_data_start + i
            _copy_row_style_and_formula(
                ee,
                INDIA_EE_DATA_START,
                dest,
                max_col=40,
                l_from=l_data_start,
                l_to=l_row,
            )


def _find_pn_row_by_label(ws: Worksheet, keyword: str, col: int = 1) -> int | None:
    key = keyword.lower()
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and key in v.lower():
            return row
    return None


def fit_india_pn_employees(wb, employee_count: int) -> dict[str, Any]:
    fx_row = _find_pn_row_by_label(wb[PN_SHEET], "FX rate") if PN_SHEET in wb.sheetnames else None
    return {"fx_row": fx_row or 28, "employee_count": employee_count}


def _pn_customer_id(pn_meta: PnMeta | dict[str, Any] | None) -> str | None:
    if pn_meta is None:
        return None
    if isinstance(pn_meta, PnMeta):
        cid = (pn_meta.customer_id or "").strip()
        return cid or None
    if isinstance(pn_meta, dict):
        cid = str(pn_meta.get("customer_id") or pn_meta.get("customerId") or "").strip()
        return cid or None
    return None


def apply_india_ee_codes(
    wb,
    employees: list[dict[str, Any]],
    *,
    employee_directory: list[dict[str, Any]] | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
) -> list[str]:
    if INDIA_EE_SHEET not in wb.sheetnames:
        return []
    ws = wb[INDIA_EE_SHEET]
    client_code = _pn_customer_id(pn_meta)
    directory = list(employee_directory or [])
    warnings: list[str] = []
    for i, emp in enumerate(employees):
        row = INDIA_EE_DATA_START + i
        # Client Code / Name：母版常引用 PN；有客户编号时写 B
        if client_code and not _cell_formula_text(ws.cell(row, 2).value):
            ws.cell(row, 2).value = client_code
        name = _norm(emp.get("Employee Name"))
        code, warn = match_ee_code([name] if name else [], directory)
        if not _cell_formula_text(ws.cell(row, 4).value):
            ws.cell(row, 4).value = code
        if warn:
            warnings.append(f"India EE 第{i + 1}人：{warn}")
    return warnings


def apply_fx(wb, *, fill_fx: bool = True, fx_row: int | None = None, convert_mapping: dict | None = None) -> tuple[float | None, dict | None]:
    """写入 PN FX：优先 =基准*系数；基准空则用网上 INR 作基准（仍写公式，点开可见）。"""
    if not fill_fx or PN_SHEET not in wb.sheetnames:
        return None, None
    from fx_policy import apply_fx_formula_to_cell_ex, fx_policy, make_pn_fx_provenance, read_fx_base_adjustment, resolve_vendor_currency

    policy = fx_policy(convert_mapping)
    mode = str(policy.get("mode") or "api_as_base").strip().lower()
    if mode == "none":
        return None, None
    row = fx_row or _find_pn_row_by_label(wb[PN_SHEET], "FX rate") or 28
    col = 2
    cell = wb[PN_SHEET].cell(row, col)

    api_base: float | None = None
    fx_source: str | None = None
    mapped_base, _adj, _legacy = read_fx_base_adjustment(convert_mapping)
    if mapped_base is None and mode in ("api_as_base", "api"):
        currency = resolve_vendor_currency(convert_mapping, str(policy.get("defaultCurrency") or "INR")) or "INR"
        api_base = float(get_usd_rate(currency))
        fx_source = f"api:{currency}*0.97"

    product, write_source = apply_fx_formula_to_cell_ex(cell, convert_mapping, api_base=api_base)
    if product is not None:
        return product, make_pn_fx_provenance(
            PN_SHEET, row, col, convert_mapping, product, write_source=write_source, fx_source=fx_source
        )
    fx = get_india_pn_fx_rate()
    cell.value = fx
    return fx, make_pn_fx_provenance(
        PN_SHEET, row, col, convert_mapping, fx, write_source="api", fx_source="api:INR*0.97"
    )


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    registry_dir: Path | None = None,
    convert_mapping: dict[str, Any] | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    global _ACTIVE_MAPPING
    _ACTIVE_MAPPING = resolve_convert_mapping("india_payroll_calc", convert_mapping)
    try:
        source_path = Path(source_path).resolve()
        output_path = Path(output_path).resolve()
        template_path = Path(template_path).resolve()
        if not source_path.is_file():
            raise FileNotFoundError(f"源文件不存在: {source_path}")
        if not template_path.is_file():
            raise FileNotFoundError(f"母版不存在: {template_path}")

        src_wb = load_workbook(source_path, data_only=False)
        try:
            l_name = find_sheet_name(list(src_wb.sheetnames), _active_mapping().get("sourceEmployeeSheet"))
            if not l_name or l_name not in src_wb.sheetnames:
                if INDIA_L_SHEET in src_wb.sheetnames:
                    l_name = INDIA_L_SHEET
                else:
                    raise ValueError(f"未找到 India-L，现有: {src_wb.sheetnames}")
            employees = parse_india_l_employees(src_wb[l_name])
        finally:
            src_wb.close()

        if not employees:
            raise ValueError("India-L 未解析到员工行")

        # 映射有七项拆分时覆盖 L 表（PDF 路径常先把 CTC 整笔进 Basic）
        from bill_convert.india_salary_split import (
            INDIA_SPLIT_PROVENANCE_FIELDS,
            apply_salary_split_to_employee,
            build_india_salary_split_cell_writes,
            parse_india_salary_splits,
            resolve_field_columns_from_headers,
        )

        warnings: list[str] = []
        if parse_india_salary_splits(_ACTIVE_MAPPING):
            for emp in employees:
                ctc = _as_float(emp.get("_ctc"))
                if ctc is None:
                    # 用当前七项合计作校验基准；有映射则仍覆盖写入
                    parts = [
                        emp.get("Basic salary"),
                        emp.get("HRA"),
                        emp.get("Telephone allowance"),
                        emp.get("LTA"),
                        emp.get("Special allowance"),
                        emp.get("Wellness Stipend"),
                    ]
                    nums = [_as_float(x) for x in parts]
                    if any(x is not None for x in nums):
                        ctc = round(sum(x or 0.0 for x in nums), 2)
                apply_salary_split_to_employee(
                    emp,
                    mapping=_ACTIVE_MAPPING,
                    ctc=ctc,
                    warnings=warnings,
                    fallback_ctc_to_basic=False,
                )

        cell_writes: list[dict[str, Any]] = []

        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)
        fx = None
        pn_fx_write = None
        applied_pn = None
        try:
            if INDIA_L_SHEET not in wb.sheetnames:
                raise ValueError(f"母版缺少 {INDIA_L_SHEET}")
            write_india_l(wb[INDIA_L_SHEET], employees)
            header_row, l_data_start = _india_l_layout(target=True)
            india_l_headers = _header_map(wb[INDIA_L_SHEET], header_row)
            split_field_cols = resolve_field_columns_from_headers(
                india_l_headers,
                fields=INDIA_SPLIT_PROVENANCE_FIELDS,
            )
            cell_writes = build_india_salary_split_cell_writes(
                employees,
                sheet=INDIA_L_SHEET,
                data_start=l_data_start,
                field_cols=split_field_cols,
            )
            expand_india_employee_rows(wb, len(employees))
            pn_layout = fit_india_pn_employees(wb, len(employees))
            if len(employees) > 1:
                warnings.append(
                    f"India PN 多人明细行扩行暂定：已扩 India/India EE（{len(employees)} 人），请人工核对 PN"
                )
            try:
                fx, pn_fx_write = apply_fx(
                    wb, fill_fx=fill_fx, fx_row=pn_layout.get("fx_row"), convert_mapping=_ACTIVE_MAPPING
                )
            except Exception as exc:
                warnings.append(f"写入 PN 汇率失败: {exc}")

            if pn_meta is not None:
                applied_pn = apply_pn_meta(
                    wb,
                    pn_meta,
                    registry_dir=registry_dir or output_path.parent,
                    reserve_invoice_number=True,
                )
            warnings.extend(
                apply_india_ee_codes(
                    wb,
                    employees,
                    employee_directory=employee_directory,
                    pn_meta=applied_pn or pn_meta,
                )
            )
            apply_luckysheet_compat(wb, pn_sheet=PN_SHEET)
            wb.save(output_path)
        finally:
            wb.close()

        postprocess_converted_xlsx(output_path)
        return {
            "ok": True,
            "engine_id": "india_payroll_calc",
            "region": "India",
            "output": str(output_path),
            "employee_count": len(employees),
            "fx_rate": fx,
            "warnings": warnings,
            "pn_meta": applied_pn.to_dict() if applied_pn else None,
            "pn_fx_write": pn_fx_write,
            "cell_writes": cell_writes,
        }
    finally:
        _ACTIVE_MAPPING = None


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="India-L → India PN")
    parser.add_argument("source", type=Path)
    parser.add_argument("-o", "--output", type=Path, default=None)
    parser.add_argument("-t", "--template", type=Path, default=None)
    parser.add_argument("--no-fx", action="store_true")
    args = parser.parse_args(argv)
    source = args.source.resolve()
    output = (args.output or source.with_name(f"PN_India_{source.stem}.xlsx")).resolve()
    template = (args.template or DEFAULT_TEMPLATE).resolve()
    result = convert(source, output, template, fill_fx=not args.no_fx)
    print(result)
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
