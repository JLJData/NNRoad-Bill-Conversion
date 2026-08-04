# -*- coding: utf-8 -*-
"""
Pakistan-L → Pakistan PN（引擎 pakistan_payroll_calc）

源: sheet「Pakistan-L」表头第 7 行、数据第 8 行起（可由 panda_work_pk ingest 产出）。
默认母版: templates/pakistan/template.xlsx
母版公式（Recurring Fee / Business Tax / PN Labor）一律保留，只写数据与扩行平移。
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from bill_convert.formula_copy import (
    copy_row_formulas as shared_copy_row_formulas,
    fix_ee_row_pakistan_refs,
    fix_pakistan_row_pakistan_ee_refs,
    shift_row_formula,
    snapshot_row_cells,
)
from bill_convert.formula_layout import (
    apply_employee_formula_styles,
    needed_example_rows_for_styles,
    tw_l_row_for_data_row,
)
from bill_convert.formula_layout import _default_example_row as default_example_row_for_mapping
from bill_convert.pakistan_business_tax import (
    BT_FEDERAL_HEADER,
    BT_SINDH_HEADER,
    business_tax_formula,
    parse_pakistan_business_tax_cfg,
)
from bill_convert.pakistan_employee_fees import lookup_employee_fee, parse_pakistan_employee_fees
from convert_mapping import find_sheet_name, resolve_convert_mapping
from fx_rate import fetch_usd_rates, get_pakistan_pn_fx_rate
from pn_meta import PnMeta, apply_pn_meta
from profiles.tw_payroll_calc.convert import match_ee_code
from region_templates import get_region_template
from xlsx_postprocess import postprocess_converted_xlsx

DEFAULT_TEMPLATE = get_region_template("Pakistan")

PK_L_SHEET = "Pakistan-L"
PK_SHEET = "Pakistan"
PK_EE_SHEET = "Pakistan EE"
PN_SHEET = "PN"

PK_L_HEADER_ROW = 7
PK_L_DATA_START = 8
PK_DATA_START = 9
PK_EE_DATA_START = 10
MAX_ROWS = 60

_DATE_FMT = "yyyy/m/d"
_ACTIVE_MAPPING: dict[str, Any] | None = None


def _active_mapping() -> dict[str, Any]:
    return (
        _ACTIVE_MAPPING
        if isinstance(_ACTIVE_MAPPING, dict)
        else resolve_convert_mapping("pakistan_payroll_calc", None)
    )


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("PKR", "").replace("\xa0", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _cell_formula_text(value: Any) -> str | None:
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        h = _norm(ws.cell(header_row, col).value)
        if h and h not in out:
            out[h] = col
    return out


def _pk_l_layout() -> tuple[int, int]:
    spec = _active_mapping().get("sourceEmployeeSheet") or {}
    header = int(spec.get("headerRow") or PK_L_HEADER_ROW)
    start = int(spec.get("dataStartRow") or PK_L_DATA_START)
    return header, start


def coerce_datetime_for_excel(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def parse_pakistan_l_employees(ws: Worksheet) -> list[dict[str, Any]]:
    header_row, data_start = _pk_l_layout()
    headers = _header_map(ws, header_row)
    if not headers:
        raise ValueError(f"「{PK_L_SHEET}」第 {header_row} 行表头为空")
    name_keys = ["Name of Employee", "Employee Name", "EE Name"]
    employees: list[dict[str, Any]] = []
    for row in range(data_start, (ws.max_row or data_start) + 1):
        name = None
        for key in name_keys:
            if key in headers:
                name = ws.cell(row, headers[key]).value
                if _norm(name):
                    break
        if not _norm(name):
            continue
        if _norm(name).upper().startswith("TOTAL"):
            continue
        row_data: dict[str, Any] = {}
        for h, col in headers.items():
            row_data[h] = ws.cell(row, col).value
        # 账期元数据
        row_data["From"] = ws.cell(2, 3).value
        row_data["To"] = ws.cell(2, 5).value
        row_data["Client"] = ws.cell(1, 3).value
        row_data["_fx_rate"] = _as_float(ws.cell(3, 6).value)
        employees.append(row_data)
    return employees


def _pk_l_formula_cols(ws: Worksheet, data_start: int) -> dict[int, str]:
    out: dict[int, str] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        text = _cell_formula_text(ws.cell(data_start, col).value)
        if text:
            out[col] = text
    return out


def _copy_row_style_and_formula(
    ws: Worksheet,
    src_row: int,
    dest_row: int,
    max_col: int = 40,
    *,
    l_from: int | None = None,
    l_to: int | None = None,
) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dest = ws.cell(dest_row, c)
        if type(dest).__name__ == "MergedCell" or type(src).__name__ == "MergedCell":
            continue
        if src.has_style:
            dest.font = copy(src.font)
            dest.border = copy(src.border)
            dest.fill = copy(src.fill)
            dest.number_format = src.number_format
            dest.protection = copy(src.protection)
            dest.alignment = copy(src.alignment)
        text = _cell_formula_text(src.value)
        if text:
            dest.value = shift_row_formula(
                text,
                src_row,
                dest_row,
                target_l_from=l_from if l_from is not None else -1,
                target_l_to=l_to if l_to is not None else -1,
                target_l_sheet=PK_L_SHEET,
            )
        elif src.value is not None and not isinstance(src.value, ArrayFormula):
            dest.value = src.value


def write_pakistan_l(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    header_row, data_start = _pk_l_layout()
    headers = _header_map(ws, header_row)
    if not headers:
        raise ValueError(f"「{PK_L_SHEET}」第 {header_row} 行表头为空")

    if employees:
        emp0 = employees[0]
        if _norm(ws.cell(1, 1).value).lower().startswith("company"):
            client = _norm(emp0.get("Client"))
            if client:
                ws.cell(1, 3).value = client
        if _norm(ws.cell(2, 1).value).lower().startswith("payroll"):
            for key, col in (("From", 3), ("To", 5)):
                dt = coerce_datetime_for_excel(emp0.get(key))
                if dt is None:
                    continue
                cell = ws.cell(2, col)
                cell.value = dt
                cell.number_format = _DATE_FMT
        fx = _as_float(emp0.get("_fx_rate"))
        if fx is not None:
            ws.cell(3, 5).value = "FX USD:PKR"
            ws.cell(3, 6).value = fx

    n = len(employees)
    formula_by_col = _pk_l_formula_cols(ws, data_start)
    max_col = max(max(headers.values(), default=1), max(formula_by_col.keys(), default=1), 18)

    for i in range(1, n):
        _copy_row_style_and_formula(
            ws,
            data_start,
            data_start + i,
            max_col=max_col,
            l_from=data_start,
            l_to=data_start + i,
        )

    last_keep = data_start + max(n, 1) - 1
    max_row = max(ws.max_row or data_start, data_start + MAX_ROWS)
    for row in range(data_start, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ == "MergedCell":
                continue
            if row <= last_keep and col in formula_by_col:
                continue
            cell.value = None

    for i in range(n):
        row = data_start + i
        for col, formula in formula_by_col.items():
            cur = ws.cell(row, col).value
            if _cell_formula_text(cur):
                continue
            ws.cell(row, col).value = shift_row_formula(
                formula,
                data_start,
                row,
                target_l_from=data_start,
                target_l_to=row,
                target_l_sheet=PK_L_SHEET,
            )

    skip_write = set(formula_by_col.keys())
    for idx, emp in enumerate(employees):
        row = data_start + idx
        for h, col in headers.items():
            if col in skip_write:
                continue
            if h not in emp or emp.get(h) is None:
                continue
            if h in ("From", "To", "Client"):
                continue
            val = emp[h]
            cell = ws.cell(row, col)
            cell.value = val

    _apply_mapped_eobi_it(ws, employees, headers, data_start)


def _apply_mapped_eobi_it(
    ws: Worksheet,
    employees: list[dict[str, Any]],
    headers: dict[str, int],
    data_start: int,
) -> None:
    """按映射为每位员工写入 E.O.B.I / IT（季度拆月后同一人多行共用）。"""
    fees = parse_pakistan_employee_fees(_active_mapping())
    if not fees:
        return
    eobi_col = headers.get("E.O.B.I") or headers.get("EOBI")
    it_col = headers.get("IT")
    if not eobi_col and not it_col:
        return
    for idx, emp in enumerate(employees):
        name = _norm(emp.get("Name of Employee") or emp.get("Employee Name"))
        fee = lookup_employee_fee(name, fees)
        if not fee:
            continue
        row = data_start + idx
        if eobi_col and "eobi" in fee:
            ws.cell(row, eobi_col).value = fee["eobi"]
        if it_col and "it" in fee:
            ws.cell(row, it_col).value = fee["it"]


def _apply_invoice_derived_business_tax(
    wb,
    employees: list[dict[str, Any]],
) -> list[str]:
    """Danfoss：用 L 上的月度 BT 系数覆盖 Pakistan!F Business Tax 公式。"""
    cfg = parse_pakistan_business_tax_cfg(_active_mapping())
    if not cfg or PK_SHEET not in wb.sheetnames:
        return []
    pk = wb[PK_SHEET]
    warnings: list[str] = []
    for idx, emp in enumerate(employees):
        sindh = _as_float(emp.get(BT_SINDH_HEADER) if emp.get(BT_SINDH_HEADER) is not None else emp.get("_bt_sindh_usd"))
        federal = _as_float(
            emp.get(BT_FEDERAL_HEADER) if emp.get(BT_FEDERAL_HEADER) is not None else emp.get("_bt_federal_usd")
        )
        row = PK_DATA_START + idx
        if sindh is None or federal is None:
            name = _norm(emp.get("Name of Employee") or emp.get("Employee Name")) or f"第{idx + 1}人"
            warnings.append(f"{name}：映射启用了发票推导 Business Tax，但缺少 BT 系数（请用含推导列的源表/PDF 转换）")
            continue
        pk.cell(row, 6).value = business_tax_formula(sindh, federal)
    return warnings


def clear_employee_row(ws: Worksheet, row: int) -> None:
    for col in range(1, (ws.max_column or 0) + 1):
        cell = ws.cell(row, col)
        if type(cell).__name__ == "MergedCell":
            continue
        cell.value = None


def count_template_employee_slots(ws: Worksheet, data_start_row: int, marker_col: int) -> int:
    n = 0
    for row in range(data_start_row, data_start_row + MAX_ROWS):
        marker = ws.cell(row, marker_col).value
        has_formula = False
        for col in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(row, col)
            if cell.data_type == "f" and cell.value:
                has_formula = True
                break
        if marker is not None or has_formula:
            n += 1
            continue
        break
    return max(n, 1)


def _pk_formula_rows() -> dict[str, int]:
    _, l_data_start = _pk_l_layout()
    return {
        "l_data_start": l_data_start,
        "main_data_start": PK_DATA_START,
        "ee_data_start": PK_EE_DATA_START,
        # 兼容旧键
        "tw_l_data_start": l_data_start,
        "tw_data_start": PK_DATA_START,
        "tw_ee_data_start": PK_EE_DATA_START,
    }


def fit_pakistan_formula_sheets(
    wb,
    employee_count: int,
    *,
    clear_excess: bool = True,
    protected_pk_rows: set[int] | None = None,
    protected_ee_rows: set[int] | None = None,
) -> None:
    n = max(int(employee_count), 1)
    mapping = _active_mapping()
    _, l_start = _pk_l_layout()
    if PK_SHEET not in wb.sheetnames or PK_EE_SHEET not in wb.sheetnames:
        return
    pk = wb[PK_SHEET]
    ee = wb[PK_EE_SHEET]
    pk_slots = count_template_employee_slots(pk, PK_DATA_START, marker_col=2)
    ee_slots = count_template_employee_slots(ee, PK_EE_DATA_START, marker_col=5)
    pk_tpl = default_example_row_for_mapping(mapping, "Pakistan", PK_DATA_START)
    ee_tpl = default_example_row_for_mapping(mapping, "Pakistan EE", PK_EE_DATA_START)
    prot_p = set(protected_pk_rows or ())
    prot_e = set(protected_ee_rows or ())
    prot_p.add(pk_tpl)
    prot_e.add(ee_tpl)
    for r in prot_p:
        pk_slots = max(pk_slots, r - PK_DATA_START + 1)
    for r in prot_e:
        ee_slots = max(ee_slots, r - PK_EE_DATA_START + 1)
    src_pk_l = tw_l_row_for_data_row(pk_tpl, data_start=PK_DATA_START, target_l_data_start=l_start)
    src_ee_l = tw_l_row_for_data_row(ee_tpl, data_start=PK_EE_DATA_START, target_l_data_start=l_start)

    if clear_excess:
        for i in range(n, pk_slots):
            clear_employee_row(pk, PK_DATA_START + i)
        for i in range(n, ee_slots):
            clear_employee_row(ee, PK_EE_DATA_START + i)

    if n > pk_slots:
        for i in range(pk_slots, n):
            dst_row = PK_DATA_START + i
            dst_l = l_start + i
            ee_row = PK_EE_DATA_START + i
            shared_copy_row_formulas(
                pk, pk_tpl, dst_row, src_pk_l, dst_l, target_l_sheet=PK_L_SHEET
            )
            fix_pakistan_row_pakistan_ee_refs(pk, dst_row, ee_row)

    if n > ee_slots:
        for i in range(ee_slots, n):
            dst_row = PK_EE_DATA_START + i
            dst_l = l_start + i
            pk_row = PK_DATA_START + i
            shared_copy_row_formulas(
                ee, ee_tpl, dst_row, src_ee_l, dst_l, target_l_sheet=PK_L_SHEET
            )
            fix_ee_row_pakistan_refs(ee, dst_row, pk_row)


def clear_excess_pakistan_formula_rows(wb, employee_count: int) -> None:
    n = max(int(employee_count), 1)
    if PK_SHEET in wb.sheetnames:
        pk = wb[PK_SHEET]
        slots = count_template_employee_slots(pk, PK_DATA_START, marker_col=2)
        for i in range(n, slots):
            clear_employee_row(pk, PK_DATA_START + i)
    if PK_EE_SHEET in wb.sheetnames:
        ee = wb[PK_EE_SHEET]
        slots = count_template_employee_slots(ee, PK_EE_DATA_START, marker_col=5)
        for i in range(n, slots):
            clear_employee_row(ee, PK_EE_DATA_START + i)


def apply_pakistan_employee_formula_styles(
    wb,
    employees: list[dict[str, Any]],
    *,
    formula_rows: dict[str, int] | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    main_snapshots: dict[int, list[dict[str, Any]]] | None = None,
    ee_snapshots: dict[int, list[dict[str, Any]]] | None = None,
) -> list[dict[str, Any]]:
    return apply_employee_formula_styles(
        wb,
        employees,
        _active_mapping(),
        formula_rows=formula_rows or _pk_formula_rows(),
        employee_directory=employee_directory,
        main_sheet=PK_SHEET,
        ee_sheet=PK_EE_SHEET,
        target_l_sheet=PK_L_SHEET,
        main_template_key="Pakistan",
        ee_template_key="Pakistan EE",
        main_example_field="pakistanExampleRow",
        ee_example_field="pakistanEeExampleRow",
        fix_main_ee_refs=fix_pakistan_row_pakistan_ee_refs,
        fix_ee_main_refs=fix_ee_row_pakistan_refs,
        main_snapshots=main_snapshots,
        ee_snapshots=ee_snapshots,
    )


def expand_pakistan_employee_rows(
    wb,
    employees: list[dict[str, Any]],
    *,
    employee_directory: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """扩行 + 按 employeeFormulaStyles 盖公式。返回 (formula_plan, warnings)。"""
    mapping = _active_mapping()
    formula_rows = _pk_formula_rows()
    warnings: list[str] = []
    if PK_SHEET not in wb.sheetnames or PK_EE_SHEET not in wb.sheetnames:
        return [], warnings

    need_pk, need_ee = needed_example_rows_for_styles(
        mapping,
        employees,
        main_template_key="Pakistan",
        ee_template_key="Pakistan EE",
        main_example_field="pakistanExampleRow",
        ee_example_field="pakistanEeExampleRow",
        main_data_start=PK_DATA_START,
        ee_data_start=PK_EE_DATA_START,
        employee_directory=employee_directory,
    )
    main_snaps = {r: snapshot_row_cells(wb[PK_SHEET], r) for r in need_pk}
    ee_snaps = {r: snapshot_row_cells(wb[PK_EE_SHEET], r) for r in need_ee}
    fit_pakistan_formula_sheets(
        wb,
        len(employees),
        clear_excess=False,
        protected_pk_rows=need_pk,
        protected_ee_rows=need_ee,
    )
    formula_plan = apply_pakistan_employee_formula_styles(
        wb,
        employees,
        formula_rows=formula_rows,
        employee_directory=employee_directory,
        main_snapshots=main_snaps,
        ee_snapshots=ee_snaps,
    )
    clear_excess_pakistan_formula_rows(wb, len(employees))
    for p in formula_plan or []:
        warnings.append(
            f"公式配对：第{p.get('index')}人 → Pakistan第{p.get('mainExampleRow')}行 / Pakistan EE第{p.get('eeExampleRow')}行"
        )
    return formula_plan or [], warnings


def _pn_customer_id(pn_meta: PnMeta | dict[str, Any] | None) -> str | None:
    if pn_meta is None:
        return None
    if isinstance(pn_meta, PnMeta):
        cid = (pn_meta.customer_id or "").strip()
        return cid or None
    if isinstance(pn_meta, dict):
        cid = str(pn_meta.get("customer_id") or pn_meta.get("customerId") or "").strip()
        return cid or None
    return None


def apply_pakistan_ee_codes(
    wb,
    employees: list[dict[str, Any]],
    *,
    employee_directory: list[dict[str, Any]] | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
) -> list[str]:
    if PK_EE_SHEET not in wb.sheetnames:
        return []
    ws = wb[PK_EE_SHEET]
    client_code = _pn_customer_id(pn_meta)
    directory = list(employee_directory or [])
    warnings: list[str] = []
    for i, emp in enumerate(employees):
        row = PK_EE_DATA_START + i
        if client_code:
            # 仅当目标非公式时写入
            if not _cell_formula_text(ws.cell(row, 2).value):
                ws.cell(row, 2).value = client_code
        name = _norm(emp.get("Name of Employee") or emp.get("Employee Name"))
        code, warn = match_ee_code([name] if name else [], directory)
        # EE Code：样例母版多为公式 PN!B9&" - "&n；有公式则不覆盖
        if not _cell_formula_text(ws.cell(row, 4).value):
            ws.cell(row, 4).value = code
        if warn:
            warnings.append(f"Pakistan EE 第{i + 1}人：{warn}")
    return warnings


def apply_fx(wb, *, fill_fx: bool = True, source_fx: float | None = None) -> float | None:
    if not fill_fx or PN_SHEET not in wb.sheetnames:
        return None
    fx = source_fx
    if fx is None:
        try:
            fx = get_pakistan_pn_fx_rate(fetch_usd_rates())
        except Exception:
            return None
    pn = wb[PN_SHEET]
    cell = pn.cell(33, 2)
    existing = cell.value
    text = _cell_formula_text(existing)
    # 母版常写成 =270.0317 这类常量公式，应用发票推算/API 汇率覆盖
    if text:
        body = text[1:].strip()
        try:
            float(body)
        except ValueError:
            # 真正业务公式则保留
            return fx
    cell.value = float(fx)
    return fx


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    registry_dir: Path | None = None,
    convert_mapping: dict[str, Any] | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    global _ACTIVE_MAPPING
    _ACTIVE_MAPPING = resolve_convert_mapping("pakistan_payroll_calc", convert_mapping)
    try:
        source_path = Path(source_path).resolve()
        output_path = Path(output_path).resolve()
        template_path = Path(template_path).resolve()
        if not source_path.is_file():
            raise FileNotFoundError(f"源文件不存在: {source_path}")
        if not template_path.is_file():
            raise FileNotFoundError(f"母版不存在: {template_path}")

        src_wb = load_workbook(source_path, data_only=False)
        try:
            l_name = find_sheet_name(list(src_wb.sheetnames), _active_mapping().get("sourceEmployeeSheet"))
            if not l_name or l_name not in src_wb.sheetnames:
                if PK_L_SHEET in src_wb.sheetnames:
                    l_name = PK_L_SHEET
                else:
                    raise ValueError(f"未找到 {PK_L_SHEET}，现有: {src_wb.sheetnames}")
            employees = parse_pakistan_l_employees(src_wb[l_name])
            source_fx = None
            if employees:
                source_fx = _as_float(employees[0].get("_fx_rate"))
            if source_fx is None and PK_L_SHEET in src_wb.sheetnames:
                source_fx = _as_float(src_wb[PK_L_SHEET].cell(3, 6).value)
        finally:
            src_wb.close()

        if not employees:
            raise ValueError("Pakistan-L 未解析到员工行")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)
        warnings: list[str] = []
        formula_plan: list[dict[str, Any]] = []
        fx = None
        applied_pn = None
        try:
            if PK_L_SHEET not in wb.sheetnames:
                raise ValueError(f"母版缺少 {PK_L_SHEET}")
            write_pakistan_l(wb[PK_L_SHEET], employees)
            formula_plan, style_warnings = expand_pakistan_employee_rows(
                wb,
                employees,
                employee_directory=employee_directory,
            )
            warnings.extend(style_warnings)
            warnings.extend(_apply_invoice_derived_business_tax(wb, employees))
            try:
                fx = apply_fx(wb, fill_fx=fill_fx, source_fx=source_fx)
            except Exception as exc:
                warnings.append(f"写入 PN 汇率失败: {exc}")

            if pn_meta is not None:
                applied_pn = apply_pn_meta(
                    wb,
                    pn_meta,
                    registry_dir=registry_dir or output_path.parent,
                    reserve_invoice_number=True,
                )
            warnings.extend(
                apply_pakistan_ee_codes(
                    wb,
                    employees,
                    employee_directory=employee_directory,
                    pn_meta=applied_pn or pn_meta,
                )
            )
            wb.save(output_path)
        finally:
            wb.close()

        postprocess_converted_xlsx(output_path)
        default_main = default_example_row_for_mapping(_ACTIVE_MAPPING or {}, "Pakistan", PK_DATA_START)
        used_mains = {
            int(p["mainExampleRow"])
            for p in (formula_plan or [])
            if p.get("mainExampleRow") is not None
        }
        formula_rows_text = ",".join(str(r) for r in sorted(used_mains)) if used_mains else None
        if not formula_plan:
            match_hint = None
        elif used_mains - {default_main}:
            match_hint = "style-row-hit"
        else:
            match_hint = "style-row-default"
        return {
            "ok": True,
            "engine_id": "pakistan_payroll_calc",
            "region": "Pakistan",
            "output": str(output_path),
            "employee_count": len(employees),
            "fx_rate": fx,
            "warnings": warnings,
            "pn_meta": applied_pn.to_dict() if applied_pn else None,
            "formula_main_rows": formula_rows_text,
            "formula_match_hint": match_hint,
        }
    finally:
        _ACTIVE_MAPPING = None


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Pakistan-L → Pakistan PN")
    parser.add_argument("source", type=Path)
    parser.add_argument("-o", "--output", type=Path, required=True)
    parser.add_argument("-t", "--template", type=Path, default=None)
    args = parser.parse_args(argv)
    tpl = args.template or DEFAULT_TEMPLATE
    result = convert(args.source, args.output, tpl)
    print(result)
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
