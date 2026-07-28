# -*- coding: utf-8 -*-
"""
香港 Vertical-L 源账单 → Hong Kong PN（引擎 hk_vertical_l）

用法:
  python -m profiles.hk_vertical_l.convert <原始 T-N 账单.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

原始账单: sheet「Hong Kong-L」第 7 行表头、第 8 行起员工数据（按表头名匹配）
默认母版: templates/hongkong/template.xlsx
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from fx_rate import fetch_usd_rates, get_hk_pn_fx_rate
from pn_meta import PnMeta, apply_pn_meta
from region_templates import get_region_template
from xlsx_convert_utils import (
    clean_value,
    coerce_datetime_for_excel,
    is_date_column_header,
    norm,
)
from xlsx_postprocess import postprocess_converted_xlsx

DEFAULT_TEMPLATE = get_region_template("Hong Kong")

HK_L_SHEET = "Hong Kong-L"
HK_SHEET = "Hong Kong"
HK_EE_SHEET = "Hong Kong EE"
PN_SHEET = "PN"

HK_L_HEADER_ROW = 7
HK_L_DATA_START_ROW = 8
HK_DATA_START_ROW = 9
HK_EE_DATA_START_ROW = 10
_DATE_FMT = "yyyy/m/d"
MAX_EMPLOYEES = 10

# 源表头 → 目标表头（同名则省略）
SOURCE_TO_TARGET_HEADER: dict[str, str] = {
    "Basic Salary": "Base Salary",
}

# 源表有、目标表没有的列（跳过）
SKIP_SOURCE_HEADERS = frozenset({"Medical Insurance Allowance"})

NAME_HEADERS = ("Name of Employee", "EE Name", "Name")


def map_source_header(source_header: str) -> str | None:
    h = norm(source_header)
    if not h or h in SKIP_SOURCE_HEADERS:
        return None
    return SOURCE_TO_TARGET_HEADER.get(h, h)


def build_header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def format_payroll_date(value: Any) -> datetime | None:
    return coerce_datetime_for_excel(value)


def normalize_company_name(value: Any) -> Any:
    if value is None:
        return None
    s = norm(value)
    m = re.search(r"\(([^)]+)\)", s)
    if m:
        return m.group(1).strip()
    return s


def read_hk_l_employees(ws: Worksheet) -> list[dict[str, Any]]:
    source_headers = build_header_map(ws, HK_L_HEADER_ROW)
    name_col = None
    for nh in NAME_HEADERS:
        if nh in source_headers:
            name_col = source_headers[nh]
            break
    if name_col is None:
        raise ValueError(f"「Hong Kong-L」第 {HK_L_HEADER_ROW} 行须包含员工姓名表头（如 Name of Employee）")

    employees: list[dict[str, Any]] = []
    for row in range(HK_L_DATA_START_ROW, (ws.max_row or 0) + 1):
        name = clean_value(ws.cell(row, name_col).value)
        if name is None:
            continue
        record: dict[str, Any] = {}
        for src_hdr, col in source_headers.items():
            target_hdr = map_source_header(src_hdr)
            if target_hdr is None:
                continue
            val = clean_value(ws.cell(row, col).value)
            if val is not None:
                if is_date_column_header(target_hdr):
                    dt = coerce_datetime_for_excel(val)
                    if dt is not None:
                        val = dt
                record[target_hdr] = val
        employees.append(record)

    if not employees:
        raise ValueError("「Hong Kong-L」中未找到有效员工行")
    if len(employees) > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {len(employees)} 超过模板上限 {MAX_EMPLOYEES}")
    return employees


def read_hk_l_meta(ws: Worksheet) -> dict[str, Any]:
    return {
        "company_name": normalize_company_name(ws.cell(1, 3).value),
        "period_from": format_payroll_date(ws.cell(2, 3).value),
        "period_to": format_payroll_date(ws.cell(2, 5).value),
        "currency": ws.cell(3, 3).value,
    }


def ensure_hk_period_date_formats(wb) -> None:
    """Hong Kong!B2/C2 引用账期；强制日期格式，避免显示成金额。"""
    try:
        hkl = wb[HK_L_SHEET]
        for row, col in ((2, 3), (2, 5)):  # C2 / E2
            hkl.cell(row, col).number_format = _DATE_FMT
    except KeyError:
        pass
    try:
        hk = wb[HK_SHEET]
        for row, col in ((2, 2), (2, 3)):  # B2 / C2
            hk.cell(row, col).number_format = _DATE_FMT
    except KeyError:
        pass


def clear_hk_l_data(ws: Worksheet, from_row: int = HK_L_DATA_START_ROW) -> None:
    max_row = max(ws.max_row or from_row, from_row + MAX_EMPLOYEES)
    max_col = ws.max_column or 1
    for row in range(from_row, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def write_hk_l(ws: Worksheet, employees: list[dict[str, Any]], meta: dict[str, Any]) -> None:
    target_headers = build_header_map(ws, HK_L_HEADER_ROW)
    if not target_headers:
        raise ValueError(f"「Hong Kong-L」第 {HK_L_HEADER_ROW} 行表头为空")

    if meta.get("company_name") is not None:
        ws.cell(1, 3).value = meta["company_name"]
    if meta.get("period_from") is not None:
        cell = ws.cell(2, 3)
        cell.value = meta["period_from"]
        cell.number_format = _DATE_FMT
    if meta.get("period_to") is not None:
        cell = ws.cell(2, 5)
        cell.value = meta["period_to"]
        cell.number_format = _DATE_FMT
    if meta.get("currency") is not None:
        ws.cell(3, 3).value = meta["currency"]

    clear_hk_l_data(ws)
    for idx, emp in enumerate(employees):
        row = HK_L_DATA_START_ROW + idx
        for hdr, val in emp.items():
            col = target_headers.get(hdr)
            if col is not None:
                out_val = val
                if is_date_column_header(hdr):
                    dt = coerce_datetime_for_excel(val)
                    if dt is not None:
                        out_val = dt
                cell = ws.cell(row, col)
                cell.value = out_val
                if is_date_column_header(hdr) and isinstance(out_val, datetime):
                    cell.number_format = _DATE_FMT


def shift_hk_formula(
    formula: str,
    from_row: int,
    to_row: int,
    hk_l_from: int,
    hk_l_to: int,
) -> str:
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    s = re.sub(
        r"'Hong Kong-L'!([A-Z]{1,3})(\d+)",
        lambda m: f"'Hong Kong-L'!{m.group(1)}{hk_l_to}",
        formula,
    )
    s = re.sub(
        rf"(?<!\$)(?<![A-Z])([A-Z]{{1,3}}){from_row}(?!\d)",
        lambda m: f"{m.group(1)}{to_row}",
        s,
    )
    return s


def copy_row_formulas(
    ws: Worksheet,
    from_row: int,
    to_row: int,
    hk_l_from: int,
    hk_l_to: int,
) -> None:
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        src = ws.cell(from_row, col)
        dst = ws.cell(to_row, col)
        if src.data_type == "f" and isinstance(src.value, str):
            dst.value = shift_hk_formula(src.value, from_row, to_row, hk_l_from, hk_l_to)
        elif src.value is not None and src.data_type != "f":
            dst.value = copy(src.value)


def expand_hk_employee_rows(ws: Worksheet, employee_count: int) -> None:
    if employee_count <= 1:
        return
    for i in range(1, employee_count):
        target_row = HK_DATA_START_ROW + i
        hk_l_row = HK_L_DATA_START_ROW + i
        copy_row_formulas(ws, HK_DATA_START_ROW, target_row, HK_L_DATA_START_ROW, hk_l_row)


def expand_hk_ee_rows(ws: Worksheet, employee_count: int) -> None:
    if employee_count <= 1:
        return
    for i in range(1, employee_count):
        target_row = HK_EE_DATA_START_ROW + i
        hk_l_row = HK_L_DATA_START_ROW + i
        hk_row = HK_DATA_START_ROW + i
        copy_row_formulas(ws, HK_EE_DATA_START_ROW, target_row, HK_L_DATA_START_ROW, hk_l_row)
        for col in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(target_row, col)
            if cell.data_type == "f" and isinstance(cell.value, str):
                cell.value = re.sub(
                    rf"Hong Kong!([A-Z]+){HK_DATA_START_ROW}(?!\d)",
                    lambda m: f"Hong Kong!{m.group(1)}{hk_row}",
                    cell.value,
                )
                cell.value = re.sub(
                    rf"'Hong Kong EE'!([A-Z]+){HK_EE_DATA_START_ROW}(?!\d)",
                    lambda m: f"'Hong Kong EE'!{m.group(1)}{target_row}",
                    cell.value,
                )


def parse_source(source_path: Path) -> dict[str, Any]:
    wb = load_workbook(source_path, data_only=True, read_only=True)
    if HK_L_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"未找到 sheet「{HK_L_SHEET}」，现有: {wb.sheetnames}")

    ws = wb[HK_L_SHEET]
    employees = read_hk_l_employees(ws)
    meta = read_hk_l_meta(ws)
    wb.close()
    return {"employees": employees, "meta": meta}


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
) -> dict[str, Any]:
    if not template_path.is_file():
        raise FileNotFoundError(f"母版不存在: {template_path}")
    if not source_path.is_file():
        raise FileNotFoundError(f"原始账单不存在: {source_path}")

    parsed = parse_source(source_path)
    employees = parsed["employees"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, rich_text=True)
    for name in (HK_L_SHEET, HK_SHEET, HK_EE_SHEET, PN_SHEET):
        if name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"母版缺少 sheet「{name}」，现有: {wb.sheetnames}")

    applied_pn: PnMeta | None = None
    if pn_meta is not None:
        applied_pn = apply_pn_meta(
            wb,
            pn_meta,
            registry_dir=registry_dir or output_path.parent,
            reserve_invoice_number=True,
        )

    write_hk_l(wb[HK_L_SHEET], employees, parsed["meta"])
    expand_hk_employee_rows(wb[HK_SHEET], len(employees))
    expand_hk_ee_rows(wb[HK_EE_SHEET], len(employees))

    rates = fetch_usd_rates()
    fx_rate = get_hk_pn_fx_rate(rates)
    wb[PN_SHEET]["B28"].value = fx_rate

    ensure_hk_period_date_formats(wb)
    wb.save(output_path)
    wb.close()
    postprocess_converted_xlsx(output_path)

    return {
        "employee_count": len(employees),
        "employee_names": [
            e.get("Name of Employee") or e.get("EE Name") or e.get("Name")
            for e in employees
        ],
        "company_name": parsed["meta"].get("company_name"),
        "period": (parsed["meta"].get("period_from"), parsed["meta"].get("period_to")),
        "fx_rate": fx_rate,
        "fx_source": "api:HKD*0.97",
        "output": str(output_path),
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Hong Kong T-N 账单 → PN N-C 自动转换")
    parser.add_argument("source", type=Path, help="原始 T-N Excel 路径")
    parser.add_argument("-o", "--output", type=Path, help="输出 PN 路径")
    parser.add_argument("-t", "--template", type=Path, default=DEFAULT_TEMPLATE, help="PN 母版路径")
    args = parser.parse_args(argv)

    source = args.source.resolve()
    output = args.output.resolve() if args.output else source.parent / f"PN_auto_{source.stem}.xlsx"

    try:
        result = convert(source, output, args.template.resolve())
    except Exception as exc:
        print(f"转换失败: {exc}", file=sys.stderr)
        return 1

    print("转换完成")
    print(f"  输出: {result['output']}")
    print(f"  公司: {result['company_name']}")
    print(f"  账期: {result['period'][0]} ~ {result['period'][1]}")
    print(f"  员工: {result['employee_count']} 人 → {result['employee_names']}")
    print(f"  汇率 PN!B28: {result['fx_rate']} ({result.get('fx_source')})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
