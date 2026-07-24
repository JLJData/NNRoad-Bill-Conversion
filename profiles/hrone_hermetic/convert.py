# -*- coding: utf-8 -*-
"""
HROne Co., Ltd. + Hermetic 账单 → PN 转换脚本

用法:
  python -m profiles.hrone_hermetic.convert <原始账单.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

默认母版: templates/china/template.xlsx
（含 PN / China / China EE / China-L 公式结构）
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from pathlib import Path
from typing import Any

from pn_meta import PnMeta, apply_pn_meta

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from fx_rate import fetch_usd_rates, get_china_pn_fx_rate
from region_templates import get_region_template
from xlsx_convert_utils import clean_value, norm
from xlsx_luckysheet_compat import apply_luckysheet_compat
from xlsx_postprocess import postprocess_converted_xlsx

DEFAULT_TEMPLATE = get_region_template("China")

CALC_SHEET_NAMES = ("计算结果",)
OTHER_FEE_NAMES = ("Other Fee",)
PAYMENT_NOTICE_NAMES = ("S-Payment Notice",)

CHINA_DATA_START_ROW = 9
CHINA_EE_DATA_START_ROW = 10
CHINA_L_DATA_START_ROW = 2
MAX_EMPLOYEES = 10


def find_sheet(wb, candidates: tuple[str, ...]) -> str | None:
    names = {n: n for n in wb.sheetnames}
    for c in candidates:
        if c in names:
            return c
    lower = {n.lower(): n for n in wb.sheetnames}
    for c in candidates:
        if c.lower() in lower:
            return lower[c.lower()]
    for n in wb.sheetnames:
        for c in candidates:
            if c in n:
                return n
    return None


def build_header_map(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def parse_expense_count(text: Any) -> int:
    if text is None:
        return 0
    m = re.search(r"(\d+)", str(text))
    return int(m.group(1)) if m else 0


def extract_by_label_col_a(ws: Worksheet, label: str, value_col: int) -> Any:
    target = norm(label)
    for row in range(1, (ws.max_row or 0) + 1):
        if norm(ws.cell(row, 1).value) == target:
            return clean_value(ws.cell(row, value_col).value)
    return None


def read_calc_employees(ws: Worksheet) -> list[dict[str, Any]]:
    headers = build_header_map(ws, 1)
    if "姓名" not in headers:
        raise ValueError("「计算结果」sheet 第 1 行须包含表头「姓名」")

    employees: list[dict[str, Any]] = []
    for row in range(2, (ws.max_row or 0) + 1):
        name = clean_value(ws.cell(row, headers["姓名"]).value)
        if name is None:
            continue
        record: dict[str, Any] = {}
        for hdr, col in headers.items():
            record[hdr] = clean_value(ws.cell(row, col).value)
        employees.append(record)

    if not employees:
        raise ValueError("「计算结果」中未找到有效员工行（姓名非空）")
    if len(employees) > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {len(employees)} 超过模板上限 {MAX_EMPLOYEES}")
    return employees


def clear_china_l_data(ws: Worksheet, from_row: int = CHINA_L_DATA_START_ROW) -> None:
    max_row = max(ws.max_row or from_row, from_row + MAX_EMPLOYEES)
    max_col = ws.max_column or 1
    for row in range(from_row, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def write_china_l(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    target_headers = build_header_map(ws, 1)
    if not target_headers:
        raise ValueError("China-L 第 1 行表头为空")

    clear_china_l_data(ws)
    for idx, emp in enumerate(employees):
        row = CHINA_L_DATA_START_ROW + idx
        for hdr, val in emp.items():
            col = target_headers.get(hdr)
            if col is not None and val is not None:
                ws.cell(row, col).value = val


def shift_formula(formula: str, from_row: int, to_row: int, china_l_from: int, china_l_to: int) -> str:
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    s = re.sub(
        r"'China-L'!([A-Z]{1,3})(\d+)",
        lambda m: f"'China-L'!{m.group(1)}{china_l_to}",
        formula,
    )
    s = re.sub(
        rf"(?<!\$)(?<![A-Z])([A-Z]{{1,3}}){from_row}(?!\d)",
        lambda m: f"{m.group(1)}{to_row}",
        s,
    )
    return s


def copy_row_formulas(ws: Worksheet, from_row: int, to_row: int, china_l_from: int, china_l_to: int) -> None:
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        src = ws.cell(from_row, col)
        dst = ws.cell(to_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
        if src.data_type == "f" and isinstance(src.value, str):
            dst.value = shift_formula(src.value, from_row, to_row, china_l_from, china_l_to)
        elif src.value is not None and src.data_type != "f":
            dst.value = copy(src.value)


def clear_employee_row(ws: Worksheet, row: int) -> None:
    for col in range(1, (ws.max_column or 0) + 1):
        ws.cell(row, col).value = None


def count_template_employee_slots(ws: Worksheet, data_start_row: int, marker_col: int) -> int:
    n = 0
    for row in range(data_start_row, data_start_row + MAX_EMPLOYEES):
        if ws.cell(row, marker_col).value is not None:
            n += 1
        else:
            break
    return max(n, 1)


def fit_china_formula_sheets(wb, employee_count: int) -> None:
    """
    母版 China / China EE 预置 1 人公式行（汇总区预留到 +9）。
    人数不足时清多余行；人数更多时复制扩展（含样式）。
    """
    n = max(int(employee_count), 1)
    china = wb["China"]
    ee = wb["China EE"]
    china_slots = count_template_employee_slots(china, CHINA_DATA_START_ROW, marker_col=3)
    ee_slots = count_template_employee_slots(ee, CHINA_EE_DATA_START_ROW, marker_col=4)

    for i in range(n, china_slots):
        clear_employee_row(china, CHINA_DATA_START_ROW + i)
    for i in range(n, ee_slots):
        clear_employee_row(ee, CHINA_EE_DATA_START_ROW + i)

    if n > china_slots:
        src_row = CHINA_DATA_START_ROW + china_slots - 1
        src_l = CHINA_L_DATA_START_ROW + china_slots - 1
        for i in range(china_slots, n):
            copy_row_formulas(
                china,
                src_row,
                CHINA_DATA_START_ROW + i,
                src_l,
                CHINA_L_DATA_START_ROW + i,
            )

    if n > ee_slots:
        src_row = CHINA_EE_DATA_START_ROW + ee_slots - 1
        src_l = CHINA_L_DATA_START_ROW + ee_slots - 1
        for i in range(ee_slots, n):
            dst_row = CHINA_EE_DATA_START_ROW + i
            china_row = CHINA_DATA_START_ROW + i
            copy_row_formulas(ee, src_row, dst_row, src_l, CHINA_L_DATA_START_ROW + i)
            for col in range(1, (ee.max_column or 0) + 1):
                cell = ee.cell(dst_row, col)
                if cell.data_type == "f" and isinstance(cell.value, str):
                    cell.value = re.sub(
                        rf"China!([A-Z]+){CHINA_DATA_START_ROW + ee_slots - 1}(?!\d)",
                        lambda m, cr=china_row: f"China!{m.group(1)}{cr}",
                        cell.value,
                    )


def apply_china_specials(
    ws: Worksheet,
    employee_count: int,
    expense_count: int,
    other_amount: Any,
    *,
    fx_row: int = 29,
) -> None:
    for i in range(employee_count):
        row = CHINA_DATA_START_ROW + i
        ws.cell(row, 9).value = f"=40*PN!$B${fx_row}*{expense_count}"
        # 同步母版 H 列里的汇率绝对行号
        h = ws.cell(row, 8).value
        if isinstance(h, str) and "PN!" in h:
            ws.cell(row, 8).value = re.sub(r"PN!\$?B\$?\d+", f"PN!$B${fx_row}", h)
        if other_amount is not None:
            ws.cell(row, 10).value = other_amount


# ---------- PN 人员扩减 ----------

PN_SHEET = "PN"
_PN_EOR_ROW = 15
_PN_LABOR_START_ROW = 16
_PN_RMB_FMT = "#,##0.00"
_PN_USD_FMT = '$#,##0.00'


def _pn_layout(employee_count: int) -> dict[str, int]:
    n = max(int(employee_count), 1)
    labor_start = _PN_LABOR_START_ROW
    expense_start = labor_start + n
    blank_row = expense_start + n
    service_row = blank_row + 1
    mgmt_row = service_row + 1
    return {
        "n": n,
        "eor_row": _PN_EOR_ROW,
        "labor_start": labor_start,
        "expense_start": expense_start,
        "blank_row": blank_row,
        "service_row": service_row,
        "mgmt_row": mgmt_row,
        "sum_end": expense_start + n - 1,
    }


def _count_pn_labor_slots(ws: Worksheet) -> int:
    n = 0
    for row in range(_PN_LABOR_START_ROW, _PN_LABOR_START_ROW + MAX_EMPLOYEES + 2):
        v = ws.cell(row, 1).value
        if isinstance(v, str) and "Labor cost" in v:
            n += 1
            continue
        break
    return max(n, 1)


def _find_pn_row_by_label(ws: Worksheet, keyword: str, col: int = 1) -> int | None:
    key = keyword.lower()
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and key in v.lower():
            return row
    return None


def _find_pn_fx_row(ws: Worksheet) -> int:
    return _find_pn_row_by_label(ws, "FX rate") or 29


def _copy_pn_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    max_col = max(ws.max_column or 6, 6)
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _ensure_merge_a_c(ws: Worksheet, row: int) -> None:
    rng = f"A{row}:C{row}"
    for m in list(ws.merged_cells.ranges):
        if m.min_row <= row <= m.max_row and m.min_col <= 1 and m.max_col >= 3:
            if m.min_row == row and m.max_row == row and m.min_col == 1 and m.max_col == 3:
                return
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    try:
        ws.merge_cells(rng)
    except ValueError:
        pass


def _collect_merges_from(ws: Worksheet, start_row: int) -> list[tuple[int, int, int, int]]:
    kept: list[tuple[int, int, int, int]] = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row < start_row:
            continue
        kept.append((m.min_row, m.min_col, m.max_row, m.max_col))
        try:
            ws.unmerge_cells(str(m))
        except ValueError:
            pass
    return kept


def _restore_merges(
    ws: Worksheet,
    merges: list[tuple[int, int, int, int]],
    row_shift: int = 0,
) -> None:
    for min_r, min_c, max_r, max_c in merges:
        nr1, nr2 = min_r + row_shift, max_r + row_shift
        if nr1 < 1 or nr2 < nr1:
            continue
        try:
            ws.merge_cells(
                start_row=nr1,
                start_column=min_c,
                end_row=nr2,
                end_column=max_c,
            )
        except ValueError:
            pass


def _shift_row_heights(ws: Worksheet, start_row: int, amount: int) -> None:
    if amount == 0:
        return
    captured: dict[int, float] = {}
    max_r = max(ws.max_row or start_row, start_row)
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            captured[r] = h
        max_r = max(max_r, r)
    for r in range(start_row, max_r + abs(amount) + 2):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    if amount > 0:
        for r in sorted(captured.keys(), reverse=True):
            ws.row_dimensions[r + amount].height = captured[r]
    else:
        for r in sorted(captured.keys()):
            new_r = r + amount
            if new_r >= start_row:
                ws.row_dimensions[new_r].height = captured[r]


def _capture_row_heights_from(ws: Worksheet, start_row: int) -> dict[int, float]:
    out: dict[int, float] = {}
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            out[r] = h
    return out


def _apply_row_heights(ws: Worksheet, heights: dict[int, float], start_row: int) -> None:
    max_r = max([start_row, *heights.keys()], default=start_row)
    for r in range(start_row, max_r + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r, h in heights.items():
        if r >= start_row:
            ws.row_dimensions[r].height = h


def _pn_insert_rows(ws: Worksheet, idx: int, amount: int, fill_style_row: int | None = None) -> None:
    if amount <= 0:
        return
    merges = _collect_merges_from(ws, idx)
    _shift_row_heights(ws, idx, amount)
    ws.insert_rows(idx, amount)
    _restore_merges(ws, merges, row_shift=amount)
    if fill_style_row is not None:
        style_row = fill_style_row + amount if fill_style_row >= idx else fill_style_row
        for r in range(idx, idx + amount):
            _copy_pn_row_style(ws, style_row, r)
            _ensure_merge_a_c(ws, r)


def _pn_delete_rows(ws: Worksheet, idx: int, amount: int) -> None:
    if amount <= 0:
        return
    merges = _collect_merges_from(ws, idx + amount)
    heights = _capture_row_heights_from(ws, idx)
    ws.delete_rows(idx, amount)
    new_heights: dict[int, float] = {}
    for r, h in heights.items():
        if r < idx:
            continue
        if r < idx + amount:
            continue
        new_heights[r - amount] = h
    _apply_row_heights(ws, new_heights, idx)
    _restore_merges(ws, merges, row_shift=-amount)


def _rewrite_pn_settlement_block(ws: Worksheet, *, fx_row: int, total_row: int) -> None:
    """
    中国母版 FX 下方：
      F{fx}=F{total}；F{fx+1..4}=China!E22..E25；Sub/Tax/Total 相对锚点。
    """
    f_eor = ws.cell(fx_row, 6)
    f_eor.value = f"=F{total_row}"
    f_eor.number_format = _PN_USD_FMT
    for i, china_e in enumerate((22, 23, 24, 25), start=1):
        cell = ws.cell(fx_row + i, 6)
        cell.value = f"=China!E{china_e}"
    sub_row = fx_row + 5
    tax_row = fx_row + 6
    grand_row = fx_row + 7
    ws.cell(sub_row, 6).value = f"=SUM(F{fx_row}:F{fx_row + 4})"
    ws.cell(tax_row, 6).value = f"=China!F6/$B${fx_row}"
    ws.cell(grand_row, 6).value = f"=SUM(F{sub_row}:F{tax_row})"
    for r in (sub_row, tax_row, grand_row):
        ws.cell(r, 6).number_format = _PN_USD_FMT


def fit_pn_employees(ws: Worksheet, employee_count: int) -> dict[str, int]:
    """按人数扩/缩 PN Labor+Expense，并重写 Service/Deposit/合计/汇率区公式。"""
    n = max(int(employee_count), 1)
    if n > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {n} 超过模板上限 {MAX_EMPLOYEES}")

    old_n = _count_pn_labor_slots(ws)
    labor_start = _PN_LABOR_START_ROW
    expense_start = labor_start + old_n
    delta = n - old_n

    if delta > 0:
        _pn_insert_rows(ws, expense_start, delta, fill_style_row=labor_start)
        expense_start += delta
        _pn_insert_rows(ws, expense_start + old_n, delta, fill_style_row=expense_start)
    elif delta < 0:
        _pn_delete_rows(ws, expense_start + n, -delta)
        _pn_delete_rows(ws, labor_start + n, -delta)

    layout = _pn_layout(n)
    expense_start = layout["expense_start"]
    blank_row = layout["blank_row"]
    eor_row = layout["eor_row"]
    sum_end = layout["sum_end"]

    for m in list(ws.merged_cells.ranges):
        if m.min_row == blank_row and m.max_row == blank_row:
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    for col in range(1, 7):
        cell = ws.cell(blank_row, col)
        if type(cell).__name__ == "MergedCell":
            continue
        cell.value = None
    _ensure_merge_a_c(ws, blank_row)

    fx_row = _find_pn_fx_row(ws)

    for i in range(n):
        china_row = CHINA_DATA_START_ROW + i
        labor_row = labor_start + i
        expense_row = expense_start + i
        if i > 0:
            _copy_pn_row_style(ws, labor_start, labor_row)
            _copy_pn_row_style(ws, expense_start, expense_row)
        _ensure_merge_a_c(ws, labor_row)
        _ensure_merge_a_c(ws, expense_row)

        ws.cell(labor_row, 1).value = (
            f'="- Labor cost for "&China!B{china_row}&"  -  "&MONTH(China!$B$2)&"-"&YEAR(China!$B$2)'
        )
        e_labor = ws.cell(labor_row, 5)
        e_labor.value = f"=China!K{china_row}+China!AQ{china_row}+China!BC{china_row}"
        e_labor.number_format = _PN_RMB_FMT
        f_labor = ws.cell(labor_row, 6)
        f_labor.value = f"=E{labor_row}/$B${fx_row}"
        f_labor.number_format = _PN_USD_FMT

        ws.cell(expense_row, 1).value = f'="- Expense claim for "&China!B{china_row}'
        e_exp = ws.cell(expense_row, 5)
        e_exp.value = f"=China!BF{china_row}"
        e_exp.number_format = _PN_RMB_FMT
        f_exp = ws.cell(expense_row, 6)
        f_exp.value = f"=E{expense_row}/$B${fx_row}"
        f_exp.number_format = _PN_USD_FMT

    e_eor = ws.cell(eor_row, 5)
    e_eor.value = f"=SUM(E{labor_start}:E{sum_end})"
    e_eor.number_format = _PN_RMB_FMT
    f_eor = ws.cell(eor_row, 6)
    f_eor.value = f"=SUM(F{labor_start}:F{sum_end})"
    f_eor.number_format = _PN_USD_FMT

    svc = _find_pn_row_by_label(ws, "Service Fee") or layout["service_row"]
    mgmt = _find_pn_row_by_label(ws, "Management Fee") or layout["mgmt_row"]
    e_svc = ws.cell(svc, 5)
    e_svc.value = f"=SUM(E{mgmt}:E{mgmt + 1})"
    e_svc.number_format = _PN_RMB_FMT
    f_svc = ws.cell(svc, 6)
    f_svc.value = f"=SUM(F{mgmt}:F{mgmt + 1})"
    f_svc.number_format = _PN_USD_FMT
    ws.cell(mgmt, 1).value = '="- Management Fee - "&MONTH(China!$B$2)&"-"&YEAR(China!$B$2)'
    e_mgmt = ws.cell(mgmt, 5)
    e_mgmt.value = "=China!G6"
    e_mgmt.number_format = _PN_RMB_FMT
    f_mgmt = ws.cell(mgmt, 6)
    f_mgmt.value = f"=E{mgmt}/$B${fx_row}"
    f_mgmt.number_format = _PN_USD_FMT
    _ensure_merge_a_c(ws, svc)
    _ensure_merge_a_c(ws, mgmt)

    deposit_hdr = _find_pn_row_by_label(ws, "Recurring Deposit")
    deposit_row = None
    if deposit_hdr:
        deposit_row = deposit_hdr + 1
        e_dh = ws.cell(deposit_hdr, 5)
        e_dh.value = f"=SUM(E{deposit_row})"
        e_dh.number_format = _PN_RMB_FMT
        f_dh = ws.cell(deposit_hdr, 6)
        f_dh.value = f"=SUM(F{deposit_row})"
        f_dh.number_format = _PN_USD_FMT
        ws.cell(deposit_row, 1).value = (
            '="- Deposit  for "&China!B9&"  -  "&MONTH(China!B2)&"-"&YEAR(China!B2)'
            if n == 1
            else '="- Deposit"'
        )
        e_dep = ws.cell(deposit_row, 5)
        e_dep.value = "=China!BI6"
        e_dep.number_format = _PN_RMB_FMT
        f_dep = ws.cell(deposit_row, 6)
        f_dep.value = f"=E{deposit_row}/$B${fx_row}"
        f_dep.number_format = _PN_USD_FMT

    total_row = _find_pn_row_by_label(ws, "EOR/PEO Service Cost")
    if total_row is None or total_row <= eor_row:
        total_row = (deposit_hdr or svc) + 3
    e_total = ws.cell(total_row, 5)
    if deposit_hdr:
        e_total.value = f"=E{eor_row}+E{svc}+E{deposit_hdr}"
    else:
        e_total.value = f"=E{eor_row}+E{svc}"
    e_total.number_format = _PN_RMB_FMT
    f_total = ws.cell(total_row, 6)
    if deposit_hdr:
        f_total.value = f"=F{eor_row}+F{svc}+F{deposit_hdr}"
    else:
        f_total.value = f"=F{eor_row}+F{svc}"
    f_total.number_format = _PN_USD_FMT

    fx_row = _find_pn_fx_row(ws)
    _rewrite_pn_settlement_block(ws, fx_row=fx_row, total_row=total_row)
    for row in range(labor_start, fx_row + 8):
        fcell = ws.cell(row, 6)
        if isinstance(fcell.value, str) and "$B$" in fcell.value:
            fcell.value = re.sub(r"\$B\$\d+", f"$B${fx_row}", fcell.value)

    layout["service_row"] = svc
    layout["mgmt_row"] = mgmt
    layout["deposit_hdr"] = deposit_hdr or 0
    layout["deposit_row"] = deposit_row or 0
    layout["total_row"] = total_row
    layout["fx_row"] = fx_row
    return layout



def retarget_pn_fx_refs(wb, fx_row: int) -> None:
    """插删 PN 行后，把 China / China EE / PN 里的 PN!$B$xx 指到新汇率行。"""
    pat = re.compile(r"PN!\$?B\$?\d+", re.IGNORECASE)
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "PN!" in v.upper() and pat.search(v):
                    cell.value = pat.sub(f"PN!$B${fx_row}", v)
                elif isinstance(v, str) and name == PN_SHEET and "$B$" in v:
                    # PN 表内 =E16/$B$29
                    if re.search(r"\$B\$\d+", v):
                        cell.value = re.sub(r"\$B\$\d+", f"$B${fx_row}", v)


def parse_source(source_path: Path) -> dict[str, Any]:
    wb = load_workbook(source_path, data_only=True, read_only=True)
    calc_name = find_sheet(wb, CALC_SHEET_NAMES)
    if not calc_name:
        wb.close()
        raise ValueError(f"未找到 sheet「计算结果」，现有: {wb.sheetnames}")

    employees = read_calc_employees(wb[calc_name])

    other_name = find_sheet(wb, OTHER_FEE_NAMES)
    payment_name = find_sheet(wb, PAYMENT_NOTICE_NAMES)

    other_amount = None
    expense_count = 0
    if other_name:
        other_ws = wb[other_name]
        other_amount = extract_by_label_col_a(other_ws, "其他", 4)
        expense_text = extract_by_label_col_a(other_ws, "报销服务费", 2)
        expense_count = parse_expense_count(expense_text)

    vendor_fx_rate = None
    if payment_name:
        vendor_fx_rate = clean_value(wb[payment_name]["C49"].value)

    wb.close()

    try:
        rates = fetch_usd_rates()
        fx_rate = get_china_pn_fx_rate(rates)
        fx_source = "api:CNY"
    except RuntimeError:
        fx_rate = vendor_fx_rate
        fx_source = "vendor:C49" if vendor_fx_rate is not None else "none"

    return {
        "employees": employees,
        "other_amount": other_amount,
        "expense_count": expense_count,
        "fx_rate": fx_rate,
        "fx_source": fx_source,
        "vendor_fx_rate": vendor_fx_rate,
    }


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
) -> dict[str, Any]:
    if not template_path.is_file():
        raise FileNotFoundError(f"母版不存在: {template_path}")
    if not source_path.is_file():
        raise FileNotFoundError(f"原始账单不存在: {source_path}")

    parsed = parse_source(source_path)
    employees = parsed["employees"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, rich_text=True)
    required = ("China-L", "China", "China EE", "PN")
    for name in required:
        if name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"母版缺少 sheet「{name}」，现有: {wb.sheetnames}")

    applied_pn: PnMeta | None = None
    if pn_meta is not None:
        applied_pn = apply_pn_meta(
            wb,
            pn_meta,
            registry_dir=registry_dir or output_path.parent,
            reserve_invoice_number=True,
        )

    write_china_l(wb["China-L"], employees)
    fit_china_formula_sheets(wb, len(employees))
    pn_layout = fit_pn_employees(wb[PN_SHEET], len(employees))
    fx_row = int(pn_layout.get("fx_row") or _find_pn_fx_row(wb[PN_SHEET]))

    if parsed["fx_rate"] is not None:
        wb[PN_SHEET].cell(fx_row, 2).value = parsed["fx_rate"]

    retarget_pn_fx_refs(wb, fx_row)
    apply_china_specials(
        wb["China"],
        len(employees),
        parsed["expense_count"],
        parsed["other_amount"],
        fx_row=fx_row,
    )

    apply_luckysheet_compat(wb, pn_sheet=PN_SHEET)

    wb.save(output_path)
    wb.close()
    # 主题填充 / 富文本；金额由前端 HyperFormula 按公式重算，不再注入 PN 缓存
    postprocess_converted_xlsx(output_path)

    return {
        "employee_count": len(employees),
        "employee_names": [e.get("姓名") for e in employees],
        "fx_rate": parsed["fx_rate"],
        "fx_source": parsed.get("fx_source"),
        "vendor_fx_rate": parsed.get("vendor_fx_rate"),
        "other_amount": parsed["other_amount"],
        "expense_count": parsed["expense_count"],
        "fx_row": fx_row,
        "output": str(output_path),
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="China 供应商账单 → PN 自动转换")
    parser.add_argument("source", type=Path, help="原始账单 Excel 路径")
    parser.add_argument("-o", "--output", type=Path, help="输出 PN 路径（默认同目录 PN_auto_*.xlsx）")
    parser.add_argument("-t", "--template", type=Path, default=DEFAULT_TEMPLATE, help="PN 母版路径")
    args = parser.parse_args(argv)

    source = args.source.resolve()
    if args.output:
        output = args.output.resolve()
    else:
        output = source.parent / f"PN_auto_{source.stem}.xlsx"

    try:
        result = convert(source, output, args.template.resolve())
    except Exception as exc:
        print(f"转换失败: {exc}", file=sys.stderr)
        return 1

    print("转换完成")
    print(f"  输出: {result['output']}")
    print(f"  员工: {result['employee_count']} 人 → {result['employee_names']}")
    print(f"  汇率 PN!B{result.get('fx_row', 29)}: {result['fx_rate']} ({result.get('fx_source')})")
    if result.get("vendor_fx_rate") is not None and result.get("fx_source") == "api:CNY":
        print(f"  供应商 C49 参考: {result['vendor_fx_rate']}")
    print(f"  Other Fee → China!J*: {result['other_amount']}")
    print(f"  报销笔数 → For Expense: {result['expense_count']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
