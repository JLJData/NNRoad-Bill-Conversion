# -*- coding: utf-8 -*-
"""分析 Payroll calculation → TW-L 字段映射，输出 mapping.json。"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[2]
PROFILE_DIR = Path(__file__).resolve().parent
SRC = ROOT / r"账单\now\Taiwan\Payroll Info_NNRoad 202603_Coral Sea -供应商账单.xlsx"
OUT = ROOT / r"账单\now\Taiwan\PN_CoralSea_N-C 20260311.xlsx"


def norm(v):
    if v is None:
        return None
    return str(v).replace("\uFEFF", "").strip()


def build_map(ws, row):
    m = {}
    for c in range(1, (ws.max_column or 0) + 1):
        h = norm(ws.cell(row, c).value)
        if h and h not in m:
            m[h] = c
    return m


def main():
    wb_src = load_workbook(SRC, data_only=True)
    wb_out = load_workbook(OUT, data_only=True)
    wb_out_f = load_workbook(OUT, data_only=False)

    pc = wb_src["Payroll calculation"]
    tw_l = wb_out["TW-L"]
    tw = wb_out_f["TW"]
    tw_ee = wb_out_f["TW EE"]

    PC_HR = 2
    TW_L_HR = 7
    TW_L_DR = 8

    pc_map = build_map(pc, PC_HR)
    out_map = build_map(tw_l, TW_L_HR)

    # explicit mapping from value comparison row 8 (first data row)
    pc_row = 3  # first data row in PC (no CN Name)
    out_row = 8
    mapping = []
    for oh, oc in out_map.items():
        out_val = tw_l.cell(out_row, oc).value
        if out_val is None:
            continue
        matched = None
        for ph, pc_col in pc_map.items():
            if pc.cell(pc_row, pc_col).value == out_val:
                matched = ph
                break
        mapping.append({"out": oh, "pc": matched, "val": out_val})

    # same for employee row 4 -> 9
    pc_row2 = 4
    out_row2 = 9
    mapping2 = []
    for oh, oc in out_map.items():
        out_val = tw_l.cell(out_row2, oc).value
        if out_val is None:
            continue
        matched = None
        for ph, pc_col in pc_map.items():
            if pc.cell(pc_row2, pc_col).value == out_val:
                matched = ph
                break
        mapping2.append({"out": oh, "pc": matched, "val": out_val})

    # TW sheet scan
    tw_info = {"meta": [], "formulas": []}
    for r in range(1, 15):
        for c in range(1, 15):
            v = tw.cell(r, c).value
            if v is not None:
                tw_info["meta"].append({"cell": f"{get_column_letter(c)}{r}", "v": str(v)[:120]})

    for r in range(8, 15):
        row_f = []
        for c in range(1, 30):
            cell = tw.cell(r, c)
            if cell.value is not None:
                item = {"col": get_column_letter(c), "v": str(cell.value)[:120]}
                if cell.data_type == "f":
                    item["f"] = True
                row_f.append(item)
        if row_f:
            tw_info["formulas"].append({"row": r, "cells": row_f})

    # TW EE
    ee_info = {"formulas": []}
    for r in range(4, 15):
        row_f = []
        for c in range(1, 40):
            cell = tw_ee.cell(r, c)
            if cell.value is not None:
                item = {"col": get_column_letter(c), "v": str(cell.value)[:120]}
                if cell.data_type == "f":
                    item["f"] = True
                row_f.append(item)
        if row_f:
            ee_info["formulas"].append({"row": r, "cells": row_f})

    # Summary sheet meta
    summary = wb_src["Summary"]
    sum_meta = []
    for r in range(1, 20):
        for c in range(1, 10):
            v = summary.cell(r, c).value
            if v is not None:
                sum_meta.append({"cell": f"{get_column_letter(c)}{r}", "v": v})

    # count employees in PC (rows with CN Name or EN Name)
    employees = []
    for row in range(PC_HR + 1, (pc.max_row or 0) + 1):
        cn = pc.cell(row, pc_map.get("CN Name", 2)).value
        en = pc.cell(row, pc_map.get("EN Name", 3)).value
        if cn or en:
            employees.append({"row": row, "cn": cn, "en": en})
        elif not any(pc.cell(row, c).value for c in range(1, (pc.max_column or 0) + 1)):
            continue

    # also rows without name but with data (summary row?)
    summary_rows = []
    for row in range(PC_HR + 1, (pc.max_row or 0) + 1):
        cn = pc.cell(row, pc_map.get("CN Name", 2)).value
        en = pc.cell(row, pc_map.get("EN Name", 3)).value
        if not cn and not en:
            has_data = any(
                isinstance(pc.cell(row, c).value, (int, float))
                for c in range(8, min(20, (pc.max_column or 0) + 1))
            )
            if has_data:
                summary_rows.append({"row": row, "preview": {h: pc.cell(row, col).value for h, col in list(pc_map.items())[:12] if pc.cell(row, col).value is not None}})

    result = {
        "mapping_row8": mapping,
        "mapping_row9": mapping2,
        "tw_info": tw_info,
        "ee_info": ee_info,
        "summary_meta": sum_meta,
        "employees": employees,
        "summary_rows": summary_rows,
        "pc_headers": list(pc_map.keys()),
        "out_headers": list(out_map.keys()),
        "out_only": sorted(set(out_map) - set(pc_map)),
        "pc_only": sorted(set(pc_map) - set(out_map)),
    }

    out_file = PROFILE_DIR / "mapping.json"
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)

    print("written", out_file)
    print("employees:", len(employees), employees)
    print("summary_rows:", summary_rows)
    print("\n=== PC→TW-L mapping (row8) ===")
    for m in mapping:
        if m["pc"]:
            print(f"  {m['out']!r} <- {m['pc']!r}")
    print("\nout_only headers:", result["out_only"])
    print("pc_only headers:", result["pc_only"])


if __name__ == "__main__":
    main()
