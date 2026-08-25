# -*- coding: utf-8 -*-
"""
台湾 Payroll calculation 源账单 → PN（引擎 tw_payroll_calc）

用法:
  python -m profiles.tw_payroll_calc.convert <原始供应商账单.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

原始账单: sheet「Payroll calculation」按表头名匹配（支持多员工）
默认母版: templates/taiwan/template.xlsx
Office 各供应商–客户差异见 convert_mapping / portal mapping_json，非本目录分客户。
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bill_convert.formula_copy import (
    copy_row_formulas as _copy_row_formulas_impl,
    fix_ee_row_tw_refs,
    fix_tw_row_tw_ee_refs,
    shift_row_formula,
)
from bill_convert.formula_layout import (
    apply_employee_formula_styles as _apply_employee_formula_styles_impl,
    resolve_formula_rows_layout,
    tw_l_row_for_data_row,
)
from bill_convert.formula_layout import _default_example_row as _default_example_row_for_mapping
from bill_convert.header_scan import find_header_row_by_markers
from bill_convert.headers import (
    build_header_cols,
    build_header_map,
    build_qualified_header_cols,
    list_qualified_header_cells,
    resolve_header_cols,
    resolve_target_col as _resolve_target_col,
)
from bill_convert.meta_period import parse_period, payroll_month_start, read_summary_meta as _read_summary_meta
from bill_convert.person import norm_person_name as _norm_person_name
from bill_convert.person import score_person_name_match as score_ee_name_match
from bill_convert.target_l_layout import (
    find_target_l_header_row,
    resolve_target_l_layout,
    resolve_target_l_sheet_name,
    target_l_auto_detect_layout,
)
from bill_convert.template_rows import clear_row_values, count_data_slots
from convert_mapping import find_sheet_name, resolve_convert_mapping
from fx_rate import fetch_usd_rates, get_tw_pn_fx_rate
from pn_meta import PnMeta, apply_pn_meta
from region_templates import get_region_template
from xlsx_convert_utils import clean_value, coerce_datetime_for_excel, is_date_column_header, norm
from xlsx_luckysheet_compat import apply_luckysheet_compat
from xlsx_postprocess import postprocess_converted_xlsx

DEFAULT_TEMPLATE = get_region_template("Taiwan")

PC_SHEET = "Payroll calculation"
SUMMARY_SHEET = "Summary"
TW_L_SHEET = "TW-L"
TW_SHEET = "TW"
TW_EE_SHEET = "TW EE"
PN_SHEET = "PN"

TW_L_HEADER_ROW = 7
TW_L_SUMMARY_ROW = 8
TW_L_DATA_START_ROW = 9
TW_L_SUMMARY_MIN_END_ROW = 14
TW_DATA_START_ROW = 9
TW_EE_DATA_START_ROW = 10
MAX_EMPLOYEES = 9
PAYROLL_DUP_MIN_COL = 14
# Pay Period / Start End：母版若误用金额格式会显示成 46,113.00，PN 的 MONTH/YEAR 仍可读序列，但页面很误导
_DATE_FMT = "yyyy/m/d"

# 源表有、但 TW-L 不直接写入的列（加班已汇总到「加班費」；Service Fee 由模板留空）
SKIP_SOURCE_HEADERS = frozenset({
    "Hr (1.34x)",
    "OT Payment (1.34x)",
    "Hr (1.67x)",
    "OT Payment (1.67x)",
    "Hr (1x)",
    "OT Payment (1x)",
    "Hr (2.67x)",
    "OT Payment (2.67x)",
    "Service Fee",
})

SICK_LEAVE_PAY_HEADER = "病假扣薪\nSick Leave\n(half pay)"
SICK_LEAVE_HOURS_HEADER = "病假時數 Sick Leave Hours"

# 列名对照仅来自 Office 保存的 columnRename；引擎不再内置默认别名。
SOURCE_TO_TARGET_HEADER: dict[str, str] = {}

NAME_HEADERS = ("CN Name", "EN Name")
PC_HEADER_KEYS = frozenset({"BU", "CN Name", "EN Name"})

_ACTIVE_MAPPING: dict[str, Any] | None = None


def _active_mapping() -> dict[str, Any]:
    if _ACTIVE_MAPPING is not None:
        return _ACTIVE_MAPPING
    return resolve_convert_mapping("tw_payroll_calc", None)


def _source_employee_spec() -> dict[str, Any]:
    spec = _active_mapping().get("sourceEmployeeSheet")
    return spec if isinstance(spec, dict) else {}


def _skip_source_headers() -> frozenset[str]:
    raw = _active_mapping().get("skipSourceHeaders")
    if raw is None:
        return frozenset()
    if not raw:
        return frozenset()
    return frozenset(norm(str(x)) for x in raw)


def _column_rename_map() -> dict[str, str]:
    extra = _active_mapping().get("columnRename") or {}
    if not isinstance(extra, dict):
        return {}
    return {
        norm(str(k)): str(v)
        for k, v in extra.items()
        if k is not None and v is not None and str(k).strip() and str(v).strip()
    }


def _explicit_rename_targets(rename: dict[str, str] | None = None) -> set[str]:
    """已被「列名对照」显式占用的目标列名（规范化）。"""
    m = rename if rename is not None else _column_rename_map()
    out: set[str] = set()
    for v in m.values():
        nv = norm(v)
        if not nv:
            continue
        out.add(nv)
        base = nv.split("#", 1)[0]
        out.add(base)
        out.add(base.rsplit("/", 1)[-1])
    return out


def map_source_header(source_header: str) -> str | None:
    """
    源列 → 目标列。
    - 有显式 columnRename：用对照结果
    - 否则同名自动匹配；但若目标名已被其它源列显式对照占用，则跳过（避免 Total 盖掉 Hours→Total）
    """
    h = norm(source_header)
    if not h or h in _skip_source_headers():
        return None
    rename = _column_rename_map()
    if h in rename:
        return rename[h]
    claimed = _explicit_rename_targets(rename)
    if h in claimed:
        return None
    base = h.split("#", 1)[0]
    child = base.rsplit("/", 1)[-1]
    if base in claimed or child in claimed:
        return None
    return h


def _target_l_auto_detect_layout() -> bool:
    return target_l_auto_detect_layout(_active_mapping())


def find_tw_l_header_row(
    ws: Worksheet,
    max_scan: int | None = None,
    marker_keys: frozenset[str] | None = None,
    sheet_label: str | None = None,
) -> int:
    mapping = _active_mapping()
    spec = mapping.get("targetL") if isinstance(mapping.get("targetL"), dict) else {}
    if marker_keys is not None:
        scan = max_scan if max_scan is not None else int(spec.get("headerScanMaxRow") or 15)
        label = sheet_label or str(spec.get("sheet") or TW_L_SHEET)
        return find_header_row_by_markers(
            ws, marker_keys=marker_keys, max_scan=scan, sheet_label=label
        )
    return find_target_l_header_row(
        ws, mapping, sheet_label=sheet_label, default_sheet_name=TW_L_SHEET
    )


def resolve_tw_l_layout(ws: Worksheet, *, sheet_label: str | None = None) -> dict[str, int]:
    return resolve_target_l_layout(
        ws,
        _active_mapping(),
        sheet_label=sheet_label,
        default_sheet_name=TW_L_SHEET,
        fallback_header_row=TW_L_HEADER_ROW,
    )


def resolve_tw_l_sheet_name(sheet_names: list[str]) -> str:
    return resolve_target_l_sheet_name(
        sheet_names, _active_mapping(), default_sheet_name=TW_L_SHEET
    )


def find_pc_header_row(
    ws: Worksheet,
    max_scan: int | None = None,
    marker_keys: frozenset[str] | None = None,
    sheet_label: str | None = None,
) -> int:
    spec = _source_employee_spec()
    scan = max_scan if max_scan is not None else int(spec.get("headerScanMaxRow") or 15)
    if marker_keys is not None:
        keys = marker_keys
    elif isinstance(spec.get("headerMarkerKeys"), list):
        keys = frozenset(norm(str(x)) for x in spec["headerMarkerKeys"])
    else:
        keys = PC_HEADER_KEYS
    label = sheet_label or str(spec.get("sheet") or PC_SHEET)
    return find_header_row_by_markers(
        ws, marker_keys=keys, max_scan=scan, sheet_label=label
    )


def resolve_target_col(cols: list[int]) -> int | None:
    dup_min = int(_source_employee_spec().get("payrollDupMinCol") or PAYROLL_DUP_MIN_COL)
    return _resolve_target_col(cols, dup_min_col=dup_min)


def format_payroll_date(value: Any) -> datetime | None:
    """账期/日期写入模板：必须是 datetime，禁止返回字符串。"""
    return coerce_datetime_for_excel(value)


def read_summary_meta(wb, meta_spec: dict[str, Any] | None = None) -> dict[str, Any]:
    return _read_summary_meta(wb, meta_spec, default_sheet=SUMMARY_SHEET)


def read_pc_employees(ws: Worksheet, header_row: int) -> list[dict[str, Any]]:
    # 源表用资格化表头（父子/同名消歧）；姓名标志列仍按子列名匹配
    qualified = list_qualified_header_cells(ws, header_row)
    source_headers = {str(h["key"]): int(h["col"]) for h in qualified}
    child_to_keys: dict[str, list[str]] = {}
    for h in qualified:
        child_to_keys.setdefault(str(h["child"]), []).append(str(h["key"]))
    spec = _source_employee_spec()
    sheet_label = str(spec.get("sheet") or PC_SHEET)
    name_keys = spec.get("nameHeaders")
    if isinstance(name_keys, list) and name_keys:
        name_header_list = tuple(str(x) for x in name_keys)
    else:
        name_header_list = NAME_HEADERS
    name_cols: list[int] = []
    for h in name_header_list:
        nk = norm(h)
        if nk in source_headers:
            name_cols.append(source_headers[nk])
            continue
        for qk in child_to_keys.get(nk, []):
            if qk in source_headers:
                name_cols.append(source_headers[qk])
                break
    if not name_cols:
        raise ValueError(f"「{sheet_label}」表头行须包含 {' 或 '.join(name_header_list)}")

    employees: list[dict[str, Any]] = []
    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        has_name = any(clean_value(ws.cell(row, col).value) for col in name_cols)
        if not has_name:
            continue

        record: dict[str, Any] = {}
        rename = _column_rename_map()
        # 先写显式对照，再写同名自动匹配，避免后者覆盖前者
        ordered = list(source_headers.items())
        explicit_first = [(s, c) for s, c in ordered if norm(s) in rename]
        auto_rest = [(s, c) for s, c in ordered if norm(s) not in rename]
        for src_hdr, col in explicit_first + auto_rest:
            target_hdr = map_source_header(src_hdr)
            if target_hdr is None:
                continue
            val = clean_value(ws.cell(row, col).value)
            if val is not None:
                if is_date_column_header(target_hdr):
                    dt = coerce_datetime_for_excel(val)
                    if dt is not None:
                        val = dt
                # 显式对照优先：自动匹配不覆盖已有键
                if target_hdr in record and norm(src_hdr) not in rename:
                    continue
                record[target_hdr] = val
        employees.append(record)

    if not employees:
        raise ValueError(f"「{sheet_label}」中未找到有效员工行（需有 {' 或 '.join(name_header_list)}）")
    if len(employees) > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {len(employees)} 超过模板上限 {MAX_EMPLOYEES}")
    return employees


def apply_sick_leave_formula(ws: Worksheet, row: int, target_cols: dict[str, list[int]]) -> None:
    """有病假时数时写入公式 =-H{row}*T{row}/2，与母版一致"""
    pay_cols = resolve_header_cols(target_cols, SICK_LEAVE_PAY_HEADER)
    if not pay_cols:
        return
    hours_cols = resolve_header_cols(target_cols, SICK_LEAVE_HOURS_HEADER)
    hours_col = resolve_target_col(hours_cols) if hours_cols else None
    hours = ws.cell(row, hours_col).value if hours_col else None
    try:
        has_hours = hours is not None and float(hours) > 0
    except (TypeError, ValueError):
        has_hours = False

    col = resolve_target_col(pay_cols)
    if col is None:
        return
    if has_hours and hours_col:
        ws.cell(row, col).value = f"=-H{row}*{get_column_letter(hours_col)}{row}/2"
    elif not has_hours:
        ws.cell(row, col).value = 0


def clear_tw_l_data(ws: Worksheet, employee_count: int, *, data_start_row: int) -> None:
    max_row = max(ws.max_row or data_start_row, data_start_row + MAX_EMPLOYEES)
    max_col = ws.max_column or 1
    for row in range(data_start_row, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def write_tw_l(ws: Worksheet, employees: list[dict[str, Any]], meta: dict[str, Any]) -> None:
    layout = resolve_tw_l_layout(ws)
    header_row = layout["header_row"]
    data_start_row = layout["data_start_row"]
    summary_row = layout["summary_row"]
    # 与映射下拉同一套资格化表头，避免「父/子」对照写不进母版
    target_cols = build_qualified_header_cols(ws, header_row)
    if not target_cols:
        target_cols = build_header_cols(ws, header_row)
    if not target_cols:
        raise ValueError(f"「{TW_L_SHEET}」第 {header_row} 行表头为空")

    if meta.get("payroll_month_start") is not None:
        cell = ws.cell(3, 3)
        cell.value = meta["payroll_month_start"]
        cell.number_format = _DATE_FMT
    if meta.get("period_from") is not None:
        cell = ws.cell(4, 3)
        cell.value = meta["period_from"]
        cell.number_format = _DATE_FMT
    if meta.get("period_to") is not None:
        cell = ws.cell(5, 3)
        cell.value = meta["period_to"]
        cell.number_format = _DATE_FMT

    clear_tw_l_data(ws, len(employees), data_start_row=data_start_row)
    for idx, emp in enumerate(employees):
        row = data_start_row + idx
        for hdr, val in emp.items():
            cols = resolve_header_cols(target_cols, str(hdr))
            col = resolve_target_col(cols) if cols else None
            if col is not None:
                out_val = val
                if is_date_column_header(hdr):
                    dt = coerce_datetime_for_excel(val)
                    if dt is not None:
                        out_val = dt
                cell = ws.cell(row, col)
                cell.value = out_val
                if is_date_column_header(hdr) and isinstance(out_val, datetime):
                    cell.number_format = _DATE_FMT
        apply_sick_leave_formula(ws, row, target_cols)

    update_tw_l_summary_formulas(
        ws,
        len(employees),
        data_start_row=data_start_row,
        summary_row=summary_row,
    )


def update_tw_l_summary_formulas(
    ws: Worksheet,
    employee_count: int,
    *,
    data_start_row: int,
    summary_row: int,
) -> None:
    end_row = max(
        data_start_row + max(employee_count, 1) - 1,
        TW_L_SUMMARY_MIN_END_ROW,
    )
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        cell = ws.cell(summary_row, col)
        if cell.data_type != "f" or not isinstance(cell.value, str):
            continue
        formula = cell.value
        if not re.search(r"SUM\([A-Z]+\d+:[A-Z]+\d+\)", formula, re.I):
            continue
        cell.value = re.sub(
            r"(SUM\([A-Z]+)(\d+)(:[A-Z]+)(\d+)(\))",
            lambda m: f"{m.group(1)}{data_start_row}{m.group(3)}{end_row}{m.group(5)}",
            formula,
            flags=re.I,
        )


def shift_tw_formula(
    formula: str,
    from_row: int,
    to_row: int,
    tw_l_from: int,
    tw_l_to: int,
) -> str:
    return shift_row_formula(
        formula,
        from_row,
        to_row,
        target_l_from=tw_l_from,
        target_l_to=tw_l_to,
        target_l_sheet=TW_L_SHEET,
    )


def clear_employee_row(ws: Worksheet, row: int) -> None:
    clear_row_values(ws, row)


def count_template_employee_slots(ws: Worksheet, data_start_row: int, marker_col: int = 2) -> int:
    return count_data_slots(ws, data_start_row, marker_col=marker_col, max_slots=MAX_EMPLOYEES)


# TW!F = Business Tax = TW-L!Total(BO) * 5%
_TW_BUSINESS_TAX_COL = 6  # F
_TW_L_TOTAL_COL = 67  # BO


def ensure_tw_period_date_formats(wb) -> None:
    """TW!B2/C2 引用 TW-L 账期；强制日期格式，避免网页显示成 46,113.00。"""
    try:
        tw_l = wb[TW_L_SHEET]
        for row, col in ((3, 3), (4, 3), (5, 3)):
            tw_l.cell(row, col).number_format = _DATE_FMT
    except KeyError:
        pass
    try:
        tw = wb[TW_SHEET]
        for row, col in ((2, 2), (2, 3)):
            tw.cell(row, col).number_format = _DATE_FMT
    except KeyError:
        pass


def apply_tw_business_tax(
    ws_tw: Worksheet,
    employee_count: int,
    *,
    tw_data_start_row: int,
    tw_l_data_start_row: int,
) -> None:
    """每人 Business Tax：TW!F = ROUND('TW-L'!BO * 5%, 0) 取整。LuckySheet 用 (5/100)。"""
    n = max(int(employee_count), 0)
    for i in range(n):
        tw_row = tw_data_start_row + i
        tw_l_row = tw_l_data_start_row + i
        cell = ws_tw.cell(tw_row, _TW_BUSINESS_TAX_COL)
        cell.value = f"=ROUND('TW-L'!BO{tw_l_row}*(5/100),0)"
        cell.number_format = "#,##0"


def match_ee_code(
    excel_names: list[str],
    directory: list[dict[str, Any]],
) -> tuple[str | None, str | None]:
    """
    从员工目录匹配 employee_code。
    返回 (code, warning)；匹配不到/歧义时 code=None 并带 warning。
    """
    names = [n for n in (_norm_person_name(x) for x in excel_names) if n]
    if not names:
        return None, "源表员工姓名为空，无法匹配 EE Code"
    if not directory:
        return None, "未提供客户员工目录，无法匹配 EE Code"

    best_score = 0
    best: list[dict[str, Any]] = []
    for row in directory:
        code = norm(row.get("employee_code") or row.get("employeeCode"))
        if not code:
            continue
        score = 0
        for field in (
            row.get("employee_name_en") or row.get("employeeNameEn"),
            row.get("employee_name") or row.get("employeeName"),
        ):
            for en in excel_names:
                score = max(score, score_ee_name_match(en, str(field or "")))
        if score <= 0:
            continue
        if score > best_score:
            best_score = score
            best = [row]
        elif score == best_score:
            best.append(row)

    if not best:
        label = " / ".join(n for n in excel_names if n)
        return None, f"未匹配到 EE Code：{label}"

    # 同分多人：优先更短英文名（更具体的全名通常更长，但同分时取最短减少歧义误伤）
    if len(best) > 1:
        # 若编码相同可合并
        codes = {norm(r.get("employee_code") or r.get("employeeCode")) for r in best}
        if len(codes) == 1:
            return next(iter(codes)), None
        label = " / ".join(n for n in excel_names if n)
        return None, f"EE Code 匹配歧义（{len(best)} 人同分）：{label}"

    return norm(best[0].get("employee_code") or best[0].get("employeeCode")), None


def apply_tw_ee_codes(
    ws_ee: Worksheet,
    employees: list[dict[str, Any]],
    directory: list[dict[str, Any]] | None,
    *,
    tw_ee_data_start_row: int = TW_EE_DATA_START_ROW,
) -> list[str]:
    """写入 TW EE!D（EE Code）；匹配失败留空并返回 warnings。"""
    warnings: list[str] = []
    dir_list = list(directory or [])
    for i, emp in enumerate(employees):
        row = tw_ee_data_start_row + i
        excel_names = [
            emp.get("EN Name"),
            emp.get("CN Name"),
        ]
        code, warn = match_ee_code(
            [str(x) for x in excel_names if x],
            dir_list,
        )
        ws_ee.cell(row, 4).value = code  # 匹配不到显式清空，避免母版残留
        if warn:
            warnings.append(f"第{i + 1}人：{warn}")
    return warnings


_PN_EOR_ROW = 15
_PN_LABOR_START_ROW = 16
# 母版默认 3 人：Labor 16-18 / Expense 19-21 / 空行 22 / Service 23 / Mgmt 24 / Total 28 / FX B31
_PN_NTD_FMT = "#,##0.00"
_PN_USD_FMT = '$#,##0.00'


def _pn_layout(employee_count: int) -> dict[str, int]:
    """按人数推算 PN 明细区行号（与母版 3 人布局同构）。"""
    n = max(int(employee_count), 0)
    labor_start = _PN_LABOR_START_ROW
    expense_start = labor_start + n
    blank_row = expense_start + n
    service_row = blank_row + 1
    mgmt_row = service_row + 1
    return {
        "n": n,
        "eor_row": _PN_EOR_ROW,
        "labor_start": labor_start,
        "expense_start": expense_start,
        "blank_row": blank_row,
        "service_row": service_row,
        "mgmt_row": mgmt_row,
        "sum_end": blank_row,  # E15=SUM(E16:E{blank})，含空行
    }


def _count_pn_labor_slots(ws: Worksheet) -> int:
    n = 0
    for row in range(_PN_LABOR_START_ROW, _PN_LABOR_START_ROW + MAX_EMPLOYEES + 2):
        v = ws.cell(row, 1).value
        if isinstance(v, str) and "Labor cost" in v:
            n += 1
            continue
        break
    return max(n, 1)


def _find_pn_row_by_label(ws: Worksheet, keyword: str, col: int = 1) -> int | None:
    key = keyword.lower()
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and key in v.lower():
            return row
    return None


def _find_pn_fx_row(ws: Worksheet) -> int:
    r = _find_pn_row_by_label(ws, "FX rate")
    return r if r else 31


def _copy_pn_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    max_col = max(ws.max_column or 6, 6)
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _ensure_merge_a_c(ws: Worksheet, row: int) -> None:
    """母版 Labor/Expense 描述列为 A:C 合并。"""
    rng = f"A{row}:C{row}"
    for m in list(ws.merged_cells.ranges):
        if m.min_row <= row <= m.max_row and m.min_col <= 1 and m.max_col >= 3:
            if m.min_row == row and m.max_row == row and m.min_col == 1 and m.max_col == 3:
                return
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    try:
        ws.merge_cells(rng)
    except ValueError:
        pass


def _collect_merges_from(ws: Worksheet, start_row: int) -> list[tuple[int, int, int, int]]:
    """收集 min_row>=start_row 的合并区，并先解除，避免与 insert/delete 错位。"""
    kept: list[tuple[int, int, int, int]] = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row < start_row:
            continue
        kept.append((m.min_row, m.min_col, m.max_row, m.max_col))
        try:
            ws.unmerge_cells(str(m))
        except ValueError:
            pass
    return kept


def _restore_merges(
    ws: Worksheet,
    merges: list[tuple[int, int, int, int]],
    row_shift: int = 0,
) -> None:
    for min_r, min_c, max_r, max_c in merges:
        nr1, nr2 = min_r + row_shift, max_r + row_shift
        if nr1 < 1 or nr2 < nr1:
            continue
        try:
            ws.merge_cells(
                start_row=nr1,
                start_column=min_c,
                end_row=nr2,
                end_column=max_c,
            )
        except ValueError:
            pass


def _shift_row_heights(ws: Worksheet, start_row: int, amount: int) -> None:
    """openpyxl insert/delete 时常不带动 row_dimensions，需手动平移行高。"""
    if amount == 0:
        return
    captured: dict[int, float] = {}
    max_r = max(ws.max_row or start_row, start_row)
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            captured[r] = h
        max_r = max(max_r, r)

    for r in range(start_row, max_r + abs(amount) + 2):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None

    if amount > 0:
        for r in sorted(captured.keys(), reverse=True):
            ws.row_dimensions[r + amount].height = captured[r]
    else:
        for r in sorted(captured.keys()):
            new_r = r + amount
            if new_r >= start_row:
                ws.row_dimensions[new_r].height = captured[r]


def _capture_row_heights_from(ws: Worksheet, start_row: int) -> dict[int, float]:
    out: dict[int, float] = {}
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            out[r] = h
    return out


def _apply_row_heights(ws: Worksheet, heights: dict[int, float], start_row: int) -> None:
    max_r = max([start_row, *heights.keys()], default=start_row)
    for r in range(start_row, max_r + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r, h in heights.items():
        if r >= start_row:
            ws.row_dimensions[r].height = h


def _pn_insert_rows(ws: Worksheet, idx: int, amount: int, fill_style_row: int | None = None) -> None:
    """插入行，并让下方合并区/行高一起下移；新行复制样板行样式。"""
    if amount <= 0:
        return
    merges = _collect_merges_from(ws, idx)
    _shift_row_heights(ws, idx, amount)
    ws.insert_rows(idx, amount)
    _restore_merges(ws, merges, row_shift=amount)
    if fill_style_row is not None:
        style_row = fill_style_row + amount if fill_style_row >= idx else fill_style_row
        for r in range(idx, idx + amount):
            _copy_pn_row_style(ws, style_row, r)
            _ensure_merge_a_c(ws, r)


def _pn_delete_rows(ws: Worksheet, idx: int, amount: int) -> None:
    """删除行，并让下方合并区/行高一起上移。"""
    if amount <= 0:
        return
    for m in list(ws.merged_cells.ranges):
        if m.max_row < idx or m.min_row >= idx + amount:
            continue
        # 与删除区间相交的合并先拆掉
        try:
            ws.unmerge_cells(str(m))
        except ValueError:
            pass
    merges = _collect_merges_from(ws, idx + amount)
    heights = _capture_row_heights_from(ws, idx)
    ws.delete_rows(idx, amount)
    # 删除区间内行高丢弃；下方行高上移 amount
    new_heights: dict[int, float] = {}
    for r, h in heights.items():
        if r < idx:
            continue
        if r < idx + amount:
            continue
        new_heights[r - amount] = h
    _apply_row_heights(ws, new_heights, idx)
    _restore_merges(ws, merges, row_shift=-amount)


def fit_pn_employees(ws: Worksheet, employee_count: int, *, tw_data_start_row: int = TW_DATA_START_ROW) -> dict[str, int]:
    """
    按实际人数扩/缩 PN 的 Labor + Expense 明细行，并重写相关公式。
    插入/删除时同步平移下方合并单元格与行高，避免 Service/合计/汇率区样式错位。
    """
    n = max(int(employee_count), 1)
    if n > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {n} 超过模板上限 {MAX_EMPLOYEES}")

    old_n = _count_pn_labor_slots(ws)
    labor_start = _PN_LABOR_START_ROW
    expense_start = labor_start + old_n
    delta = n - old_n

    if delta > 0:
        # 在首个 Expense 前插入 Labor 行（样式跟上一行 Labor）
        _pn_insert_rows(ws, expense_start, delta, fill_style_row=labor_start)
        expense_start += delta
        # 在 Expense 块末尾后插入 Expense 行
        _pn_insert_rows(ws, expense_start + old_n, delta, fill_style_row=expense_start)
    elif delta < 0:
        # 先删多余 Expense，再删多余 Labor
        _pn_delete_rows(ws, expense_start + n, -delta)
        _pn_delete_rows(ws, labor_start + n, -delta)

    layout = _pn_layout(n)
    expense_start = layout["expense_start"]
    blank_row = layout["blank_row"]
    service_row = layout["service_row"]
    mgmt_row = layout["mgmt_row"]
    eor_row = layout["eor_row"]
    sum_end = layout["sum_end"]

    # 空行：清内容但保留已随行移动的样式/行高
    for m in list(ws.merged_cells.ranges):
        if m.min_row == blank_row and m.max_row == blank_row:
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    for col in range(1, 7):
        cell = ws.cell(blank_row, col)
        if type(cell).__name__ == "MergedCell":
            continue
        cell.value = None
    _ensure_merge_a_c(ws, blank_row)

    fx_row = _find_pn_fx_row(ws)

    for i in range(n):
        tw_row = tw_data_start_row + i
        labor_row = labor_start + i
        expense_row = expense_start + i

        if i > 0:
            _copy_pn_row_style(ws, labor_start, labor_row)
            _copy_pn_row_style(ws, expense_start, expense_row)

        _ensure_merge_a_c(ws, labor_row)
        _ensure_merge_a_c(ws, expense_row)

        ws.cell(labor_row, 1).value = (
            f'="- Labor cost for "&TW!B{tw_row}&"  -  "&MONTH(TW!$B$2)&"-"&YEAR(TW!$B$2)'
        )
        e_labor = ws.cell(labor_row, 5)
        e_labor.value = f"=TW!J{tw_row}+TW!Z{tw_row}"
        e_labor.number_format = _PN_NTD_FMT
        f_labor = ws.cell(labor_row, 6)
        f_labor.value = f"=E{labor_row}/$B${fx_row}"
        f_labor.number_format = _PN_USD_FMT

        ws.cell(expense_row, 1).value = f'="- Expense claim for "&TW!B{tw_row}'
        e_exp = ws.cell(expense_row, 5)
        e_exp.value = f"=TW!AB{tw_row}"
        e_exp.number_format = _PN_NTD_FMT
        f_exp = ws.cell(expense_row, 6)
        f_exp.value = f"=E{expense_row}/$B${fx_row}"
        f_exp.number_format = _PN_USD_FMT

    # EOR 汇总（简单数字格式，避免 LuckySheet 会计格式空白）
    e_eor = ws.cell(eor_row, 5)
    e_eor.value = f"=SUM(E{labor_start}:E{sum_end})"
    e_eor.number_format = _PN_NTD_FMT
    f_eor = ws.cell(eor_row, 6)
    f_eor.value = f"=SUM(F{labor_start}:F{sum_end})"
    f_eor.number_format = _PN_USD_FMT

    svc = _find_pn_row_by_label(ws, "Service Fee") or service_row
    mgmt = _find_pn_row_by_label(ws, "Management Fee") or mgmt_row
    e_svc = ws.cell(svc, 5)
    e_svc.value = f"=SUM(E{mgmt}:E{mgmt + 1})"
    e_svc.number_format = _PN_NTD_FMT
    f_svc = ws.cell(svc, 6)
    f_svc.value = f"=SUM(F{mgmt}:F{mgmt + 1})"
    f_svc.number_format = _PN_USD_FMT
    if isinstance(ws.cell(mgmt, 1).value, str) or ws.cell(mgmt, 1).value is None:
        ws.cell(mgmt, 1).value = f'="- Management Fee - "&MONTH(TW!B2)&"-"&YEAR(TW!B2)'
    e_mgmt = ws.cell(mgmt, 5)
    e_mgmt.value = "=TW!G6"
    e_mgmt.number_format = _PN_NTD_FMT
    f_mgmt = ws.cell(mgmt, 6)
    f_mgmt.value = f"=E{mgmt}/$B${fx_row}"
    f_mgmt.number_format = _PN_USD_FMT
    _ensure_merge_a_c(ws, svc)
    _ensure_merge_a_c(ws, mgmt)

    total_row = None
    for row in range(eor_row + 1, (ws.max_row or 0) + 1):
        ev = ws.cell(row, 5).value
        if isinstance(ev, str) and ev.startswith("=") and "E15" in ev.replace("$", ""):
            total_row = row
            break
        av = ws.cell(row, 1).value
        if isinstance(av, str) and "EOR/PEO Service Cost" in av:
            total_row = row
            break
    if total_row is None:
        total_row = svc + 5

    e_total = ws.cell(total_row, 5)
    e_total.value = f"=E{eor_row}+E{svc}"
    e_total.number_format = _PN_NTD_FMT
    f_total = ws.cell(total_row, 6)
    f_total.value = f"=F{eor_row}+F{svc}"
    f_total.number_format = _PN_USD_FMT
    # 合计行母版是 A:D 合并
    for m in list(ws.merged_cells.ranges):
        if m.min_row == total_row and m.max_row == total_row and m.min_col == 1:
            break
    else:
        try:
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        except ValueError:
            pass

    for row in range(labor_start, sum_end + 1):
        for col, fmt in ((5, _PN_NTD_FMT), (6, _PN_USD_FMT)):
            cell = ws.cell(row, col)
            if cell.value is not None:
                cell.number_format = fmt

    # 插删后重找 FX；openpyxl 不会改写公式文本里的绝对行号，需整段重写结算区
    fx_row = _find_pn_fx_row(ws)
    _rewrite_pn_settlement_block(ws, fx_row=fx_row, total_row=total_row)
    for row in range(labor_start, fx_row + 8):
        fcell = ws.cell(row, 6)
        if isinstance(fcell.value, str) and "$B$" in fcell.value:
            fcell.value = re.sub(r"\$B\$\d+", f"$B${fx_row}", fcell.value)
        if isinstance(fcell.value, str) and "PN!B" in fcell.value.replace("$", ""):
            fcell.value = re.sub(r"PN!\$?B\$?\d+", f"PN!B{fx_row}", fcell.value)
    layout["service_row"] = svc
    layout["mgmt_row"] = mgmt
    layout["total_row"] = total_row
    layout["fx_row"] = fx_row
    return layout


def _rewrite_pn_settlement_block(ws: Worksheet, *, fx_row: int, total_row: int) -> None:
    """
    母版 FX 下方结算块相对布局（以 fx_row 为锚）：
      F{fx}     = F{total}                         EOR USD
      F{fx+1..4}= TW!E22..E25                      Other Fee 透传
      F{fx+5}   = SUM(F{fx}:F{fx+4})               Sub Total
      F{fx+6}   = ROUNDUP(TW!F6/PN!B{fx},2)+0.01   Tax
      F{fx+7}   = SUM(F{fx+5}:F{fx+6})             Total
    插删行后公式文本不会自动更新，必须按新行号重写。
    """
    # EOR USD 引用合计行
    f_eor = ws.cell(fx_row, 6)
    f_eor.value = f"=F{total_row}"
    f_eor.number_format = _PN_USD_FMT

    # Other Fee 透传（TW 行号固定，与人数无关）
    for i, tw_e in enumerate((22, 23, 24, 25), start=1):
        cell = ws.cell(fx_row + i, 6)
        cell.value = f"=TW!E{tw_e}"

    sub_row = fx_row + 5
    tax_row = fx_row + 6
    grand_row = fx_row + 7

    f_sub = ws.cell(sub_row, 6)
    f_sub.value = f"=SUM(F{fx_row}:F{fx_row + 4})"
    f_sub.number_format = _PN_USD_FMT

    f_tax = ws.cell(tax_row, 6)
    f_tax.value = f"=ROUNDUP(TW!F6/PN!B{fx_row},2)+0.01"
    f_tax.number_format = _PN_USD_FMT

    f_grand = ws.cell(grand_row, 6)
    f_grand.value = f"=SUM(F{sub_row}:F{tax_row})"
    f_grand.number_format = _PN_USD_FMT



def copy_row_formulas(
    ws: Worksheet,
    from_row: int,
    to_row: int,
    tw_l_from: int,
    tw_l_to: int,
) -> None:
    _copy_row_formulas_impl(
        ws, from_row, to_row, tw_l_from, tw_l_to, target_l_sheet=TW_L_SHEET
    )


def _default_example_row(sheet_key: str, data_start_fallback: int) -> int:
    return _default_example_row_for_mapping(
        _active_mapping(), sheet_key, data_start_fallback
    )


def resolve_tw_formula_rows_layout(wb, tw_l_data_start: int) -> dict[str, int]:
    return resolve_formula_rows_layout(
        wb,
        _active_mapping(),
        tw_l_data_start,
        tw_sheet=TW_SHEET,
        tw_ee_sheet=TW_EE_SHEET,
        fallback_tw_data_start=TW_DATA_START_ROW,
        fallback_tw_ee_data_start=TW_EE_DATA_START_ROW,
        target_l_sheet=TW_L_SHEET,
    )


def apply_employee_formula_styles(
    wb,
    employees: list[dict[str, Any]],
    *,
    formula_rows: dict[str, int],
    employee_directory: list[dict[str, Any]] | None = None,
) -> None:
    _apply_employee_formula_styles_impl(
        wb,
        employees,
        _active_mapping(),
        formula_rows=formula_rows,
        employee_directory=employee_directory,
        tw_sheet=TW_SHEET,
        tw_ee_sheet=TW_EE_SHEET,
        target_l_sheet=TW_L_SHEET,
    )


def fit_formula_sheets(
    wb,
    employee_count: int,
    client_id_prefix: str,
    *,
    formula_rows: dict[str, int],
    clear_excess: bool = True,
) -> None:
    """母版已预置 N 人公式行：人数>N 时复制扩展；人数≤N 时可清除多余行。

    注意：若随后还要从母版「示例行」（如第 10 行第二种公式）复制到实际员工行，
    须先 apply_employee_formula_styles，再 clear_excess——否则人数不足时会先清掉示例行。
    """
    tw_l_data_start_row = formula_rows["tw_l_data_start"]
    tw_data_start = formula_rows["tw_data_start"]
    tw_ee_data_start = formula_rows["tw_ee_data_start"]
    tw = wb[TW_SHEET]
    ee = wb[TW_EE_SHEET]
    tw_slots = count_template_employee_slots(tw, tw_data_start, marker_col=2)
    ee_slots = count_template_employee_slots(ee, tw_ee_data_start, marker_col=5)
    tw_tpl_src = _default_example_row("TW", tw_data_start)
    ee_tpl_src = _default_example_row("TW EE", tw_ee_data_start)
    src_tw_l_tpl = tw_l_row_for_data_row(
        tw_tpl_src, data_start=tw_data_start, target_l_data_start=tw_l_data_start_row
    )
    src_ee_l_tpl = tw_l_row_for_data_row(
        ee_tpl_src, data_start=tw_ee_data_start, target_l_data_start=tw_l_data_start_row
    )

    if clear_excess:
        for i in range(employee_count, tw_slots):
            clear_employee_row(tw, tw_data_start + i)

        for i in range(employee_count, ee_slots):
            clear_employee_row(ee, tw_ee_data_start + i)

    if employee_count > tw_slots:
        src_row = tw_tpl_src
        src_tw_l = src_tw_l_tpl
        for i in range(tw_slots, employee_count):
            dst_row = tw_data_start + i
            dst_tw_l = tw_l_data_start_row + i
            ee_row = tw_ee_data_start + i
            copy_row_formulas(tw, src_row, dst_row, src_tw_l, dst_tw_l)
            fix_tw_row_tw_ee_refs(tw, dst_row, ee_row)

    if employee_count > ee_slots:
        src_row = ee_tpl_src
        src_tw_l = src_ee_l_tpl
        for i in range(ee_slots, employee_count):
            dst_row = tw_ee_data_start + i
            dst_tw_l = tw_l_data_start_row + i
            tw_row = tw_data_start + i
            copy_row_formulas(ee, src_row, dst_row, src_tw_l, dst_tw_l)
            ee.cell(dst_row, 4).value = None
            fix_ee_row_tw_refs(ee, dst_row, tw_row)


def clear_excess_formula_rows(
    wb,
    employee_count: int,
    *,
    formula_rows: dict[str, int],
) -> None:
    """清除 TW / TW EE 上超出本次员工人数的预置公式行。"""
    tw_data_start = formula_rows["tw_data_start"]
    tw_ee_data_start = formula_rows["tw_ee_data_start"]
    tw = wb[TW_SHEET]
    ee = wb[TW_EE_SHEET]
    tw_slots = count_template_employee_slots(tw, tw_data_start, marker_col=2)
    ee_slots = count_template_employee_slots(ee, tw_ee_data_start, marker_col=5)
    for i in range(employee_count, tw_slots):
        clear_employee_row(tw, tw_data_start + i)
    for i in range(employee_count, ee_slots):
        clear_employee_row(ee, tw_ee_data_start + i)


def parse_source(source_path: Path) -> dict[str, Any]:
    mapping = _active_mapping()
    wb = load_workbook(source_path, data_only=True, read_only=True)
    src_spec = mapping.get("sourceEmployeeSheet") or {}
    pc_name = find_sheet_name(list(wb.sheetnames), src_spec if isinstance(src_spec, dict) else None)
    if not pc_name:
        names = wb.sheetnames
        wb.close()
        want = (src_spec.get("sheet") if isinstance(src_spec, dict) else None) or PC_SHEET
        raise ValueError(f"未找到 sheet「{want}」，现有: {names}")

    ws = wb[pc_name]
    header_row = find_pc_header_row(ws, sheet_label=pc_name)
    employees = read_pc_employees(ws, header_row)
    meta = read_summary_meta(wb, mapping.get("sourceMetaSheet"))
    wb.close()
    return {"employees": employees, "meta": meta, "pc_header_row": header_row, "pc_sheet": pc_name}


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    global _ACTIVE_MAPPING
    _ACTIVE_MAPPING = resolve_convert_mapping("tw_payroll_calc", convert_mapping)
    try:
        return _convert_impl(
            source_path,
            output_path,
            template_path,
            pn_meta=pn_meta,
            registry_dir=registry_dir,
            employee_directory=employee_directory,
        )
    finally:
        _ACTIVE_MAPPING = None


def _convert_impl(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if not template_path.is_file():
        raise FileNotFoundError(f"母版不存在: {template_path}")
    if not source_path.is_file():
        raise FileNotFoundError(f"原始账单不存在: {source_path}")

    parsed = parse_source(source_path)
    employees = parsed["employees"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, rich_text=True)
    tw_l_name = resolve_tw_l_sheet_name(list(wb.sheetnames))
    for name in (TW_SHEET, TW_EE_SHEET, PN_SHEET):
        if name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"母版缺少 sheet「{name}」，现有: {wb.sheetnames}")

    tw_l_ws = wb[tw_l_name]
    tw_l_layout = resolve_tw_l_layout(tw_l_ws, sheet_label=tw_l_name)
    tw_l_data_start = tw_l_layout["data_start_row"]
    formula_rows = resolve_tw_formula_rows_layout(wb, tw_l_data_start)

    applied_pn: PnMeta | None = None
    if pn_meta is not None:
        applied_pn = apply_pn_meta(
            wb,
            pn_meta,
            registry_dir=registry_dir or output_path.parent,
            reserve_invoice_number=True,
        )

    write_tw_l(tw_l_ws, employees, parsed["meta"])
    client_prefix = (
        applied_pn.customer_id
        if applied_pn
        else norm(wb[PN_SHEET]["B9"].value) or "CUS1516"
    )
    # 先扩展（不清多余行）→ 按员工配对从示例行复制公式 → 再清多余行。
    # 若先清，人数 < 母版槽位时会把「第 10 行第二种公式」等示例行清掉，导致配对无法生效。
    fit_formula_sheets(
        wb, len(employees), client_prefix, formula_rows=formula_rows, clear_excess=False
    )
    apply_employee_formula_styles(
        wb, employees, formula_rows=formula_rows, employee_directory=employee_directory
    )
    clear_excess_formula_rows(wb, len(employees), formula_rows=formula_rows)
    apply_tw_business_tax(
        wb[TW_SHEET],
        len(employees),
        tw_data_start_row=formula_rows["tw_data_start"],
        tw_l_data_start_row=formula_rows["tw_l_data_start"],
    )
    ee_warnings = apply_tw_ee_codes(
        wb[TW_EE_SHEET],
        employees,
        employee_directory,
        tw_ee_data_start_row=formula_rows["tw_ee_data_start"],
    )
    pn_layout = fit_pn_employees(wb[PN_SHEET], len(employees), tw_data_start_row=formula_rows["tw_data_start"])

    summary_fx = parsed["meta"].get("exchange_rate")
    if summary_fx is not None:
        fx_rate = float(summary_fx)
        fx_source = "summary:Exchange rate"
    else:
        rates = fetch_usd_rates()
        fx_rate = get_tw_pn_fx_rate(rates)
        fx_source = "api:TWD"
    fx_row = pn_layout.get("fx_row") or _find_pn_fx_row(wb[PN_SHEET])
    wb[PN_SHEET].cell(fx_row, 2).value = fx_rate
    pn_layout["fx_row"] = fx_row
    from fx_policy import make_pn_fx_provenance

    write_source = "api" if str(fx_source or "").startswith("api:") else "mapping"
    if str(fx_source or "").startswith("summary:"):
        write_source = "mapping"
    pn_fx_write = make_pn_fx_provenance(
        PN_SHEET,
        int(fx_row),
        2,
        _ACTIVE_MAPPING,
        float(fx_rate),
        write_source=write_source,
        fx_source=str(fx_source or ""),
    )
    ensure_tw_period_date_formats(wb)
    apply_luckysheet_compat(wb, pn_sheet=PN_SHEET)

    wb.save(output_path)
    wb.close()
    # 主题填充 / 富文本；金额由前端 HyperFormula 按公式重算，不再注入 PN 缓存
    postprocess_converted_xlsx(output_path)

    warnings = list(ee_warnings)
    if summary_fx is None:
        warnings.append("Summary 未找到有效 Exchange rate，已回退 API TWD 汇率")

    return {
        "employee_count": len(employees),
        "employee_names": [
            e.get("CN Name") or e.get("EN Name") for e in employees
        ],
        "company_name": parsed["meta"].get("company_name"),
        "period": (parsed["meta"].get("period_from"), parsed["meta"].get("period_to")),
        "fx_rate": fx_rate,
        "fx_source": fx_source,
        "summary_exchange_rate": parsed["meta"].get("exchange_rate"),
        "fx_row": fx_row,
        "output": str(output_path),
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
        "warnings": warnings,
        "pn_fx_write": pn_fx_write,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="台湾 People Search 账单 → PN N-C 自动转换")
    parser.add_argument("source", type=Path, help="原始供应商 Excel 路径")
    parser.add_argument("-o", "--output", type=Path, help="输出 PN 路径")
    parser.add_argument("-t", "--template", type=Path, default=DEFAULT_TEMPLATE, help="PN 母版路径")
    args = parser.parse_args(argv)

    source = args.source.resolve()
    output = args.output.resolve() if args.output else source.parent / f"PN_auto_{source.stem}.xlsx"

    try:
        result = convert(source, output, args.template.resolve())
    except Exception as exc:
        print(f"转换失败: {exc}", file=sys.stderr)
        return 1

    print("转换完成")
    print(f"  输出: {result['output']}")
    print(f"  公司: {result['company_name']}")
    print(f"  账期: {result['period'][0]} ~ {result['period'][1]}")
    print(f"  员工: {result['employee_count']} 人 → {result['employee_names']}")
    print(f"  汇率 PN!B{result.get('fx_row', 31)}: {result['fx_rate']} ({result.get('fx_source')})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
