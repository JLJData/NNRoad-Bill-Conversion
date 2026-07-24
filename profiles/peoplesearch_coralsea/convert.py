# -*- coding: utf-8 -*-
"""
People Search + Coral Sea 台湾账单 → PN N-C 转换脚本

用法:
  python -m profiles.peoplesearch_coralsea.convert <原始供应商账单.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

原始账单: sheet「Payroll calculation」按表头名匹配（支持多员工）
默认母版: templates/taiwan/template.xlsx
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from fx_rate import fetch_usd_rates, get_tw_pn_fx_rate
from pn_meta import PnMeta, apply_pn_meta
from region_templates import get_region_template
from xlsx_formula_cache import inject_formula_cached_values
from xlsx_richtext_fix import migrate_inlinestr_richtext_to_shared_strings
from xlsx_theme_fill_fix import materialize_theme_fills

DEFAULT_TEMPLATE = get_region_template("Taiwan")

PC_SHEET = "Payroll calculation"
SUMMARY_SHEET = "Summary"
TW_L_SHEET = "TW-L"
TW_SHEET = "TW"
TW_EE_SHEET = "TW EE"
PN_SHEET = "PN"

TW_L_HEADER_ROW = 7
TW_L_SUMMARY_ROW = 8
TW_L_DATA_START_ROW = 9
TW_L_SUMMARY_MIN_END_ROW = 14
TW_DATA_START_ROW = 9
TW_EE_DATA_START_ROW = 10
MAX_EMPLOYEES = 9
PAYROLL_DUP_MIN_COL = 14

# 源表有、但 TW-L 不直接写入的列（加班已汇总到「加班費」；Service Fee 由模板留空）
SKIP_SOURCE_HEADERS = frozenset({
    "Hr (1.34x)",
    "OT Payment (1.34x)",
    "Hr (1.67x)",
    "OT Payment (1.67x)",
    "Hr (1x)",
    "OT Payment (1x)",
    "Hr (2.67x)",
    "OT Payment (2.67x)",
    "Service Fee",
})

SICK_LEAVE_PAY_HEADER = "病假扣薪\nSick Leave\n(half pay)"
SICK_LEAVE_HOURS_HEADER = "病假時數 Sick Leave Hours"

# 源表头 → 目标 TW-L 表头（按名称匹配，非位置）
SOURCE_TO_TARGET_HEADER: dict[str, str] = {
    "時薪 Hourly Rate": "Full Pay/Hourly Rate",
    "時數\nHours Worked": "Employment day/Hours Worked",
    "健保級距\nInsured Salary Grading - HI": "健保投保級距Insured Salary Grading - HI",
    "Position ": "Position",
}

NAME_HEADERS = ("CN Name", "EN Name")
PC_HEADER_KEYS = frozenset({"BU", "CN Name", "EN Name"})


def norm(text: Any) -> str:
    if text is None:
        return ""
    return str(text).replace("\uFEFF", "").strip()


def map_source_header(source_header: str) -> str | None:
    h = norm(source_header)
    if not h or h in SKIP_SOURCE_HEADERS:
        return None
    return SOURCE_TO_TARGET_HEADER.get(h, h)


def find_pc_header_row(ws: Worksheet, max_scan: int = 15) -> int:
    for row in range(1, max_scan + 1):
        hits = 0
        for col in range(1, (ws.max_column or 0) + 1):
            if norm(ws.cell(row, col).value) in PC_HEADER_KEYS:
                hits += 1
        if hits >= 2:
            return row
    raise ValueError(
        f"「{PC_SHEET}」前 {max_scan} 行未找到表头行（需含 BU / CN Name / EN Name）"
    )


def build_header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def build_header_cols(ws: Worksheet, header_row: int) -> dict[str, list[int]]:
    mapping: dict[str, list[int]] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key:
            mapping.setdefault(key, []).append(col)
    return mapping


def resolve_target_col(cols: list[int]) -> int | None:
    if not cols:
        return None
    if len(cols) == 1:
        return cols[0]
    payroll = [c for c in cols if c >= PAYROLL_DUP_MIN_COL]
    return max(payroll) if payroll else max(cols)


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    s = norm(value)
    if s in ("", "#N/A", "#REF!", "#VALUE!", "-"):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    # 千分位逗号 / 中文逗号：4,149 → 4149（否则会写成文本，LuckySheet 公式算不动）
    compact = s.replace(",", "").replace("，", "").replace(" ", "")
    try:
        if re.fullmatch(r"-?\d+(\.\d+)?", compact):
            return float(compact) if "." in compact else int(compact)
    except ValueError:
        pass
    return value


def format_payroll_date(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        d = value.date()
        return f"{d.year}/{d.month}/{d.day}"
    if isinstance(value, date):
        return f"{value.year}/{value.month}/{value.day}"
    s = norm(value)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            d = datetime.strptime(s, fmt).date()
            return f"{d.year}/{d.month}/{d.day}"
        except ValueError:
            continue
    return value


def parse_period(value: Any, payroll_month: Any = None) -> tuple[Any, Any]:
    """解析 Summary 中 Period 如 3/1-3/31"""
    if value is None:
        return None, None
    s = norm(value)
    m = re.match(r"(\d{1,2})/(\d{1,2})\s*-\s*(\d{1,2})/(\d{1,2})", s)
    if not m:
        return None, None

    year = None
    if payroll_month is not None:
        if isinstance(payroll_month, datetime):
            year = payroll_month.year
        elif isinstance(payroll_month, date):
            year = payroll_month.year
        else:
            ps = norm(payroll_month)
            for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d"):
                try:
                    year = datetime.strptime(ps[:19], fmt).year
                    break
                except ValueError:
                    continue
    if year is None:
        year = datetime.now().year

    m1, d1, m2, d2 = (int(m.group(i)) for i in range(1, 5))
    return (
        format_payroll_date(date(year, m1, d1)),
        format_payroll_date(date(year, m2, d2)),
    )


def payroll_month_start(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        d = value.date().replace(day=1)
        return datetime(d.year, d.month, d.day)
    if isinstance(value, date):
        d = value.replace(day=1)
        return datetime(d.year, d.month, d.day)
    s = norm(value)
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            d = datetime.strptime(s[:19], fmt).date().replace(day=1)
            return datetime(d.year, d.month, d.day)
        except ValueError:
            continue
    return value


def read_pc_employees(ws: Worksheet, header_row: int) -> list[dict[str, Any]]:
    source_headers = build_header_map(ws, header_row)
    name_cols = [source_headers[h] for h in NAME_HEADERS if h in source_headers]
    if not name_cols:
        raise ValueError(f"「{PC_SHEET}」表头行须包含 CN Name 或 EN Name")

    employees: list[dict[str, Any]] = []
    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        has_name = any(clean_value(ws.cell(row, col).value) for col in name_cols)
        if not has_name:
            continue

        record: dict[str, Any] = {}
        for src_hdr, col in source_headers.items():
            target_hdr = map_source_header(src_hdr)
            if target_hdr is None:
                continue
            val = clean_value(ws.cell(row, col).value)
            if val is not None:
                record[target_hdr] = val
        employees.append(record)

    if not employees:
        raise ValueError(f"「{PC_SHEET}」中未找到有效员工行（需有 CN Name 或 EN Name）")
    if len(employees) > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {len(employees)} 超过模板上限 {MAX_EMPLOYEES}")
    return employees


def read_summary_meta(wb) -> dict[str, Any]:
    if SUMMARY_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SUMMARY_SHEET]
    meta: dict[str, Any] = {}
    for row in range(1, 20):
        label = norm(ws.cell(row, 1).value)
        val = ws.cell(row, 2).value
        if label == "Client":
            meta["company_name"] = val
        elif label == "Payroll Month":
            meta["payroll_month"] = val
        elif label == "Period":
            meta["period_raw"] = val
    period_from, period_to = parse_period(meta.get("period_raw"), meta.get("payroll_month"))
    meta["period_from"] = period_from
    meta["period_to"] = period_to
    meta["payroll_month_start"] = payroll_month_start(meta.get("payroll_month"))
    return meta


def apply_sick_leave_formula(ws: Worksheet, row: int, target_cols: dict[str, list[int]]) -> None:
    """有病假时数时写入公式 =-H{row}*T{row}/2，与母版一致"""
    pay_cols = target_cols.get(SICK_LEAVE_PAY_HEADER)
    if not pay_cols:
        return
    hours_cols = target_cols.get(SICK_LEAVE_HOURS_HEADER)
    hours_col = resolve_target_col(hours_cols) if hours_cols else None
    hours = ws.cell(row, hours_col).value if hours_col else None
    try:
        has_hours = hours is not None and float(hours) > 0
    except (TypeError, ValueError):
        has_hours = False

    col = resolve_target_col(pay_cols)
    if col is None:
        return
    if has_hours and hours_col:
        ws.cell(row, col).value = f"=-H{row}*{get_column_letter(hours_col)}{row}/2"
    elif not has_hours:
        ws.cell(row, col).value = 0


def clear_tw_l_data(ws: Worksheet, employee_count: int) -> None:
    max_row = max(ws.max_row or TW_L_DATA_START_ROW, TW_L_DATA_START_ROW + MAX_EMPLOYEES)
    max_col = ws.max_column or 1
    for row in range(TW_L_DATA_START_ROW, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def write_tw_l(ws: Worksheet, employees: list[dict[str, Any]], meta: dict[str, Any]) -> None:
    target_cols = build_header_cols(ws, TW_L_HEADER_ROW)
    if not target_cols:
        raise ValueError(f"「{TW_L_SHEET}」第 {TW_L_HEADER_ROW} 行表头为空")

    if meta.get("payroll_month_start") is not None:
        ws.cell(3, 3).value = meta["payroll_month_start"]
    if meta.get("period_from") is not None:
        ws.cell(4, 3).value = meta["period_from"]
    if meta.get("period_to") is not None:
        ws.cell(5, 3).value = meta["period_to"]

    clear_tw_l_data(ws, len(employees))
    for idx, emp in enumerate(employees):
        row = TW_L_DATA_START_ROW + idx
        for hdr, val in emp.items():
            cols = target_cols.get(hdr)
            col = resolve_target_col(cols) if cols else None
            if col is not None:
                ws.cell(row, col).value = val
        apply_sick_leave_formula(ws, row, target_cols)

    update_tw_l_summary_formulas(ws, len(employees))


def update_tw_l_summary_formulas(ws: Worksheet, employee_count: int) -> None:
    end_row = max(
        TW_L_DATA_START_ROW + max(employee_count, 1) - 1,
        TW_L_SUMMARY_MIN_END_ROW,
    )
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        cell = ws.cell(TW_L_SUMMARY_ROW, col)
        if cell.data_type != "f" or not isinstance(cell.value, str):
            continue
        formula = cell.value
        if not re.search(r"SUM\([A-Z]+\d+:[A-Z]+\d+\)", formula, re.I):
            continue
        cell.value = re.sub(
            r"(SUM\([A-Z]+)(\d+)(:[A-Z]+)(\d+)(\))",
            lambda m: f"{m.group(1)}{TW_L_DATA_START_ROW}{m.group(3)}{end_row}{m.group(5)}",
            formula,
            flags=re.I,
        )


def shift_tw_formula(
    formula: str,
    from_row: int,
    to_row: int,
    tw_l_from: int,
    tw_l_to: int,
) -> str:
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    placeholders: list[str] = []

    def stash_external(match: re.Match[str]) -> str:
        placeholders.append(match.group(0))
        return f"__EXT{len(placeholders) - 1}__"

    # 先暂存跨 sheet 引用，避免本 sheet 行号替换误改 'TW-L'!N10 中的 10
    s = re.sub(r"'[^']+'!\$?[A-Z]{1,3}\$?\d+", stash_external, formula)
    s = re.sub(
        rf"(?<!\$)(?<![A-Z])([A-Z]{{1,3}}){from_row}(?!\d)",
        lambda m: f"{m.group(1)}{to_row}",
        s,
    )
    for idx, ref in enumerate(placeholders):
        ref = re.sub(
            r"'TW-L'!([A-Z]{1,3})(\d+)",
            lambda m: f"'TW-L'!{m.group(1)}{tw_l_to}",
            ref,
        )
        s = s.replace(f"__EXT{idx}__", ref)
    return s


def clear_employee_row(ws: Worksheet, row: int) -> None:
    for col in range(1, (ws.max_column or 0) + 1):
        ws.cell(row, col).value = None


# TW!F = Business Tax = TW-L!Total(BO) * 5%
_TW_BUSINESS_TAX_COL = 6  # F
_TW_L_TOTAL_COL = 67  # BO


def apply_tw_business_tax(ws_tw: Worksheet, employee_count: int) -> None:
    """每人 Business Tax：TW!F = ROUND('TW-L'!BO * 5%, 0) 取整。LuckySheet 用 (5/100)。"""
    n = max(int(employee_count), 0)
    for i in range(n):
        tw_row = TW_DATA_START_ROW + i
        tw_l_row = TW_L_DATA_START_ROW + i
        cell = ws_tw.cell(tw_row, _TW_BUSINESS_TAX_COL)
        cell.value = f"=ROUND('TW-L'!BO{tw_l_row}*(5/100),0)"
        cell.number_format = "#,##0"


def _norm_person_name(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).replace("\u3000", " ").strip().lower()
    return re.sub(r"\s+", " ", s)


def _name_tokens(value: str) -> list[str]:
    return [t for t in _norm_person_name(value).split(" ") if t]


def score_ee_name_match(excel_name: str, candidate_name: str) -> int:
    """
    评分：精确 100；一方包含另一方 80；Excel 全部 token 都在系统名中 70（Pin Han ⊆ Pin Han Wang）。
    """
    a = _norm_person_name(excel_name)
    b = _norm_person_name(candidate_name)
    if not a or not b:
        return 0
    if a == b:
        return 100
    if a in b or b in a:
        return 80
    ta, tb = _name_tokens(a), set(_name_tokens(b))
    if ta and all(t in tb for t in ta):
        return 70
    return 0


def match_ee_code(
    excel_names: list[str],
    directory: list[dict[str, Any]],
) -> tuple[str | None, str | None]:
    """
    从员工目录匹配 employee_code。
    返回 (code, warning)；匹配不到/歧义时 code=None 并带 warning。
    """
    names = [n for n in (_norm_person_name(x) for x in excel_names) if n]
    if not names:
        return None, "源表员工姓名为空，无法匹配 EE Code"
    if not directory:
        return None, "未提供客户员工目录，无法匹配 EE Code"

    best_score = 0
    best: list[dict[str, Any]] = []
    for row in directory:
        code = norm(row.get("employee_code") or row.get("employeeCode"))
        if not code:
            continue
        score = 0
        for field in (
            row.get("employee_name_en") or row.get("employeeNameEn"),
            row.get("employee_name") or row.get("employeeName"),
        ):
            for en in excel_names:
                score = max(score, score_ee_name_match(en, str(field or "")))
        if score <= 0:
            continue
        if score > best_score:
            best_score = score
            best = [row]
        elif score == best_score:
            best.append(row)

    if not best:
        label = " / ".join(n for n in excel_names if n)
        return None, f"未匹配到 EE Code：{label}"

    # 同分多人：优先更短英文名（更具体的全名通常更长，但同分时取最短减少歧义误伤）
    if len(best) > 1:
        # 若编码相同可合并
        codes = {norm(r.get("employee_code") or r.get("employeeCode")) for r in best}
        if len(codes) == 1:
            return next(iter(codes)), None
        label = " / ".join(n for n in excel_names if n)
        return None, f"EE Code 匹配歧义（{len(best)} 人同分）：{label}"

    return norm(best[0].get("employee_code") or best[0].get("employeeCode")), None


def apply_tw_ee_codes(
    ws_ee: Worksheet,
    employees: list[dict[str, Any]],
    directory: list[dict[str, Any]] | None,
) -> list[str]:
    """写入 TW EE!D（EE Code）；匹配失败留空并返回 warnings。"""
    warnings: list[str] = []
    dir_list = list(directory or [])
    for i, emp in enumerate(employees):
        row = TW_EE_DATA_START_ROW + i
        excel_names = [
            emp.get("EN Name"),
            emp.get("CN Name"),
        ]
        code, warn = match_ee_code(
            [str(x) for x in excel_names if x],
            dir_list,
        )
        ws_ee.cell(row, 4).value = code  # 匹配不到显式清空，避免母版残留
        if warn:
            warnings.append(f"第{i + 1}人：{warn}")
    return warnings


_AMP_PLUS_RE = re.compile(r"&\s*\+")
_DASH_AMP_QUOTE_RE = re.compile(r'="-\s*"&"([^"]*)"')

# PN!E16 = TW!J9+TW!Z9；J=SUM(K:Y) 对应 TW-L 列（R 在模板里被加两次）
_PN_J_TW_L_COLS = ("N", "AC", "O", "P", "Q", "R", "R", "U", "V", "W", "X", "Y", "AA", "AB")
_PN_Z_TW_L_COL = "BL"  # ER Insurance
_PN_AB_TW_L_COL = "Z"  # Reimb Claim
_PN_EOR_ROW = 15
_PN_LABOR_START_ROW = 16
# 母版默认 3 人：Labor 16-18 / Expense 19-21 / 空行 22 / Service 23 / Mgmt 24 / Total 28 / FX B31
_PN_NTD_FMT = "#,##0.00"
_PN_USD_FMT = '$#,##0.00'
_MGMT_RATE = 0.12
_MGMT_MIN = 4500.0
_EXPENSE_OP_FEE = 1500.0


def _pn_layout(employee_count: int) -> dict[str, int]:
    """按人数推算 PN 明细区行号（与母版 3 人布局同构）。"""
    n = max(int(employee_count), 0)
    labor_start = _PN_LABOR_START_ROW
    expense_start = labor_start + n
    blank_row = expense_start + n
    service_row = blank_row + 1
    mgmt_row = service_row + 1
    return {
        "n": n,
        "eor_row": _PN_EOR_ROW,
        "labor_start": labor_start,
        "expense_start": expense_start,
        "blank_row": blank_row,
        "service_row": service_row,
        "mgmt_row": mgmt_row,
        "sum_end": blank_row,  # E15=SUM(E16:E{blank})，含空行
    }


def _count_pn_labor_slots(ws: Worksheet) -> int:
    n = 0
    for row in range(_PN_LABOR_START_ROW, _PN_LABOR_START_ROW + MAX_EMPLOYEES + 2):
        v = ws.cell(row, 1).value
        if isinstance(v, str) and "Labor cost" in v:
            n += 1
            continue
        break
    return max(n, 1)


def _find_pn_row_by_label(ws: Worksheet, keyword: str, col: int = 1) -> int | None:
    key = keyword.lower()
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and key in v.lower():
            return row
    return None


def _find_pn_fx_row(ws: Worksheet) -> int:
    r = _find_pn_row_by_label(ws, "FX rate")
    return r if r else 31


def _copy_pn_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    max_col = max(ws.max_column or 6, 6)
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _ensure_merge_a_c(ws: Worksheet, row: int) -> None:
    """母版 Labor/Expense 描述列为 A:C 合并。"""
    rng = f"A{row}:C{row}"
    for m in list(ws.merged_cells.ranges):
        if m.min_row <= row <= m.max_row and m.min_col <= 1 and m.max_col >= 3:
            if m.min_row == row and m.max_row == row and m.min_col == 1 and m.max_col == 3:
                return
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    try:
        ws.merge_cells(rng)
    except ValueError:
        pass


def _collect_merges_from(ws: Worksheet, start_row: int) -> list[tuple[int, int, int, int]]:
    """收集 min_row>=start_row 的合并区，并先解除，避免与 insert/delete 错位。"""
    kept: list[tuple[int, int, int, int]] = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row < start_row:
            continue
        kept.append((m.min_row, m.min_col, m.max_row, m.max_col))
        try:
            ws.unmerge_cells(str(m))
        except ValueError:
            pass
    return kept


def _restore_merges(
    ws: Worksheet,
    merges: list[tuple[int, int, int, int]],
    row_shift: int = 0,
) -> None:
    for min_r, min_c, max_r, max_c in merges:
        nr1, nr2 = min_r + row_shift, max_r + row_shift
        if nr1 < 1 or nr2 < nr1:
            continue
        try:
            ws.merge_cells(
                start_row=nr1,
                start_column=min_c,
                end_row=nr2,
                end_column=max_c,
            )
        except ValueError:
            pass


def _shift_row_heights(ws: Worksheet, start_row: int, amount: int) -> None:
    """openpyxl insert/delete 时常不带动 row_dimensions，需手动平移行高。"""
    if amount == 0:
        return
    captured: dict[int, float] = {}
    max_r = max(ws.max_row or start_row, start_row)
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            captured[r] = h
        max_r = max(max_r, r)

    for r in range(start_row, max_r + abs(amount) + 2):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None

    if amount > 0:
        for r in sorted(captured.keys(), reverse=True):
            ws.row_dimensions[r + amount].height = captured[r]
    else:
        for r in sorted(captured.keys()):
            new_r = r + amount
            if new_r >= start_row:
                ws.row_dimensions[new_r].height = captured[r]


def _capture_row_heights_from(ws: Worksheet, start_row: int) -> dict[int, float]:
    out: dict[int, float] = {}
    for r in list(ws.row_dimensions.keys()):
        if not isinstance(r, int) or r < start_row:
            continue
        h = ws.row_dimensions[r].height
        if h is not None:
            out[r] = h
    return out


def _apply_row_heights(ws: Worksheet, heights: dict[int, float], start_row: int) -> None:
    max_r = max([start_row, *heights.keys()], default=start_row)
    for r in range(start_row, max_r + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r, h in heights.items():
        if r >= start_row:
            ws.row_dimensions[r].height = h


def _pn_insert_rows(ws: Worksheet, idx: int, amount: int, fill_style_row: int | None = None) -> None:
    """插入行，并让下方合并区/行高一起下移；新行复制样板行样式。"""
    if amount <= 0:
        return
    merges = _collect_merges_from(ws, idx)
    _shift_row_heights(ws, idx, amount)
    ws.insert_rows(idx, amount)
    _restore_merges(ws, merges, row_shift=amount)
    if fill_style_row is not None:
        style_row = fill_style_row + amount if fill_style_row >= idx else fill_style_row
        for r in range(idx, idx + amount):
            _copy_pn_row_style(ws, style_row, r)
            _ensure_merge_a_c(ws, r)


def _pn_delete_rows(ws: Worksheet, idx: int, amount: int) -> None:
    """删除行，并让下方合并区/行高一起上移。"""
    if amount <= 0:
        return
    for m in list(ws.merged_cells.ranges):
        if m.max_row < idx or m.min_row >= idx + amount:
            continue
        # 与删除区间相交的合并先拆掉
        try:
            ws.unmerge_cells(str(m))
        except ValueError:
            pass
    merges = _collect_merges_from(ws, idx + amount)
    heights = _capture_row_heights_from(ws, idx)
    ws.delete_rows(idx, amount)
    # 删除区间内行高丢弃；下方行高上移 amount
    new_heights: dict[int, float] = {}
    for r, h in heights.items():
        if r < idx:
            continue
        if r < idx + amount:
            continue
        new_heights[r - amount] = h
    _apply_row_heights(ws, new_heights, idx)
    _restore_merges(ws, merges, row_shift=-amount)


def fit_pn_employees(ws: Worksheet, employee_count: int) -> dict[str, int]:
    """
    按实际人数扩/缩 PN 的 Labor + Expense 明细行，并重写相关公式。
    插入/删除时同步平移下方合并单元格与行高，避免 Service/合计/汇率区样式错位。
    """
    n = max(int(employee_count), 1)
    if n > MAX_EMPLOYEES:
        raise ValueError(f"员工数 {n} 超过模板上限 {MAX_EMPLOYEES}")

    old_n = _count_pn_labor_slots(ws)
    labor_start = _PN_LABOR_START_ROW
    expense_start = labor_start + old_n
    delta = n - old_n

    if delta > 0:
        # 在首个 Expense 前插入 Labor 行（样式跟上一行 Labor）
        _pn_insert_rows(ws, expense_start, delta, fill_style_row=labor_start)
        expense_start += delta
        # 在 Expense 块末尾后插入 Expense 行
        _pn_insert_rows(ws, expense_start + old_n, delta, fill_style_row=expense_start)
    elif delta < 0:
        # 先删多余 Expense，再删多余 Labor
        _pn_delete_rows(ws, expense_start + n, -delta)
        _pn_delete_rows(ws, labor_start + n, -delta)

    layout = _pn_layout(n)
    expense_start = layout["expense_start"]
    blank_row = layout["blank_row"]
    service_row = layout["service_row"]
    mgmt_row = layout["mgmt_row"]
    eor_row = layout["eor_row"]
    sum_end = layout["sum_end"]

    # 空行：清内容但保留已随行移动的样式/行高
    for m in list(ws.merged_cells.ranges):
        if m.min_row == blank_row and m.max_row == blank_row:
            try:
                ws.unmerge_cells(str(m))
            except ValueError:
                pass
    for col in range(1, 7):
        cell = ws.cell(blank_row, col)
        if type(cell).__name__ == "MergedCell":
            continue
        cell.value = None
    _ensure_merge_a_c(ws, blank_row)

    fx_row = _find_pn_fx_row(ws)

    for i in range(n):
        tw_row = TW_DATA_START_ROW + i
        labor_row = labor_start + i
        expense_row = expense_start + i

        if i > 0:
            _copy_pn_row_style(ws, labor_start, labor_row)
            _copy_pn_row_style(ws, expense_start, expense_row)

        _ensure_merge_a_c(ws, labor_row)
        _ensure_merge_a_c(ws, expense_row)

        ws.cell(labor_row, 1).value = (
            f'="- Labor cost for "&TW!B{tw_row}&"  -  "&MONTH(TW!$B$2)&"-"&YEAR(TW!$B$2)'
        )
        e_labor = ws.cell(labor_row, 5)
        e_labor.value = f"=TW!J{tw_row}+TW!Z{tw_row}"
        e_labor.number_format = _PN_NTD_FMT
        f_labor = ws.cell(labor_row, 6)
        f_labor.value = f"=E{labor_row}/$B${fx_row}"
        f_labor.number_format = _PN_USD_FMT

        ws.cell(expense_row, 1).value = f'="- Expense claim for "&TW!B{tw_row}'
        e_exp = ws.cell(expense_row, 5)
        e_exp.value = f"=TW!AB{tw_row}"
        e_exp.number_format = _PN_NTD_FMT
        f_exp = ws.cell(expense_row, 6)
        f_exp.value = f"=E{expense_row}/$B${fx_row}"
        f_exp.number_format = _PN_USD_FMT

    # EOR 汇总（简单数字格式，避免 LuckySheet 会计格式空白）
    e_eor = ws.cell(eor_row, 5)
    e_eor.value = f"=SUM(E{labor_start}:E{sum_end})"
    e_eor.number_format = _PN_NTD_FMT
    f_eor = ws.cell(eor_row, 6)
    f_eor.value = f"=SUM(F{labor_start}:F{sum_end})"
    f_eor.number_format = _PN_USD_FMT

    svc = _find_pn_row_by_label(ws, "Service Fee") or service_row
    mgmt = _find_pn_row_by_label(ws, "Management Fee") or mgmt_row
    e_svc = ws.cell(svc, 5)
    e_svc.value = f"=SUM(E{mgmt}:E{mgmt + 1})"
    e_svc.number_format = _PN_NTD_FMT
    f_svc = ws.cell(svc, 6)
    f_svc.value = f"=SUM(F{mgmt}:F{mgmt + 1})"
    f_svc.number_format = _PN_USD_FMT
    if isinstance(ws.cell(mgmt, 1).value, str) or ws.cell(mgmt, 1).value is None:
        ws.cell(mgmt, 1).value = f'="- Management Fee - "&MONTH(TW!B2)&"-"&YEAR(TW!B2)'
    e_mgmt = ws.cell(mgmt, 5)
    e_mgmt.value = "=TW!G6"
    e_mgmt.number_format = _PN_NTD_FMT
    f_mgmt = ws.cell(mgmt, 6)
    f_mgmt.value = f"=E{mgmt}/$B${fx_row}"
    f_mgmt.number_format = _PN_USD_FMT
    _ensure_merge_a_c(ws, svc)
    _ensure_merge_a_c(ws, mgmt)

    total_row = None
    for row in range(eor_row + 1, (ws.max_row or 0) + 1):
        ev = ws.cell(row, 5).value
        if isinstance(ev, str) and ev.startswith("=") and "E15" in ev.replace("$", ""):
            total_row = row
            break
        av = ws.cell(row, 1).value
        if isinstance(av, str) and "EOR/PEO Service Cost" in av:
            total_row = row
            break
    if total_row is None:
        total_row = svc + 5

    e_total = ws.cell(total_row, 5)
    e_total.value = f"=E{eor_row}+E{svc}"
    e_total.number_format = _PN_NTD_FMT
    f_total = ws.cell(total_row, 6)
    f_total.value = f"=F{eor_row}+F{svc}"
    f_total.number_format = _PN_USD_FMT
    # 合计行母版是 A:D 合并
    for m in list(ws.merged_cells.ranges):
        if m.min_row == total_row and m.max_row == total_row and m.min_col == 1:
            break
    else:
        try:
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        except ValueError:
            pass

    for row in range(labor_start, sum_end + 1):
        for col, fmt in ((5, _PN_NTD_FMT), (6, _PN_USD_FMT)):
            cell = ws.cell(row, col)
            if cell.value is not None:
                cell.number_format = fmt

    # 插删后重找 FX；openpyxl 不会改写公式文本里的绝对行号，需整段重写结算区
    fx_row = _find_pn_fx_row(ws)
    _rewrite_pn_settlement_block(ws, fx_row=fx_row, total_row=total_row)
    for row in range(labor_start, fx_row + 8):
        fcell = ws.cell(row, 6)
        if isinstance(fcell.value, str) and "$B$" in fcell.value:
            fcell.value = re.sub(r"\$B\$\d+", f"$B${fx_row}", fcell.value)
        if isinstance(fcell.value, str) and "PN!B" in fcell.value.replace("$", ""):
            fcell.value = re.sub(r"PN!\$?B\$?\d+", f"PN!B{fx_row}", fcell.value)
    layout["service_row"] = svc
    layout["mgmt_row"] = mgmt
    layout["total_row"] = total_row
    layout["fx_row"] = fx_row
    return layout


def _rewrite_pn_settlement_block(ws: Worksheet, *, fx_row: int, total_row: int) -> None:
    """
    母版 FX 下方结算块相对布局（以 fx_row 为锚）：
      F{fx}     = F{total}                         EOR USD
      F{fx+1..4}= TW!E22..E25                      Other Fee 透传
      F{fx+5}   = SUM(F{fx}:F{fx+4})               Sub Total
      F{fx+6}   = ROUNDUP(TW!F6/PN!B{fx},2)+0.01   Tax
      F{fx+7}   = SUM(F{fx+5}:F{fx+6})             Total
    插删行后公式文本不会自动更新，必须按新行号重写。
    """
    # EOR USD 引用合计行
    f_eor = ws.cell(fx_row, 6)
    f_eor.value = f"=F{total_row}"
    f_eor.number_format = _PN_USD_FMT

    # Other Fee 透传（TW 行号固定，与人数无关）
    for i, tw_e in enumerate((22, 23, 24, 25), start=1):
        cell = ws.cell(fx_row + i, 6)
        cell.value = f"=TW!E{tw_e}"

    sub_row = fx_row + 5
    tax_row = fx_row + 6
    grand_row = fx_row + 7

    f_sub = ws.cell(sub_row, 6)
    f_sub.value = f"=SUM(F{fx_row}:F{fx_row + 4})"
    f_sub.number_format = _PN_USD_FMT

    f_tax = ws.cell(tax_row, 6)
    f_tax.value = f"=ROUNDUP(TW!F6/PN!B{fx_row},2)+0.01"
    f_tax.number_format = _PN_USD_FMT

    f_grand = ws.cell(grand_row, 6)
    f_grand.value = f"=SUM(F{sub_row}:F{tax_row})"
    f_grand.number_format = _PN_USD_FMT


def _cell_num(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("="):
            return 0.0
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return 0.0
    return 0.0


def _tw_l_sick_pay(ws_twl: Worksheet, row: int) -> float:
    """TW-L!U 常为 =-H*T/2，openpyxl 不会算，需手算。"""
    from openpyxl.utils import column_index_from_string as ci

    u = ws_twl.cell(row, ci("U")).value
    if isinstance(u, str) and u.startswith("="):
        h = _cell_num(ws_twl.cell(row, ci("H")).value)
        t = _cell_num(ws_twl.cell(row, ci("T")).value)
        return (-h * t / 2.0) if t else 0.0
    return _cell_num(u)


def _tw_l_amount(ws_twl: Worksheet, row: int, letter: str) -> float:
    from openpyxl.utils import column_index_from_string as ci

    if letter == "U":
        return _tw_l_sick_pay(ws_twl, row)
    return _cell_num(ws_twl.cell(row, ci(letter)).value)


def _compute_employee_pn_parts(ws_twl: Worksheet, row: int) -> tuple[float, float, float]:
    """返回 (labor_ntd, expense_ntd, management_fee_ntd)。"""
    j = sum(_tw_l_amount(ws_twl, row, c) for c in _PN_J_TW_L_COLS)
    z = _tw_l_amount(ws_twl, row, _PN_Z_TW_L_COL)
    ab = _tw_l_amount(ws_twl, row, _PN_AB_TW_L_COL)
    labor = j + z
    fee_h = max(_MGMT_MIN, labor * _MGMT_RATE)
    fee_i = _EXPENSE_OP_FEE if ab > 0 else 0.0
    return labor, ab, fee_h + fee_i


def _write_pn_amount(ws_pn: Worksheet, row: int, ntd: float, fx: float) -> None:
    e = ws_pn.cell(row, 5)
    f = ws_pn.cell(row, 6)
    e.value = float(ntd)
    e.number_format = _PN_NTD_FMT
    f.value = (float(ntd) / fx) if fx else 0.0
    f.number_format = _PN_USD_FMT


def compute_pn_amount_cache(
    wb,
    employee_count: int,
    layout: dict[str, int] | None = None,
) -> tuple[dict[str, dict[str, float]], dict[str, float]]:
    """
    按母版公式语义计算 PN!E/F 金额，用于写入 xlsx 公式缓存值。
    不覆盖单元格公式：Excel 里改 TW-L / 汇率后仍可重算联动；
    LuckySheet 则靠缓存 <v> 显示 NTD/USD。
    """
    ws_twl = wb[TW_L_SHEET]
    ws_pn = wb[PN_SHEET]
    lay = layout or _pn_layout(employee_count)
    fx_row = int(lay.get("fx_row") or _find_pn_fx_row(ws_pn))
    fx = _cell_num(ws_pn.cell(fx_row, 2).value) or 1.0

    labor_total = 0.0
    expense_total = 0.0
    mgmt_total = 0.0
    cell_cache: dict[str, float] = {}

    def put(row: int, ntd: float) -> None:
        usd = (float(ntd) / fx) if fx else 0.0
        cell_cache[f"E{row}"] = float(ntd)
        cell_cache[f"F{row}"] = usd

    n = int(lay.get("n") or employee_count)
    labor_start = int(lay["labor_start"])
    expense_start = int(lay["expense_start"])
    for i in range(n):
        tw_row = TW_L_DATA_START_ROW + i
        labor, expense, mgmt = _compute_employee_pn_parts(ws_twl, tw_row)
        labor_total += labor
        expense_total += expense
        mgmt_total += mgmt
        put(labor_start + i, labor)
        put(expense_start + i, expense)

    eor = labor_total + expense_total
    eor_row = int(lay["eor_row"])
    svc_row = int(lay["service_row"])
    mgmt_row = int(lay["mgmt_row"])
    total_row = int(lay.get("total_row") or (svc_row + 5))
    put(eor_row, eor)
    put(mgmt_row, mgmt_total)
    put(svc_row, mgmt_total)
    put(total_row, eor + mgmt_total)
    put(mgmt_row + 1, 0.0)

    summary = {
        "fx": fx,
        "eor_ntd": eor,
        "mgmt_ntd": mgmt_total,
        "total_ntd": eor + mgmt_total,
        "layout": lay,
    }
    return {PN_SHEET: cell_cache}, summary


def fix_pn_illegal_concat_formulas(ws: Worksheet) -> int:
    """母版误写 `="- "&+"Expense...`：Excel 能算，LuckySheet/部分引擎会 #VALUE!。"""
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=") and _AMP_PLUS_RE.search(v)):
                continue
            nv = _AMP_PLUS_RE.sub("&", v)
            nv = _DASH_AMP_QUOTE_RE.sub(r'="- \1"', nv)
            if nv != v:
                cell.value = nv
                n += 1
    return n


_SHEET_TITLE_SELF_REF_RE = re.compile(
    r'CELL\s*\(\s*"filename"\s*,\s*\$?A\$?1\s*\)',
    re.IGNORECASE,
)


def flatten_sheet_title_self_refs(wb) -> int:
    """
    母版 TW / TW EE 的 A1 用 =MID(CELL(\"filename\",A1),...) 取表名。
    Excel 允许这种自引用；LuckySheet 重算时会弹
    「公式不可引用其本身的单元格，会导致计算结果不准确」。
    转换时写成静态表名即可（缓存值本就是表名）。
    """
    n = 0
    for name in wb.sheetnames:
        cell = wb[name].cell(1, 1)
        v = cell.value
        if isinstance(v, str) and v.startswith("=") and _SHEET_TITLE_SELF_REF_RE.search(v):
            cell.value = name
            n += 1
    return n


def count_template_employee_slots(ws: Worksheet, data_start_row: int, marker_col: int = 2) -> int:
    n = 0
    for row in range(data_start_row, data_start_row + MAX_EMPLOYEES):
        if ws.cell(row, marker_col).value is not None:
            n += 1
        else:
            break
    return max(n, 1)


def fit_formula_sheets(wb, employee_count: int, client_id_prefix: str) -> None:
    """母版已预置 N 人公式行：人数≤N 时只清除多余行；人数>N 时才复制扩展。"""
    tw = wb[TW_SHEET]
    ee = wb[TW_EE_SHEET]
    # TW!B 有员工标记；TW EE 的 D(EE Code) 母版常为空，改用 E(姓名公式) 计数
    tw_slots = count_template_employee_slots(tw, TW_DATA_START_ROW, marker_col=2)
    ee_slots = count_template_employee_slots(ee, TW_EE_DATA_START_ROW, marker_col=5)

    for i in range(employee_count, tw_slots):
        clear_employee_row(tw, TW_DATA_START_ROW + i)

    for i in range(employee_count, ee_slots):
        clear_employee_row(ee, TW_EE_DATA_START_ROW + i)

    if employee_count > tw_slots:
        src_row = TW_DATA_START_ROW + tw_slots - 1
        src_tw_l = TW_L_DATA_START_ROW + tw_slots - 1
        for i in range(tw_slots, employee_count):
            dst_row = TW_DATA_START_ROW + i
            dst_tw_l = TW_L_DATA_START_ROW + i
            ee_row = TW_EE_DATA_START_ROW + i
            copy_row_formulas(tw, src_row, dst_row, src_tw_l, dst_tw_l)
            for col in range(1, (tw.max_column or 0) + 1):
                cell = tw.cell(dst_row, col)
                if cell.data_type == "f" and isinstance(cell.value, str):
                    cell.value = re.sub(
                        rf"'TW EE'!([A-Z]+){ee_row - 1}(?!\d)",
                        lambda m: f"'TW EE'!{m.group(1)}{ee_row}",
                        cell.value,
                    )

    if employee_count > ee_slots:
        src_row = TW_EE_DATA_START_ROW + ee_slots - 1
        src_tw_l = TW_L_DATA_START_ROW + ee_slots - 1
        for i in range(ee_slots, employee_count):
            dst_row = TW_EE_DATA_START_ROW + i
            dst_tw_l = TW_L_DATA_START_ROW + i
            tw_row = TW_DATA_START_ROW + i
            copy_row_formulas(ee, src_row, dst_row, src_tw_l, dst_tw_l)
            # EE Code 由 apply_tw_ee_codes 按主数据匹配写入，扩行时先留空
            ee.cell(dst_row, 4).value = None
            for col in range(1, (ee.max_column or 0) + 1):
                cell = ee.cell(dst_row, col)
                if cell.data_type == "f" and isinstance(cell.value, str):
                    cell.value = re.sub(
                        rf"TW!([A-Z]+){tw_row - 1}(?!\d)",
                        lambda m: f"TW!{m.group(1)}{tw_row}",
                        cell.value,
                    )


def copy_row_formulas(
    ws: Worksheet,
    from_row: int,
    to_row: int,
    tw_l_from: int,
    tw_l_to: int,
) -> None:
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        src = ws.cell(from_row, col)
        dst = ws.cell(to_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
        if src.data_type == "f" and isinstance(src.value, str):
            dst.value = shift_tw_formula(src.value, from_row, to_row, tw_l_from, tw_l_to)
        elif src.value is not None and src.data_type != "f":
            dst.value = copy(src.value)


def parse_source(source_path: Path) -> dict[str, Any]:
    wb = load_workbook(source_path, data_only=True, read_only=True)
    if PC_SHEET not in wb.sheetnames:
        names = wb.sheetnames
        wb.close()
        raise ValueError(f"未找到 sheet「{PC_SHEET}」，现有: {names}")

    ws = wb[PC_SHEET]
    header_row = find_pc_header_row(ws)
    employees = read_pc_employees(ws, header_row)
    meta = read_summary_meta(wb)
    wb.close()
    return {"employees": employees, "meta": meta, "pc_header_row": header_row}


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if not template_path.is_file():
        raise FileNotFoundError(f"母版不存在: {template_path}")
    if not source_path.is_file():
        raise FileNotFoundError(f"原始账单不存在: {source_path}")

    parsed = parse_source(source_path)
    employees = parsed["employees"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, rich_text=True)
    for name in (TW_L_SHEET, TW_SHEET, TW_EE_SHEET, PN_SHEET):
        if name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"母版缺少 sheet「{name}」，现有: {wb.sheetnames}")

    applied_pn: PnMeta | None = None
    if pn_meta is not None:
        applied_pn = apply_pn_meta(
            wb,
            pn_meta,
            registry_dir=registry_dir or output_path.parent,
            reserve_invoice_number=True,
        )

    write_tw_l(wb[TW_L_SHEET], employees, parsed["meta"])
    client_prefix = (
        applied_pn.customer_id
        if applied_pn
        else norm(wb[PN_SHEET]["B9"].value) or "CUS1516"
    )
    fit_formula_sheets(wb, len(employees), client_prefix)
    apply_tw_business_tax(wb[TW_SHEET], len(employees))
    ee_warnings = apply_tw_ee_codes(wb[TW_EE_SHEET], employees, employee_directory)
    pn_layout = fit_pn_employees(wb[PN_SHEET], len(employees))

    rates = fetch_usd_rates()
    fx_rate = get_tw_pn_fx_rate(rates)
    fx_row = pn_layout.get("fx_row") or _find_pn_fx_row(wb[PN_SHEET])
    wb[PN_SHEET].cell(fx_row, 2).value = fx_rate
    pn_layout["fx_row"] = fx_row
    fix_pn_illegal_concat_formulas(wb[PN_SHEET])
    flatten_sheet_title_self_refs(wb)
    pn_cache, pn_amounts = compute_pn_amount_cache(wb, len(employees), pn_layout)

    wb.save(output_path)
    wb.close()
    # openpyxl rich_text 会把标题写成 inlineStr，LuckyExcel 预览只读首个 <t>；迁回 sharedStrings
    migrate_inlinestr_richtext_to_shared_strings(output_path)
    # 主题填充落地 RGB + 补 applyFill，避免预览/Excel 丢浅蓝底
    materialize_theme_fills(output_path)
    # 保留 PN 金额公式，只补缓存值：预览能显示，Excel 改 TW-L/汇率仍可重算
    inject_formula_cached_values(output_path, pn_cache)

    return {
        "employee_count": len(employees),
        "employee_names": [
            e.get("CN Name") or e.get("EN Name") for e in employees
        ],
        "company_name": parsed["meta"].get("company_name"),
        "period": (parsed["meta"].get("period_from"), parsed["meta"].get("period_to")),
        "fx_rate": fx_rate,
        "fx_source": "api:TWD",
        "pn_amounts": pn_amounts,
        "output": str(output_path),
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
        "warnings": ee_warnings,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="台湾 People Search 账单 → PN N-C 自动转换")
    parser.add_argument("source", type=Path, help="原始供应商 Excel 路径")
    parser.add_argument("-o", "--output", type=Path, help="输出 PN 路径")
    parser.add_argument("-t", "--template", type=Path, default=DEFAULT_TEMPLATE, help="PN 母版路径")
    args = parser.parse_args(argv)

    source = args.source.resolve()
    output = args.output.resolve() if args.output else source.parent / f"PN_auto_{source.stem}.xlsx"

    try:
        result = convert(source, output, args.template.resolve())
    except Exception as exc:
        print(f"转换失败: {exc}", file=sys.stderr)
        return 1

    print("转换完成")
    print(f"  输出: {result['output']}")
    print(f"  公司: {result['company_name']}")
    print(f"  账期: {result['period'][0]} ~ {result['period'][1]}")
    print(f"  员工: {result['employee_count']} 人 → {result['employee_names']}")
    fx_row = (result.get("pn_amounts") or {}).get("layout", {}).get("fx_row", 31)
    print(f"  汇率 PN!B{fx_row}: {result['fx_rate']} ({result.get('fx_source')})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
