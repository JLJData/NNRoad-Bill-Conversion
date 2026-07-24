# -*- coding: utf-8 -*-
"""分析 People Search + Coral Sea 账单结构，输出 structure.json。"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[2]
PROFILE_DIR = Path(__file__).resolve().parent
SRC = ROOT / r"账单\now\Taiwan\Payroll Info_NNRoad 202603_Coral Sea -供应商账单.xlsx"
OUT = ROOT / r"账单\now\Taiwan\PN_CoralSea_N-C 20260311.xlsx"
OUT_FILE = PROFILE_DIR / "structure.json"


def norm(v):
    if v is None:
        return None
    return str(v).replace("\uFEFF", "").strip()


def find_header_row(ws, keywords, max_scan=30):
    for row in range(1, max_scan + 1):
        vals = [norm(ws.cell(row, c).value) for c in range(1, (ws.max_column or 0) + 1)]
        hits = sum(1 for v in vals if v and any(k.lower() in v.lower() for k in keywords))
        if hits >= 2:
            return row, vals
    return None, []


def scan_headers(ws, header_row, max_col=None):
    mc = max_col or (ws.max_column or 0)
    headers = []
    for c in range(1, mc + 1):
        h = norm(ws.cell(header_row, c).value)
        if h:
            headers.append({"col": get_column_letter(c), "col_idx": c, "header": h})
    return headers


def scan_data_rows(ws, header_row, name_headers, max_rows=20):
    hdr_map = {}
    for c in range(1, (ws.max_column or 0) + 1):
        h = norm(ws.cell(header_row, c).value)
        if h and h not in hdr_map:
            hdr_map[h] = c
    name_col = None
    for nh in name_headers:
        if nh in hdr_map:
            name_col = hdr_map[nh]
            break
    rows = []
    for row in range(header_row + 1, header_row + 1 + max_rows):
        if name_col:
            name = ws.cell(row, name_col).value
            if name is None or norm(name) == "":
                continue
        preview = {}
        for h, c in list(hdr_map.items())[:15]:
            v = ws.cell(row, c).value
            if v is not None:
                preview[h] = v
        rows.append({"row": row, "preview": preview})
    return rows


def scan_row_formulas(ws, row, max_col=80):
    cells = []
    for c in range(1, min(max_col, (ws.max_column or 0) + 1)):
        cell = ws.cell(row, c)
        v = cell.value
        if v is not None:
            item = {"col": get_column_letter(c), "v": str(v)[:200]}
            if cell.data_type == "f":
                item["formula"] = str(v)
            cells.append(item)
    return cells


def scan_meta_cells(ws, rows=10, cols=10):
    meta = []
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            v = ws.cell(r, c).value
            if v is not None and norm(v):
                meta.append({"cell": f"{get_column_letter(c)}{r}", "value": v})
    return meta


def diff_tw_l(src_hdrs, out_hdrs):
    src_map = {h["header"]: h for h in src_hdrs}
    out_map = {h["header"]: h for h in out_hdrs}
    all_keys = sorted(set(src_map) | set(out_map))
    mapping = []
    for k in all_keys:
        s = src_map.get(k)
        o = out_map.get(k)
        mapping.append({
            "header": k,
            "in_src": s is not None,
            "in_out": o is not None,
            "src_col": s["col"] if s else None,
            "out_col": o["col"] if o else None,
            "same_name": s is not None and o is not None,
        })
    return mapping


def main():
    r = {}
    wb_src = load_workbook(SRC, data_only=True)
    wb_out = load_workbook(OUT, data_only=True)
    wb_src_f = load_workbook(SRC, data_only=False)
    wb_out_f = load_workbook(OUT, data_only=False)

    r["src_sheets"] = wb_src.sheetnames
    r["out_sheets"] = wb_out.sheetnames

    # Payroll calculation sheet
    pc_name = None
    for sn in wb_src.sheetnames:
        if "payroll" in sn.lower() and "calc" in sn.lower():
            pc_name = sn
            break
    r["payroll_calc_sheet"] = pc_name

    if pc_name:
        ws_pc = wb_src[pc_name]
        hr, _ = find_header_row(ws_pc, ["Name", "Employee", "Salary", "Basic"])
        r["pc_header_row"] = hr
        r["pc_headers"] = scan_headers(ws_pc, hr) if hr else []
        r["pc_data_rows"] = scan_data_rows(ws_pc, hr, ["Name of Employee", "EE Name", "Name", "Employee Name"]) if hr else []
        r["pc_meta"] = scan_meta_cells(ws_pc)

    # TW-L sheet in source and output
    tw_l_name = "TW-L" if "TW-L" in wb_src.sheetnames else None
    r["tw_l_sheet"] = tw_l_name
    if tw_l_name:
        ws = wb_src[tw_l_name]
        hr, _ = find_header_row(ws, ["Name", "Employee", "Salary"])
        r["src_tw_l_header_row"] = hr
        r["src_tw_l_headers"] = scan_headers(ws, hr) if hr else []
        r["src_tw_l_data"] = scan_data_rows(ws, hr, ["Name of Employee", "EE Name", "Name"]) if hr else []
        r["src_tw_l_meta"] = scan_meta_cells(ws, 6, 8)

    ws_out_tw_l = wb_out["TW-L"] if "TW-L" in wb_out.sheetnames else None
    if ws_out_tw_l:
        hr = 7  # try HK pattern
        hdr_test = scan_headers(ws_out_tw_l, hr)
        if len(hdr_test) < 3:
            hr, _ = find_header_row(ws_out_tw_l, ["Name", "Employee", "Salary"])
        r["out_tw_l_header_row"] = hr
        r["out_tw_l_headers"] = scan_headers(ws_out_tw_l, hr)
        r["out_tw_l_data"] = scan_data_rows(ws_out_tw_l, hr, ["Name of Employee", "EE Name", "Name"])
        r["out_tw_l_meta"] = scan_meta_cells(ws_out_tw_l, 6, 8)

    if tw_l_name and ws_out_tw_l:
        r["tw_l_header_mapping"] = diff_tw_l(r["src_tw_l_headers"], r["out_tw_l_headers"])

    # Compare PC headers vs TW-L headers (source)
    if pc_name and tw_l_name:
        pc_hdrs = {h["header"] for h in r.get("pc_headers", [])}
        tw_hdrs = {h["header"] for h in r.get("src_tw_l_headers", [])}
        r["pc_only_headers"] = sorted(pc_hdrs - tw_hdrs)
        r["tw_l_only_headers"] = sorted(tw_hdrs - pc_hdrs)
        r["common_headers"] = sorted(pc_hdrs & tw_hdrs)

    # TW sheet formulas
    if "TW" in wb_out_f.sheetnames:
        ws_tw = wb_out_f["TW"]
        hr_tw, _ = find_header_row(ws_tw, ["Name", "Employee"])
        r["tw_header_row"] = hr_tw
        r["tw_headers"] = scan_headers(ws_tw, hr_tw) if hr_tw else []
        data_row = (hr_tw or 7) + 1
        r["tw_data_row_formulas"] = scan_row_formulas(ws_tw, data_row)

    # TW EE sheet
    if "TW EE" in wb_out_f.sheetnames:
        ws_ee = wb_out_f["TW EE"]
        hr_ee, _ = find_header_row(ws_ee, ["Name", "Employee"])
        r["tw_ee_header_row"] = hr_ee
        data_row_ee = (hr_ee or 9) + 1
        r["tw_ee_data_row_formulas"] = scan_row_formulas(ws_ee, data_row_ee)

    # PN key cells
    if "PN" in wb_out_f.sheetnames:
        for addr in ["B8", "B28", "B29", "C49"]:
            r[f"out_pn_{addr}"] = wb_out_f["PN"][addr].value if addr in ["B8", "B28", "B29"] or True else None
        try:
            pn = wb_out_f["PN"]
            pn_keys = []
            for row in range(1, 35):
                for col in range(1, 6):
                    v = pn.cell(row, col).value
                    if v and isinstance(v, str) and ("rate" in v.lower() or "fx" in v.lower() or "汇率" in v):
                        pn_keys.append({"cell": f"{get_column_letter(col)}{row}", "label": v, "value": pn.cell(row, col+1).value})
            r["pn_fx_cells"] = pn_keys
        except Exception as e:
            r["pn_fx_error"] = str(e)

    # Value diff TW-L src vs out (data_only)
    if tw_l_name and ws_out_tw_l:
        diffs = []
        src_map = {h["header"]: h for h in r.get("src_tw_l_headers", [])}
        out_map = {h["header"]: h for h in r.get("out_tw_l_headers", [])}
        ws_s = wb_src[tw_l_name]
        ws_o = ws_out_tw_l
        hr_s = r["src_tw_l_header_row"]
        hr_o = r["out_tw_l_header_row"]
        for row_idx in range(5):
            sr = hr_s + 1 + row_idx
            orow = hr_o + 1 + row_idx
            for hdr in set(src_map) & set(out_map):
                sc = src_map[hdr]["col_idx"]
                oc = out_map[hdr]["col_idx"]
                vs = ws_s.cell(sr, sc).value
                vo = ws_o.cell(orow, oc).value
                if vs != vo and (vs is not None or vo is not None):
                    diffs.append({"header": hdr, "row": row_idx + 1, "src": vs, "out": vo})
        r["tw_l_value_diffs"] = diffs[:50]

    # Compare PC data vs OUT TW-L data
    if pc_name and ws_out_tw_l:
        pc_ws = wb_src[pc_name]
        hr_pc = r["pc_header_row"]
        hr_out = r["out_tw_l_header_row"]
        pc_map = {h["header"]: h["col_idx"] for h in r.get("pc_headers", [])}
        out_map = {h["header"]: h["col_idx"] for h in r.get("out_tw_l_headers", [])}
        pc_to_out = []
        for hdr in out_map:
            if hdr in pc_map:
                pc_to_out.append({"out_header": hdr, "pc_col": get_column_letter(pc_map[hdr]), "out_col": get_column_letter(out_map[hdr]), "match": "same_name"})
        r["pc_to_out_same_name"] = pc_to_out

        # renamed headers: compare first employee row values
        if r.get("pc_data_rows") and r.get("out_tw_l_data"):
            pc_row = r["pc_data_rows"][0]["row"]
            out_row = r["out_tw_l_data"][0]["row"]
            rename_candidates = []
            for oh, oc in out_map.items():
                if oh in pc_map:
                    continue
                out_val = ws_out_tw_l.cell(out_row, oc).value
                if out_val is None:
                    continue
                for ph, pc in pc_map.items():
                    pv = pc_ws.cell(pc_row, pc).value
                    if pv == out_val and pv is not None:
                        rename_candidates.append({"out_header": oh, "pc_header": ph, "value": pv})
            r["rename_candidates"] = rename_candidates

    wb_src.close()
    wb_out.close()
    wb_src_f.close()
    wb_out_f.close()

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(r, f, ensure_ascii=False, indent=2, default=str)

    print("written", OUT_FILE)
    print("src sheets:", r["src_sheets"])
    print("out sheets:", r["out_sheets"])
    print("PC sheet:", r.get("payroll_calc_sheet"), "header row:", r.get("pc_header_row"))
    print("PC employees:", len(r.get("pc_data_rows", [])))
    print("TW-L src header row:", r.get("src_tw_l_header_row"), "out:", r.get("out_tw_l_header_row"))
    print("TW-L src headers:", len(r.get("src_tw_l_headers", [])), "out:", len(r.get("out_tw_l_headers", [])))
    print("rename candidates:", r.get("rename_candidates", [])[:10])
    print("common headers PC/TW-L:", len(r.get("common_headers", [])))


if __name__ == "__main__":
    main()
