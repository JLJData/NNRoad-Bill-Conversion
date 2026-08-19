# -*- coding: utf-8 -*-
"""
UK-L 竖表源账单 → UK PN（引擎 uk_payroll_calc）

用法:
  python -m profiles.uk_payroll_calc.convert <源.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

源账单: sheet「UK-L」标签→金额（B 列）；可由 PDF ingest（eor_uk / topsource_uk）产出。
列名别名按 pdfProfileId overlay（convert_mapping.PROFILE_MAPPING_OVERLAYS）。
默认母版: templates/uk/template.xlsx
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from convert_mapping import find_sheet_name, resolve_convert_mapping
from fx_rate import fetch_usd_rates, get_uk_gbp_per_usd
from pn_meta import PnMeta, apply_pn_meta
from region_templates import get_region_template
from xlsx_postprocess import postprocess_converted_xlsx
from bill_convert.uk_layout import (
    ensure_uk_l_count,
    expand_uk_employee_rows,
    list_uk_l_sheets,
)

DEFAULT_TEMPLATE = get_region_template("UK")

UK_L_SHEET = "UK-L"
UK_SHEET = "UK"
UK_EE_SHEET = "UK EE"
PN_SHEET = "PN"

# 写入母版时用的标签 → 单元格（金额在 B 列）
_LABEL_CELLS: dict[str, str] = {
    "Gross Salary": "B7",
    "Holiday Pay": "B8",
    "Car Allowance": "B9",
    "ER' NIC": "B10",
    "ER' Pension (Auto Enrolment)": "B11",
    "PAYE (Estimated)": "B14",
    "EE'NIC": "B15",
    "EE' Pension (Auto Enrolment)": "B16",
    "App Levy": "B22",
    "Payment Fees": "B26",
}

_ACTIVE_MAPPING: dict[str, Any] | None = None
_TITLE_RE = re.compile(
    r"^\s*(.+?)\s*-?\s*Salary Calculation\s+for\s+FY\s+(.+?)\s*$",
    re.IGNORECASE,
)


def _active_mapping() -> dict[str, Any]:
    return (
        _ACTIVE_MAPPING
        if isinstance(_ACTIVE_MAPPING, dict)
        else resolve_convert_mapping("uk_payroll_calc", None)
    )


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("£", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _apply_uk_label_rename(label: str, rename: dict[str, Any] | None) -> str:
    if not label or not isinstance(rename, dict) or not rename:
        return label
    if label in rename:
        return str(rename[label])
    low = {str(k).strip().lower(): v for k, v in rename.items() if k is not None}
    hit = low.get(label.lower())
    return str(hit) if hit else label


def read_uk_l(ws: Worksheet) -> dict[str, Any]:
    """读取竖表：标签→金额，以及标题中的员工名 / FY。"""
    rename = _active_mapping().get("columnRename")
    amounts: dict[str, float] = {}
    for row in range(1, (ws.max_row or 40) + 1):
        label = _norm(ws.cell(row, 1).value)
        if not label:
            continue
        num = _as_float(ws.cell(row, 2).value)
        if num is not None:
            amounts[_apply_uk_label_rename(label, rename)] = num

    employee_name = None
    fy = None
    title = _norm(ws["A3"].value)
    m = _TITLE_RE.match(title) if title else None
    if m:
        employee_name = m.group(1).strip()
        fy = m.group(2).strip()

    fx = _as_float(ws["D24"].value)
    return {
        "amounts": amounts,
        "employee_name": employee_name,
        "fy": fy,
        "title": title,
        "fx_rate": fx,
    }


def parse_source(source_path: Path) -> list[dict[str, Any]]:
    """读取全部 UK-L / UK-L (2)… 员工竖表。"""
    mapping = _active_mapping()
    src_spec = (
        mapping.get("sourceEmployeeSheet")
        if isinstance(mapping.get("sourceEmployeeSheet"), dict)
        else {}
    )
    wb = load_workbook(source_path, data_only=False)
    names = list_uk_l_sheets(list(wb.sheetnames))
    if not names:
        name = find_sheet_name(list(wb.sheetnames), src_spec)
        names = [name] if name else []
    if not names:
        existing = wb.sheetnames
        wb.close()
        raise ValueError(f"未找到 sheet「{src_spec.get('sheet') or UK_L_SHEET}」，现有: {existing}")

    employees: list[dict[str, Any]] = []
    for i, sheet_name in enumerate(names):
        parsed = read_uk_l(wb[sheet_name])
        parsed["sheet_name"] = sheet_name
        if not parsed.get("employee_name") and UK_SHEET in wb.sheetnames:
            parsed["employee_name"] = _norm(wb[UK_SHEET].cell(9 + i, 2).value) or None
        employees.append(parsed)
    wb.close()
    return employees


def write_uk_l(ws: Worksheet, parsed: dict[str, Any]) -> None:
    name = parsed.get("employee_name") or "Employee"
    fy = parsed.get("fy") or "YY-YY"
    # TopSource 样例无连字符；EOR 有「-」
    ws["A3"] = f"{name} Salary Calculation for FY {fy}"
    amounts: dict[str, float] = parsed.get("amounts") or {}
    # Holiday / Holiday Pay 兼容
    if "Holiday Pay" not in amounts and "Holiday" in amounts:
        amounts = {**amounts, "Holiday Pay": amounts["Holiday"]}
    for label, addr in _LABEL_CELLS.items():
        if label in amounts:
            ws[addr] = float(amounts[label])


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
    convert_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    global _ACTIVE_MAPPING
    _ACTIVE_MAPPING = resolve_convert_mapping("uk_payroll_calc", convert_mapping)
    try:
        return _convert_impl(
            source_path,
            output_path,
            template_path,
            pn_meta=pn_meta,
            registry_dir=registry_dir,
        )
    finally:
        _ACTIVE_MAPPING = None


def _convert_impl(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    registry_dir: Path | None = None,
) -> dict[str, Any]:
    if not template_path.is_file():
        raise FileNotFoundError(f"母版不存在: {template_path}")
    if not source_path.is_file():
        raise FileNotFoundError(f"原始账单不存在: {source_path}")

    employees = parse_source(source_path)
    if not employees:
        raise ValueError("未解析到任何 UK-L 员工")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, rich_text=True)
    for name in (UK_L_SHEET, UK_SHEET, PN_SHEET):
        if name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"母版缺少 sheet「{name}」，现有: {wb.sheetnames}")

    applied_pn: PnMeta | None = None
    if pn_meta is not None:
        applied_pn = apply_pn_meta(
            wb,
            pn_meta,
            registry_dir=registry_dir or output_path.parent,
            reserve_invoice_number=True,
        )

    n = len(employees)
    sheet_names = ensure_uk_l_count(wb, n)
    expand_uk_employee_rows(wb, n, sheet_names)

    warnings: list[str] = []
    emp_names: list[str] = []
    for i, parsed in enumerate(employees):
        write_uk_l(wb[sheet_names[i]], parsed)
        emp_name = parsed.get("employee_name") or f"Employee {i + 1}"
        emp_names.append(emp_name)
        wb[UK_SHEET].cell(9 + i, 2).value = emp_name
        amounts = parsed.get("amounts") or {}
        if not amounts.get("Gross Salary"):
            warnings.append(
                f"{sheet_names[i]}（{emp_name}）Gross Salary 为空/0，请按截图人工补齐 GBP 明细"
            )

    fx_rate = None
    fx_source = None
    fact_store_updates: dict[str, Any] = {}
    mapping = _ACTIVE_MAPPING or resolve_convert_mapping("uk_payroll_calc", convert_mapping)
    from fx_policy import (
        UK_VENDOR_BILL_FX_FACT,
        api_fx_for_currency,
        build_fx_fact_update,
        fx_policy,
        read_shared_fx,
        resolve_vendor_currency,
    )

    policy = fx_policy(mapping)
    mode = str(policy.get("mode") or "api").strip().lower()
    persist_key = str(policy.get("persistFactKey") or "").strip() or None
    fact_key = str(policy.get("factKey") or persist_key or UK_VENDOR_BILL_FX_FACT).strip()
    currency = resolve_vendor_currency(mapping, str(policy.get("defaultCurrency") or "GBP")) or "GBP"
    adjustment = float(policy.get("adjustment") or 1.0)
    invert = bool(policy.get("invert", True))

    # 1) 共享 fact / 本批注入（EOR 读 TopSource 同源）
    shared = read_shared_fx(mapping, fact_key)
    # 2) 源表已有 D24（TopSource vendor-to-source 已写入）
    src_fx = None
    if employees:
        src_fx = _as_float(employees[0].get("fx_rate"))
    if src_fx is None and sheet_names:
        try:
            src_fx = _as_float(wb[sheet_names[0]]["D24"].value)
        except Exception:
            src_fx = None

    try:
        if mode == "shared_fact" and shared is not None:
            fx_rate = float(shared)
            fx_source = f"shared_fact:{fact_key}"
        elif mode in ("vendor_bill", "shared_fact") and src_fx is not None and src_fx > 0:
            fx_rate = float(src_fx)
            fx_source = "source:UK-L!D24"
        elif shared is not None and mode == "vendor_bill":
            fx_rate = float(shared)
            fx_source = f"shared_fact:{fact_key}"
        elif mode == "none":
            fx_rate = None
            fx_source = "none"
        else:
            # api / fallback
            fx_rate = api_fx_for_currency(currency, adjustment=adjustment, invert=invert)
            fx_source = f"api:{'1/' if invert else ''}{currency}"
        if fx_rate is not None:
            for name in sheet_names:
                wb[name]["D24"] = fx_rate
            if persist_key and mode == "vendor_bill":
                fact_store_updates.update(
                    build_fx_fact_update(persist_key, fx_rate, source=fx_source)
                )
    except Exception as exc:
        if src_fx is not None:
            for name in sheet_names:
                wb[name]["D24"] = float(src_fx)
            fx_rate = float(src_fx)
            fx_source = "source:UK-L!D24"
            warnings.append(f"在线汇率失败，沿用源表 D24={src_fx}: {exc}")
        else:
            warnings.append(f"写入 UK-L!D24 汇率失败: {exc}")

    wb.save(output_path)
    wb.close()
    postprocess_converted_xlsx(output_path)

    return {
        "employee_count": n,
        "employee_names": emp_names,
        "fx_rate": fx_rate,
        "fx_source": fx_source,
        "output": str(output_path),
        "pn_meta": applied_pn.to_dict() if applied_pn else None,
        "warnings": warnings,
        "uk_l_sheets": sheet_names,
        "fact_store_updates": fact_store_updates,
    }

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="UK-L → UK PN")
    parser.add_argument("source", type=Path)
    parser.add_argument("-o", "--output", type=Path)
    parser.add_argument("-t", "--template", type=Path, default=DEFAULT_TEMPLATE)
    args = parser.parse_args(argv)

    src = args.source.resolve()
    out = (
        args.output.resolve()
        if args.output
        else src.parent / f"UK_PN_{src.stem}.xlsx"
    )
    try:
        result = convert(src, out, Path(args.template).resolve())
    except Exception as exc:
        print(f"失败: {exc}", file=sys.stderr)
        return 1
    print("完成")
    print(f"  输出: {result['output']}")
    print(f"  员工: {result.get('employee_names')}")
    print(f"  FX: {result.get('fx_rate')} ({result.get('fx_source')})")
    for w in result.get("warnings") or []:
        print(f"  ! {w}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
