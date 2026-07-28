# -*- coding: utf-8 -*-
"""分析 HROne + Hermetic 各 Sheet 明细与公式，输出 detail.json。"""
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
PROFILE_DIR = Path(__file__).resolve().parent
OUT = ROOT / r"账单\now\China\PN_Hermetic Solutions Group_N-C 20260318.xlsx"
SRC = ROOT / r"账单\now\China\HROne HK Payment Notice-NNRoad-Hermetic Solutions-202603C.xlsx"
from region_templates import get_region_template

TEMPLATE = get_region_template("China")
OUT_FILE = PROFILE_DIR / "detail.json"


def scan_china_l(ws):
    """China-L: row1=field names, row2=values (horizontal KV layout)."""
    pairs = []
    for c in range(1, (ws.max_column or 0) + 1):
        k = ws.cell(1, c).value
        v = ws.cell(2, c).value
        if k is not None:
            pairs.append({"col": get_column_letter(c), "field": str(k).strip(), "value": v})
    return pairs


def scan_china_sheet_full(ws):
    """Scan China sheet - find all headers and formulas in rows 1-15."""
    result = {"rows": []}
    for r in range(1, 16):
        cells = []
        for c in range(1, min(80, (ws.max_column or 0) + 1)):
            cell = ws.cell(r, c)
            v = cell.value
            if v is not None and str(v).strip():
                item = {"col": get_column_letter(c), "v": str(v)[:120]}
                if cell.data_type == "f":
                    item["formula"] = str(v)
                cells.append(item)
        if cells:
            result["rows"].append({"row": r, "cells": cells})
    return result


def scan_other_fee_full(ws):
    result = {}
    for addr in ["C15", "D19", "B15", "A15", "A19", "B19", "C19"]:
        c = ws[addr]
        result[addr] = {"value": c.value, "formula": c.value if c.data_type == "f" else None}
    # scan rows 12-22 cols A-E
    rows = []
    for r in range(12, 23):
        row = []
        for col in range(1, 6):
            v = ws.cell(r, col).value
            if v is not None:
                row.append({get_column_letter(col): str(v)[:80]})
        if row:
            rows.append({"row": r, "cells": row})
    result["context"] = rows
    return result


def scan_pn(ws):
    result = {"B29": ws["B29"].value}
    for addr in ["B8", "B29", "C49"]:
        c = ws[addr]
        result[addr] = {"value": c.value, "formula": c.value if c.data_type == "f" else None}
    return result


def scan_china_ee_full(ws):
    rows = []
    for r in range(1, 25):
        cells = []
        for c in range(1, 30):
            cell = ws.cell(r, c)
            v = cell.value
            if v is not None:
                item = {"col": get_column_letter(c), "v": str(v)[:100]}
                if cell.data_type == "f":
                    item["formula"] = str(v)
                cells.append(item)
        if cells:
            rows.append({"row": r, "cells": cells})
    return rows


def main():
    r = {}

    wb = load_workbook(SRC, data_only=False)
    r["other_fee"] = scan_other_fee_full(wb["Other Fee"])
    r["payment_notice"] = scan_pn(wb["S-Payment Notice"])
    wb.close()

    wb = load_workbook(SRC, data_only=True)
    r["other_fee_values"] = {"C15": wb["Other Fee"]["C15"].value, "D19": wb["Other Fee"]["D19"].value}
    r["payment_notice_C49_value"] = wb["S-Payment Notice"]["C49"].value
    wb.close()

    wb = load_workbook(OUT, data_only=False)
    r["china_l"] = scan_china_l(wb["China-L"])
    r["china"] = scan_china_sheet_full(wb["China"])
    r["pn"] = scan_pn(wb["PN"])
    r["china_ee"] = scan_china_ee_full(wb["China EE"])
    wb.close()

    wb = load_workbook(TEMPLATE, data_only=False)
    r["template_sheets"] = wb.sheetnames
    if "China-L" in wb.sheetnames or any("China" in s for s in wb.sheetnames):
        for s in wb.sheetnames:
            if "China-L" in s:
                r["template_china_l_row1_sample"] = [wb[s].cell(1, c).value for c in range(1, 11)]
    wb.close()

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(r, f, ensure_ascii=False, indent=2, default=str)
    print("ok")


if __name__ == "__main__":
    main()
