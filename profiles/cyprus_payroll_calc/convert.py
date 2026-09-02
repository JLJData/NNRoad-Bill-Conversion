# -*- coding: utf-8 -*-
"""
Cyprus-L → Cyprus PN（引擎 cyprus_payroll_calc）

用法:
  python -m profiles.cyprus_payroll_calc.convert <源.xlsx> [-o 输出.xlsx] [-t 母版.xlsx]

源账单: sheet「Cyprus-L」（可由 at_technical_cyprus ingest 产出）。
默认母版: templates/cyprus/template.xlsx

原则：PN / Cyprus / Cyprus EE 以母版公式为准；只写 Cyprus-L 数据与必要元数据。
Recurring Fee：mapping.cyprusRecurringFee 有值才写；格子见 mapping.fixedValueWrites。
"""
from __future__ import annotations

import argparse
import calendar
import re
import shutil
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from bill_convert.fixed_value_writes import apply_fixed_value_writes
from bill_convert.convert_checks import merge_warnings, parse_cell_ref, resolve_col_with_fallback, sanity_check_convert_result
from bill_convert.headers import build_header_map
from bill_convert.formula_copy import shift_row_formula
from bill_convert.formula_layout import sort_employees_by_code
from convert_mapping import find_sheet_name, resolve_convert_mapping
from fx_rate import fetch_usd_rates, get_cyprus_pn_fx_rate
from pn_meta import PnMeta, apply_pn_meta
from profiles.tw_payroll_calc.convert import match_ee_code
from region_templates import get_region_template
from xlsx_convert_utils import coerce_datetime_for_excel, is_date_column_header
from xlsx_luckysheet_compat import apply_luckysheet_compat
from xlsx_postprocess import postprocess_converted_xlsx

DEFAULT_TEMPLATE = get_region_template("Cyprus")

CYPRUS_L_SHEET = "Cyprus-L"
CYPRUS_SHEET = "Cyprus"
CYPRUS_EE_SHEET = "Cyprus EE"
PN_SHEET = "PN"

CYPRUS_L_HEADER_ROW = 7
CYPRUS_L_DATA_START = 8
CYPRUS_L_PERIOD_ROW = 2
CYPRUS_DATA_START = 9
CYPRUS_EE_DATA_START = 10
MAX_EMPLOYEES = 20
_DATE_FMT = "yyyy/m/d"

# Cyprus-L 固定列（样例母版）
COL_EE_CODE = 1
COL_NAME = 2
COL_BASE = 3
COL_ER_CONTRIB = 12
COL_LIABILITY = 13
COL_EE_SI = 14
COL_EE_TAX = 15
COL_EE_NHS = 16
COL_EXPENSE = 10

# 表头优先；找不到时回退固定列（默认母版兼容）
_CYPRUS_FIELD_HEADERS: dict[str, tuple[list[str], int]] = {
    "ee_code": (["No. of EE", "EE Code", "Employee Code"], COL_EE_CODE),
    "name": (["Name of Employee", "Employee Name", "Name of EE"], COL_NAME),
    "base": (["Base salary", "Base Salary"], COL_BASE),
    "er_contrib": (["Employer's contributions", "Employers contributions"], COL_ER_CONTRIB),
    "liability": (
        [
            "Employer's & Public Liability",
            "Employers & Public Liability",
            "Employer's & Public Liabilit",  # 默认母版表头被 Excel 截断
        ],
        COL_LIABILITY,
    ),
    "ee_si": (["Employee's Social Insurance", "Employees Social Insurance"], COL_EE_SI),
    "ee_tax": (["Employee's tax", "Employee tax"], COL_EE_TAX),
    "ee_nhs": (["Employee - N.H.S.-SI", "Employee - N.H.S. - SI"], COL_EE_NHS),
    "expense": (["Expense Reimbursment", "Expense Reimbursement"], COL_EXPENSE),
}


def _cyprus_meta_cells() -> tuple[tuple[int, int], tuple[int, int]]:
    """账期起止格：mapping.metaCells.periodFrom / periodTo，默认 C2 / E2。"""
    mapping = _active_mapping()
    meta = mapping.get("metaCells") if isinstance(mapping.get("metaCells"), dict) else {}
    period_from = parse_cell_ref(meta.get("periodFrom"), default_row=CYPRUS_L_PERIOD_ROW, default_col=3)
    period_to = parse_cell_ref(meta.get("periodTo"), default_row=CYPRUS_L_PERIOD_ROW, default_col=5)
    return period_from, period_to


def _resolve_cyprus_cols(ws, header_row: int, warnings: list[str] | None = None) -> dict[str, int]:
    header_map = build_header_map(ws, header_row)
    try:
        for k, v in _header_map(ws, header_row).items():
            header_map.setdefault(k, v)
    except Exception:
        pass
    mapping = _active_mapping()
    custom = mapping.get("fieldHeaders") if isinstance(mapping.get("fieldHeaders"), dict) else {}
    out: dict[str, int] = {}
    for field, (names, fallback) in _CYPRUS_FIELD_HEADERS.items():
        use_names = names
        if isinstance(custom.get(field), list) and custom.get(field):
            use_names = [str(x) for x in custom[field] if x]
        col = resolve_col_with_fallback(
            header_map,
            field=field,
            header_names=use_names,
            fallback=fallback,
            warnings=warnings,
            strict=(field == "name"),
        )
        if col is not None:
            out[field] = col
    return out


_ACTIVE_MAPPING: dict[str, Any] | None = None

_MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def _active_mapping() -> dict[str, Any]:
    return (
        _ACTIVE_MAPPING
        if isinstance(_ACTIVE_MAPPING, dict)
        else resolve_convert_mapping("cyprus_payroll_calc", None)
    )


def _cyprus_l_layout(*, target: bool = False) -> tuple[int, int]:
    mapping = _active_mapping()
    key = "targetL" if target else "sourceEmployeeSheet"
    spec = mapping.get(key) if isinstance(mapping.get(key), dict) else {}
    if not spec and target:
        spec = (
            mapping.get("sourceEmployeeSheet")
            if isinstance(mapping.get("sourceEmployeeSheet"), dict)
            else {}
        )
    header = int(spec.get("headerRow") or CYPRUS_L_HEADER_ROW)
    data_start = int(spec.get("dataStartRow") or CYPRUS_L_DATA_START)
    return header, data_start


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _cell_formula_text(value: Any) -> str | None:
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _set_cell_value(cell, value: Any) -> None:
    if isinstance(value, float):
        cell.value = round(value, 6) if abs(value) >= 1e-9 else 0
    else:
        cell.value = value


def _header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        h = _norm(ws.cell(header_row, col).value)
        if h and h not in out:
            out[h] = col
    return out


def _cyprus_l_formula_cols(ws: Worksheet, data_start: int) -> dict[int, str]:
    out: dict[int, str] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        text = _cell_formula_text(ws.cell(data_start, col).value)
        if text:
            out[col] = text
    return out


def parse_pay_period_label(label: Any) -> tuple[int, int] | None:
    text = _norm(label)
    if not text:
        return None
    m = re.search(
        r"(jan|january|feb|february|mar|march|apr|april|may|jun|june|jul|july|"
        r"aug|august|sep|sept|september|oct|october|nov|november|dec|december)"
        r"[\s\-']*(\d{2,4})",
        text,
        flags=re.I,
    )
    if not m:
        m2 = re.search(r"(\d{1,2})\s*/\s*(\d{4})", text)
        if m2:
            return int(m2.group(2)), int(m2.group(1))
        return None
    month = _MONTHS.get(m.group(1).lower())
    if not month:
        return None
    year = int(m.group(2))
    if year < 100:
        year += 2000
    return year, month


def period_bounds_from_label(label: Any) -> tuple[date, date] | None:
    parsed = parse_pay_period_label(label)
    if not parsed:
        return None
    year, month = parsed
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    return start, end


def looks_like_cyprus_l_workbook(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            return CYPRUS_L_SHEET in wb.sheetnames
        finally:
            wb.close()
    except Exception:
        return False


def parse_cyprus_l_employees(ws: Worksheet, warnings: list[str] | None = None) -> list[dict[str, Any]]:
    header_row, data_start = _cyprus_l_layout(target=False)
    cols = _resolve_cyprus_cols(ws, header_row, warnings)
    headers = _header_map(ws, header_row)
    (pf_row, pf_col), (pt_row, pt_col) = _cyprus_meta_cells()
    period_from = ws.cell(pf_row, pf_col).value
    period_to = ws.cell(pt_row, pt_col).value

    name_col = cols["name"]
    ee_col = cols.get("ee_code", COL_EE_CODE)
    field_pairs = [
        ("Base salary", cols.get("base", COL_BASE)),
        ("Employer's contributions", cols.get("er_contrib", COL_ER_CONTRIB)),
        ("Employer's & Public Liability", cols.get("liability", COL_LIABILITY)),
        ("Employee's Social Insurance", cols.get("ee_si", COL_EE_SI)),
        ("Employee's tax", cols.get("ee_tax", COL_EE_TAX)),
        ("Employee - N.H.S.-SI", cols.get("ee_nhs", COL_EE_NHS)),
        ("Expense Reimbursment", cols.get("expense", COL_EXPENSE)),
    ]
    used_cols = set(cols.values())

    employees: list[dict[str, Any]] = []
    max_row = max(ws.max_row or data_start, data_start)
    for row in range(data_start, max_row + 1):
        name = _norm(ws.cell(row, name_col).value)
        if not name:
            continue
        emp: dict[str, Any] = {
            "Employee Name": name,
            "Name of Employee": name,
            "_period_from": period_from,
            "_period_to": period_to,
            "From": period_from,
            "To": period_to,
        }
        ee_code = _norm(ws.cell(row, ee_col).value)
        if ee_code:
            emp["No. of EE"] = ee_code
            emp["_ee_code"] = ee_code
        for key, col in field_pairs:
            val = ws.cell(row, col).value
            if _cell_formula_text(val):
                continue
            if val is None or val == "":
                continue
            emp[key] = val
        for h, col in headers.items():
            if col in used_cols:
                continue
            val = ws.cell(row, col).value
            if _cell_formula_text(val) or val is None or val == "":
                continue
            emp[h] = val
        employees.append(emp)
    return employees



def write_cyprus_l_period(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    if not employees:
        return
    emp0 = employees[0]
    start = emp0.get("_period_from") or emp0.get("From")
    end = emp0.get("_period_to") or emp0.get("To")
    if start is None and end is None:
        label = emp0.get("_period_label") or emp0.get("Pay Period")
        bounds = period_bounds_from_label(label)
        if bounds:
            start, end = bounds
    (pf_row, pf_col), (pt_row, pt_col) = _cyprus_meta_cells()
    for (row, col), val in (((pf_row, pf_col), start), ((pt_row, pt_col), end)):
        if val is None:
            continue
        cell = ws.cell(row, col)
        dt = coerce_datetime_for_excel(val)
        if dt is not None:
            cell.value = dt
            cell.number_format = _DATE_FMT
        else:
            cell.value = val



def write_cyprus_l(ws: Worksheet, employees: list[dict[str, Any]]) -> None:
    """只覆盖数据列，母版公式列保留/扩行复制。"""
    _, data_start = _cyprus_l_layout(target=True)
    write_cyprus_l_period(ws, employees)

    n = len(employees)
    formula_by_col = _cyprus_l_formula_cols(ws, data_start)
    max_col = max(ws.max_column or 20, 20)

    for i in range(1, n):
        _copy_row_style_and_formula(
            ws,
            data_start,
            data_start + i,
            max_col=max_col,
            l_from=data_start,
            l_to=data_start + i,
        )

    last_keep = data_start + max(n, 1) - 1
    max_row = max(ws.max_row or data_start, data_start + MAX_EMPLOYEES)
    for row in range(data_start, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ == "MergedCell":
                continue
            if row <= last_keep and col in formula_by_col:
                continue
            cell.value = None

    for i in range(n):
        row = data_start + i
        for col, formula in formula_by_col.items():
            cur = ws.cell(row, col).value
            if _cell_formula_text(cur):
                continue
            ws.cell(row, col).value = shift_row_formula(
                formula,
                data_start,
                row,
                target_l_from=data_start,
                target_l_to=row,
                target_l_sheet=CYPRUS_L_SHEET,
            )

    header_row, _ = _cyprus_l_layout(target=True)
    cols = _resolve_cyprus_cols(ws, header_row, None)
    name_col = cols.get("name", COL_NAME)
    ee_col = cols.get("ee_code", COL_EE_CODE)
    field_cols = {
        "No. of EE": ee_col,
        "Name of Employee": name_col,
        "Employee Name": name_col,
        "Base salary": cols.get("base", COL_BASE),
        "Employer's contributions": cols.get("er_contrib", COL_ER_CONTRIB),
        "Employer's & Public Liability": cols.get("liability", COL_LIABILITY),
        "Employee's Social Insurance": cols.get("ee_si", COL_EE_SI),
        "Employee's tax": cols.get("ee_tax", COL_EE_TAX),
        "Employee - N.H.S.-SI": cols.get("ee_nhs", COL_EE_NHS),
        "Expense Reimbursment": cols.get("expense", COL_EXPENSE),
    }

    for idx, emp in enumerate(employees):
        row = data_start + idx
        name = _norm(emp.get("Employee Name") or emp.get("Name of Employee"))
        if name:
            ws.cell(row, name_col).value = name
        ee_code = _norm(emp.get("No. of EE") or emp.get("_ee_code"))
        if ee_code and ee_col not in formula_by_col:
            ws.cell(row, ee_col).value = ee_code
        for key, col in field_cols.items():
            if key in ("Name of Employee", "Employee Name", "No. of EE"):
                continue
            if col in formula_by_col:
                continue
            if key not in emp:
                continue
            val = emp[key]
            if val is None:
                continue
            cell = ws.cell(row, col)
            if is_date_column_header(key):
                dt = coerce_datetime_for_excel(val)
                if dt is not None:
                    cell.value = dt
                    cell.number_format = _DATE_FMT
                    continue
            _set_cell_value(cell, val)


def _copy_row_style_and_formula(
    ws: Worksheet,
    src_row: int,
    dest_row: int,
    max_col: int = 40,
    *,
    l_from: int | None = None,
    l_to: int | None = None,
) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dest = ws.cell(dest_row, c)
        if type(dest).__name__ == "MergedCell" or type(src).__name__ == "MergedCell":
            continue
        if src.has_style:
            dest.font = copy(src.font)
            dest.border = copy(src.border)
            dest.fill = copy(src.fill)
            dest.number_format = src.number_format
            dest.alignment = copy(src.alignment)
        text = _cell_formula_text(src.value)
        if text and l_from is not None and l_to is not None:
            dest.value = shift_row_formula(
                text,
                src_row,
                dest_row,
                target_l_from=l_from,
                target_l_to=l_to,
                target_l_sheet=CYPRUS_L_SHEET,
            )
        elif text:
            dest.value = shift_row_formula(text, src_row, dest_row)
        else:
            dest.value = None


def _retarget_l_refs(formula: str, l_from: int, l_to: int) -> str:
    if not (isinstance(formula, str) and formula.startswith("=")):
        return formula
    return re.sub(
        rf"('Cyprus-L'!\$?[A-Z]{{1,3}})\$?{l_from}(?!\d)",
        lambda m: f"{m.group(1)}{l_to}",
        formula,
        flags=re.I,
    )


def _retarget_ee_refs(formula: str, ee_from: int, ee_to: int) -> str:
    if not (isinstance(formula, str) and formula.startswith("=")):
        return formula
    return re.sub(
        rf"('Cyprus EE'!\$?[A-Z]{{1,3}})\$?{ee_from}(?!\d)",
        lambda m: f"{m.group(1)}{ee_to}",
        formula,
        flags=re.I,
    )


def expand_cyprus_employee_rows(wb, employee_count: int) -> None:
    n = max(int(employee_count), 1)
    _, l_data_start = _cyprus_l_layout(target=True)
    if CYPRUS_SHEET in wb.sheetnames:
        cyprus = wb[CYPRUS_SHEET]
        for i in range(1, n):
            dest = CYPRUS_DATA_START + i
            l_row = l_data_start + i
            ee_row = CYPRUS_EE_DATA_START + i
            _copy_row_style_and_formula(
                cyprus,
                CYPRUS_DATA_START,
                dest,
                max_col=40,
                l_from=l_data_start,
                l_to=l_row,
            )
            for c in range(1, 41):
                cell = cyprus.cell(dest, c)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = _retarget_l_refs(cell.value, l_data_start, l_row)
                    cell.value = _retarget_ee_refs(cell.value, CYPRUS_EE_DATA_START, ee_row)

    if CYPRUS_EE_SHEET in wb.sheetnames:
        ee = wb[CYPRUS_EE_SHEET]
        for i in range(1, n):
            dest = CYPRUS_EE_DATA_START + i
            l_row = l_data_start + i
            _copy_row_style_and_formula(
                ee,
                CYPRUS_EE_DATA_START,
                dest,
                max_col=40,
                l_from=l_data_start,
                l_to=l_row,
            )
            for c in range(1, 41):
                cell = ee.cell(dest, c)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = _retarget_l_refs(cell.value, l_data_start, l_row)


def set_recurring_fees(wb, employees: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Cyprus Recurring Fee：格子见 mapping.fixedValueWrites（convert_mapping）。"""
    return apply_fixed_value_writes(wb, employees, _active_mapping())


def _find_pn_row_by_label(ws: Worksheet, keyword: str, col: int = 1) -> int | None:
    key = keyword.lower()
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and key in v.lower():
            return row
    return None


def fit_cyprus_pn_employees(wb, employee_count: int) -> dict[str, Any]:
    fx_row = _find_pn_row_by_label(wb[PN_SHEET], "FX rate") if PN_SHEET in wb.sheetnames else None
    return {"fx_row": fx_row or 28, "employee_count": employee_count}


def _pn_customer_id(pn_meta: PnMeta | dict[str, Any] | None) -> str | None:
    if pn_meta is None:
        return None
    if isinstance(pn_meta, PnMeta):
        cid = (pn_meta.customer_id or "").strip()
        return cid or None
    if isinstance(pn_meta, dict):
        cid = str(pn_meta.get("customer_id") or pn_meta.get("customerId") or "").strip()
        return cid or None
    return None


def apply_cyprus_ee_codes(
    wb,
    employees: list[dict[str, Any]],
    *,
    employee_directory: list[dict[str, Any]] | None = None,
    pn_meta: PnMeta | dict[str, Any] | None = None,
) -> list[str]:
    """员工库匹配 EE Code → Cyprus-L!A + Cyprus EE!D；Client Code → Cyprus EE!B。"""
    warnings: list[str] = []
    client_code = _pn_customer_id(pn_meta)
    directory = list(employee_directory or [])
    _, l_data_start = _cyprus_l_layout(target=True)

    for i, emp in enumerate(employees):
        name = _norm(emp.get("Employee Name") or emp.get("Name of Employee"))
        code, warn = match_ee_code([name] if name else [], directory)
        if code:
            emp["No. of EE"] = code
            emp["_ee_code"] = code
            if CYPRUS_L_SHEET in wb.sheetnames:
                wb[CYPRUS_L_SHEET].cell(l_data_start + i, COL_EE_CODE).value = code
        if warn:
            warnings.append(f"Cyprus EE 第{i + 1}人：{warn}")

        if CYPRUS_EE_SHEET not in wb.sheetnames:
            continue
        ee = wb[CYPRUS_EE_SHEET]
        row = CYPRUS_EE_DATA_START + i
        if client_code and not _cell_formula_text(ee.cell(row, 2).value):
            ee.cell(row, 2).value = client_code
        if code and not _cell_formula_text(ee.cell(row, 4).value):
            ee.cell(row, 4).value = code
    return warnings


def apply_fx(wb, *, fill_fx: bool = True, fx_row: int | None = None, convert_mapping: dict | None = None) -> tuple[float | None, dict | None]:
    """Cyprus：默认不写汇率（fxPolicy.mode=none）。"""
    if not fill_fx or PN_SHEET not in wb.sheetnames:
        return None, None
    from fx_policy import fx_policy, make_pn_fx_provenance

    mode = str(fx_policy(convert_mapping).get("mode") or "none").strip().lower()
    if mode == "none":
        return None, None
    # 仅当显式改成其它模式时才写
    rates = fetch_usd_rates()
    fx = get_cyprus_pn_fx_rate(rates)
    row = fx_row or _find_pn_row_by_label(wb[PN_SHEET], "FX rate") or 28
    col = 2
    cell = wb[PN_SHEET].cell(row, col)
    if _cell_formula_text(cell.value):
        return fx, None
    cell.value = fx
    return fx, make_pn_fx_provenance(
        PN_SHEET, row, col, convert_mapping, fx, write_source="api", fx_source="api:EUR*0.97"
    )


def convert(
    source_path: Path,
    output_path: Path,
    template_path: Path,
    *,
    pn_meta: PnMeta | dict[str, Any] | None = None,
    employee_directory: list[dict[str, Any]] | None = None,
    registry_dir: Path | None = None,
    convert_mapping: dict[str, Any] | None = None,
    fill_fx: bool = True,
) -> dict[str, Any]:
    global _ACTIVE_MAPPING
    _ACTIVE_MAPPING = resolve_convert_mapping("cyprus_payroll_calc", convert_mapping)
    try:
        source_path = Path(source_path).resolve()
        output_path = Path(output_path).resolve()
        template_path = Path(template_path).resolve()
        if not source_path.is_file():
            raise FileNotFoundError(f"源文件不存在: {source_path}")
        if not template_path.is_file():
            raise FileNotFoundError(f"母版不存在: {template_path}")

        src_wb = load_workbook(source_path, data_only=False)
        try:
            l_name = find_sheet_name(list(src_wb.sheetnames), _active_mapping().get("sourceEmployeeSheet"))
            if not l_name or l_name not in src_wb.sheetnames:
                if CYPRUS_L_SHEET in src_wb.sheetnames:
                    l_name = CYPRUS_L_SHEET
                else:
                    raise ValueError(f"未找到 Cyprus-L，现有: {src_wb.sheetnames}")
            parse_warnings: list[str] = []
            employees = parse_cyprus_l_employees(src_wb[l_name], parse_warnings)
        finally:
            src_wb.close()

        if not employees:
            raise ValueError("Cyprus-L 未解析到员工行")
        sort_employees_by_code(employees, employee_directory)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)
        warnings: list[str] = list(parse_warnings)
        fx = None
        pn_fx_write = None
        applied_pn = None
        try:
            if CYPRUS_L_SHEET not in wb.sheetnames:
                raise ValueError(f"母版缺少 {CYPRUS_L_SHEET}")
            write_cyprus_l(wb[CYPRUS_L_SHEET], employees)
            expand_cyprus_employee_rows(wb, len(employees))
            fixed_value_writes = set_recurring_fees(wb, employees)
            pn_layout = fit_cyprus_pn_employees(wb, len(employees))
            if len(employees) > 1:
                warnings.append(
                    f"Cyprus PN 多人明细行扩行暂定：已扩 Cyprus/Cyprus EE（{len(employees)} 人），请人工核对 PN"
                )
            try:
                fx, pn_fx_write = apply_fx(
                    wb, fill_fx=fill_fx, fx_row=pn_layout.get("fx_row"), convert_mapping=_ACTIVE_MAPPING
                )
            except Exception as exc:
                warnings.append(f"写入 PN 汇率失败: {exc}")

            if pn_meta is not None:
                applied_pn = apply_pn_meta(
                    wb,
                    pn_meta,
                    registry_dir=registry_dir or output_path.parent,
                    reserve_invoice_number=True,
                )
            warnings.extend(
                apply_cyprus_ee_codes(
                    wb,
                    employees,
                    employee_directory=employee_directory,
                    pn_meta=applied_pn or pn_meta,
                )
            )

            apply_luckysheet_compat(wb, pn_sheet=PN_SHEET)
            wb.save(output_path)
        finally:
            wb.close()

        postprocess_converted_xlsx(output_path)
        warnings = merge_warnings(
            warnings,
            sanity_check_convert_result({"employee_count": len(employees), "warnings": warnings}),
        )
        return {
            "ok": True,
            "engine_id": "cyprus_payroll_calc",
            "region": "Cyprus",
            "output": str(output_path),
            "employee_count": len(employees),
            "fx_rate": fx,
            "warnings": warnings,
            "pn_meta": applied_pn.to_dict() if applied_pn else None,
            "fixed_value_writes": fixed_value_writes,
            "pn_fx_write": pn_fx_write,
        }
    finally:
        _ACTIVE_MAPPING = None


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Cyprus-L → Cyprus PN")
    parser.add_argument("source", type=Path)
    parser.add_argument("-o", "--output", type=Path, default=None)
    parser.add_argument("-t", "--template", type=Path, default=None)
    parser.add_argument("--no-fx", action="store_true")
    args = parser.parse_args(argv)
    source = args.source.resolve()
    output = (args.output or source.with_name(f"PN_Cyprus_{source.stem}.xlsx")).resolve()
    template = (args.template or DEFAULT_TEMPLATE).resolve()
    result = convert(source, output, template, fill_fx=not args.no_fx)
    print(result)
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
