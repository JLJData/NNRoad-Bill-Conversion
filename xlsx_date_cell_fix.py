# -*- coding: utf-8 -*-
"""
转换结果 xlsx：仅处理「模板/convert 已标日期格式」的单元格，把文本/带 ' 的值规范为 datetime。
不对 General 格式的数字做日期猜测（避免误伤金额/编号）。
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from xlsx_convert_utils import coerce_datetime_for_excel
from xlsx_keep_images import require_pillow

_DATE_FMT_HINT = re.compile(r"y{1,4}|m/d|d/m|d-mmm|h:mm|年|月", re.I)


def _is_excel_date_format(fmt: str | None) -> bool:
    if not fmt or str(fmt).strip().lower() == "general":
        return False
    head = str(fmt).split(";")[0]
    if re.match(r"^[\s#0,]*\.?0*%?$", head.replace("$", "")) and not _DATE_FMT_HINT.search(head):
        return False
    return bool(_DATE_FMT_HINT.search(head))


def _parse_date_text(text: str) -> datetime | None:
    s = text.strip().lstrip("'").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d %H:%M:%S"):
        try:
            dt = datetime.strptime(s, fmt)
            return datetime(dt.year, dt.month, dt.day)
        except ValueError:
            continue
    return None


def normalize_template_date_cells(xlsx_path: Path | str) -> int:
    path = Path(xlsx_path)
    if not path.is_file():
        return 0
    require_pillow()
    wb = load_workbook(path)
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                fmt = cell.number_format or ""
                if not _is_excel_date_format(fmt):
                    continue
                val = cell.value
                # 公式格只保留公式，勿把缓存/文本误写成 datetime（会变成「公式变数值」）
                if getattr(cell, "data_type", None) == "f":
                    continue
                if isinstance(val, str) and val.startswith("="):
                    continue
                if isinstance(val, datetime):
                    if val.hour or val.minute or val.second:
                        cell.value = datetime(val.year, val.month, val.day)
                        changed += 1
                    continue
                if isinstance(val, date) and not isinstance(val, datetime):
                    cell.value = datetime(val.year, val.month, val.day)
                    changed += 1
                    continue
                if isinstance(val, str):
                    dt = _parse_date_text(val)
                    if dt is not None:
                        cell.value = dt
                        changed += 1
                        continue
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    dt = coerce_datetime_for_excel(val)
                    if dt is not None:
                        cell.value = dt
                        changed += 1
    # From/To 等日期列表头：强制日期格式（含公式列，避免 Lucky/HF 显示成 46235）
    changed += _ensure_date_header_column_formats(wb)
    if changed:
        wb.save(path)
    return changed


def _ensure_date_header_column_formats(wb) -> int:
    """按表头识别 From/To 等日期列，给数据行打上 yyyy/m/d（不改公式本身）。"""
    from xlsx_convert_utils import is_date_column_header, norm

    changed = 0
    date_fmt = "yyyy/m/d"
    for ws in wb.worksheets:
        date_cols: dict[int, int] = {}  # col -> header_row
        max_scan_r = min(ws.max_row or 0, 12)
        max_c = min(ws.max_column or 0, 80)
        for r in range(1, max_scan_r + 1):
            for c in range(1, max_c + 1):
                raw = ws.cell(r, c).value
                if raw is None:
                    continue
                if isinstance(raw, str) and raw.startswith("="):
                    continue
                if is_date_column_header(str(raw)):
                    # 避免把整段「Pay Period」合并区误标；From/To 单列表头才收
                    h = norm(raw).lower()
                    if h in ("from", "to", "period from", "period to") or h.endswith(" date") or "start date" in h or "end date" in h:
                        date_cols[c] = r
        if not date_cols:
            continue
        max_r = ws.max_row or 0
        for c, hdr_r in date_cols.items():
            for r in range(hdr_r + 1, max_r + 1):
                cell = ws.cell(r, c)
                # 跳过空行/合计标签行
                if cell.value is None and not cell.number_format:
                    continue
                if _is_excel_date_format(cell.number_format or ""):
                    continue
                # 有公式或数值/日期才刷格式
                has_f = getattr(cell, "data_type", None) == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                )
                if has_f or isinstance(cell.value, (int, float, datetime, date)) or (
                    isinstance(cell.value, str) and _parse_date_text(cell.value)
                ):
                    if (cell.number_format or "").strip().lower() != date_fmt:
                        cell.number_format = date_fmt
                        changed += 1
    return changed
